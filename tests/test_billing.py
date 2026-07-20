from __future__ import annotations

import asyncio
import json
import tempfile
import unittest
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import create_engine, delete, select
from sqlalchemy.orm import sessionmaker

from rental_manager.database import Base
from rental_manager.main import apartment_month_state, bot_dialog_messages_payload, build_all_debts_breakdown, build_dashboard, create_manual_payment, create_move_out_utility_lines, delete_utility_bill, ensure_utility_advance_drafts_for_bills, expense_period_summary, issue_utility_bill, month_dashboard_summary, move_out, owner_charge_status_label, owner_expected_ip_for_charge, panel_role_for_pin, payment_receipt_document, preview_issue_utility_bill, process_move_out_notifications, provider_reading_statuses_for_month, render_message_text, rent_report, resolve_broadcast_recipients, tenant_payment_history, transfer_lease, update_payment_receipt, utility_bill_for_month
from rental_manager.main import apply_database_import_payload, current_database_snapshot, inspect_database_import_payload, parse_database_import_bytes
from rental_manager.main import api_performance, app_state, record_perf_request
from rental_manager.main import dashboard_income_trend
from rental_manager.models import AiConversation, AiMessage, AppSetting, Apartment, Expense, Lease, MessageLog, Meter, MeterReading, PaymentReceipt, RentalObject, RentCharge, Tenant, UtilityBill, UtilityBillLine, UtilityService, Tariff
from rental_manager.services.billing import (
    IGNORE_LEASE_MARK,
    LEGACY_IMPORT_MARK,
    allocate_odn,
    calculate_tiered_cost,
    calculate_utility_bill,
    effective_due_date,
    full_months_lived,
    generate_rent_charges,
    next_due_after,
    reading_value_at,
    status_for_amount,
    update_rent_charge_status,
)
from rental_manager.services.seed import seed_if_empty
from rental_manager.security.pins import PIN_SETTING_KEYS, hash_pin


class DatabaseTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        db_path = Path(self.tmp.name) / "test.db"
        self.engine = create_engine(f"sqlite:///{db_path.as_posix()}", future=True)
        Base.metadata.create_all(self.engine)
        self.Session = sessionmaker(bind=self.engine, autoflush=False, autocommit=False, future=True)

    def tearDown(self) -> None:
        self.engine.dispose()
        self.tmp.cleanup()

    def seed(self):
        session = self.Session()
        seed_if_empty(session)
        return session


class DummyRequest:
    headers: dict[str, str] = {}
    cookies: dict[str, str] = {}

    def __init__(self, role: str = "owner") -> None:
        self.state = type("State", (), {"panel_role": role})()


class DatabaseImportExportTests(DatabaseTestCase):
    def test_database_snapshot_roundtrip_restores_records(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Дом для экспорта", short_code="ЭКС")
            tenant = Tenant(full_name="Тестовый жилец")
            session.add_all([rental_object, tenant, AppSetting(key="telegram_owner_chat_id", value="123")])
            session.flush()
            apartment = Apartment(object_id=rental_object.id, name="ЭКС1", sort_order=1, odn_share_percent=100, active=True)
            session.add(apartment)
            session.flush()
            lease = Lease(apartment_id=apartment.id, tenant_id=tenant.id, start_date=date(2026, 5, 1), payment_day=1, ip_amount=20000, personal_amount=3000)
            session.add(lease)
            session.commit()

            snapshot = current_database_snapshot(session)
            inspection = inspect_database_import_payload(snapshot)

            session.execute(delete(Lease))
            session.execute(delete(Apartment))
            session.execute(delete(Tenant))
            session.execute(delete(RentalObject))
            session.execute(delete(AppSetting))
            session.commit()

            result = apply_database_import_payload(session, snapshot)
            session.commit()

            restored_lease = session.scalar(select(Lease).where(Lease.id == lease.id))
            restored_setting = session.get(AppSetting, "telegram_owner_chat_id")

        self.assertEqual(inspection["format"], "rental-manager-db-export")
        self.assertGreaterEqual(inspection["total_rows"], 5)
        self.assertEqual(result["counts"]["leases"], 1)
        self.assertIsNotNone(restored_lease)
        self.assertEqual(restored_setting.value, "123")


class AppStatePerformanceTests(DatabaseTestCase):
    def test_app_state_can_return_selected_sections_only(self) -> None:
        with self.seed() as session:
            request = DummyRequest("owner")

            bootstrap_only = app_state(request, sections="bootstrap", session=session)
            self.assertEqual(set(bootstrap_only), {"bootstrap"})
            self.assertIn("dashboard", bootstrap_only["bootstrap"])
            self.assertEqual(bootstrap_only["bootstrap"]["objects"], [])
            self.assertEqual(bootstrap_only["bootstrap"]["leases"], [])
            self.assertEqual(bootstrap_only["bootstrap"]["meters"], [])

            selected = app_state(request, sections="rent_charges,tariffs", session=session)
            self.assertEqual(set(selected), {"rent_charges", "tariffs"})
            self.assertIsInstance(selected["rent_charges"], list)
            self.assertIsInstance(selected["tariffs"], list)

            registry = app_state(request, sections="registry", session=session)
            self.assertEqual(set(registry), {"registry"})
            self.assertGreater(len(registry["registry"]["objects"]), 0)
            self.assertIsInstance(registry["registry"]["leases"], list)
            self.assertGreater(len(registry["registry"]["meters"]), 0)

    def test_performance_snapshot_exposes_request_and_ai_summary(self) -> None:
        record_perf_request(
            method="GET",
            path="/api/app-state",
            status_code=200,
            duration_ms=42,
            user_agent="unit-test",
        )
        with self.seed() as session:
            snapshot = api_performance(session)

        self.assertIn("routes", snapshot)
        self.assertIn("ai_summary", snapshot)
        self.assertTrue(any(item["route"] == "GET /api/app-state" for item in snapshot["routes"]))

    def test_guest_bootstrap_contains_only_safe_executive_metrics(self) -> None:
        with self.seed() as session:
            payload = app_state(DummyRequest("guest"), sections="bootstrap", session=session)["bootstrap"]

        dashboard = payload["dashboard"]
        self.assertIn("month_summary", dashboard)
        self.assertEqual(len(dashboard["income_trend"]), 6)
        self.assertIn("summary_counts", dashboard)
        self.assertNotIn("rent_overdue", dashboard)
        self.assertEqual(payload["objects"], [])

    def test_parse_database_import_bytes_rejects_wrong_payload(self) -> None:
        with self.assertRaises(HTTPException):
            parse_database_import_bytes(b'{"format":"nope","version":1,"tables":{}}')

    def test_panel_pin_hashes_resolve_owner_and_guest_roles(self) -> None:
        with self.Session() as session:
            session.add_all([
                AppSetting(key=PIN_SETTING_KEYS["owner"], value=hash_pin("735194")),
                AppSetting(key=PIN_SETTING_KEYS["guest"], value=hash_pin("846205")),
            ])
            session.commit()
            self.assertEqual(panel_role_for_pin(session, "735194"), "owner")
            self.assertEqual(panel_role_for_pin(session, "846205"), "guest")
            self.assertIsNone(panel_role_for_pin(session, "0000"))


class BroadcastMessageTests(DatabaseTestCase):
    def test_broadcast_recipients_skip_unlinked_and_duplicate_chats(self) -> None:
        with self.seed() as session:
            apartments = session.scalars(select(Apartment).order_by(Apartment.object_id, Apartment.sort_order).limit(3)).all()
            tenants = [Tenant(full_name=f"Жилец {index}") for index in range(1, 4)]
            session.add_all(tenants)
            session.flush()
            leases = [
                Lease(apartment_id=apartments[index].id, tenant_id=tenants[index].id, start_date=date(2026, 5, 1), payment_day=1, ip_amount=10000, personal_amount=0)
                for index in range(3)
            ]
            session.add_all(leases)
            session.flush()
            session.add(
                AppSetting(
                    key="telegram_tenant_links",
                    value=json.dumps({str(tenants[0].id): "100", str(tenants[1].id): "100"}),
                )
            )
            session.flush()

            resolved = resolve_broadcast_recipients(session, {"all": True})

        self.assertEqual([item["lease"].id for item in resolved["recipients"]], [leases[0].id])
        self.assertEqual([item["status"] for item in resolved["skipped"]], ["duplicate_chat", "unlinked"])

    def test_broadcast_recipients_accept_explicit_lease_selection(self) -> None:
        with self.seed() as session:
            apartment = session.scalar(select(Apartment).order_by(Apartment.object_id, Apartment.sort_order).limit(1))
            tenant = Tenant(full_name="Адресный жилец")
            session.add(tenant)
            session.flush()
            lease = Lease(apartment_id=apartment.id, tenant_id=tenant.id, start_date=date(2026, 5, 1), payment_day=1, ip_amount=10000, personal_amount=0)
            session.add(lease)
            session.flush()
            session.add(AppSetting(key="telegram_tenant_links", value=json.dumps({str(tenant.id): "200"})))
            session.flush()

            resolved = resolve_broadcast_recipients(session, {"lease_ids": [lease.id]})

        self.assertEqual(len(resolved["recipients"]), 1)
        self.assertEqual(resolved["recipients"][0]["chat_id"], "200")


async def read_streaming_response_bytes(response) -> bytes:
    payload = b""
    async for chunk in response.body_iterator:
        payload += chunk
    return payload


class RentScheduleTests(DatabaseTestCase):
    def test_missing_payment_day_moves_to_first_of_next_month(self) -> None:
        self.assertEqual(effective_due_date(2026, 2, 29), date(2026, 3, 1))
        self.assertEqual(effective_due_date(2026, 2, 30), date(2026, 3, 1))
        self.assertEqual(effective_due_date(2026, 2, 31), date(2026, 3, 1))

    def test_next_due_returns_same_month_real_day_after_shifted_date(self) -> None:
        self.assertEqual(next_due_after(date(2026, 1, 31), 31), date(2026, 3, 1))
        self.assertEqual(next_due_after(date(2026, 3, 1), 31), date(2026, 3, 31))
        self.assertEqual(next_due_after(date(2026, 4, 30), 31), date(2026, 5, 1))
        self.assertEqual(next_due_after(date(2026, 5, 1), 31), date(2026, 5, 31))

    def test_generated_rent_charge_periods_for_31st_payment_day(self) -> None:
        with self.seed() as session:
            apartment = session.get(Apartment, 1)
            tenant = Tenant(full_name="Tenant")
            session.add(tenant)
            session.flush()
            lease = Lease(
                apartment_id=apartment.id,
                tenant_id=tenant.id,
                start_date=date(2026, 1, 31),
                payment_day=31,
                ip_amount=20000,
                personal_amount=5000,
            )
            session.add(lease)
            session.flush()

            created = generate_rent_charges(session, until=date(2026, 4, 5))
            charges = lease.rent_charges

        self.assertEqual(created, 3)
        self.assertEqual([charge.due_date for charge in charges], [date(2026, 1, 31), date(2026, 3, 1), date(2026, 3, 31)])
        self.assertEqual(charges[0].period_end, date(2026, 2, 28))
        self.assertEqual(charges[1].period_end, date(2026, 3, 30))

    def test_inactive_apartment_does_not_generate_new_rent_charges(self) -> None:
        with self.seed() as session:
            apartment = session.get(Apartment, 1)
            apartment.active = False
            tenant = Tenant(full_name="Paused tenant")
            session.add(tenant)
            session.flush()
            lease = Lease(
                apartment_id=apartment.id,
                tenant_id=tenant.id,
                start_date=date(2026, 4, 14),
                payment_day=14,
                ip_amount=20000,
                personal_amount=5000,
            )
            session.add(lease)
            session.flush()

            created = generate_rent_charges(session, until=date(2026, 6, 30))

        self.assertEqual(created, 0)

    def test_generate_rent_charges_does_not_scan_every_historic_charge(self) -> None:
        with self.seed() as session:
            inactive_apartment = Apartment(object_id=1, name="Архив", active=False)
            tenant = Tenant(full_name="Архивный жилец")
            session.add_all([inactive_apartment, tenant])
            session.flush()
            inactive_lease = Lease(
                apartment_id=inactive_apartment.id,
                tenant_id=tenant.id,
                start_date=date(2025, 1, 1),
                payment_day=1,
                active=False,
                ip_amount=10000,
                personal_amount=0,
            )
            session.add(inactive_lease)
            session.flush()
            historic_charge = RentCharge(
                lease_id=inactive_lease.id,
                period_start=date(2025, 1, 1),
                period_end=date(2025, 1, 31),
                due_date=date(2025, 1, 1),
                ip_due=10000,
                personal_due=0,
                status="pending",
            )
            session.add(historic_charge)
            session.flush()
            historic_charge_id = historic_charge.id

            with patch("rental_manager.services.billing.update_rent_charge_status", wraps=update_rent_charge_status) as status_mock:
                generate_rent_charges(session, until=date(2026, 9, 30))

            updated_ids = {
                call.args[0].id
                for call in status_mock.call_args_list
                if call.args and getattr(call.args[0], "id", None)
            }

        self.assertNotIn(historic_charge_id, updated_ids)

    def test_statuses_for_two_part_rent_payment(self) -> None:
        charge = self._charge(due_date=date(2026, 4, 14), ip_due=20000, personal_due=5000)
        self.assertEqual(update_rent_charge_status(charge, today=date(2026, 4, 14)), "pending")
        self.assertEqual(update_rent_charge_status(charge, today=date(2026, 4, 15)), "overdue")

        charge.ip_paid = 20000
        self.assertEqual(update_rent_charge_status(charge, today=date(2026, 4, 15)), "partial")

        charge.personal_paid = 5000
        self.assertEqual(update_rent_charge_status(charge, today=date(2026, 4, 15)), "paid")

        charge.personal_paid = 6000
        self.assertEqual(update_rent_charge_status(charge, today=date(2026, 4, 15)), "paid_ahead")

    def test_total_rent_status_does_not_become_paid_ahead_without_real_overpayment(self) -> None:
        charge = self._charge(due_date=date(2026, 3, 2), period_end=date(2026, 4, 1), ip_due=20000, personal_due=0)
        charge.ip_paid = 20000

        self.assertEqual(update_rent_charge_status(charge, today=date(2026, 3, 16)), "paid")

    def test_active_deferral_overrides_partial_debt(self) -> None:
        charge = self._charge(due_date=date(2026, 4, 14), ip_due=20000, personal_due=5000)
        charge.ip_paid = 20000
        charge.deferral_until = date(2026, 4, 25)

        self.assertEqual(update_rent_charge_status(charge, today=date(2026, 4, 23)), "deferred")

    def test_legacy_imported_rent_before_cutoff_does_not_require_phone_transfer(self) -> None:
        tenant = Tenant(full_name="Legacy Tenant", notes=LEGACY_IMPORT_MARK)
        lease = Lease(tenant=tenant, notes=LEGACY_IMPORT_MARK)
        charge = self._charge(due_date=date(2026, 3, 14), ip_due=20000, personal_due=5000)
        charge.lease = lease
        charge.ip_paid = 20000

        self.assertEqual(update_rent_charge_status(charge, today=date(2026, 4, 1)), "paid")
        self.assertEqual(charge.personal_paid, 5000)

    def test_current_rent_after_cutoff_still_requires_phone_transfer(self) -> None:
        tenant = Tenant(full_name="Fresh Tenant", notes=LEGACY_IMPORT_MARK)
        lease = Lease(tenant=tenant, notes=LEGACY_IMPORT_MARK)
        charge = self._charge(due_date=date(2026, 4, 14), ip_due=20000, personal_due=5000)
        charge.lease = lease
        charge.ip_paid = 20000

        self.assertEqual(update_rent_charge_status(charge, today=date(2026, 4, 15)), "partial")
        self.assertEqual(float(charge.personal_paid or 0), 0.0)

    def test_owner_report_treats_expense_fund_payment_as_reduced_ip_expectation(self) -> None:
        from rental_manager.models import PaymentReceipt

        charge = self._charge(due_date=date(2026, 4, 14), ip_due=20000, personal_due=0)
        charge.receipts = [
            PaymentReceipt(amount=20000, channel="expense_fund", status="accepted", paid_at=date(2026, 4, 16)),
        ]
        charge.ip_paid = 20000

        self.assertEqual(owner_expected_ip_for_charge(charge), 0.0)
        self.assertEqual(owner_charge_status_label(charge), "оплачено, ушло на расходы")

    def test_owner_report_status_does_not_show_paid_ahead_without_overpayment(self) -> None:
        charge = self._charge(due_date=date(2026, 3, 2), period_end=date(2026, 4, 1), ip_due=20000, personal_due=0)
        charge.ip_paid = 20000

        self.assertEqual(owner_charge_status_label(charge), "оплачено")

    def test_full_months_lived_uses_payment_periods(self) -> None:
        lease = Lease(start_date=date(2026, 4, 14), payment_day=14)
        self.assertEqual(full_months_lived(lease, date(2026, 5, 13)), 1)
        self.assertEqual(full_months_lived(lease, date(2026, 5, 14)), 1)
        self.assertEqual(full_months_lived(lease, date(2026, 6, 13)), 2)

    @staticmethod
    def _charge(**kwargs):
        from rental_manager.models import RentCharge

        defaults = {
            "lease_id": 1,
            "period_start": date(2026, 4, 14),
            "period_end": date(2026, 5, 13),
            "due_date": date(2026, 4, 14),
            "ip_due": 0,
            "personal_due": 0,
        }
        defaults.update(kwargs)
        return RentCharge(**defaults)


class UtilityBillingTests(DatabaseTestCase):
    def test_expense_period_summary_keeps_opening_balance_and_sign(self) -> None:
        with self.seed() as session:
            session.add_all(
                [
                    Expense(expense_date=date(2025, 12, 20), amount=20000, source_funds="expense_fund_income", category="Пополнение"),
                    Expense(expense_date=date(2025, 12, 25), amount=11000, source_funds="personal", category="Ремонт"),
                    Expense(expense_date=date(2026, 1, 5), amount=3000, source_funds="personal", category="Кран"),
                    Expense(expense_date=date(2026, 1, 10), amount=2000, source_funds="expense_fund_income", category="Пополнение"),
                ]
            )
            session.commit()

            summary = expense_period_summary(session, date(2026, 1, 1), date(2026, 1, 31))

        self.assertEqual(summary["opening_balance"], -9000)
        self.assertEqual(summary["period_spent"], 3000)
        self.assertEqual(summary["period_income"], 2000)
        self.assertEqual(summary["closing_balance"], -8000)

    def test_tiered_cost_is_applied_to_object_consumption(self) -> None:
        tiers = json.dumps(
            [
                {"limit": 3900, "price": 4.18},
                {"limit": 6000, "price": 6.01},
                {"limit": None, "price": 7.48},
            ]
        )
        self.assertEqual(calculate_tiered_cost(7000, tiers), 36403.0)

    def test_odn_is_reallocated_between_occupied_apartments_by_day(self) -> None:
        a1 = Apartment(id=1, odn_share_percent=20)
        a2 = Apartment(id=2, odn_share_percent=30)
        a3 = Apartment(id=3, odn_share_percent=50)
        leases = {
            1: Lease(start_date=date(2026, 4, 1), end_date=date(2026, 4, 1)),
            2: Lease(start_date=date(2026, 4, 1), end_date=None),
            3: None,
        }

        allocated = allocate_odn([a1, a2, a3], leases, date(2026, 4, 1), date(2026, 4, 3), 100)

        self.assertAlmostEqual(allocated[1], 20.0)
        self.assertAlmostEqual(allocated[2], 80.0)
        self.assertAlmostEqual(allocated[3], 0.0)

    def test_exact_utility_bill_uses_average_object_price(self) -> None:
        with self.seed() as session:
            obj = session.get(RentalObject, 1)
            apartment = session.get(Apartment, 1)
            service = session.scalar(select(UtilityService).where(UtilityService.object_id == obj.id, UtilityService.kind == "electricity"))
            self._lease(session, apartment)
            self._read_all_meters(session, service, object_start=1000, object_end=2200, apartment_end_values={apartment.id: 110})

            bill, warnings = calculate_utility_bill(session, service.id, date(2026, 4, 1), date(2026, 5, 1), allow_estimate=False)

        self.assertEqual(warnings, [])
        self.assertEqual(bill.total_consumption, 1200.0)
        self.assertEqual(bill.total_cost, 5016.0)
        self.assertEqual(bill.average_unit_price, 4.18)
        self.assertEqual(len(bill.lines), 1)
        self.assertEqual(bill.lines[0].personal_consumption, 100.0)
        self.assertEqual(bill.lines[0].odn_consumption, 1100.0)
        self.assertEqual(bill.lines[0].total_amount, 5016.0)

    def test_utility_bill_fails_when_apartment_consumption_exceeds_object(self) -> None:
        with self.seed() as session:
            obj = session.get(RentalObject, 1)
            apartment = session.get(Apartment, 1)
            service = session.scalar(select(UtilityService).where(UtilityService.object_id == obj.id, UtilityService.kind == "electricity"))
            self._lease(session, apartment)
            self._read_all_meters(session, service, object_start=1000, object_end=1050, apartment_end_values={apartment.id: 200})

            with self.assertRaisesRegex(ValueError, "квартирных"):
                calculate_utility_bill(session, service.id, date(2026, 4, 1), date(2026, 5, 1), allow_estimate=False)

    def test_reading_interpolation_and_forecast_are_marked_estimated(self) -> None:
        with self.seed() as session:
            meter = session.get(Meter, 1)
            session.add(MeterReading(meter_id=meter.id, reading_date=date(2026, 4, 1), value=1000))
            session.add(MeterReading(meter_id=meter.id, reading_date=date(2026, 4, 11), value=1100))
            session.add(MeterReading(meter_id=meter.id, reading_date=date(2026, 4, 21), value=1220))
            session.flush()

            interpolated = reading_value_at(session, meter.id, date(2026, 4, 16), allow_estimate=True)
            forecast = reading_value_at(session, meter.id, date(2026, 4, 25), allow_estimate=True)

        self.assertTrue(interpolated.estimated)
        self.assertEqual(interpolated.note, "interpolated")
        self.assertAlmostEqual(interpolated.value, 1160.0)
        self.assertTrue(forecast.estimated)
        self.assertEqual(forecast.note, "forecast")
        self.assertAlmostEqual(forecast.value, 1268.0)

    def test_utility_bill_is_split_when_tenant_changes_inside_period(self) -> None:
        with self.seed() as session:
            obj = session.get(RentalObject, 1)
            apartment = session.get(Apartment, 1)
            service = session.scalar(select(UtilityService).where(UtilityService.object_id == obj.id, UtilityService.kind == "electricity"))
            self._lease(session, apartment, end_date=date(2026, 4, 3), full_name="First tenant")
            self._lease(session, apartment, start_date=date(2026, 4, 4), full_name="Second tenant")
            self._read_all_meters(session, service, object_start=1000, object_end=1310, apartment_end_values={apartment.id: 40})

            bill, warnings = calculate_utility_bill(session, service.id, date(2026, 4, 1), date(2026, 5, 1), allow_estimate=False)

        self.assertTrue(any("04.04.2026" in warning for warning in warnings))
        self.assertEqual(len(bill.lines), 2)
        self.assertEqual([line.note for line in bill.lines], ["01.04.2026 -> 04.04.2026 (3 дн.)", "04.04.2026 -> 01.05.2026 (27 дн.)"])
        self.assertAlmostEqual(sum(line.personal_consumption for line in bill.lines), 30.0)
        self.assertAlmostEqual(sum(line.odn_consumption for line in bill.lines), 280.0)

    def test_utility_bill_starts_tenant_segment_from_move_in_date(self) -> None:
        with self.seed() as session:
            obj = session.get(RentalObject, 1)
            apartment = session.get(Apartment, 1)
            service = session.scalar(select(UtilityService).where(UtilityService.object_id == obj.id, UtilityService.kind == "electricity"))
            self._lease(session, apartment, start_date=date(2026, 4, 10), full_name="Late move-in")
            self._read_all_meters(session, service, object_start=1000, object_end=1300, apartment_end_values={apartment.id: 40})

            bill, warnings = calculate_utility_bill(session, service.id, date(2026, 4, 1), date(2026, 5, 1), allow_estimate=True)

        self.assertTrue(any("10.04.2026" in warning for warning in warnings))
        self.assertEqual(len(bill.lines), 1)
        self.assertEqual(bill.lines[0].note, "10.04.2026 -> 01.05.2026 (21 дн.)")
        self.assertAlmostEqual(bill.lines[0].personal_consumption, 21.0)
        self.assertAlmostEqual(bill.lines[0].odn_consumption, 189.0)

    def test_utility_bill_skips_interval_that_was_already_issued(self) -> None:
        with self.seed() as session:
            obj = session.get(RentalObject, 1)
            apartment = session.get(Apartment, 1)
            service = session.scalar(select(UtilityService).where(UtilityService.object_id == obj.id, UtilityService.kind == "electricity"))
            lease = self._lease(session, apartment)
            self._read_all_meters(session, service, object_start=1000, object_end=1310, apartment_end_values={apartment.id: 40})

            issued_bill = UtilityBill(
                service_id=service.id,
                period_start=date(2026, 4, 1),
                period_end=date(2026, 4, 10),
                status="issued",
                total_consumption=93,
                apartment_consumption=9,
                odn_consumption=84,
                total_cost=388.74,
                average_unit_price=4.18,
            )
            issued_bill.lines.append(
                UtilityBillLine(
                    apartment_id=apartment.id,
                    lease_id=lease.id,
                    personal_consumption=9,
                    odn_consumption=84,
                    total_amount=388.74,
                    paid_amount=0,
                    status="issued",
                    note="01.04.2026 -> 10.04.2026 (9 дн.)",
                )
            )
            session.add(issued_bill)
            session.flush()

            bill, warnings = calculate_utility_bill(session, service.id, date(2026, 4, 1), date(2026, 5, 1), allow_estimate=False)

        self.assertEqual(warnings[-1], "Пропущено уже выставленных сегментов: 1.")
        self.assertEqual(len(bill.lines), 1)
        self.assertEqual(bill.lines[0].note, "10.04.2026 -> 01.05.2026 (21 дн.)")

    def test_delete_issued_utility_bill_without_payments_detaches_logs(self) -> None:
        with self.seed() as session:
            obj = session.get(RentalObject, 1)
            apartment = session.get(Apartment, 1)
            service = session.scalar(select(UtilityService).where(UtilityService.object_id == obj.id, UtilityService.kind == "electricity"))
            lease = self._lease(session, apartment)
            issued_bill = UtilityBill(
                service_id=service.id,
                period_start=date(2026, 4, 1),
                period_end=date(2026, 5, 1),
                status="issued",
                total_consumption=100,
                apartment_consumption=20,
                odn_consumption=80,
                total_cost=418,
                average_unit_price=4.18,
            )
            issued_bill.lines.append(
                UtilityBillLine(
                    apartment_id=apartment.id,
                    lease_id=lease.id,
                    personal_consumption=20,
                    odn_consumption=80,
                    total_amount=418,
                    paid_amount=0,
                    status="issued",
                    note="01.04.2026 -> 01.05.2026 (30 дн.)",
                )
            )
            session.add(issued_bill)
            session.flush()
            log = MessageLog(lease_id=lease.id, utility_line_id=issued_bill.lines[0].id, template_key="message_utility_bill", text="sent")
            session.add(log)
            session.commit()

            result = delete_utility_bill(issued_bill.id, session)
            preserved_log = session.get(MessageLog, log.id)
            deleted_bill = session.get(UtilityBill, issued_bill.id)

        self.assertTrue(result["ok"])
        self.assertEqual(result["deleted_status"], "issued")
        self.assertIsNone(deleted_bill)
        self.assertIsNotNone(preserved_log)
        self.assertIsNone(preserved_log.utility_line_id)

    def test_delete_utility_bill_blocks_when_payment_is_linked(self) -> None:
        with self.seed() as session:
            obj = session.get(RentalObject, 1)
            apartment = session.get(Apartment, 1)
            service = session.scalar(select(UtilityService).where(UtilityService.object_id == obj.id, UtilityService.kind == "electricity"))
            lease = self._lease(session, apartment)
            issued_bill = UtilityBill(
                service_id=service.id,
                period_start=date(2026, 4, 1),
                period_end=date(2026, 5, 1),
                status="issued",
                total_consumption=100,
                apartment_consumption=20,
                odn_consumption=80,
                total_cost=418,
                average_unit_price=4.18,
            )
            issued_line = UtilityBillLine(
                apartment_id=apartment.id,
                lease_id=lease.id,
                personal_consumption=20,
                odn_consumption=80,
                total_amount=418,
                paid_amount=100,
                status="partial",
                note="01.04.2026 -> 01.05.2026 (30 дн.)",
            )
            issued_bill.lines.append(issued_line)
            session.add(issued_bill)
            session.flush()
            session.add(
                PaymentReceipt(
                    lease_id=lease.id,
                    utility_line_id=issued_line.id,
                    apartment_id=apartment.id,
                    amount=100,
                    channel="utilities",
                    paid_at=datetime(2026, 4, 20, 12, 0),
                    source="manual",
                    status="accepted",
                )
            )
            session.commit()

            with self.assertRaises(HTTPException) as ctx:
                delete_utility_bill(issued_bill.id, session)

        self.assertEqual(ctx.exception.status_code, 400)

    def test_move_out_utility_lines_are_created_as_forecasted_issued_bill(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Тестовый дом", short_code="ТД")
            apartment = Apartment(name="ТД1", sort_order=1, odn_share_percent=100, active=True, object=rental_object)
            tenant = Tenant(full_name="Съезжающий жилец", active=True)
            lease = Lease(
                apartment=apartment,
                tenant=tenant,
                start_date=date(2026, 4, 1),
                end_date=date(2026, 4, 10),
                payment_day=1,
                ip_amount=10000,
                personal_amount=2000,
                active=False,
            )
            service = UtilityService(
                object=rental_object,
                kind="electricity",
                name="Электричество",
                provider_due_day=24,
                resident_due_days=7,
                active=True,
            )
            session.add_all([rental_object, apartment, tenant, lease, service])
            session.flush()
            session.add(Tariff(service_id=service.id, starts_on=date(2026, 1, 1), tiers_json=json.dumps([{"limit": None, "price": 4.18}])))
            object_meter = Meter(service_id=service.id, object_id=rental_object.id, scope="object", name="Общий", active=True)
            apartment_meter = Meter(service_id=service.id, object_id=rental_object.id, apartment_id=apartment.id, scope="apartment", name="ТД1", active=True)
            session.add_all([object_meter, apartment_meter])
            session.flush()
            session.add_all(
                [
                    MeterReading(meter_id=object_meter.id, reading_date=date(2026, 4, 1), value=1000),
                    MeterReading(meter_id=object_meter.id, reading_date=date(2026, 4, 20), value=1200),
                    MeterReading(meter_id=apartment_meter.id, reading_date=date(2026, 4, 1), value=10),
                    MeterReading(meter_id=apartment_meter.id, reading_date=date(2026, 4, 20), value=30),
                ]
            )
            session.commit()

            lines, warnings = create_move_out_utility_lines(session, lease)

        self.assertEqual(len(lines), 1)
        self.assertTrue(warnings)
        line = lines[0]
        self.assertEqual(line.status, "issued")
        self.assertEqual(line.bill.status, "issued")
        self.assertTrue(line.bill.is_forecast)
        self.assertTrue(line.bill.provider_paid)
        self.assertEqual(line.bill.period_start, date(2026, 4, 1))
        self.assertEqual(line.bill.period_end, date(2026, 4, 10))
        self.assertEqual(line.note, "01.04.2026 -> 10.04.2026 (9 дн.)")
        self.assertGreater(line.total_amount, 0)

    def test_move_out_utility_uses_last_issued_period_not_latest_draft(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Draft test object", short_code="DTO")
            apartment = Apartment(name="DTO1", sort_order=1, odn_share_percent=100, active=True, object=rental_object)
            tenant = Tenant(full_name="Draft tenant", active=True)
            lease = Lease(
                apartment=apartment,
                tenant=tenant,
                start_date=date(2026, 4, 1),
                end_date=date(2026, 4, 10),
                payment_day=1,
                ip_amount=10000,
                personal_amount=2000,
                active=False,
            )
            service = UtilityService(object=rental_object, kind="electricity", name="Draft electricity", provider_due_day=24, resident_due_days=7, active=True)
            session.add_all([rental_object, apartment, tenant, lease, service])
            session.flush()
            session.add(Tariff(service_id=service.id, starts_on=date(2026, 1, 1), tiers_json=json.dumps([{"limit": None, "price": 4.18}])))
            object_meter = Meter(service_id=service.id, object_id=rental_object.id, scope="object", name="Object", active=True)
            apartment_meter = Meter(service_id=service.id, object_id=rental_object.id, apartment_id=apartment.id, scope="apartment", name="Apartment", active=True)
            session.add_all([object_meter, apartment_meter])
            session.flush()
            session.add_all(
                [
                    MeterReading(meter_id=object_meter.id, reading_date=date(2026, 4, 1), value=1000),
                    MeterReading(meter_id=object_meter.id, reading_date=date(2026, 4, 20), value=1200),
                    MeterReading(meter_id=apartment_meter.id, reading_date=date(2026, 4, 1), value=10),
                    MeterReading(meter_id=apartment_meter.id, reading_date=date(2026, 4, 20), value=30),
                ]
            )
            session.flush()

            issued_bill = UtilityBill(
                service_id=service.id,
                period_start=date(2026, 4, 1),
                period_end=date(2026, 4, 5),
                status="issued",
                due_date=date(2026, 4, 12),
                total_consumption=5,
                apartment_consumption=5,
                odn_consumption=0,
                total_cost=100,
                average_unit_price=4.18,
            )
            issued_bill.lines.append(
                UtilityBillLine(
                    apartment_id=apartment.id,
                    lease_id=lease.id,
                    personal_consumption=5,
                    odn_consumption=0,
                    total_amount=100,
                    paid_amount=0,
                    status="issued",
                    note="01.04.2026 -> 05.04.2026 (4 d.)",
                )
            )
            draft_bill = UtilityBill(
                service_id=service.id,
                period_start=date(2026, 4, 5),
                period_end=date(2026, 4, 8),
                status="draft",
                total_consumption=3,
                apartment_consumption=3,
                odn_consumption=0,
                total_cost=50,
                average_unit_price=4.18,
            )
            draft_bill.lines.append(
                UtilityBillLine(
                    apartment_id=apartment.id,
                    lease_id=lease.id,
                    personal_consumption=3,
                    odn_consumption=0,
                    total_amount=50,
                    paid_amount=0,
                    status="draft",
                    note="05.04.2026 -> 08.04.2026 (3 d.)",
                )
            )
            session.add_all([issued_bill, draft_bill])
            session.commit()

            lines, warnings = create_move_out_utility_lines(session, lease)

        self.assertEqual(len(lines), 1)
        self.assertTrue(warnings)
        self.assertEqual(lines[0].bill.period_start, date(2026, 4, 5))
        self.assertEqual(lines[0].bill.period_end, date(2026, 4, 10))

    def test_move_out_notifications_send_once_and_create_logs(self) -> None:
        with self.Session() as session:
            session.add_all(
                [
                    AppSetting(key="telegram_bot_token", value="token"),
                    AppSetting(key="telegram_owner_chat_id", value="999"),
                ]
            )
            rental_object = RentalObject(name="Тестовый дом 2", short_code="Т2")
            apartment = Apartment(name="Т21", sort_order=1, odn_share_percent=100, active=True, object=rental_object)
            tenant = Tenant(full_name="Жилец на выезд", active=True)
            lease = Lease(
                apartment=apartment,
                tenant=tenant,
                start_date=date(2026, 4, 1),
                end_date=date(2026, 4, 10),
                payment_day=1,
                ip_amount=10000,
                personal_amount=2000,
                active=False,
            )
            service = UtilityService(object=rental_object, kind="electricity", name="Электричество", provider_due_day=24, resident_due_days=7, active=True)
            session.add_all([rental_object, apartment, tenant, lease, service])
            session.flush()
            session.add(AppSetting(key="telegram_tenant_links", value=json.dumps({str(tenant.id): "123"})))
            session.add(Tariff(service_id=service.id, starts_on=date(2026, 1, 1), tiers_json=json.dumps([{"limit": None, "price": 4.18}])))
            object_meter = Meter(service_id=service.id, object_id=rental_object.id, scope="object", name="Общий", active=True)
            apartment_meter = Meter(service_id=service.id, object_id=rental_object.id, apartment_id=apartment.id, scope="apartment", name="Т21", active=True)
            session.add_all([object_meter, apartment_meter])
            session.flush()
            session.add_all(
                [
                    MeterReading(meter_id=object_meter.id, reading_date=date(2026, 4, 1), value=1000),
                    MeterReading(meter_id=object_meter.id, reading_date=date(2026, 4, 20), value=1200),
                    MeterReading(meter_id=apartment_meter.id, reading_date=date(2026, 4, 1), value=10),
                    MeterReading(meter_id=apartment_meter.id, reading_date=date(2026, 4, 20), value=30),
                ]
            )
            session.commit()

            with patch("rental_manager.main.send_telegram_text", return_value={"ok": True}) as mocked_send:
                summary = process_move_out_notifications(session, date(2026, 4, 10))
                session.commit()
                repeat = process_move_out_notifications(session, date(2026, 4, 10))

            self.assertEqual(summary["processed"], 1)
            self.assertEqual(summary["tenant_sent"], 1)
            self.assertEqual(summary["owner_sent"], 1)
            self.assertEqual(repeat["processed"], 0)
            self.assertEqual(mocked_send.call_count, 2)
            logs = session.scalars(select(MessageLog).where(MessageLog.lease_id == lease.id)).all()

        self.assertTrue(any(log.template_key == "message_utility_bill" for log in logs))

    def test_move_out_summary_ignores_rent_debts_before_cutoff(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Дом выезда", short_code="ДВ")
            apartment = Apartment(name="ДВ1", sort_order=1, odn_share_percent=100, active=True, object=rental_object)
            tenant = Tenant(full_name="Старый жилец", active=True)
            lease = Lease(
                apartment=apartment,
                tenant=tenant,
                start_date=date(2026, 1, 1),
                payment_day=1,
                ip_amount=10000,
                personal_amount=2000,
                active=True,
            )
            session.add_all([rental_object, apartment, tenant, lease, AppSetting(key="notification_cutoff_date", value="2026-03-01")])
            session.flush()
            generate_rent_charges(session, until=date(2026, 4, 10))

            result = move_out(lease.id, {"end_date": "2026-04-10"}, session)

        self.assertEqual(result["summary"]["rent_debt"], 24000.0)

    def test_transfer_lease_moves_tenant_to_new_apartment_and_updates_rent(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Дом переезда", short_code="ДП")
            old_apartment = Apartment(name="ДП1", sort_order=1, odn_share_percent=50, active=True, object=rental_object)
            new_apartment = Apartment(name="ДП2", sort_order=2, odn_share_percent=50, active=True, object=rental_object)
            tenant = Tenant(full_name="Переезжающий жилец", active=True)
            lease = Lease(
                apartment=old_apartment,
                tenant=tenant,
                start_date=date(2026, 4, 1),
                payment_day=1,
                ip_amount=10000,
                personal_amount=2000,
                active=True,
            )
            session.add_all([rental_object, old_apartment, new_apartment, tenant, lease])
            session.flush()
            generate_rent_charges(session, until=date(2026, 5, 31))
            future_charge = session.scalar(
                select(RentCharge).where(RentCharge.lease_id == lease.id, RentCharge.due_date == date(2026, 5, 1))
            )
            self.assertIsNotNone(future_charge)
            assert future_charge is not None
            session.add(
                MessageLog(
                    lease_id=lease.id,
                    rent_charge_id=future_charge.id,
                    template_key="message_rent_due",
                    text="already sent",
                )
            )
            session.flush()

            result = transfer_lease(
                lease.id,
                {
                    "apartment_id": new_apartment.id,
                    "transfer_date": "2026-04-10",
                    "ip_amount": "15 000",
                    "personal_amount": "2 500,50",
                },
                session,
            )

            old_lease = session.get(Lease, result["old_lease"]["id"])
            new_lease = session.get(Lease, result["new_lease"]["id"])
            summary = {
                "old_end_date": old_lease.end_date,
                "old_active": old_lease.active,
                "new_tenant_id": new_lease.tenant_id,
                "old_tenant_id": old_lease.tenant_id,
                "new_apartment_id": new_lease.apartment_id,
                "target_apartment_id": new_apartment.id,
                "new_start_date": new_lease.start_date,
                "new_ip_amount": new_lease.ip_amount,
                "new_personal_amount": new_lease.personal_amount,
                "tenant_active": new_lease.tenant.active,
                "old_has_future_charge": any(charge.due_date > old_lease.end_date for charge in old_lease.rent_charges),
                "old_future_charges_deleted": session.scalar(
                    select(RentCharge.id).where(
                        RentCharge.lease_id == old_lease.id,
                        RentCharge.due_date > old_lease.end_date,
                    )
                )
                is None,
                "message_log_detached": session.scalar(select(MessageLog).where(MessageLog.text == "already sent")).rent_charge_id is None,
            }

        self.assertEqual(summary["old_end_date"], date(2026, 4, 9))
        self.assertFalse(summary["old_active"])
        self.assertEqual(summary["new_tenant_id"], summary["old_tenant_id"])
        self.assertEqual(summary["new_apartment_id"], summary["target_apartment_id"])
        self.assertEqual(summary["new_start_date"], date(2026, 4, 10))
        self.assertEqual(summary["new_ip_amount"], 15000.0)
        self.assertEqual(summary["new_personal_amount"], 2500.5)
        self.assertTrue(summary["tenant_active"])
        self.assertFalse(summary["old_has_future_charge"])
        self.assertTrue(summary["old_future_charges_deleted"])
        self.assertTrue(summary["message_log_detached"])

    def test_transfer_lease_splits_utility_bill_between_old_and_new_apartment(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Дом коммунального переезда", short_code="ДКП")
            old_apartment = Apartment(name="ДКП1", sort_order=1, odn_share_percent=50, active=True, object=rental_object)
            new_apartment = Apartment(name="ДКП2", sort_order=2, odn_share_percent=50, active=True, object=rental_object)
            tenant = Tenant(full_name="Жилец с коммунальным переездом", active=True)
            lease = Lease(
                apartment=old_apartment,
                tenant=tenant,
                start_date=date(2026, 4, 1),
                payment_day=1,
                ip_amount=10000,
                personal_amount=2000,
                active=True,
            )
            service = UtilityService(
                object=rental_object,
                kind="electricity",
                name="Электричество",
                provider_due_day=24,
                resident_due_days=7,
                active=True,
            )
            session.add_all([rental_object, old_apartment, new_apartment, tenant, lease, service])
            session.flush()
            session.add(Tariff(service_id=service.id, starts_on=date(2026, 1, 1), tiers_json=json.dumps([{"limit": None, "price": 1.0}])))
            object_meter = Meter(service_id=service.id, object_id=rental_object.id, scope="object", name="Общий", active=True)
            old_meter = Meter(service_id=service.id, object_id=rental_object.id, apartment_id=old_apartment.id, scope="apartment", name="ДКП1", active=True)
            new_meter = Meter(service_id=service.id, object_id=rental_object.id, apartment_id=new_apartment.id, scope="apartment", name="ДКП2", active=True)
            session.add_all([object_meter, old_meter, new_meter])
            session.flush()
            session.add_all(
                [
                    MeterReading(meter_id=object_meter.id, reading_date=date(2026, 4, 1), value=0),
                    MeterReading(meter_id=object_meter.id, reading_date=date(2026, 4, 10), value=90),
                    MeterReading(meter_id=object_meter.id, reading_date=date(2026, 5, 1), value=300),
                    MeterReading(meter_id=old_meter.id, reading_date=date(2026, 4, 1), value=0),
                    MeterReading(meter_id=old_meter.id, reading_date=date(2026, 4, 10), value=90),
                    MeterReading(meter_id=old_meter.id, reading_date=date(2026, 5, 1), value=90),
                    MeterReading(meter_id=new_meter.id, reading_date=date(2026, 4, 1), value=0),
                    MeterReading(meter_id=new_meter.id, reading_date=date(2026, 4, 10), value=0),
                    MeterReading(meter_id=new_meter.id, reading_date=date(2026, 5, 1), value=210),
                ]
            )
            session.flush()

            result = transfer_lease(lease.id, {"apartment_id": new_apartment.id, "transfer_date": "2026-04-10"}, session)
            new_lease_id = result["new_lease"]["id"]
            old_lease_id = result["old_lease"]["id"]
            old_apartment_id = old_apartment.id
            new_apartment_id = new_apartment.id
            bill, warnings = calculate_utility_bill(session, service.id, date(2026, 4, 1), date(2026, 5, 1), allow_estimate=True)

        self.assertEqual(warnings, [])
        self.assertEqual(len(bill.lines), 2)
        by_lease = {line.lease_id: line for line in bill.lines}
        self.assertEqual(by_lease[old_lease_id].apartment_id, old_apartment_id)
        self.assertEqual(by_lease[old_lease_id].note, "01.04.2026 -> 10.04.2026 (9 дн.)")
        self.assertEqual(by_lease[old_lease_id].total_amount, 90.0)
        self.assertEqual(by_lease[new_lease_id].apartment_id, new_apartment_id)
        self.assertEqual(by_lease[new_lease_id].note, "10.04.2026 -> 01.05.2026 (21 дн.)")
        self.assertEqual(by_lease[new_lease_id].total_amount, 210.0)

    def test_all_debts_message_groups_rent_and_utility_by_month(self) -> None:
        with self.seed() as session:
            obj = session.get(RentalObject, 1)
            apartment = session.get(Apartment, 1)
            service = session.scalar(select(UtilityService).where(UtilityService.object_id == obj.id, UtilityService.kind == "electricity"))
            lease = self._lease(session, apartment)
            from rental_manager.models import RentCharge

            feb_charge = RentCharge(
                lease_id=lease.id,
                period_start=date(2026, 2, 1),
                period_end=date(2026, 2, 28),
                due_date=date(2026, 2, 14),
                ip_due=20000,
                personal_due=2200,
                ip_paid=20000,
                personal_paid=0,
                status="partial",
            )
            mar_charge = RentCharge(
                lease_id=lease.id,
                period_start=date(2026, 3, 1),
                period_end=date(2026, 3, 31),
                due_date=date(2026, 3, 14),
                ip_due=20000,
                personal_due=2200,
                ip_paid=0,
                personal_paid=2200,
                status="partial",
            )
            feb_bill = UtilityBill(
                service_id=service.id,
                period_start=date(2026, 1, 15),
                period_end=date(2026, 2, 15),
                status="issued",
                total_consumption=0,
                apartment_consumption=0,
                odn_consumption=0,
                total_cost=14535,
                average_unit_price=0,
            )
            feb_bill.lines.append(
                UtilityBillLine(
                    apartment_id=apartment.id,
                    lease_id=lease.id,
                    total_amount=14535,
                    paid_amount=0,
                    status="overdue",
                    due_date=date(2026, 2, 20),
                    note="15.01.2026 -> 15.02.2026 (31 дн.)",
                )
            )
            mar_bill = UtilityBill(
                service_id=service.id,
                period_start=date(2026, 2, 15),
                period_end=date(2026, 3, 15),
                status="issued",
                total_consumption=0,
                apartment_consumption=0,
                odn_consumption=0,
                total_cost=11560,
                average_unit_price=0,
            )
            mar_bill.lines.append(
                UtilityBillLine(
                    apartment_id=apartment.id,
                    lease_id=lease.id,
                    total_amount=11560,
                    paid_amount=10000,
                    status="partial",
                    due_date=date(2026, 3, 20),
                    note="15.02.2026 -> 15.03.2026 (28 дн.)",
                )
            )
            session.add_all([feb_charge, mar_charge, feb_bill, mar_bill])
            session.flush()

            breakdown = build_all_debts_breakdown(session, lease, today=date(2026, 4, 24))
            text = render_message_text(session, "message_all_debts", lease)

        self.assertIn("Февраль:", breakdown)
        self.assertIn("ИП оплачено", breakdown)
        self.assertIn("перевод по номеру телефона не поступил", breakdown)
        self.assertIn("Март:", breakdown)
        self.assertIn("ИП не оплачено", breakdown)
        self.assertIn("Остаток", breakdown)
        self.assertIn("отправьте соответствующий чек в этот чат в виде документа PDF", text)

    def test_utility_message_lists_detailed_lines(self) -> None:
        with self.seed() as session:
            gas_service = session.scalar(select(UtilityService).where(UtilityService.kind == "gas"))
            water_service = session.scalar(select(UtilityService).where(UtilityService.kind == "water"))
            apartment = gas_service.object.apartments[0]
            lease = self._lease(session, apartment)

            gas_bill = UtilityBill(
                service_id=gas_service.id,
                period_start=date(2026, 3, 1),
                period_end=date(2026, 3, 31),
                status="issued",
                total_consumption=0,
                apartment_consumption=0,
                odn_consumption=0,
                total_cost=1200,
                average_unit_price=0,
                due_date=date(2026, 4, 24),
            )
            gas_line = UtilityBillLine(
                apartment_id=apartment.id,
                lease_id=lease.id,
                total_amount=1200,
                paid_amount=0,
                status="issued",
                due_date=date(2026, 4, 24),
                note="01.03.2026 -> 31.03.2026 (30 дн.)",
            )
            gas_bill.lines.append(gas_line)

            water_bill = UtilityBill(
                service_id=water_service.id,
                period_start=date(2026, 3, 1),
                period_end=date(2026, 3, 31),
                status="issued",
                total_consumption=0,
                apartment_consumption=0,
                odn_consumption=0,
                total_cost=800,
                average_unit_price=0,
                due_date=date(2026, 4, 24),
            )
            water_line = UtilityBillLine(
                apartment_id=apartment.id,
                lease_id=lease.id,
                total_amount=800,
                paid_amount=0,
                status="issued",
                due_date=date(2026, 4, 24),
                note="01.03.2026 -> 31.03.2026 (30 дн.)",
            )
            water_bill.lines.append(water_line)
            session.add_all([gas_bill, water_bill])
            session.flush()

            text = render_message_text(
                session,
                "message_utility_bill",
                lease,
                line=gas_line,
                utility_lines_override=[gas_line, water_line],
                utility_due_date_override=date(2026, 4, 24),
            )

        self.assertIn("Сформированы счета по коммунальным платежам", text)
        self.assertIn("газ за период 01.03.2026 -> 31.03.2026", text)
        self.assertIn("вода за период 01.03.2026 -> 31.03.2026", text)
        self.assertIn("Всего:", text)
        self.assertIn("Коммуналка оплачивается переводом по номеру", text)

    def test_utility_message_uses_tenant_segment_period(self) -> None:
        with self.seed() as session:
            service = session.scalar(select(UtilityService).where(UtilityService.kind == "electricity"))
            apartment = service.object.apartments[0]
            lease = self._lease(session, apartment, start_date=date(2026, 5, 5))
            bill = UtilityBill(
                service_id=service.id,
                period_start=date(2026, 4, 20),
                period_end=date(2026, 5, 30),
                status="issued",
                total_consumption=0,
                apartment_consumption=0,
                odn_consumption=0,
                total_cost=900,
                average_unit_price=0,
                due_date=date(2026, 6, 5),
            )
            line = UtilityBillLine(
                apartment_id=apartment.id,
                lease_id=lease.id,
                total_amount=900,
                paid_amount=0,
                status="issued",
                due_date=date(2026, 6, 5),
                note="05.05.2026 -> 30.05.2026 (25 дн.)",
            )
            bill.lines.append(line)
            session.add(bill)
            session.flush()

            text = render_message_text(session, "message_utility_bill", lease, line=line, utility_lines_override=[line], utility_due_date_override=date(2026, 6, 5))

        self.assertIn("05.05.2026 -> 30.05.2026", text)
        self.assertNotIn("20.04.2026 -> 30.05.2026", text)

    @staticmethod
    def _lease(session, apartment: Apartment, start_date: date = date(2026, 4, 1), end_date: date | None = None, full_name: str | None = None) -> Lease:
        tenant = Tenant(full_name=full_name or f"Tenant {apartment.id}")
        session.add(tenant)
        session.flush()
        lease = Lease(
            apartment_id=apartment.id,
            tenant_id=tenant.id,
            start_date=start_date,
            end_date=end_date,
            payment_day=1,
            ip_amount=10000,
            personal_amount=2000,
        )
        session.add(lease)
        session.flush()
        return lease

    @staticmethod
    def _read_all_meters(
        session,
        service: UtilityService,
        object_start: float,
        object_end: float,
        apartment_end_values: dict[int, float],
    ) -> None:
        meters = session.scalars(select(Meter).where(Meter.service_id == service.id)).all()
        for meter in meters:
            if meter.scope == "object":
                start_value = object_start
                end_value = object_end
            else:
                start_value = 10
                end_value = apartment_end_values.get(meter.apartment_id, 10)
            session.add(MeterReading(meter_id=meter.id, reading_date=date(2026, 4, 1), value=start_value))
            session.add(MeterReading(meter_id=meter.id, reading_date=date(2026, 5, 1), value=end_value))
        session.flush()


class PaymentReceiptReviewTests(DatabaseTestCase):
    def test_suspicious_receipt_can_be_assigned_and_accepted(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Дом проверки", short_code="ДП")
            apartment = Apartment(name="ДП1", sort_order=1, odn_share_percent=100, active=True, object=rental_object)
            tenant = Tenant(full_name="Проверяемый жилец")
            lease = Lease(
                apartment=apartment,
                tenant=tenant,
                start_date=date(2026, 6, 1),
                payment_day=15,
                ip_amount=20000,
                personal_amount=0,
            )
            charge = RentCharge(
                lease=lease,
                period_start=date(2026, 6, 15),
                period_end=date(2026, 7, 14),
                due_date=date(2026, 6, 15),
                ip_due=20000,
                personal_due=0,
                status="overdue",
            )
            receipt = PaymentReceipt(
                lease_id=None,
                apartment_id=apartment.id,
                amount=20000,
                channel="unknown",
                paid_at=datetime(2026, 6, 16, 12, 0),
                source="telegram",
                status="suspicious",
                notes="нужна ручная проверка",
            )
            session.add_all([rental_object, apartment, tenant, lease, charge, receipt])
            session.commit()

            result = update_payment_receipt(
                receipt.id,
                {
                    "target_kind": "rent",
                    "rent_charge_id": charge.id,
                    "channel": "ip",
                    "status": "accepted",
                    "notes": "проверено владельцем",
                },
                session,
            )
            session.refresh(charge)
            session.refresh(receipt)

        self.assertEqual(result["status"], "accepted")
        self.assertEqual(receipt.status, "accepted")
        self.assertEqual(receipt.rent_charge_id, charge.id)
        self.assertEqual(receipt.channel, "ip")
        self.assertEqual(charge.ip_paid, 20000)
        self.assertEqual(charge.status, "paid")

    def test_suspicious_receipt_cannot_be_accepted_without_target(self) -> None:
        with self.Session() as session:
            receipt = PaymentReceipt(
                amount=1000,
                channel="unknown",
                paid_at=datetime(2026, 6, 16, 12, 0),
                source="telegram",
                status="suspicious",
            )
            session.add(receipt)
            session.commit()

            with self.assertRaises(HTTPException) as error:
                update_payment_receipt(receipt.id, {"status": "accepted"}, session)

        self.assertEqual(error.exception.status_code, 400)

    def test_future_rent_payment_can_create_and_remove_linked_expense(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Дом авансов", short_code="ДА")
            apartment = Apartment(name="ДА1", sort_order=1, odn_share_percent=100, active=True, object=rental_object)
            tenant = Tenant(full_name="Жилец с авансом")
            lease = Lease(
                apartment=apartment,
                tenant=tenant,
                start_date=date(2026, 7, 1),
                payment_day=5,
                ip_amount=20000,
                personal_amount=5000,
                active=True,
            )
            session.add_all([rental_object, apartment, tenant, lease])
            session.commit()

            result = create_manual_payment(
                {
                    "lease_id": lease.id,
                    "kind": "rent",
                    "target_month": 10,
                    "target_year": 2026,
                    "channel": "personal",
                    "amount": 3000,
                    "paid_at": "2026-07-20T12:30:00",
                    "is_expense": True,
                },
                session=session,
            )
            receipt_id = result["created"][0]["id"]
            receipt = session.get(PaymentReceipt, receipt_id)
            expense = session.scalar(select(Expense).where(Expense.source_funds == "rental_budget"))

            self.assertEqual(receipt.rent_charge.due_date, date(2026, 10, 5))
            self.assertTrue(receipt.is_expense)
            self.assertEqual(receipt.channel, "personal")
            self.assertIsNotNone(expense)
            self.assertEqual(expense.apartment_id, apartment.id)
            self.assertEqual(expense.expense_date, date(2026, 7, 20))
            self.assertEqual(expense.amount, 3000)

            update_payment_receipt(receipt.id, {"is_expense": False}, session=session)
            linked_expense = session.scalar(select(Expense).where(Expense.source_funds == "rental_budget"))

        self.assertIsNone(linked_expense)

    def test_tenant_payment_history_combines_old_and_current_apartments(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Баня", short_code="Б")
            old_apartment = Apartment(name="Баня3", sort_order=3, odn_share_percent=50, active=True, object=rental_object)
            current_apartment = Apartment(name="Баня4", sort_order=4, odn_share_percent=50, active=True, object=rental_object)
            tenant = Tenant(full_name="Никита")
            old_lease = Lease(
                apartment=old_apartment,
                tenant=tenant,
                start_date=date(2026, 1, 1),
                end_date=date(2026, 7, 10),
                payment_day=5,
                ip_amount=20000,
                personal_amount=5000,
                active=False,
            )
            current_lease = Lease(
                apartment=current_apartment,
                tenant=tenant,
                start_date=date(2026, 7, 11),
                payment_day=5,
                ip_amount=20000,
                personal_amount=5000,
                active=True,
            )
            session.add_all([rental_object, old_apartment, current_apartment, tenant, old_lease, current_lease])
            session.flush()
            session.add_all([
                PaymentReceipt(lease_id=old_lease.id, apartment_id=old_apartment.id, amount=1000, channel="personal", status="accepted", paid_at=datetime(2026, 7, 1, 12, 0)),
                PaymentReceipt(lease_id=current_lease.id, apartment_id=current_apartment.id, amount=2000, channel="ip", status="accepted", paid_at=datetime(2026, 7, 15, 12, 0)),
            ])
            session.commit()

            history = tenant_payment_history(tenant.id, session=session)

        self.assertEqual(history["lease_id"], current_lease.id)
        self.assertEqual(history["tenant_id"], tenant.id)
        self.assertEqual(history["apartment"], "Баня, Баня3 → Баня, Баня4")
        self.assertEqual({item["lease_id"] for item in history["receipts"]}, {old_lease.id, current_lease.id})

    def test_bot_dialog_keeps_assistant_messages_after_tenant_move(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Баня", short_code="Б")
            old_apartment = Apartment(name="Баня3", sort_order=3, odn_share_percent=50, active=True, object=rental_object)
            current_apartment = Apartment(name="Баня4", sort_order=4, odn_share_percent=50, active=True, object=rental_object)
            tenant = Tenant(full_name="Никита")
            old_lease = Lease(apartment=old_apartment, tenant=tenant, start_date=date(2026, 1, 1), end_date=date(2026, 7, 10), payment_day=5, active=False)
            current_lease = Lease(apartment=current_apartment, tenant=tenant, start_date=date(2026, 7, 11), payment_day=5, active=True)
            session.add_all([rental_object, old_apartment, current_apartment, tenant, old_lease, current_lease])
            session.flush()
            session.add(AppSetting(key="telegram_tenant_links", value=json.dumps({str(tenant.id): "456"})))
            conversation = AiConversation(chat_id="456", role="tenant", lease_id=old_lease.id, tenant_id=tenant.id)
            session.add(conversation)
            session.flush()
            session.add(AiMessage(conversation_id=conversation.id, lease_id=old_lease.id, role="assistant", channel="telegram", text="Сообщение бота до переезда"))
            session.commit()

            payload = bot_dialog_messages_payload(session, f"lease:{current_lease.id}")

        self.assertTrue(any(item["author"] == "Бот" and item["text"] == "Сообщение бота до переезда" for item in payload["messages"]))

    def test_receipt_document_is_served_inline_only_from_receipt_storage(self) -> None:
        with self.Session() as session:
            receipt = PaymentReceipt(
                amount=1000,
                channel="personal",
                paid_at=datetime(2026, 7, 20, 12, 0),
                source="telegram",
                status="suspicious",
                file_path="data/telegram_receipts/check.pdf",
            )
            session.add(receipt)
            session.commit()
            test_root = Path(self.tmp.name)
            stored_file = test_root / "data" / "telegram_receipts" / "check.pdf"
            stored_file.parent.mkdir(parents=True, exist_ok=True)
            stored_file.write_bytes(b"%PDF-1.4\n%%EOF")

            with patch("rental_manager.main.ROOT_DIR", test_root):
                response = payment_receipt_document(receipt.id, session=session)
                self.assertEqual(Path(response.path), stored_file)
                self.assertIn("inline", response.headers["content-disposition"])

                receipt.file_path = "outside.pdf"
                session.commit()
                with self.assertRaises(HTTPException) as error:
                    payment_receipt_document(receipt.id, session=session)

        self.assertEqual(error.exception.status_code, 404)


class DashboardCutoffTests(DatabaseTestCase):
    def test_dashboard_keeps_old_rent_debt_after_tenant_transfer(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Дом переезда", short_code="ДП")
            old_apartment = Apartment(name="1", sort_order=1, odn_share_percent=50, active=True, object=rental_object)
            current_apartment = Apartment(name="2", sort_order=2, odn_share_percent=50, active=True, object=rental_object)
            tenant = Tenant(full_name="Переехавший жилец")
            old_lease = Lease(
                apartment=old_apartment,
                tenant=tenant,
                start_date=date(2026, 1, 1),
                end_date=date(2026, 5, 31),
                payment_day=1,
                ip_amount=10000,
                personal_amount=0,
                active=False,
            )
            current_lease = Lease(
                apartment=current_apartment,
                tenant=tenant,
                start_date=date(2026, 6, 1),
                payment_day=1,
                ip_amount=10000,
                personal_amount=0,
                active=True,
            )
            old_charge = RentCharge(
                lease=old_lease,
                period_start=date(2026, 5, 1),
                period_end=date(2026, 5, 31),
                due_date=date(2026, 5, 1),
                ip_due=10000,
                personal_due=0,
                ip_paid=0,
                personal_paid=0,
                status="overdue",
            )
            session.add_all([rental_object, old_apartment, current_apartment, tenant, old_lease, current_lease, old_charge])
            session.commit()

            dashboard = build_dashboard(session)

        item = next(entry for entry in dashboard["rent_overdue"] if entry["id"] == old_charge.id)
        self.assertEqual(item["tenant_id"], tenant.id)
        self.assertFalse(item["lease_active"])
        self.assertEqual(item["debt"], 10000)
        self.assertEqual(dashboard["expected_receipts"]["rent"], 10000)

    def test_apartment_month_state_skips_ignored_lease(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Дом-игнор", short_code="ДИ")
            apartment = Apartment(name="ДИ1", sort_order=1, odn_share_percent=100, active=True, object=rental_object)
            tenant = Tenant(full_name="Старая арендаторша")
            lease = Lease(
                apartment=apartment,
                tenant=tenant,
                start_date=date(2026, 4, 1),
                end_date=None,
                payment_day=1,
                ip_amount=20000,
                personal_amount=3000,
                notes=IGNORE_LEASE_MARK,
            )
            session.add_all([rental_object, apartment, tenant, lease])
            session.commit()

            state = apartment_month_state(apartment, date(2026, 4, 1), date(2026, 4, 30))

        self.assertEqual(state["tenant_names"], "нет жильца")
        self.assertTrue(state["vacant_full"])

    def test_rent_report_hides_ignored_lease_charge(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Дом-отчёт", short_code="ДО")
            apartment = Apartment(name="ДО1", sort_order=1, odn_share_percent=100, active=True, object=rental_object)
            tenant = Tenant(full_name="Лишний хвост")
            lease = Lease(
                apartment=apartment,
                tenant=tenant,
                start_date=date(2026, 1, 1),
                end_date=None,
                payment_day=14,
                ip_amount=20000,
                personal_amount=3000,
                notes=IGNORE_LEASE_MARK,
            )
            charge = RentCharge(
                lease=lease,
                period_start=date(2026, 4, 14),
                period_end=date(2026, 5, 13),
                due_date=date(2026, 4, 14),
                ip_due=20000,
                personal_due=3000,
                status="overdue",
            )
            session.add_all([rental_object, apartment, tenant, lease, charge])
            session.commit()

            response = rent_report(start="2026-04-01", end="2026-04-30", session=session)
            workbook_bytes = asyncio.run(read_streaming_response_bytes(response))

        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet = workbook["Аренда"]
        values = [cell for row in sheet.iter_rows(values_only=True) for cell in row if isinstance(cell, str)]
        self.assertNotIn("Лишний хвост", values)

    def test_dashboard_ignores_debts_before_cutoff_date(self) -> None:
        with self.Session() as session:
            rental_object = RentalObject(name="Тестовый дом", short_code="ТД")
            tenant = Tenant(full_name="Старый хвост")
            session.add_all([rental_object, tenant, AppSetting(key="notification_cutoff_date", value="2026-01-01")])
            session.flush()

            apartment = Apartment(object_id=rental_object.id, name="ТД1", sort_order=1, odn_share_percent=100, active=True)
            session.add(apartment)
            session.flush()

            lease = Lease(
                apartment_id=apartment.id,
                tenant_id=tenant.id,
                start_date=date(2025, 1, 1),
                payment_day=14,
                ip_amount=20000,
                personal_amount=3000,
            )
            session.add(lease)
            session.flush()

            session.add_all(
                [
                    RentCharge(
                        lease_id=lease.id,
                        period_start=date(2025, 12, 14),
                        period_end=date(2026, 1, 13),
                        due_date=date(2025, 12, 14),
                        ip_due=20000,
                        personal_due=3000,
                        status="overdue",
                    ),
                    RentCharge(
                        lease_id=lease.id,
                        period_start=date(2026, 1, 14),
                        period_end=date(2026, 2, 13),
                        due_date=date(2026, 1, 14),
                        ip_due=20000,
                        personal_due=3000,
                        status="overdue",
                    ),
                ]
            )
            session.commit()

            dashboard = build_dashboard(session)

        self.assertEqual(len(dashboard["rent_overdue"]), 1)
        self.assertEqual(dashboard["rent_overdue"][0]["due_date"], "2026-01-14")


class UtilityAdvanceTests(DatabaseTestCase):
    def _seed_bill(self, session, total_amount: float = 1000) -> tuple[Lease, UtilityBill, UtilityBillLine]:
        rental_object = RentalObject(name="Дом авансов", short_code="ДА")
        apartment = Apartment(name="ДА1", sort_order=1, odn_share_percent=100, active=True, object=rental_object)
        tenant = Tenant(full_name="Жилец с авансом", active=True)
        lease = Lease(
            apartment=apartment,
            tenant=tenant,
            start_date=date(2026, 6, 1),
            payment_day=1,
            ip_amount=30000,
            personal_amount=5000,
            active=True,
        )
        service = UtilityService(
            object=rental_object,
            kind="electricity",
            name="Электричество",
            provider_due_day=24,
            resident_due_days=7,
            active=True,
        )
        bill = UtilityBill(
            service=service,
            period_start=date(2026, 6, 1),
            period_end=date(2026, 6, 30),
            status="draft",
            total_consumption=100,
            apartment_consumption=100,
            odn_consumption=0,
            total_cost=total_amount,
            average_unit_price=10,
        )
        line = UtilityBillLine(
            apartment=apartment,
            lease=lease,
            personal_consumption=100,
            odn_consumption=0,
            total_amount=total_amount,
            paid_amount=0,
            status="draft",
        )
        bill.lines.append(line)
        session.add_all([rental_object, apartment, tenant, lease, service, bill])
        session.flush()
        return lease, bill, line

    def test_manual_utility_advance_is_applied_when_bill_is_issued(self) -> None:
        with self.Session() as session:
            lease, bill, line = self._seed_bill(session, total_amount=1000)
            create_manual_payment(
                {
                    "lease_id": lease.id,
                    "kind": "utility_advance",
                    "amount": 600,
                    "paid_at": "2026-06-10T12:00:00",
                    "notes": "аванс за июнь",
                },
                session=session,
            )

            payload = issue_utility_bill(bill.id, session=session)
            session.refresh(line)
            receipts = session.scalars(select(PaymentReceipt).where(PaymentReceipt.lease_id == lease.id).order_by(PaymentReceipt.id)).all()

        self.assertEqual(payload["applied_advances"], 600)
        self.assertEqual(line.paid_amount, 600)
        self.assertEqual(line.status, "partial")
        self.assertEqual(len(receipts), 1)
        self.assertEqual(receipts[0].channel, "utilities")
        self.assertEqual(receipts[0].source, "utility_advance")
        self.assertEqual(receipts[0].utility_line_id, line.id)

    def test_utility_advance_keeps_leftover_balance_after_issue(self) -> None:
        with self.Session() as session:
            lease, bill, line = self._seed_bill(session, total_amount=1000)
            create_manual_payment(
                {
                    "lease_id": lease.id,
                    "kind": "utility_advance",
                    "amount": 1400,
                    "paid_at": "2026-06-10T12:00:00",
                },
                session=session,
            )

            payload = issue_utility_bill(bill.id, session=session)
            session.refresh(line)
            receipts = session.scalars(select(PaymentReceipt).where(PaymentReceipt.lease_id == lease.id).order_by(PaymentReceipt.id)).all()
            leftover = next(receipt for receipt in receipts if receipt.channel == "utility_advance")
            applied = next(receipt for receipt in receipts if receipt.source == "utility_advance" and receipt.utility_line_id == line.id)

        self.assertEqual(payload["applied_advances"], 1000)
        self.assertEqual(line.paid_amount, 1000)
        self.assertEqual(line.status, "paid")
        self.assertEqual(leftover.amount, 400)
        self.assertEqual(applied.amount, 1000)

    def test_issue_utility_bill_reports_partial_telegram_failures(self) -> None:
        with self.Session() as session:
            lease, bill, line = self._seed_bill(session, total_amount=1000)
            second_apartment = Apartment(
                object=lease.apartment.object,
                name="2",
                sort_order=2,
                odn_share_percent=100,
                active=True,
            )
            second_tenant = Tenant(full_name="Second tenant", active=True)
            second_lease = Lease(
                apartment=second_apartment,
                tenant=second_tenant,
                start_date=date(2026, 6, 1),
                payment_day=1,
                ip_amount=10000,
                personal_amount=0,
                active=True,
            )
            second_line = UtilityBillLine(
                apartment=second_apartment,
                lease=second_lease,
                personal_consumption=100,
                odn_consumption=0,
                total_amount=700,
                paid_amount=0,
                status="draft",
            )
            bill.lines.append(second_line)
            session.add_all([second_apartment, second_tenant, second_lease])
            session.flush()
            session.add(AppSetting(key="telegram_tenant_links", value=json.dumps({str(lease.tenant_id): "100", str(second_tenant.id): "200"})))
            session.commit()

            def fake_send(_session, chat_id, _text, _keyboard=None):
                if str(chat_id) == "200":
                    raise HTTPException(400, "Telegram API sendMessage failed: Not Found")
                return {"ok": True}

            with patch("rental_manager.main.send_telegram_text", side_effect=fake_send):
                payload = issue_utility_bill(bill.id, session=session)
                session.refresh(bill)
                session.refresh(line)
                session.refresh(second_line)

        self.assertEqual(payload["sent"], 1)
        self.assertEqual(payload["failed"], 1)
        self.assertEqual(payload["failed_recipients"][0]["tenant"], "Second tenant")
        self.assertEqual(bill.status, "issued")
        self.assertEqual(line.status, "issued")
        self.assertEqual(second_line.status, "issued")

    def test_issue_utility_bill_keeps_draft_when_all_telegram_sends_fail(self) -> None:
        with self.Session() as session:
            lease, bill, line = self._seed_bill(session, total_amount=1000)
            session.add(AppSetting(key="telegram_tenant_links", value=json.dumps({str(lease.tenant_id): "100"})))
            session.commit()

            with patch("rental_manager.main.send_telegram_text", side_effect=HTTPException(400, "Telegram API sendMessage failed: Not Found")):
                with self.assertRaises(HTTPException) as exc:
                    issue_utility_bill(bill.id, session=session)
                self.assertIn("Не отправлено ни одного сообщения", str(exc.exception.detail))
                session.rollback()
                session.refresh(bill)
                session.refresh(line)

        self.assertEqual(bill.status, "draft")
        self.assertEqual(line.status, "draft")

    def test_month_summary_tracks_salary_bills_and_advances(self) -> None:
        with self.Session() as session:
            lease, bill, line = self._seed_bill(session, total_amount=1000)
            create_manual_payment(
                {
                    "lease_id": lease.id,
                    "kind": "utility_advance",
                    "amount": 500,
                    "paid_at": "2026-06-05T12:00:00",
                },
                session=session,
            )
            issue_utility_bill(bill.id, session=session)
            charge = session.scalar(select(RentCharge).where(RentCharge.lease_id == lease.id, RentCharge.due_date == date(2026, 6, 1)))
            if not charge:
                charge = RentCharge(
                    lease=lease,
                    period_start=date(2026, 6, 1),
                    period_end=date(2026, 6, 30),
                    due_date=date(2026, 6, 1),
                )
                session.add(charge)
            charge.ip_due = 30000
            charge.personal_due = 5000
            charge.ip_paid = 0
            charge.personal_paid = 2500
            charge.status = "partial"
            session.commit()

            summary = month_dashboard_summary(session, 2026, 6, today=date(2026, 6, 21))

        self.assertEqual(summary["salary_due"], 5000)
        self.assertEqual(summary["salary_paid"], 2500)
        self.assertTrue(summary["utility_bills_issued"])
        self.assertFalse(summary["utility_provider_paid"])
        self.assertEqual(summary["bill_payment_due"], 1000)
        self.assertEqual(summary["bill_payment_paid"], 500)
        self.assertEqual(summary["advance_paid"], 500)
        self.assertEqual(summary["advance_due"], 1000)
        self.assertEqual(summary["occupied"], 1)

    def test_dashboard_income_trend_uses_only_accepted_receipts(self) -> None:
        with self.Session() as session:
            session.add_all(
                [
                    PaymentReceipt(amount=1500, channel="personal", paid_at=datetime(2026, 6, 5, 12, 0), status="accepted"),
                    PaymentReceipt(amount=9000, channel="personal", paid_at=datetime(2026, 6, 6, 12, 0), status="rejected"),
                ]
            )
            session.commit()

            trend = dashboard_income_trend(session, today=date(2026, 7, 11), months=3)

        self.assertEqual([item["period"] for item in trend], ["2026-05", "2026-06", "2026-07"])
        self.assertEqual([item["amount"] for item in trend], [0, 1500, 0])


    def test_auto_advance_draft_uses_apartment_history_and_current_bill(self) -> None:
        with self.Session() as session:
            lease, bill, _line = self._seed_bill(session, total_amount=1000)
            historical_bill = UtilityBill(
                service=bill.service,
                period_start=date(2025, 6, 30),
                period_end=date(2025, 7, 30),
                status="issued",
                total_consumption=90,
                apartment_consumption=90,
                odn_consumption=0,
                total_cost=900,
                average_unit_price=10,
            )
            historical_bill.lines.append(
                UtilityBillLine(
                    apartment=lease.apartment,
                    lease=lease,
                    personal_consumption=90,
                    odn_consumption=0,
                    total_amount=900,
                    paid_amount=900,
                    status="paid",
                )
            )
            session.add(historical_bill)
            session.flush()

            advance_bills = ensure_utility_advance_drafts_for_bills(session, [bill])
            session.flush()

            self.assertEqual(len(advance_bills), 1)
            advance_line = advance_bills[0].lines[0]
            self.assertEqual(advance_bills[0].bill_type, "advance")
            self.assertEqual(advance_bills[0].period_start, date(2026, 6, 30))
            self.assertEqual(advance_bills[0].period_end, date(2026, 7, 30))
            self.assertEqual(advance_line.line_type, "advance")
            self.assertAlmostEqual(advance_line.total_amount, 917.5)
            self.assertIn("Аванс коммуналки", advance_line.note)

    def test_auto_advance_caps_non_winter_history_spike_against_current_bill(self) -> None:
        with self.Session() as session:
            lease, bill, _line = self._seed_bill(session, total_amount=1475.54)
            historical_bill = UtilityBill(
                service=bill.service,
                period_start=date(2025, 6, 1),
                period_end=date(2025, 6, 30),
                status="issued",
                total_consumption=4500,
                apartment_consumption=4500,
                odn_consumption=0,
                total_cost=44948.18,
                average_unit_price=10,
            )
            historical_bill.lines.append(
                UtilityBillLine(
                    apartment=lease.apartment,
                    lease=lease,
                    personal_consumption=4500,
                    odn_consumption=0,
                    total_amount=44948.18,
                    paid_amount=44948.18,
                    status="paid",
                )
            )
            session.add(historical_bill)
            session.flush()

            advance_bills = ensure_utility_advance_drafts_for_bills(session, [bill])
            session.flush()

            advance_line = advance_bills[0].lines[0]
            metadata = json.loads(advance_line.metadata_json)
            self.assertAlmostEqual(advance_line.total_amount, 2951.08)
            self.assertTrue(metadata["forecast_capped"])
            self.assertGreater(metadata["uncapped_forecast"], 29000)
            self.assertEqual(metadata["forecast_cap_multiplier"], 2.0)

    def test_issue_preview_refreshes_existing_bad_advance_draft(self) -> None:
        with self.Session() as session:
            lease, bill, _line = self._seed_bill(session, total_amount=1475.54)
            historical_bill = UtilityBill(
                service=bill.service,
                period_start=date(2025, 6, 1),
                period_end=date(2025, 6, 30),
                status="issued",
                total_consumption=4500,
                apartment_consumption=4500,
                odn_consumption=0,
                total_cost=44948.18,
                average_unit_price=10,
            )
            historical_bill.lines.append(
                UtilityBillLine(
                    apartment=lease.apartment,
                    lease=lease,
                    personal_consumption=4500,
                    odn_consumption=0,
                    total_amount=44948.18,
                    paid_amount=44948.18,
                    status="paid",
                )
            )
            advance_bill = UtilityBill(
                service=bill.service,
                bill_type="advance",
                period_start=date(2026, 6, 30),
                period_end=date(2026, 7, 30),
                status="draft",
                total_consumption=0,
                apartment_consumption=0,
                odn_consumption=0,
                total_cost=29732.76,
                average_unit_price=0,
                is_forecast=True,
                provider_paid=True,
            )
            advance_line = UtilityBillLine(
                apartment=lease.apartment,
                lease=lease,
                line_type="advance",
                personal_consumption=0,
                odn_consumption=0,
                total_amount=29732.76,
                paid_amount=0,
                status="draft",
            )
            advance_bill.lines.append(advance_line)
            session.add_all([historical_bill, advance_bill])
            session.flush()

            preview_issue_utility_bill(bill.id, session=session)
            session.refresh(advance_line)

            self.assertAlmostEqual(advance_line.total_amount, 2951.08)
            metadata = json.loads(advance_line.metadata_json)
            self.assertTrue(metadata["forecast_capped"])

    def test_auto_advance_is_prorated_until_planned_move_out(self) -> None:
        with self.Session() as session:
            lease, bill, _line = self._seed_bill(session, total_amount=1000)
            lease.end_date = bill.period_end + timedelta(days=10)
            session.flush()

            advance_bills = ensure_utility_advance_drafts_for_bills(session, [bill])
            session.flush()

            self.assertEqual(len(advance_bills), 1)
            advance_line = advance_bills[0].lines[0]
            self.assertAlmostEqual(advance_line.total_amount, 333.33)
            self.assertIn("30.06.2026 -> 10.07.2026", advance_line.note)
            metadata = json.loads(advance_line.metadata_json)
            self.assertEqual(metadata["line_period_end"], "2026-07-10")
            self.assertAlmostEqual(metadata["period_factor"], 0.3333)

    def test_issue_preview_groups_object_drafts_and_advance(self) -> None:
        with self.Session() as session:
            lease, bill, _line = self._seed_bill(session, total_amount=1000)
            water = UtilityService(
                object=lease.apartment.object,
                kind="water",
                name="Вода",
                provider_due_day=24,
                resident_due_days=7,
                active=True,
            )
            water_bill = UtilityBill(
                service=water,
                period_start=bill.period_start,
                period_end=bill.period_end,
                status="draft",
                total_consumption=50,
                apartment_consumption=50,
                odn_consumption=0,
                total_cost=500,
                average_unit_price=10,
            )
            water_bill.lines.append(
                UtilityBillLine(
                    apartment=lease.apartment,
                    lease=lease,
                    personal_consumption=50,
                    odn_consumption=0,
                    total_amount=500,
                    paid_amount=0,
                    status="draft",
                )
            )
            session.add(water_bill)
            session.flush()
            advance_bills = ensure_utility_advance_drafts_for_bills(session, [bill, water_bill])
            session.flush()

            preview = preview_issue_utility_bill(bill.id, session=session)

            self.assertEqual(set(preview["bill_ids"]), {bill.id, water_bill.id, advance_bills[0].id})
            self.assertEqual(len(preview["targets"]), 1)
            self.assertIn("электричество", preview["targets"][0]["text"].lower())
            self.assertIn("вода", preview["targets"][0]["text"].lower())
            self.assertIn("аванс коммуналки", preview["targets"][0]["text"].lower())



class DashboardIssuedUtilityTests(DatabaseTestCase):
    def test_provider_reading_due_alert_starts_three_days_before_deadline(self) -> None:
        with self.seed() as session:
            service = session.scalar(select(UtilityService).where(UtilityService.kind == "electricity"))
            service.provider_reading_due_day = 20
            session.commit()

            statuses = provider_reading_statuses_for_month(session, 2026, 5, today=date(2026, 5, 17))
            target = next(item for item in statuses if item["service_id"] == service.id)
            self.assertEqual(target["status"], "pending")
            self.assertTrue(target["alert"])
            self.assertEqual(target["due_date"], "2026-05-20")

            meter = session.scalar(select(Meter).where(Meter.service_id == service.id, Meter.scope == "object"))
            session.add(MeterReading(meter_id=meter.id, reading_date=date(2026, 5, 18), value=1234))
            session.commit()

            statuses = provider_reading_statuses_for_month(session, 2026, 5, today=date(2026, 5, 19))
            target = next(item for item in statuses if item["service_id"] == service.id)
            self.assertEqual(target["status"], "paid")
            self.assertFalse(target["alert"])
            self.assertEqual(target["reading_date"], "2026-05-18")

    def test_utility_bill_month_is_based_on_period_start(self) -> None:
        with self.seed() as session:
            service = session.scalar(select(UtilityService).where(UtilityService.kind == "electricity"))
            bill = UtilityBill(
                service_id=service.id,
                period_start=date(2026, 4, 1),
                period_end=date(2026, 5, 1),
                status="issued",
                total_consumption=0,
                apartment_consumption=0,
                odn_consumption=0,
                total_cost=1000,
                average_unit_price=0,
                due_date=date(2026, 5, 8),
                provider_paid=True,
                provider_paid_at=datetime(2026, 5, 8, 12, 0),
            )
            session.add(bill)
            session.commit()

            april_start, april_end = date(2026, 4, 1), date(2026, 4, 30)
            may_start, may_end = date(2026, 5, 1), date(2026, 5, 31)

            self.assertEqual(utility_bill_for_month(session, service, april_start, april_end).id, bill.id)
            self.assertIsNone(utility_bill_for_month(session, service, may_start, may_end))

    def test_dashboard_includes_issued_utility_lines(self) -> None:
        with self.seed() as session:
            service = session.scalar(select(UtilityService).where(UtilityService.kind == "electricity"))
            apartment = service.object.apartments[0]
            tenant = Tenant(full_name="Жилец с выставленной коммуналкой")
            session.add(tenant)
            session.flush()

            lease = Lease(
                apartment_id=apartment.id,
                tenant_id=tenant.id,
                start_date=date(2026, 4, 1),
                payment_day=14,
                ip_amount=20000,
                personal_amount=3000,
            )
            session.add(lease)
            session.flush()

            bill = UtilityBill(
                service_id=service.id,
                period_start=date(2026, 4, 1),
                period_end=date(2026, 4, 20),
                status="issued",
                total_consumption=0,
                apartment_consumption=0,
                odn_consumption=0,
                total_cost=1800,
                average_unit_price=0,
                due_date=date(2026, 4, 27),
            )
            bill.lines.append(
                UtilityBillLine(
                    apartment_id=apartment.id,
                    lease_id=lease.id,
                    total_amount=1800,
                    paid_amount=0,
                    status="issued",
                    due_date=date(2026, 4, 27),
                    note="01.04.2026 -> 20.04.2026 (19 дн.)",
                )
            )
            session.add(bill)
            session.commit()

            dashboard = build_dashboard(session)

        self.assertEqual(len(dashboard["utility_issued"]), 1)
        self.assertEqual(dashboard["utility_issued"][0]["tenant"], "Жилец с выставленной коммуналкой")
        self.assertEqual(dashboard["expected_receipts"]["utility"], 1800)


class StatusHelperTests(unittest.TestCase):
    def test_amount_statuses(self) -> None:
        self.assertEqual(status_for_amount(100, 0), "pending")
        self.assertEqual(status_for_amount(100, 99), "partial")
        self.assertEqual(status_for_amount(100, 100), "paid")
        self.assertEqual(status_for_amount(100, 101), "paid_ahead")
        self.assertEqual(status_for_amount(0, 0), "paid")


if __name__ == "__main__":
    unittest.main()
