from __future__ import annotations

import calendar
from contextlib import asynccontextmanager
from collections import deque
import hashlib
import hmac
import json
import os
import queue
import re
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from pathlib import Path
import threading
import time as time_module
from typing import Any, cast
from zoneinfo import ZoneInfo

from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, func, or_, select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload, selectinload

from rental_manager.database import Base, DATABASE_URL, ROOT_DIR, SessionLocal, get_session, init_db
from rental_manager.models import (
    AgentActionProposal,
    AgentMemory,
    AgentTask,
    AgentTenantState,
    AiActionLog,
    AiConversation,
    AiFeatureUsageDaily,
    AiMessage,
    AiSkill,
    AiUsageDaily,
    Apartment,
    Expense,
    Lease,
    ManualDebt,
    MessageLog,
    Meter,
    MeterReading,
    PaymentReceipt,
    PaymentSituation,
    HermesAgentRun,
    OperationalCase,
    OwnerBriefing,
    OwnerBriefingItem,
    OwnerCommitment,
    OwnerPreference,
    RentalObject,
    RentCharge,
    Tariff,
    Tenant,
    TenantStrategyProfile,
    UtilityAdvanceLedger,
    UtilityAdvanceSetting,
    UtilityAdvanceSettingHistory,
    UtilityBill,
    UtilityBillLine,
    UtilityService,
    utc_now,
    AppSetting,
    ProcessedTelegramUpdate,
)
from rental_manager.security.pins import (
    LEGACY_PIN_SETTING_KEYS,
    PIN_SETTING_KEYS,
    configured_environment_pin_hash,
    hash_pin,
    is_pin_hash,
    pin_is_compromised,
    verify_pin,
)
from rental_manager.security.secrets import decrypt_secret, encrypt_secret, secret_is_encrypted
from rental_manager.security.sessions import (
    clear_login_failures,
    csrf_is_valid,
    find_session,
    issue_session,
    login_fingerprint,
    login_retry_after,
    record_login_failure,
    revoke_other_sessions,
    revoke_session,
)
from rental_manager.observability.logging import get_logger, redact_log_text
from rental_manager.services.ai_context import owner_context_text, tenant_context_text
from rental_manager.services.ai_policy import (
    AI_BUDGET_EXCEEDED_TEXT,
    AI_DAILY_LIMIT_EXCEEDED_TEXT,
    AI_DISABLED_TEXT,
    AI_UNAVAILABLE_TEXT,
    AUDIT_USER_PROMPT,
    OWNER_AGENT_SYSTEM_PROMPT,
    SUPERVISOR_SYSTEM_PROMPT,
    TENANT_AGENT_SYSTEM_PROMPT,
    TENANT_DEBT_FOLLOWUP_SYSTEM_PROMPT,
    TENANT_ESCALATION_TEXT,
    clean_ai_response,
    estimate_tokens,
    tenant_question_needs_owner,
)
from rental_manager.services.agent_protocol import AgentEnvelope, action_confirmation_keyboard, parse_agent_envelope
from rental_manager.services.ai_providers import (
    AiProvider,
    AiProviderConfigError,
    build_provider_runtime,
    normalize_deepseek_model,
    provider_chain,
)
from rental_manager.services.billing import (
    IGNORE_LEASE_MARK,
    RENT_GENERATION_START,
    active_lease_for_apartment,
    calculate_utility_bill,
    full_months_lived,
    generate_rent_charges,
    money,
    parse_date,
    resident_due_date,
    status_for_amount,
    update_rent_charge_status,
    update_utility_line_status,
)
from rental_manager.services.deepseek_client import DeepSeekClient, DeepSeekClientError, DeepSeekResult
from rental_manager.services.owner_ai_tools import build_owner_read_tools_context
from rental_manager.services.owner_operations import (
    OWNER_OPERATION_SPECS,
    owner_operation_label,
    owner_operation_preview,
    validate_owner_operation,
)
from rental_manager.services.hermes.briefing import (
    briefing_keyboard,
    briefing_preview,
    build_owner_briefing,
    mark_briefing_sent,
    snooze_briefing,
)
from rental_manager.services.hermes.cases import (
    OPEN_CASE_STATUSES,
    case_label,
    case_snapshot,
    close_case as hermes_close_case,
    reconcile_operational_cases,
    resolve_case_alias,
    snooze_case as hermes_snooze_case,
)
from rental_manager.services.hermes.control_center import (
    case_details as hermes_case_details,
    control_center_snapshot,
    list_cases as hermes_list_cases,
    overview as hermes_overview,
    serialize_commitment,
    serialize_preference,
    serialize_proposal,
    serialize_run,
    serialize_skill,
    serialize_strategy,
)
from rental_manager.services.hermes.events import emit_domain_event, install_domain_event_listeners
from rental_manager.services.hermes.memory import (
    capture_owner_preference as capture_structured_owner_preference,
    complete_owner_commitment,
    create_owner_commitment,
    is_commitment_phrase,
    postpone_owner_commitment,
    reconcile_owner_commitments,
    update_conversation_summary,
)
from rental_manager.services.hermes.reminders import (
    classify_tenant_intent,
    record_reminder_outcome,
    reminder_allowed as hermes_reminder_allowed,
    tenant_acknowledgement,
    tenant_response_policy,
    update_latest_reminder_outcome,
)
from rental_manager.services.hermes.runtime import (
    ContextManifest,
    FEATURE_OUTPUT_LIMITS,
    build_owner_context as build_hermes_owner_context,
    build_tenant_context as build_hermes_tenant_context,
    classify_owner_intent,
    complete_agent_run,
    create_agent_run,
    create_grouped_deferral_proposal,
    execute_grouped_deferral,
    plan_grouped_deferral,
    record_feature_usage,
    usage_summary as hermes_usage_summary,
)
from rental_manager.services.hermes.safety import ACTION_SAFETY_REGISTRY
from rental_manager.services.hermes.skills import (
    activate_skill,
    create_skill_draft,
    disable_skill,
    dry_run_skill,
    rollback_skill,
    skill_card,
    version_skill,
)
from rental_manager.services.payment_allocation import (
    EPS,
    build_rent_plan,
    create_rent_receipts,
    create_utility_receipts,
    describe_rent_allocation_decision,
    recalculate_lease_balances,
    utility_line_candidates,
)
from rental_manager.services.receipt_matching import (
    detect_receipt_channel,
    normalize_telegram_handle,
    receipt_validation_issues,
)
from rental_manager.services.receipt_parser import parse_receipt_file
from rental_manager.services.seed import seed_if_empty, seed_release_baseline_if_empty
from rental_manager.services.telegram_bot import (
    TelegramApiError,
    answer_callback_query,
    app_keyboard,
    build_reports_message,
    build_status_message,
    clear_inline_keyboard,
    copy_message,
    download_telegram_file,
    normalize_bot_token,
    owner_commands,
    parse_command,
    send_message,
    tenant_keyboard,
    telegram_file_info,
    telegram_api_request,
)


APP_BUILD_MARKER = "progressive-boot-2026-07-05"
LOGGER = get_logger("rental_manager")


@asynccontextmanager
async def lifespan(_: FastAPI):
    startup()
    start_telegram_workers()
    try:
        yield
    finally:
        stop_telegram_workers()


app = FastAPI(title="Rental Manager", version="0.2.0", lifespan=lifespan)
app.add_middleware(GZipMiddleware, minimum_size=1024)
app.mount("/static", StaticFiles(directory=ROOT_DIR / "static"), name="static")


PERF_STARTED_AT = time_module.time()
PERF_SLOW_REQUEST_MS = int(os.environ.get("RENTAL_MANAGER_PERF_SLOW_MS", "1500") or 1500)
PERF_LOCK = threading.Lock()
PERF_RECENT_REQUESTS: deque[dict[str, Any]] = deque(maxlen=180)
PERF_SLOW_REQUESTS: deque[dict[str, Any]] = deque(maxlen=90)
PERF_SECTION_EVENTS: deque[dict[str, Any]] = deque(maxlen=120)
PERF_BACKGROUND_EVENTS: deque[dict[str, Any]] = deque(maxlen=120)
PERF_REQUEST_AGGREGATES: dict[str, dict[str, Any]] = {}
PERF_USER_AGENTS: dict[str, dict[str, Any]] = {}


def runtime_log(area: str, message: str) -> None:
    LOGGER.info(redact_log_text(message), extra={"area": area})


def normalized_perf_path(path: str) -> str:
    if path.startswith("/static/"):
        return "/static/*"
    return re.sub(r"/\d+(?=/|$)", "/:id", path)


def compact_user_agent(value: str) -> str:
    value = re.sub(r"\s+", " ", str(value or "")).strip()
    if not value:
        return "-"
    return value[:160]


def record_perf_request(
    *,
    method: str,
    path: str,
    status_code: int,
    duration_ms: int,
    user_agent: str,
) -> None:
    if path.startswith("/static/") or path == "/favicon.ico":
        return
    route = f"{method.upper()} {normalized_perf_path(path)}"
    now = utc_now().isoformat()
    agent = compact_user_agent(user_agent)
    event = {
        "at": now,
        "route": route,
        "path": path[:180],
        "status": int(status_code or 0),
        "duration_ms": int(duration_ms),
        "user_agent": agent,
    }
    with PERF_LOCK:
        PERF_RECENT_REQUESTS.appendleft(event)
        if duration_ms >= PERF_SLOW_REQUEST_MS or status_code >= 500:
            PERF_SLOW_REQUESTS.appendleft(event)
        aggregate = PERF_REQUEST_AGGREGATES.setdefault(
            route,
            {
                "route": route,
                "count": 0,
                "total_ms": 0,
                "max_ms": 0,
                "slow": 0,
                "errors": 0,
                "last_ms": 0,
                "last_at": now,
            },
        )
        aggregate["count"] += 1
        aggregate["total_ms"] += duration_ms
        aggregate["max_ms"] = max(int(aggregate["max_ms"]), duration_ms)
        aggregate["slow"] += 1 if duration_ms >= PERF_SLOW_REQUEST_MS else 0
        aggregate["errors"] += 1 if status_code >= 500 else 0
        aggregate["last_ms"] = duration_ms
        aggregate["last_at"] = now
        agent_stats = PERF_USER_AGENTS.setdefault(
            agent,
            {"user_agent": agent, "count": 0, "total_ms": 0, "last_at": now},
        )
        agent_stats["count"] += 1
        agent_stats["total_ms"] += duration_ms
        agent_stats["last_at"] = now


def record_perf_section(area: str, label: str, duration_ms: int, detail: dict[str, Any] | None = None) -> None:
    event = {
        "at": utc_now().isoformat(),
        "area": area,
        "label": label,
        "duration_ms": int(duration_ms),
        "detail": detail or {},
    }
    with PERF_LOCK:
        PERF_SECTION_EVENTS.appendleft(event)


def record_background_event(area: str, duration_ms: int, status: str = "ok", detail: dict[str, Any] | None = None) -> None:
    event = {
        "at": utc_now().isoformat(),
        "area": area,
        "status": status,
        "duration_ms": int(duration_ms),
        "detail": detail or {},
    }
    with PERF_LOCK:
        PERF_BACKGROUND_EVENTS.appendleft(event)


def performance_snapshot() -> dict[str, Any]:
    with PERF_LOCK:
        routes = []
        for item in PERF_REQUEST_AGGREGATES.values():
            count = max(1, int(item["count"]))
            routes.append(
                {
                    **item,
                    "avg_ms": round(float(item["total_ms"]) / count, 1),
                }
            )
        agents = []
        for item in PERF_USER_AGENTS.values():
            count = max(1, int(item["count"]))
            agents.append(
                {
                    **item,
                    "avg_ms": round(float(item["total_ms"]) / count, 1),
                }
            )
        return {
            "started_at": datetime.fromtimestamp(PERF_STARTED_AT, tz=timezone.utc).isoformat(),
            "uptime_seconds": int(time_module.time() - PERF_STARTED_AT),
            "slow_threshold_ms": PERF_SLOW_REQUEST_MS,
            "routes": sorted(routes, key=lambda item: (item["total_ms"], item["max_ms"]), reverse=True)[:30],
            "user_agents": sorted(agents, key=lambda item: item["count"], reverse=True)[:20],
            "recent_requests": list(PERF_RECENT_REQUESTS)[:60],
            "slow_requests": list(PERF_SLOW_REQUESTS)[:40],
            "section_events": list(PERF_SECTION_EVENTS)[:40],
            "background_events": list(PERF_BACKGROUND_EVENTS)[:40],
        }


DEFAULT_SETTINGS = {
    "color_palette": "premium",
    "app_base_url": "",
    "telegram_owner_chat_id": "",
    "notifications_enabled": False,
    "notification_cutoff_date": "",
    "automation_rent_due_cadence": "daily_evening",
    "automation_rent_overdue_cadence": "daily_evening",
    "automation_utility_cadence": "daily_evening",
    "ai_enabled": False,
    "ai_tenant_free_text_enabled": True,
    "ai_supervisor_enabled": True,
    "ai_supervisor_cadence": "daily",
    "ai_supervisor_weekday": "0",
    "ai_supervisor_time": "10:00",
    "ai_supervisor_model": "deepseek-v4-flash",
    "ai_supervisor_max_tokens": "1200",
    "ai_action_confirmation_ttl_hours": "48",
    "ai_daily_call_limit": "100",
    "ai_max_output_tokens": "1500",
    "ai_usd_rub_rate": "100",
    "ai_owner_instructions": "",
    "ai_tenant_instructions": "",
    "ai_audit_instructions": "",
    "deepseek_model": "deepseek-v4-flash",
    "ai_monthly_budget_rub": "1000",
    "hermes_briefing_enabled": True,
    "hermes_auto_level_one_enabled": False,
    "hermes_auto_template_reminders": True,
    "hermes_mass_action_pin_threshold": "20",
    "ai_tenant_mode": "auto",
    "ai_feature_owner_chat_daily_limit": "50",
    "ai_feature_tenant_chat_daily_limit": "50",
    "ai_feature_case_details_daily_limit": "20",
    "ai_feature_skill_builder_daily_limit": "10",
    "ai_feature_deep_audit_daily_limit": "5",
    "ai_feature_key_test_daily_limit": "5",
    "ai_output_chars_daily_briefing": "800",
    "ai_output_chars_case_details": "1200",
    "ai_output_chars_owner_chat": "800",
    "ai_output_chars_tenant_chat": "500",
    "ai_output_chars_skill_proposal": "1000",
    "ip_recipient_name": "",
    "ip_recipient_inn": "",
    "ip_recipient_ogrnip": "",
    "ip_recipient_account": "",
    "ip_recipient_bank": "",
    "ip_recipient_bik": "",
    "ip_recipient_correspondent_account": "",
    "ip_recipient_bank_inn": "",
    "ip_recipient_bank_kpp": "",
    "personal_recipient_name": "",
    "personal_recipient_phone": "",
    "personal_recipient_bank": "",
    "message_rent_upcoming": (
        "Здравствуйте! Напоминаю, что {rent_due_date} ожидается оплата аренды за период {rent_period}.\n\n"
        "К оплате: {total_due}\n"
        "• {ip_due} — на расчётный счёт ИП {ip_recipient_name_text}, счёт {ip_recipient_account_text}\n"
        "• {personal_due} — переводом по номеру {personal_recipient_phone_text}, {personal_recipient_bank_text}\n\n"
        "После каждого перевода отправьте чек PDF-документом в этот чат."
    ),
    "message_rent_due": (
        "Сегодня день оплаты аренды за период {rent_period}.\n\n"
        "К оплате: {total_due}\n"
        "• {ip_due} — на расчётный счёт ИП {ip_recipient_name_text}, счёт {ip_recipient_account_text}\n"
        "• {personal_due} — переводом по номеру {personal_recipient_phone_text}, {personal_recipient_bank_text}\n\n"
        "После оплаты отправьте чек PDF-документом. Если переводы выполнялись отдельно, отправьте оба чека."
    ),
    "message_rent_overdue": (
        "Оплата аренды пока не подтверждена.\n\n"
        "Текущий долг: {debt}\n"
        "• на ИП {ip_recipient_name_text}, счёт {ip_recipient_account_text}: {ip_due}\n"
        "• по номеру {personal_recipient_phone_text}, {personal_recipient_bank_text}: {personal_due}\n"
        "Период: {rent_period}\n\n"
        "Сообщите статус оплаты или отправьте чек PDF-документом, если перевод уже выполнен."
    ),
    "message_rent_status_request": (
        "Оплата аренды всё ещё не подтверждена.\n\n"
        "Долг: {debt}\n"
        "• на ИП {ip_recipient_name_text}, счёт {ip_recipient_account_text}: {ip_due}\n"
        "• по номеру {personal_recipient_phone_text}, {personal_recipient_bank_text}: {personal_due}\n\n"
        "Пожалуйста, выберите актуальный статус. После ответа автоматические напоминания будут поставлены на паузу."
    ),
    "message_utility_bill": (
        "Сформированы счета по коммунальным платежам:\n{utility_debt_details}\n\n"
        "Всего: {utility_total}. Срок оплаты — до {utility_due_date}.\n"
        "Коммуналка оплачивается переводом по номеру {personal_recipient_phone_text}, "
        "{personal_recipient_bank_text}. Учтённые авансы уже вычтены из суммы.\n\n"
        "После оплаты отправьте чек PDF-документом в этот чат."
    ),
    "message_utility_overdue": (
        "Коммунальные платежи пока не подтверждены:\n{utility_debt_details}\n\n"
        "Остаток: {utility_total}. Оплата переводом по номеру {personal_recipient_phone_text}, "
        "{personal_recipient_bank_text}.\n\n"
        "Сообщите статус или отправьте чек PDF-документом."
    ),
    "message_all_debts": (
        "Здравствуйте. Сейчас в системе открыты следующие платежи:\n{all_debts_breakdown}\n\n"
        "Аренда оплачивается двумя назначениями: указанная часть — на расчётный счёт ИП, "
        "дополнительная часть — переводом по номеру телефона. Коммуналка оплачивается по номеру телефона.\n"
        "После каждого перевода отправьте соответствующий чек в этот чат в виде документа PDF."
    ),
    "message_receipt_received": (
        "Чек получен. Если сумма, получатель и назначение совпали, платёж будет зачтён автоматически. "
        "Если проверка найдёт расхождение, чек получит владелец."
    ),
    "message_receipt_review": "Чек получен, но автоматически зачесть его не получилось. Передал владельцу на проверку; напоминания по этой ситуации остановлены.",
    "message_receipt_duplicate": "Этот чек уже есть в системе. Повторно засчитывать его не буду.",
    "message_owner_receipt_alert": "🧾 Новый чек от {tenant_name}\n🏠 Квартира: {apartment}\n💰 Сумма: {amount}\n📌 Канал: {channel}\n⚙️ Статус: {receipt_status}\n{receipt_summary}",
}
SECRET_SETTINGS = {
    "telegram_bot_token": "",
    "telegram_webhook_secret": "",
    "deepseek_api_key": "",
}
PIN_INPUT_KEYS = {"panel_owner_pin_code", "panel_guest_pin_code"}
ALL_SETTINGS = {**DEFAULT_SETTINGS, **SECRET_SETTINGS}
INTERNAL_SETTINGS = {
    "telegram_tenant_links": "{}",
    "telegram_consent_chat_ids": "[]",
    "ignored_lease_ids": "[]",
    "accepted_monthly_reports": "[]",
    "notified_monthly_reports": "[]",
    "owner_due_digest_dates": "[]",
    "owner_payment_digest_dates": "[]",
    "processed_move_out_notifications": "[]",
}
BOOLEAN_SETTINGS = {
    "notifications_enabled",
    "ai_enabled",
    "ai_tenant_free_text_enabled",
    "ai_supervisor_enabled",
    "hermes_briefing_enabled",
    "hermes_auto_level_one_enabled",
    "hermes_auto_template_reminders",
}
ENV_SETTING_KEYS = {
    "app_base_url": "APP_BASE_URL",
    "telegram_owner_chat_id": "TELEGRAM_OWNER_CHAT_ID",
    "telegram_bot_token": "TELEGRAM_BOT_TOKEN",
    "telegram_webhook_secret": "TELEGRAM_WEBHOOK_SECRET",
    "ai_enabled": "AI_ENABLED",
    "ai_tenant_free_text_enabled": "AI_TENANT_FREE_TEXT_ENABLED",
    "ai_supervisor_enabled": "AI_SUPERVISOR_ENABLED",
    "ai_supervisor_cadence": "AI_SUPERVISOR_CADENCE",
    "ai_supervisor_weekday": "AI_SUPERVISOR_WEEKDAY",
    "ai_supervisor_time": "AI_SUPERVISOR_TIME",
    "ai_supervisor_model": "AI_SUPERVISOR_MODEL",
    "ai_supervisor_max_tokens": "AI_SUPERVISOR_MAX_TOKENS",
    "ai_action_confirmation_ttl_hours": "AI_ACTION_CONFIRMATION_TTL_HOURS",
    "ai_daily_call_limit": "AI_DAILY_CALL_LIMIT",
    "ai_max_output_tokens": "AI_MAX_OUTPUT_TOKENS",
    "ai_usd_rub_rate": "AI_USD_RUB_RATE",
    "deepseek_model": "DEEPSEEK_MODEL",
    "ai_monthly_budget_rub": "AI_MONTHLY_BUDGET_RUB",
    "hermes_briefing_enabled": "HERMES_BRIEFING_ENABLED",
    "hermes_auto_level_one_enabled": "HERMES_AUTO_LEVEL_ONE_ENABLED",
    "hermes_auto_template_reminders": "HERMES_AUTO_TEMPLATE_REMINDERS",
    "ai_tenant_mode": "AI_TENANT_MODE",
    "deepseek_api_key": "DEEPSEEK_API_KEY",
}
VALID_REMINDER_CADENCES = {"twice_daily", "daily_evening", "every_two_days", "never"}
VALID_AI_SUPERVISOR_CADENCES = {"daily", "weekdays", "weekly"}
AI_INTEGER_SETTING_LIMITS = {
    "ai_supervisor_weekday": (0, 6, "День автоаудита"),
    "ai_supervisor_max_tokens": (100, 4000, "Лимит ответа автоаудита"),
    "ai_action_confirmation_ttl_hours": (1, 168, "Срок подтверждения действия"),
    "ai_daily_call_limit": (0, 100000, "Дневной лимит AI-запросов"),
    "ai_max_output_tokens": (100, 8000, "Общий лимит ответа AI"),
    "hermes_mass_action_pin_threshold": (1, 10000, "Порог массового критического действия"),
    "ai_feature_owner_chat_daily_limit": (0, 100000, "Лимит owner chat"),
    "ai_feature_tenant_chat_daily_limit": (0, 100000, "Лимит tenant chat"),
    "ai_feature_case_details_daily_limit": (0, 100000, "Лимит case details"),
    "ai_feature_skill_builder_daily_limit": (0, 100000, "Лимит skill builder"),
    "ai_feature_deep_audit_daily_limit": (0, 100000, "Лимит deep audit"),
    "ai_feature_key_test_daily_limit": (0, 100000, "Лимит проверки ключа"),
    "ai_output_chars_daily_briefing": (200, 2000, "Лимит ежедневной сводки"),
    "ai_output_chars_case_details": (300, 4000, "Лимит подробностей кейса"),
    "ai_output_chars_owner_chat": (200, 4000, "Лимит owner chat"),
    "ai_output_chars_tenant_chat": (100, 2000, "Лимит tenant chat"),
    "ai_output_chars_skill_proposal": (300, 3000, "Лимит карточки навыка"),
}
AI_DECIMAL_SETTING_LIMITS = {
    "ai_monthly_budget_rub": (0.0, 10_000_000.0, "Месячный AI-бюджет"),
    "ai_usd_rub_rate": (1.0, 1000.0, "Курс доллара для оценки AI"),
}
AI_INSTRUCTION_SETTING_KEYS = {
    "ai_owner_instructions",
    "ai_tenant_instructions",
    "ai_audit_instructions",
}
REMINDER_CADENCE_KEYS = {
    "message_rent_due": "automation_rent_due_cadence",
    "message_rent_overdue": "automation_rent_overdue_cadence",
    "message_utility_bill": "automation_utility_cadence",
}
LEASE_AUTOMATION_TEMPLATES = ("message_rent_due", "message_rent_overdue", "message_utility_bill")
EXPENSE_FUND_RECEIPT_MARK = "PAYMENT_RECEIPT:"
UTILITY_ADVANCE_CHANNEL = "utility_advance"
UTILITY_ADVANCE_SOURCE = "utility_advance"
UTILITY_LINE_USAGE = "usage"
UTILITY_LINE_ADVANCE = "advance"
UTILITY_BILL_USAGE = "utility"
UTILITY_BILL_ADVANCE = "advance"
UTILITY_ADVANCE_CURRENT_CAP_DEFAULT = 2.0
UTILITY_ADVANCE_CURRENT_CAP_SHOULDER = 5.0
UTILITY_ADVANCE_CURRENT_CAP_WINTER = 8.0
UTILITY_ADVANCE_CURRENT_CAP_BUFFER = 1000.0
REMINDER_WORKER_STARTED = False
REMINDER_WORKER_INTERVAL_SECONDS = 900
TELEGRAM_PROCESSED_UPDATE_LIMIT = 500
TELEGRAM_UPDATE_LOCK = threading.Lock()
TELEGRAM_QUEUE_MAX = max(16, int(os.environ.get("TELEGRAM_QUEUE_MAX", "256") or 256))
TELEGRAM_WORKER_COUNT = max(1, min(8, int(os.environ.get("TELEGRAM_WORKER_COUNT", "2") or 2)))
TELEGRAM_UPDATE_QUEUE: queue.Queue[tuple[dict[str, Any], float | None]] = queue.Queue(maxsize=TELEGRAM_QUEUE_MAX)
TELEGRAM_WORKER_STOP = threading.Event()
TELEGRAM_WORKERS: list[threading.Thread] = []
TELEGRAM_RATE_LOCK = threading.Lock()
TELEGRAM_RATE_BUCKETS: dict[str, deque[float]] = {}
LOCAL_TZ = ZoneInfo("Asia/Novosibirsk")
DEFAULT_FALLBACK_KEYS = {
    "ip_recipient_name",
    "ip_recipient_inn",
    "ip_recipient_ogrnip",
    "ip_recipient_account",
    "ip_recipient_bank",
    "ip_recipient_bik",
    "ip_recipient_correspondent_account",
    "ip_recipient_bank_inn",
    "ip_recipient_bank_kpp",
}
AI_MODEL_PRICES_USD_PER_M = {
    "deepseek-v4-flash": (0.14, 0.28),
    "deepseek-v4-pro": (0.435, 0.87),
}
AI_DEFAULT_USD_RUB_RATE = 100.0

REPORT_START_MONTH = date(2026, 1, 1)
PANEL_AUTH_COOKIE = "rental_manager_panel_session"
PANEL_CSRF_COOKIE = "rental_manager_csrf"
PANEL_AUTH_MAX_AGE_SECONDS = 60 * 60 * 24 * 30
PANEL_ALLOWED_GUEST_TAB_PATHS = {"/api/bootstrap", "/api/app-state"}
MOBILE_APK_PATH = ROOT_DIR / "android" / "RentalManager" / "build" / "rental-manager-mobile.apk"
MONTH_NAMES = [
    "январь",
    "февраль",
    "март",
    "апрель",
    "май",
    "июнь",
    "июль",
    "август",
    "сентябрь",
    "октябрь",
    "ноябрь",
    "декабрь",
]


def startup() -> None:
    started = time_module.perf_counter()
    status = "ok"
    try:
        providers = ",".join(provider.value for provider in provider_chain())
    except AiProviderConfigError as exc:
        providers = f"invalid:{exc}"
    runtime_log(
        "BOOT",
        "app_marker="
        f"{APP_BUILD_MARKER} ai_providers={providers} "
        f"deepseek_env_key_configured={str(bool(os.environ.get('DEEPSEEK_API_KEY', '').strip())).lower()}",
    )
    install_domain_event_listeners()
    init_db()
    with SessionLocal() as session:
        ensure_database_schema(session)
        seeded_release = seed_release_baseline_if_empty(session)
        if not seeded_release:
            seed_if_empty(session)
        ensure_runtime_defaults(session)
        validate_production_security(session)
        generate_rent_charges(session)
        reconcile_operational_cases(session)
        reconcile_owner_commitments(session)
        session.commit()
    queue_startup_maintenance()
    queue_runtime_telegram_webhook_refresh()
    record_background_event(
        "startup",
        int((time_module.perf_counter() - started) * 1000),
        status=status,
        detail={"providers": providers},
    )


def reminder_worker_loop() -> None:
    # Первый сон нужен, чтобы редеплой не отправлял пачку сообщений в ту же секунду.
    time_module.sleep(60)
    while True:
        started = time_module.perf_counter()
        summary: dict[str, Any] = {}
        status = "ok"
        try:
            with SessionLocal() as session:
                generate_rent_charges(session)
                summary = run_due_reminders(session)
                notify_available_monthly_reports(session)
                session.commit()
        except Exception as exc:
            status = "failed"
            runtime_log("REMINDERS", f"background worker failed error={exc!r}")
            summary = {"error": str(exc)[:300]}
            try:
                with SessionLocal() as event_session:
                    emit_domain_event(
                        event_session,
                        "automation_failed",
                        entity_type="ReminderWorker",
                        entity_id="scheduled",
                        source="reminder_worker",
                        payload={"error": f"{type(exc).__name__}: {exc}"[:1000]},
                        idempotency_key=f"automation:reminder-worker:{date.today().isoformat()}:{type(exc).__name__}",
                    )
                    event_session.commit()
            except Exception:
                pass
        finally:
            record_background_event(
                "reminder_worker",
                int((time_module.perf_counter() - started) * 1000),
                status=status,
                detail=summary,
            )
        time_module.sleep(REMINDER_WORKER_INTERVAL_SECONDS)


def start_reminder_worker() -> None:
    global REMINDER_WORKER_STARTED
    if REMINDER_WORKER_STARTED:
        return
    REMINDER_WORKER_STARTED = True
    thread = threading.Thread(target=reminder_worker_loop, name="rental-reminders", daemon=True)
    thread.start()


def startup_maintenance_background() -> None:
    started = time_module.perf_counter()
    status = "ok"
    detail: dict[str, Any] = {}
    try:
        with SessionLocal() as session:
            detail["owner_preferences_backfilled"] = backfill_owner_style_preferences(session)
            detail["monthly_report_notifications"] = notify_available_monthly_reports(session)
            session.commit()
    except Exception as exc:
        status = "failed"
        detail["error"] = repr(exc)[:240]
        runtime_log("BOOT", f"startup maintenance failed error={exc!r}")
    finally:
        record_background_event(
            "startup_maintenance",
            int((time_module.perf_counter() - started) * 1000),
            status=status,
            detail=detail,
        )


def queue_startup_maintenance() -> None:
    thread = threading.Thread(
        target=startup_maintenance_background,
        name="startup-maintenance",
        daemon=True,
    )
    thread.start()


SCHEMA_ADDITIONS = {
    "utility_services": [
        ("provider_reading_due_day", "INTEGER NOT NULL DEFAULT 20"),
        ("provider_due_day", "INTEGER NOT NULL DEFAULT 24"),
        ("resident_due_days", "INTEGER NOT NULL DEFAULT 7"),
    ],
    "utility_bills": [
        ("bill_type", "VARCHAR(40) NOT NULL DEFAULT 'utility'"),
    ],
    "utility_bill_lines": [
        ("line_type", "VARCHAR(40) NOT NULL DEFAULT 'usage'"),
        ("metadata_json", "TEXT NOT NULL DEFAULT '{}'"),
    ],
    "agent_action_proposals": [
        ("error_text", "TEXT NOT NULL DEFAULT ''"),
        ("case_id", "INTEGER"),
        ("safety_level", "INTEGER NOT NULL DEFAULT 2"),
        ("idempotency_key", "VARCHAR(160) NOT NULL DEFAULT ''"),
        ("group_key", "VARCHAR(160) NOT NULL DEFAULT ''"),
        ("validation_hash", "VARCHAR(128) NOT NULL DEFAULT ''"),
        ("confirmed_by", "VARCHAR(80) NOT NULL DEFAULT ''"),
    ],
}


def ensure_database_schema(session: Session) -> None:
    bind = session.get_bind()
    if bind is not None:
        Base.metadata.create_all(bind=bind)
    ensure_schema_additions(session)


def ensure_schema_additions(session: Session) -> None:
    bind = session.get_bind()
    dialect = bind.dialect.name if bind else ""

    if dialect == "sqlite":
        for table_name, additions in SCHEMA_ADDITIONS.items():
            columns = {row[1] for row in session.execute(text(f"PRAGMA table_info({table_name})")).all()}
            for column_name, ddl in additions:
                if column_name not in columns:
                    session.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {ddl}"))
        session.commit()
        return

    for table_name, additions in SCHEMA_ADDITIONS.items():
        for column_name, ddl in additions:
            exists = session.execute(
                text(
                    """
                    SELECT 1
                    FROM information_schema.columns
                    WHERE table_name = :table_name
                      AND column_name = :column_name
                    LIMIT 1
                    """
                ),
                {"table_name": table_name, "column_name": column_name},
            ).first()
            if not exists:
                session.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {ddl}"))
    session.commit()


PERFORMANCE_INDEXES = [
    ("ix_rm_leases_active_apartment", "leases (active, apartment_id)"),
    ("ix_rm_rent_charges_lease_due", "rent_charges (lease_id, due_date)"),
    ("ix_rm_rent_charges_due_status", "rent_charges (due_date, status)"),
    ("ix_rm_payment_receipts_status_paid", "payment_receipts (status, paid_at)"),
    ("ix_rm_payment_receipts_lease_paid", "payment_receipts (lease_id, paid_at)"),
    ("ix_rm_payment_receipts_rent_status", "payment_receipts (rent_charge_id, status)"),
    ("ix_rm_payment_receipts_utility_status", "payment_receipts (utility_line_id, status)"),
    ("ix_rm_message_logs_rent_created", "message_logs (rent_charge_id, created_at)"),
    ("ix_rm_message_logs_utility_created", "message_logs (utility_line_id, created_at)"),
    ("ix_rm_message_logs_lease_created", "message_logs (lease_id, created_at)"),
    ("ix_rm_message_logs_chat_created", "message_logs (recipient_chat_id, created_at)"),
    ("ix_rm_meter_readings_meter_date", "meter_readings (meter_id, reading_date)"),
    ("ix_rm_meters_service_scope_active", "meters (service_id, scope, active)"),
    ("ix_rm_utility_bills_service_period", "utility_bills (service_id, period_start, period_end)"),
    ("ix_rm_utility_bills_status_provider", "utility_bills (status, provider_paid)"),
    ("ix_rm_utility_bill_lines_bill", "utility_bill_lines (bill_id)"),
    ("ix_rm_utility_bill_lines_lease_due", "utility_bill_lines (lease_id, due_date)"),
    ("ix_rm_utility_bill_lines_status_due", "utility_bill_lines (status, due_date)"),
    ("ix_rm_utility_bill_lines_type_due", "utility_bill_lines (line_type, due_date)"),
    ("ix_rm_utility_advance_ledger_apartment", "utility_advance_ledger (apartment_id, created_at)"),
    ("ix_rm_utility_advance_ledger_line", "utility_advance_ledger (utility_line_id)"),
    ("ix_rm_utility_advance_ledger_receipt", "utility_advance_ledger (payment_receipt_id)"),
    ("ix_rm_manual_debts_active_lease_due", "manual_debts (active, lease_id, due_date)"),
    ("ix_rm_expenses_source_compensation", "expenses (source_funds, compensation_status)"),
    ("ix_rm_ai_conversations_chat_role", "ai_conversations (chat_id, role, status)"),
    ("ix_rm_ai_messages_conversation_created", "ai_messages (conversation_id, created_at)"),
    ("ix_rm_ai_usage_daily_date", "ai_usage_daily (usage_date, provider, model)"),
    ("ix_rm_payment_situations_lease_status", "payment_situations (lease_id, status, mode)"),
    ("ix_rm_payment_situations_kind_reference", "payment_situations (kind, reference_id)"),
]


def ensure_performance_indexes(session: Session) -> None:
    for name, target in PERFORMANCE_INDEXES:
        session.execute(text(f"CREATE INDEX IF NOT EXISTS {name} ON {target}"))
    session.commit()


@app.get("/")
def index() -> FileResponse:
    return FileResponse(ROOT_DIR / "static" / "index.html")


@app.get("/healthz")
def healthz() -> dict[str, str]:
    return {
        "status": "ok",
        "build_marker": APP_BUILD_MARKER,
        "ai_provider": "deepseek",
    }


@app.get("/health")
def health() -> dict[str, str]:
    return healthz()


@app.get("/mobile-app.apk")
def mobile_app_apk() -> FileResponse:
    if not MOBILE_APK_PATH.exists():
        raise HTTPException(status_code=404, detail="APK ещё не собран")
    return FileResponse(
        MOBILE_APK_PATH,
        media_type="application/vnd.android.package-archive",
        filename="rental-manager-mobile.apk",
    )


def is_public_path(path: str) -> bool:
    return path in {
        "/",
        "/health",
        "/healthz",
        "/mobile-app.apk",
        "/api/auth/status",
        "/api/auth/pin",
        "/api/integrations/telegram/webhook",
    } or path.startswith("/static/")


def guest_api_allowed(method: str, path: str) -> bool:
    normalized_method = method.upper()
    if normalized_method not in {"GET", "HEAD"}:
        return False
    if path in PANEL_ALLOWED_GUEST_TAB_PATHS:
        return True
    return path.startswith("/api/reports/")


@app.middleware("http")
async def performance_monitor_middleware(request: Request, call_next):
    started = time_module.perf_counter()
    response = None
    status_code = 500
    try:
        response = await call_next(request)
        status_code = response.status_code
        response.headers["X-Rental-Perf-Ms"] = str(int((time_module.perf_counter() - started) * 1000))
        return response
    finally:
        duration_ms = int((time_module.perf_counter() - started) * 1000)
        record_perf_request(
            method=request.method,
            path=request.url.path,
            status_code=status_code,
            duration_ms=duration_ms,
            user_agent=request.headers.get("user-agent", ""),
        )


@app.middleware("http")
async def panel_auth_middleware(request: Request, call_next):
    path = request.url.path
    if is_public_path(path):
        return await call_next(request)
    if not path.startswith("/api/"):
        return await call_next(request)
    with SessionLocal() as auth_session:
        token = str(request.cookies.get(PANEL_AUTH_COOKIE) or "").strip()
        panel_session = find_session(auth_session, token)
        role = panel_session.role if panel_session else None
        if role and request.method.upper() in {"POST", "PUT", "PATCH", "DELETE"}:
            csrf_token = request.headers.get("X-CSRF-Token", "")
            if not csrf_is_valid(panel_session, csrf_token):
                return JSONResponse(status_code=403, content={"detail": "CSRF-проверка не пройдена."})
    request.state.panel_role = role
    if not role:
        return JSONResponse(status_code=401, content={"detail": "Введите PIN-код для доступа в пульт."})
    if role != "owner" and not guest_api_allowed(request.method, path):
        return JSONResponse(status_code=403, content={"detail": "Гостевой PIN-код даёт только обзор и скачивание отчётов."})
    return await call_next(request)


def current_month_range() -> tuple[date, date]:
    today = date.today()
    start = today.replace(day=1)
    if today.month == 12:
        end = date(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(today.year, today.month + 1, 1) - timedelta(days=1)
    return start, end


def month_range(year: int, month: int) -> tuple[date, date]:
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(year, month + 1, 1) - timedelta(days=1)
    return start, end


def iter_generated_report_months(today: date | None = None) -> list[tuple[int, int]]:
    today = today or date.today()
    last_complete = today.replace(day=1) - timedelta(days=1)
    if last_complete < REPORT_START_MONTH:
        return []

    months: list[tuple[int, int]] = []
    year = REPORT_START_MONTH.year
    month = REPORT_START_MONTH.month
    while date(year, month, 1) <= last_complete.replace(day=1):
        months.append((year, month))
        if month == 12:
            year += 1
            month = 1
        else:
            month += 1
    return months


def iter_dashboard_report_entries(today: date | None = None) -> list[dict[str, Any]]:
    today = today or date.today()
    entries = [
        {"year": year, "month": month, "kind": "full"}
        for year, month in iter_generated_report_months(today)
    ]
    if today.day >= 25 and today.replace(day=1) >= REPORT_START_MONTH:
        entries.append({"year": today.year, "month": today.month, "kind": "preliminary"})
    return entries


def monthly_report_key(year: int, month: int, kind: str) -> str:
    return f"{kind}:{year:04d}-{month:02d}"


def accepted_monthly_report_keys(session: Session) -> set[str]:
    raw = get_internal_json(session, "accepted_monthly_reports", [])
    return {str(item) for item in raw if isinstance(item, str)} if isinstance(raw, list) else set()


def mark_monthly_report_accepted(session: Session, year: int, month: int, kind: str = "full") -> None:
    keys = accepted_monthly_report_keys(session)
    keys.add(monthly_report_key(year, month, kind))
    set_internal_json(session, "accepted_monthly_reports", sorted(keys))


def notified_monthly_report_keys(session: Session) -> set[str]:
    raw = get_internal_json(session, "notified_monthly_reports", [])
    return {str(item) for item in raw if isinstance(item, str)} if isinstance(raw, list) else set()


def setting_bool_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def ensure_runtime_defaults(session: Session) -> None:
    changed = False
    for role, legacy_key in LEGACY_PIN_SETTING_KEYS.items():
        legacy = session.get(AppSetting, legacy_key)
        target_key = PIN_SETTING_KEYS[role]
        target = session.get(AppSetting, target_key)
        if legacy:
            raw_pin = str(legacy.value or "").strip()
            migrated_hash = ""
            if is_pin_hash(raw_pin):
                migrated_hash = raw_pin
            elif raw_pin and not pin_is_compromised(raw_pin):
                migrated_hash = hash_pin(raw_pin)
            if migrated_hash and not target:
                session.add(AppSetting(key=target_key, value=migrated_hash))
            session.delete(legacy)
            changed = True
    if not session.get(AppSetting, "notifications_enabled"):
        session.add(AppSetting(key="notifications_enabled", value="0"))
        changed = True
    if not session.get(AppSetting, "notification_cutoff_date"):
        session.add(AppSetting(key="notification_cutoff_date", value=date.today().isoformat()))
        changed = True
    deepseek_model = session.get(AppSetting, "deepseek_model")
    legacy_model = session.get(AppSetting, "hermes_model_default")
    if not deepseek_model and legacy_model:
        model = "deepseek-v4-pro" if "pro" in str(legacy_model.value or "").lower() else "deepseek-v4-flash"
        session.add(AppSetting(key="deepseek_model", value=model))
        changed = True
    legacy_supervisor_hour = session.get(AppSetting, "ai_supervisor_hour")
    if not session.get(AppSetting, "ai_supervisor_time") and (
        legacy_supervisor_hour or os.environ.get("AI_SUPERVISOR_HOUR", "").strip()
    ):
        raw_hour = os.environ.get("AI_SUPERVISOR_HOUR", "") or str(legacy_supervisor_hour.value or "10")
        try:
            hour = min(23, max(0, int(raw_hour)))
        except ValueError:
            hour = 10
        session.add(AppSetting(key="ai_supervisor_time", value=f"{hour:02d}:00"))
        changed = True
    for legacy_key in ("hermes_api_base_url", "hermes_model_default", "hermes_model_audit", "hermes_api_key"):
        legacy_setting = session.get(AppSetting, legacy_key)
        if legacy_setting:
            session.delete(legacy_setting)
            changed = True
    for legacy_key in ("ai_supervisor_hour", "ai_dialog_reminders_enabled"):
        legacy_setting = session.get(AppSetting, legacy_key)
        if legacy_setting:
            session.delete(legacy_setting)
            changed = True
    old_owner_receipt_alert = "Новый чек от {tenant_name}. Квартира: {apartment}. Сумма: {amount}. Канал: {channel}. Статус: {receipt_status}. {receipt_summary}"
    owner_receipt_setting = session.get(AppSetting, "message_owner_receipt_alert")
    if owner_receipt_setting and owner_receipt_setting.value == old_owner_receipt_alert:
        owner_receipt_setting.value = DEFAULT_SETTINGS["message_owner_receipt_alert"]
        changed = True
    old_message_defaults = {
        "message_rent_due": "Здравствуйте. Напоминаю: по квартире {apartment} сегодня ожидается оплата аренды. ИП: {ip_due}. Перевод: {personal_due}. Итого: {total_due}.",
        "message_rent_overdue": "Здравствуйте. По квартире {apartment} сейчас просрочена аренда. Долг: {debt}. ИП: {ip_due}. Перевод: {personal_due}.",
        "message_utility_bill": "Здравствуйте. Сформированы счета по коммунальным платежам.\n{utility_debt_details}\n\nВсего: {utility_total}. Срок оплаты до {utility_due_date}. Оплата переводом на {personal_recipient_phone_text} {personal_recipient_bank_text}.\n\nОтправьте чеки в этот чат в виде документа, скриншоты больше не принимаются.\n(История платежей в приложении банка -> чек об операции -> сохранить или отправить)",
        "message_all_debts": "Здравствуйте. Уведомляем вас о наличии следующих задолженностей:\n{all_debts_breakdown}\n\nОтправьте чеки в этот чат в виде документа, скриншоты больше не принимаются.\n(История платежей в приложении банка -> чек об операции -> сохранить или отправить)",
        "message_receipt_received": "Чек получил и сохранил. Если всё совпало, я это отмечу. Если нет, передам владельцу на ручную проверку.",
        "message_receipt_review": "Чек получил, но там есть вопросы. Я уже отправил его владельцу на проверку.",
    }
    for key, old_value in old_message_defaults.items():
        setting = session.get(AppSetting, key)
        if setting and setting.value == old_value:
            setting.value = str(DEFAULT_SETTINGS[key])
            changed = True
    if changed:
        session.flush()


def production_mode() -> bool:
    return os.environ.get("RENTAL_MANAGER_ENV", "development").strip().lower() in {"prod", "production"}


def validate_production_security(session: Session) -> None:
    if not production_mode():
        return
    owner_hash = configured_environment_pin_hash("owner") or get_setting_value(session, PIN_SETTING_KEYS["owner"])
    if not owner_hash:
        raise RuntimeError("Production requires PANEL_OWNER_PIN or a configured owner PIN hash.")
    if not os.environ.get("TELEGRAM_WEBHOOK_SECRET", "").strip() and get_setting_value(session, "telegram_bot_token"):
        raise RuntimeError("Production Telegram integration requires TELEGRAM_WEBHOOK_SECRET.")


def get_setting_value(session: Session, key: str) -> str:
    dynamic_key = key.startswith("lease_") and "_cadence_" in key
    pin_hash_key = key in PIN_SETTING_KEYS.values()
    if key not in ALL_SETTINGS and key not in INTERNAL_SETTINGS and not dynamic_key and not pin_hash_key:
        return ""
    env_name = ENV_SETTING_KEYS.get(key)
    if env_name and env_name in os.environ:
        env_value = str(os.environ.get(env_name) or "")
        if env_value or key not in SECRET_SETTINGS:
            return env_value
    row = session.get(AppSetting, key)
    if row:
        value = row.value
        return decrypt_secret(value) if key in SECRET_SETTINGS else value
    if key in ALL_SETTINGS:
        return str(ALL_SETTINGS[key])
    if dynamic_key:
        return ""
    if pin_hash_key:
        return ""
    return str(INTERNAL_SETTINGS[key])


def get_internal_json(session: Session, key: str, fallback: Any) -> Any:
    raw = get_setting_value(session, key)
    try:
        return json.loads(str(raw))
    except (TypeError, json.JSONDecodeError):
        return fallback


def set_internal_json(session: Session, key: str, value: Any) -> None:
    if key not in INTERNAL_SETTINGS:
        return
    setting = session.get(AppSetting, key)
    if not setting:
        setting = AppSetting(key=key)
        session.add(setting)
    setting.value = json.dumps(value, ensure_ascii=False)


def mark_telegram_update_seen(session: Session, update_id: Any) -> bool:
    if update_id in {None, ""}:
        return True
    try:
        with session.begin_nested():
            session.add(ProcessedTelegramUpdate(update_id=str(update_id)))
            session.flush()
    except IntegrityError:
        return False
    return True


def get_settings(session: Session) -> dict[str, str | bool]:
    settings: dict[str, str | bool] = {}
    for key in DEFAULT_SETTINGS:
        raw = get_setting_value(session, key)
        if key in DEFAULT_FALLBACK_KEYS and raw == "":
            raw = str(DEFAULT_SETTINGS[key])
        settings[key] = setting_bool_value(raw) if key in BOOLEAN_SETTINGS else raw
    for key in SECRET_SETTINGS:
        settings[f"{key}_configured"] = bool(get_setting_value(session, key))
    for role, input_key in (("owner", "panel_owner_pin_code"), ("guest", "panel_guest_pin_code")):
        settings[f"{input_key}_configured"] = bool(
            configured_environment_pin_hash(role) or get_setting_value(session, PIN_SETTING_KEYS[role])
        )
    return settings


def normalize_ai_setting(key: str, value: Any) -> str:
    if key in {"deepseek_model", "ai_supervisor_model"}:
        try:
            return normalize_deepseek_model(str(value))
        except AiProviderConfigError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
    if key == "ai_supervisor_cadence":
        cadence = str(value or "").strip().lower()
        if cadence not in VALID_AI_SUPERVISOR_CADENCES:
            raise HTTPException(status_code=400, detail="Периодичность автоаудита задана некорректно.")
        return cadence
    if key == "ai_tenant_mode":
        mode = str(value or "").strip().lower()
        if mode not in {"auto", "draft_only", "finance_only"}:
            raise HTTPException(400, "Режим Tenant AI должен быть auto, draft_only или finance_only")
        return mode
    if key == "ai_supervisor_time":
        match = re.fullmatch(r"([01]\d|2[0-3]):([0-5]\d)", str(value or "").strip())
        if not match:
            raise HTTPException(status_code=400, detail="Время автоаудита должно быть в формате ЧЧ:ММ.")
        return f"{match.group(1)}:{match.group(2)}"
    if key in AI_INTEGER_SETTING_LIMITS:
        minimum, maximum, label = AI_INTEGER_SETTING_LIMITS[key]
        try:
            number = int(str(value).strip())
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=f"{label}: требуется целое число.") from exc
        if number < minimum or number > maximum:
            raise HTTPException(status_code=400, detail=f"{label}: допустимо от {minimum} до {maximum}.")
        return str(number)
    if key in AI_DECIMAL_SETTING_LIMITS:
        minimum, maximum, label = AI_DECIMAL_SETTING_LIMITS[key]
        try:
            number = float(str(value).strip().replace(",", "."))
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=f"{label}: требуется число.") from exc
        if number < minimum or number > maximum:
            raise HTTPException(status_code=400, detail=f"{label}: допустимо от {minimum:g} до {maximum:g}.")
        return f"{number:g}"
    if key in AI_INSTRUCTION_SETTING_KEYS:
        instructions = str(value or "").strip()
        if len(instructions) > 6000:
            raise HTTPException(status_code=400, detail="Дополнительные инструкции AI не должны превышать 6000 символов.")
        return instructions
    return str(value)


def save_settings(session: Session, payload: dict[str, Any], preserve_panel_token: str = "") -> dict[str, str | bool]:
    allowed = set(ALL_SETTINGS) | PIN_INPUT_KEYS
    pin_changed = False
    for key, value in payload.items():
        if key not in allowed:
            continue
        if key in SECRET_SETTINGS and value in {"", None}:
            continue
        if key in PIN_INPUT_KEYS and value in {"", None}:
            continue
        if key in PIN_INPUT_KEYS:
            role = "owner" if key == "panel_owner_pin_code" else "guest"
            target_key = PIN_SETTING_KEYS[role]
            setting = session.get(AppSetting, target_key)
            if not setting:
                setting = AppSetting(key=target_key)
                session.add(setting)
            setting.value = hash_pin(str(value))
            pin_changed = True
            continue
        setting = session.get(AppSetting, key)
        if not setting:
            setting = AppSetting(key=key)
            session.add(setting)
        if key in BOOLEAN_SETTINGS:
            setting.value = "1" if setting_bool_value(value) else "0"
        elif key == "notification_cutoff_date" and not value:
            setting.value = date.today().isoformat()
        elif key == "telegram_bot_token":
            try:
                setting.value = encrypt_secret(normalize_bot_token(str(value)))
            except RuntimeError as exc:
                raise HTTPException(status_code=503, detail=str(exc)) from exc
        elif key.startswith("ai_") or key == "deepseek_model":
            setting.value = normalize_ai_setting(key, value)
        elif key in SECRET_SETTINGS:
            try:
                setting.value = encrypt_secret(str(value))
            except RuntimeError as exc:
                raise HTTPException(status_code=503, detail=str(exc)) from exc
        else:
            setting.value = str(value)
    if pin_changed:
        revoke_other_sessions(session, preserve_panel_token)
    session.commit()
    return get_settings(session)


def public_settings(session: Session) -> dict[str, str | bool]:
    settings = get_settings(session)
    return {
        "color_palette": settings.get("color_palette", DEFAULT_SETTINGS["color_palette"]),
    }


def panel_session_record(session: Session, token: str) -> dict[str, Any] | None:
    row = find_session(session, token)
    if not row:
        return None
    return {"role": row.role, "created_at": row.created_at, "last_seen_at": row.last_seen_at}


def panel_role_from_request(request: Request, session: Session) -> str | None:
    token = str(request.cookies.get(PANEL_AUTH_COOKIE) or "").strip()
    if not token:
        return None
    row = find_session(session, token)
    if not row:
        return None
    return str(row.role or "").strip().lower() or None


def issue_panel_session(
    session: Session,
    role: str,
    user_agent: str = "",
) -> str:
    return issue_session(session, role, user_agent).token


def revoke_panel_session(session: Session, token: str) -> None:
    revoke_session(session, token)


def panel_role_for_pin(session: Session, pin_code: str) -> str | None:
    pin_value = str(pin_code or "").strip()
    if not pin_value:
        return None
    for role in ("owner", "guest"):
        pin_hash = configured_environment_pin_hash(role) or get_setting_value(session, PIN_SETTING_KEYS[role])
        if verify_pin(pin_hash, pin_value):
            return role
    return None


DB_EXPORT_FORMAT = "rental-manager-db-export"
DB_EXPORT_VERSION = 2
DB_IMPORT_CONFIRM_TEXT = "ИМПОРТ"
DB_BACKUP_CONFIRM_TEXT = "ЗАЩИЩЕННАЯ КОПИЯ"
EXPORT_ALWAYS_EXCLUDED_TABLES = {"panel_sessions", "panel_login_attempts", "processed_telegram_updates"}
EXPORT_SECRET_SETTING_KEYS = set(SECRET_SETTINGS) | set(PIN_SETTING_KEYS.values())


def table_model_map() -> dict[str, Any]:
    return {
        "rental_objects": RentalObject,
        "apartments": Apartment,
        "tenants": Tenant,
        "leases": Lease,
        "rent_charges": RentCharge,
        "payment_receipts": PaymentReceipt,
        "message_logs": MessageLog,
        "ai_conversations": AiConversation,
        "ai_messages": AiMessage,
        "ai_action_logs": AiActionLog,
        "agent_memories": AgentMemory,
        "agent_action_proposals": AgentActionProposal,
        "agent_tenant_states": AgentTenantState,
        "payment_situations": PaymentSituation,
        "agent_tasks": AgentTask,
        "ai_usage_daily": AiUsageDaily,
        "utility_services": UtilityService,
        "meters": Meter,
        "meter_readings": MeterReading,
        "tariffs": Tariff,
        "utility_bills": UtilityBill,
        "utility_bill_lines": UtilityBillLine,
        "manual_debts": ManualDebt,
        "expenses": Expense,
        "app_settings": AppSetting,
    }


def export_scalar(value: Any) -> Any:
    if isinstance(value, datetime):
        return {"__type__": "datetime", "value": value.isoformat()}
    if isinstance(value, date):
        return {"__type__": "date", "value": value.isoformat()}
    return value


def import_scalar(column, value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, dict) and value.get("__type__") == "datetime":
        return datetime.fromisoformat(str(value.get("value") or ""))
    if isinstance(value, dict) and value.get("__type__") == "date":
        return date.fromisoformat(str(value.get("value") or ""))
    try:
        python_type = getattr(column.type, "python_type", None)
    except NotImplementedError:
        python_type = None
    if python_type is bool:
        return bool(value)
    if python_type is int:
        return int(value)
    if python_type is float:
        return float(value)
    return value


def current_database_snapshot(session: Session, *, include_secrets: bool = False) -> dict[str, Any]:
    tables: dict[str, list[dict[str, Any]]] = {}
    counts: dict[str, int] = {}
    for table in Base.metadata.sorted_tables:
        if table.name in EXPORT_ALWAYS_EXCLUDED_TABLES:
            continue
        rows = session.execute(select(table)).mappings().all()
        if table.name == "app_settings" and not include_secrets:
            rows = [row for row in rows if str(row.get("key") or "") not in EXPORT_SECRET_SETTING_KEYS]
        if table.name == "app_settings" and include_secrets:
            rows = [
                row
                for row in rows
                if str(row.get("key") or "") not in SECRET_SETTINGS
                or secret_is_encrypted(str(row.get("value") or ""))
            ]
        serialized_rows = [
            {column.name: export_scalar(row[column.name]) for column in table.columns}
            for row in rows
        ]
        tables[table.name] = serialized_rows
        counts[table.name] = len(serialized_rows)
    return {
        "format": DB_EXPORT_FORMAT,
        "version": DB_EXPORT_VERSION,
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "database_url_hint": "postgres" if DATABASE_URL.startswith("postgresql") else "sqlite",
        "scope": "protected" if include_secrets else "safe",
        "tables": tables,
        "counts": counts,
    }


def parse_database_import_bytes(raw_bytes: bytes) -> dict[str, Any]:
    if not raw_bytes:
        raise HTTPException(400, "Файл импорта пуст.")
    try:
        payload = json.loads(raw_bytes.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise HTTPException(400, "Файл импорта должен быть JSON-экспортом из приложения.") from exc
    if not isinstance(payload, dict):
        raise HTTPException(400, "Файл импорта должен содержать JSON-объект верхнего уровня.")
    if payload.get("format") != DB_EXPORT_FORMAT:
        raise HTTPException(400, "Неузнаваемый формат импорта. Нужен экспорт именно из этого приложения.")
    if int(payload.get("version") or 0) not in {1, DB_EXPORT_VERSION}:
        raise HTTPException(400, "Версия формата импорта не поддерживается.")
    tables = payload.get("tables")
    if not isinstance(tables, dict):
        raise HTTPException(400, "В файле импорта нет блока tables.")
    return payload


def inspect_database_import_payload(payload: dict[str, Any]) -> dict[str, Any]:
    table_names = {table.name for table in Base.metadata.sorted_tables}
    tables = payload.get("tables") or {}
    unknown_tables = sorted(name for name in tables.keys() if name not in table_names)
    missing_tables = sorted(name for name in table_names if name not in tables)
    counts = {
        name: len(rows) if isinstance(rows, list) else 0
        for name, rows in tables.items()
        if name in table_names
    }
    total_rows = sum(counts.values())
    if total_rows <= 0:
        raise HTTPException(400, "В импорте нет записей для восстановления.")
    warnings: list[str] = []
    if missing_tables:
        warnings.append("В файле нет части таблиц: " + ", ".join(missing_tables))
    if unknown_tables:
        warnings.append("Есть лишние таблицы, они будут проигнорированы: " + ", ".join(unknown_tables))
    if total_rows < 10:
        warnings.append("Очень мало записей. Перед импортом лучше ещё раз проверить, тот ли это файл.")
    return {
        "format": payload.get("format"),
        "version": payload.get("version"),
        "exported_at": payload.get("exported_at"),
        "database_url_hint": payload.get("database_url_hint") or "",
        "counts": counts,
        "total_rows": total_rows,
        "missing_tables": missing_tables,
        "unknown_tables": unknown_tables,
        "warnings": warnings,
    }


def backup_export_path() -> Path:
    target_dir = ROOT_DIR / "data" / "backups"
    target_dir.mkdir(parents=True, exist_ok=True)
    return target_dir / f"db-backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"


def save_database_backup(session: Session) -> Path:
    snapshot = current_database_snapshot(session, include_secrets=True)
    target = backup_export_path()
    target.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    return target


def reset_database_sequences(session: Session) -> None:
    if not DATABASE_URL.startswith("postgresql"):
        return
    for table in Base.metadata.sorted_tables:
        integer_pk_columns = [column for column in table.columns if column.primary_key and getattr(column.type, "python_type", None) is int]
        for column in integer_pk_columns:
            session.execute(
                text(
                    f"SELECT setval(pg_get_serial_sequence('{table.name}', '{column.name}'), "
                    f"COALESCE((SELECT MAX({column.name}) FROM {table.name}), 1), "
                    f"COALESCE((SELECT MAX({column.name}) FROM {table.name}), 0) > 0)"
                )
            )


def apply_database_import_payload(session: Session, payload: dict[str, Any]) -> dict[str, Any]:
    table_lookup = {table.name: table for table in Base.metadata.sorted_tables}
    tables = payload.get("tables") or {}
    protected_scope = payload.get("scope") == "protected" and int(payload.get("version") or 0) >= 2
    preserved_secret_settings = session.execute(
        select(AppSetting.__table__).where(AppSetting.key.in_(EXPORT_SECRET_SETTING_KEYS))
    ).mappings().all()
    for table in reversed(Base.metadata.sorted_tables):
        if table.name in EXPORT_ALWAYS_EXCLUDED_TABLES:
            continue
        session.execute(table.delete())
    session.flush()
    inserted_counts: dict[str, int] = {}
    for table in Base.metadata.sorted_tables:
        rows = tables.get(table.name)
        if table.name == "app_settings":
            rows = [
                row
                for row in (rows or [])
                if isinstance(row, dict)
                and (
                    str(row.get("key") or "") not in EXPORT_SECRET_SETTING_KEYS
                    or (protected_scope and secret_is_encrypted(str(row.get("value") or "")))
                )
            ]
            if not protected_scope:
                rows.extend(dict(row) for row in preserved_secret_settings)
        if not isinstance(rows, list) or not rows:
            inserted_counts[table.name] = 0
            continue
        prepared_rows = []
        for row in rows:
            if not isinstance(row, dict):
                raise HTTPException(400, f"Таблица {table.name} содержит битую строку импорта.")
            prepared_rows.append(
                {
                    column.name: import_scalar(column, row.get(column.name))
                    for column in table.columns
                    if column.name in row
                }
            )
        session.execute(table_lookup[table.name].insert(), prepared_rows)
        inserted_counts[table.name] = len(prepared_rows)
    reset_database_sequences(session)
    return {"counts": inserted_counts, "total_rows": sum(inserted_counts.values())}


def telegram_token(session: Session) -> str:
    return normalize_bot_token(get_setting_value(session, "telegram_bot_token"))


def telegram_secret(session: Session) -> str:
    return get_setting_value(session, "telegram_webhook_secret").strip()


def telegram_owner_chat_id(session: Session) -> str:
    return get_setting_value(session, "telegram_owner_chat_id").strip()


def telegram_owner_ids(session: Session) -> set[str]:
    configured = {
        item.strip()
        for item in os.environ.get("OWNER_TELEGRAM_IDS", "").replace(";", ",").split(",")
        if item.strip()
    }
    owner_id = telegram_owner_chat_id(session)
    if owner_id:
        configured.add(owner_id)
    return configured


def app_base_url(session: Session) -> str:
    return get_setting_value(session, "app_base_url").strip().rstrip("/")


def normalize_app_base_url(value: Any) -> str:
    return str(value or "").strip().rstrip("/")


def save_app_base_url_if_changed(session: Session, value: Any) -> str:
    base_url = normalize_app_base_url(value)
    if not base_url.startswith("https://"):
        return app_base_url(session)
    current = app_base_url(session)
    if base_url == current:
        return current
    setting = session.get(AppSetting, "app_base_url")
    if not setting:
        setting = AppSetting(key="app_base_url")
        session.add(setting)
    setting.value = base_url
    session.flush()
    return base_url


def owner_chat_allowed(session: Session, chat_id: int | str) -> bool:
    return str(chat_id) in telegram_owner_ids(session)


def owner_message_allowed(session: Session, message: dict[str, Any]) -> bool:
    chat = message.get("chat") or {}
    sender = message.get("from") or {}
    return owner_chat_allowed(session, chat.get("id") or "") or owner_chat_allowed(session, sender.get("id") or "")


def notifications_enabled(session: Session) -> bool:
    return setting_bool_value(get_setting_value(session, "notifications_enabled"))


def notification_cutoff_date(session: Session) -> date:
    raw = get_setting_value(session, "notification_cutoff_date").strip()
    if raw:
        return parse_date(raw, date.today())
    return date.today()


def configured_notification_cutoff_date(session: Session) -> date | None:
    setting = session.get(AppSetting, "notification_cutoff_date")
    raw = (setting.value or "").strip() if setting else ""
    if not raw:
        return None
    try:
        return parse_date(raw, date.today())
    except ValueError:
        return None


def debt_visible_by_cutoff(due_date: date | None, cutoff: date | None) -> bool:
    if not cutoff or not due_date:
        return True
    return due_date >= cutoff


def allocation_cutoff_date(session: Session) -> date:
    cutoff = configured_notification_cutoff_date(session)
    if not cutoff or cutoff < RENT_GENERATION_START:
        return RENT_GENERATION_START
    return cutoff


def ignored_lease_ids(session: Session) -> set[int]:
    raw = get_setting_value(session, "ignored_lease_ids") or "[]"
    try:
        values = json.loads(raw)
    except json.JSONDecodeError:
        return set()
    if not isinstance(values, list):
        return set()
    result: set[int] = set()
    for value in values:
        try:
            result.add(int(value))
        except (TypeError, ValueError):
            continue
    return result


def save_ignored_lease_ids(session: Session, values: set[int]) -> None:
    setting = session.get(AppSetting, "ignored_lease_ids")
    if not setting:
        setting = AppSetting(key="ignored_lease_ids")
        session.add(setting)
    setting.value = json.dumps(sorted(values), ensure_ascii=False)


def lease_ignored(session: Session | None, lease_id: int | None) -> bool:
    if not session or not lease_id:
        return False
    lease = session.get(Lease, int(lease_id))
    return int(lease_id) in ignored_lease_ids(session) or bool(lease and IGNORE_LEASE_MARK in (lease.notes or ""))


def set_lease_ignored(session: Session, lease_id: int, ignored: bool) -> None:
    values = ignored_lease_ids(session)
    if ignored:
        values.add(int(lease_id))
    else:
        values.discard(int(lease_id))
    save_ignored_lease_ids(session, values)
    lease = session.get(Lease, lease_id)
    if lease:
        notes = lease.notes or ""
        if ignored and IGNORE_LEASE_MARK not in notes:
            lease.notes = (notes + "\n" + IGNORE_LEASE_MARK).strip()
        if not ignored and IGNORE_LEASE_MARK in notes:
            lease.notes = "\n".join(line for line in notes.splitlines() if line.strip() != IGNORE_LEASE_MARK).strip()


def operational_lease(session: Session, lease: Lease | None) -> bool:
    return bool(lease and lease.active and lease.apartment.active and not lease_ignored(session, lease.id))


def reminder_cadence(session: Session, template_key: str, lease_id: int | None = None) -> str:
    # Сначала проверяем per-lease настройку
    if lease_id:
        lease_cadence = get_lease_cadence(session, lease_id, template_key)
        if lease_cadence:
            return lease_cadence
    # Иначе используем глобальную настройку
    key = REMINDER_CADENCE_KEYS.get(template_key, "")
    if not key:
        return "twice_daily"
    value = str(get_setting_value(session, key) or "").strip()
    if value in VALID_REMINDER_CADENCES:
        return value
    return str(DEFAULT_SETTINGS[key])


def get_lease_cadence(session: Session, lease_id: int, template_key: str) -> str | None:
    """Получить индивидуальную настройку cadence для арендатора."""
    key = f"lease_{lease_id}_cadence_{template_key}"
    value = get_setting_value(session, key)
    if value and value.strip() in VALID_REMINDER_CADENCES:
        return value.strip()
    return None


def set_lease_cadence(session: Session, lease_id: int, template_key: str, cadence: str) -> None:
    """Сохранить индивидуальную настройку cadence для арендатора."""
    key = f"lease_{lease_id}_cadence_{template_key}"
    setting = session.get(AppSetting, key)
    if not setting:
        setting = AppSetting(key=key)
        session.add(setting)
    setting.value = cadence


def clear_lease_cadence(session: Session, lease_id: int, template_key: str) -> None:
    """Удалить индивидуальную настройку cadence (сбросить на значение по умолчанию)."""
    key = f"lease_{lease_id}_cadence_{template_key}"
    setting = session.get(AppSetting, key)
    if setting:
        session.delete(setting)


def get_lease_automation(session: Session, lease_id: int) -> dict[str, str]:
    return {
        template_key: get_lease_cadence(session, lease_id, template_key) or "inherit"
        for template_key in LEASE_AUTOMATION_TEMPLATES
    }


def save_lease_automation(session: Session, lease_id: int, payload: dict[str, Any]) -> dict[str, str]:
    for template_key in LEASE_AUTOMATION_TEMPLATES:
        value = str(payload.get(template_key) or "inherit").strip()
        if value in {"", "inherit"}:
            clear_lease_cadence(session, lease_id, template_key)
            continue
        if value not in VALID_REMINDER_CADENCES:
            raise HTTPException(400, "Некорректная частота уведомлений")
        set_lease_cadence(session, lease_id, template_key, value)
    return get_lease_automation(session, lease_id)


def cadence_label(value: str) -> str:
    return {
        "twice_daily": "2 раза в день",
        "daily_evening": "вечером каждый день",
        "every_two_days": "раз в два дня",
        "never": "выключено",
    }.get(value, value)


def local_now() -> datetime:
    return datetime.now(LOCAL_TZ)


def to_local_time(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(LOCAL_TZ)


def cadence_slots_for_day(day: date, cadence: str) -> list[datetime]:
    if cadence == "twice_daily":
        return [
            datetime.combine(day, time(hour=10), tzinfo=LOCAL_TZ),
            datetime.combine(day, time(hour=20), tzinfo=LOCAL_TZ),
        ]
    return [datetime.combine(day, time(hour=20), tzinfo=LOCAL_TZ)]


def next_reminder_slot(cadence: str, latest: datetime | None, now: datetime | None = None) -> datetime:
    now = now or local_now()
    if cadence == "never":
        return datetime.max.replace(tzinfo=LOCAL_TZ)
    latest_local = to_local_time(latest)
    check_day = now.date()
    for _ in range(8):
        for slot in cadence_slots_for_day(check_day, cadence):
            if slot <= now:
                if latest_local is None or latest_local < slot:
                    if cadence != "every_two_days":
                        return slot
                    if latest_local is None or latest_local <= slot - timedelta(days=2):
                        return slot
            else:
                if cadence != "every_two_days":
                    return slot
                if latest_local is None:
                    return slot
                next_allowed = latest_local + timedelta(days=2)
                if slot >= next_allowed:
                    return slot
        check_day += timedelta(days=1)
    return datetime.combine(now.date() + timedelta(days=7), time(hour=20), tzinfo=LOCAL_TZ)


def cadence_allows_send(cadence: str, latest: datetime | None, now: datetime | None = None) -> bool:
    if cadence == "never":
        return False
    now = now or local_now()
    slot = next_reminder_slot(cadence, latest, now)
    return slot <= now


def reminder_schedule_meta(session: Session, template_key: str, latest: MessageLog | None, lease_id: int | None = None) -> dict[str, Any]:
    cadence = reminder_cadence(session, template_key, lease_id)
    if cadence == "never":
        return {
            "cadence": cadence,
            "cadence_label": cadence_label(cadence),
            "next_auto_at": None,
        }
    next_slot = next_reminder_slot(cadence, latest.created_at if latest else None)
    return {
        "cadence": cadence,
        "cadence_label": cadence_label(cadence),
        "next_auto_at": next_slot.isoformat(),
    }


def message_template(session: Session, key: str) -> str:
    return get_setting_value(session, key)


def get_tenant_links(session: Session) -> dict[str, str]:
    raw = get_setting_value(session, "telegram_tenant_links") or "{}"
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    if not isinstance(parsed, dict):
        return {}
    return {str(key): str(value) for key, value in parsed.items() if value}


def tenant_consent_chat_ids(session: Session) -> set[str]:
    raw = get_internal_json(session, "telegram_consent_chat_ids", [])
    if not isinstance(raw, list):
        return set()
    return {str(item) for item in raw if str(item).strip()}


def tenant_consent_sent(session: Session, chat_id: int | str | None) -> bool:
    if not chat_id:
        return False
    return str(chat_id) in tenant_consent_chat_ids(session)


def mark_tenant_consent_sent(session: Session, chat_id: int | str) -> None:
    items = tenant_consent_chat_ids(session)
    items.add(str(chat_id))
    set_internal_json(session, "telegram_consent_chat_ids", sorted(items))


def save_tenant_links(session: Session, links: dict[str, str]) -> None:
    setting = session.get(AppSetting, "telegram_tenant_links")
    if not setting:
        setting = AppSetting(key="telegram_tenant_links")
        session.add(setting)
    setting.value = json.dumps(links, ensure_ascii=False)


def lease_chat_id(session: Session, lease: Lease) -> str:
    links = get_tenant_links(session)
    return links.get(str(lease.tenant_id), "")


def tenant_chat_linked(session: Session, lease: Lease) -> bool:
    return bool(lease_chat_id(session, lease))


def normalize_apartment_code(value: str) -> str:
    return re.sub(r"[^a-zа-я0-9]+", "", (value or "").lower().replace("ё", "е"))


def normalize_phone(value: str) -> str:
    digits = re.sub(r"\D+", "", value or "")
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    if len(digits) == 10:
        digits = "7" + digits
    return digits


def lease_by_chat_id(session: Session, chat_id: int | str | None) -> Lease | None:
    if not chat_id:
        return None
    chat_ref = str(chat_id)
    links = get_tenant_links(session)
    tenant_id = None
    for raw_tenant_id, linked_chat in links.items():
        if linked_chat != chat_ref:
            continue
        try:
            tenant_id = int(raw_tenant_id)
        except (TypeError, ValueError):
            continue
        break
    if tenant_id is None:
        return None
    return session.scalar(
        select(Lease)
        .options(
            joinedload(Lease.apartment).joinedload(Apartment.object),
            joinedload(Lease.tenant),
        )
        .where(Lease.tenant_id == tenant_id, Lease.active.is_(True))
        .order_by(Lease.start_date.desc(), Lease.id.desc())
        .limit(1)
    )


def maybe_link_tenant_chat(session: Session, message: dict[str, Any]) -> Lease | None:
    sender = message.get("from") or {}
    chat = message.get("chat") or {}
    chat_id = chat.get("id")
    linked = lease_by_chat_id(session, chat_id)
    if linked:
        return linked
    contact = message.get("contact") or {}
    contact_phone = normalize_phone(contact.get("phone_number") or "")
    contact_user_id = str(contact.get("user_id") or "")
    sender_user_id = str(sender.get("id") or "")
    if chat_id and contact_phone and contact_user_id and contact_user_id == sender_user_id:
        for lease in session.scalars(select(Lease).where(Lease.active.is_(True))).all():
            if normalize_phone(lease.tenant.phone) == contact_phone:
                links = get_tenant_links(session)
                links[str(lease.tenant_id)] = str(chat_id)
                save_tenant_links(session, links)
                return lease
    handles = {
        normalize_telegram_handle(sender.get("username") or ""),
        normalize_telegram_handle(chat.get("username") or ""),
    }
    handles.discard("")
    if not chat_id or not handles:
        return None

    for lease in session.scalars(select(Lease).where(Lease.active.is_(True))).all():
        if normalize_telegram_handle(lease.tenant.telegram) in handles:
            links = get_tenant_links(session)
            links[str(lease.tenant_id)] = str(chat_id)
            save_tenant_links(session, links)
            return lease
    return None


def find_lease_by_receipt_purpose(session: Session, purpose: str) -> Lease | None:
    normalized = normalize_apartment_code(purpose)
    if not normalized:
        return None
    for lease in session.scalars(select(Lease).where(Lease.active.is_(True))).all():
        apartment_code = normalize_apartment_code(lease.apartment.name)
        if apartment_code and apartment_code in normalized:
            return lease
    return None


def fill_template(template: str, context: dict[str, Any]) -> str:
    text = template or ""
    for key, value in context.items():
        text = text.replace("{" + key + "}", str(value if value is not None else ""))
    return text


def reminder_label(template_key: str) -> str:
    return {
        "message_rent_due": "аренда сегодня",
        "message_rent_upcoming": "аренда через 3 дня",
        "message_rent_overdue": "долг по аренде",
        "message_rent_status_request": "запрос статуса аренды",
        "message_utility_bill": "коммуналка",
        "message_utility_overdue": "просроченная коммуналка",
        "message_all_debts": "все долги",
        "custom": "свой текст",
    }.get(template_key, template_key or "сообщение")


def serialize_message_log(log: MessageLog | None) -> dict[str, Any] | None:
    if not log:
        return None
    return {
        "id": log.id,
        "template_key": log.template_key,
        "label": reminder_label(log.template_key),
        "status": log.status,
        "created_at": log.created_at.isoformat(),
        "note": log.note,
    }


def latest_message_log(
    session: Session | None,
    *,
    rent_charge_id: int | None = None,
    utility_line_id: int | None = None,
) -> MessageLog | None:
    if not session:
        return None
    cached = session.info.get("latest_message_logs")
    if rent_charge_id is not None and isinstance(cached, dict) and rent_charge_id in cached.get("rent", {}):
        return cached["rent"][rent_charge_id]
    if utility_line_id is not None and isinstance(cached, dict) and utility_line_id in cached.get("utility", {}):
        return cached["utility"][utility_line_id]
    query = select(MessageLog)
    if rent_charge_id is not None:
        query = query.where(MessageLog.rent_charge_id == rent_charge_id)
    if utility_line_id is not None:
        query = query.where(MessageLog.utility_line_id == utility_line_id)
    return session.scalar(query.order_by(MessageLog.created_at.desc()).limit(1))


def prefetch_latest_message_logs(
    session: Session,
    *,
    rent_charge_ids: list[int] | tuple[int, ...] | set[int] = (),
    utility_line_ids: list[int] | tuple[int, ...] | set[int] = (),
) -> None:
    cache = session.info.setdefault("latest_message_logs", {"rent": {}, "utility": {}})
    rent_cache: dict[int, MessageLog | None] = cache.setdefault("rent", {})
    utility_cache: dict[int, MessageLog | None] = cache.setdefault("utility", {})

    missing_rent_ids = sorted({int(value) for value in rent_charge_ids if value and int(value) not in rent_cache})
    if missing_rent_ids:
        for value in missing_rent_ids:
            rent_cache[value] = None
        logs = session.scalars(
            select(MessageLog)
            .where(MessageLog.rent_charge_id.in_(missing_rent_ids))
            .order_by(MessageLog.rent_charge_id, MessageLog.created_at.desc(), MessageLog.id.desc())
        ).all()
        for log in logs:
            if log.rent_charge_id is not None and rent_cache.get(log.rent_charge_id) is None:
                rent_cache[log.rent_charge_id] = log

    missing_utility_ids = sorted({int(value) for value in utility_line_ids if value and int(value) not in utility_cache})
    if missing_utility_ids:
        for value in missing_utility_ids:
            utility_cache[value] = None
        logs = session.scalars(
            select(MessageLog)
            .where(MessageLog.utility_line_id.in_(missing_utility_ids))
            .order_by(MessageLog.utility_line_id, MessageLog.created_at.desc(), MessageLog.id.desc())
        ).all()
        for log in logs:
            if log.utility_line_id is not None and utility_cache.get(log.utility_line_id) is None:
                utility_cache[log.utility_line_id] = log


def reminder_meta(
    session: Session | None,
    lease: Lease,
    due_date: date | None,
    *,
    template_key: str | None = None,
    rent_charge_id: int | None = None,
    utility_line_id: int | None = None,
) -> dict[str, Any]:
    latest = latest_message_log(session, rent_charge_id=rent_charge_id, utility_line_id=utility_line_id)
    chat_id = lease_chat_id(session, lease) if session else ""
    cutoff = notification_cutoff_date(session) if session else date.today()
    auto_enabled = notifications_enabled(session) if session else False
    due = due_date or date.today()
    eligible = bool(chat_id) and due >= cutoff
    block_reason = ""
    if not chat_id:
        block_reason = "ждём /start"
    elif due < cutoff:
        block_reason = "старый долг до запуска"
    elif not auto_enabled:
        block_reason = "авто выключено"
    schedule = reminder_schedule_meta(session, template_key, latest, lease.id) if session and template_key else None
    return {
        "linked": bool(chat_id),
        "auto_enabled": auto_enabled,
        "eligible_auto": eligible and auto_enabled,
        "cutoff_date": cutoff.isoformat(),
        "block_reason": block_reason,
        "latest": serialize_message_log(latest),
        "schedule": schedule,
    }


def serialize_object(obj: RentalObject) -> dict[str, Any]:
    return {
        "id": obj.id,
        "name": obj.name,
        "short_code": obj.short_code,
        "notes": obj.notes,
        "active": obj.active,
        "apartments": [serialize_apartment(apartment) for apartment in sorted(obj.apartments, key=lambda item: item.sort_order)],
        "services": [serialize_service(service) for service in obj.services if service.active],
    }


def current_active_lease(apartment: Apartment, today: date | None = None) -> Lease | None:
    today = today or date.today()
    leases = sorted(apartment.leases, key=lambda item: item.start_date, reverse=True)
    return next(
        (
            lease
            for lease in leases
            if lease.active and lease.start_date <= today and (lease.end_date is None or lease.end_date >= today)
            and IGNORE_LEASE_MARK not in (lease.notes or "")
        ),
        None,
    )


def serialize_apartment(apartment: Apartment) -> dict[str, Any]:
    active_lease = current_active_lease(apartment)
    advance_setting = apartment.utility_advance_setting
    return {
        "id": apartment.id,
        "object_id": apartment.object_id,
        "object_name": apartment.object.name if apartment.object else "",
        "name": apartment.name,
        "sort_order": apartment.sort_order,
        "odn_share_percent": apartment.odn_share_percent,
        "active": apartment.active,
        "active_lease_id": active_lease.id if active_lease else None,
        "active_tenant": active_lease.tenant.full_name if active_lease else "",
        "utility_advance_override": (
            money(advance_setting.amount_override)
            if advance_setting and advance_setting.amount_override is not None
            else None
        ),
        "utility_advance_note": advance_setting.note if advance_setting else "",
    }


def serialize_tenant(tenant: Tenant) -> dict[str, Any]:
    return {
        "id": tenant.id,
        "full_name": tenant.full_name,
        "phone": tenant.phone,
        "telegram": tenant.telegram,
        "whatsapp": tenant.whatsapp,
        "notes": tenant.notes,
        "active": tenant.active,
    }


def serialize_lease(lease: Lease, session: Session | None = None) -> dict[str, Any]:
    return {
        "id": lease.id,
        "apartment_id": lease.apartment_id,
        "apartment": lease.apartment.name,
        "object": lease.apartment.object.name,
        "apartment_active": lease.apartment.active,
        "tenant_id": lease.tenant_id,
        "tenant": lease.tenant.full_name,
        "phone": lease.tenant.phone,
        "telegram": lease.tenant.telegram,
        "whatsapp": lease.tenant.whatsapp,
        "start_date": lease.start_date.isoformat(),
        "end_date": lease.end_date.isoformat() if lease.end_date else None,
        "payment_day": lease.payment_day,
        "ip_amount": lease.ip_amount,
        "personal_amount": lease.personal_amount,
        "deposit_amount": lease.deposit_amount,
        "deposit_location": lease.deposit_location,
        "deposit_terms": lease.deposit_terms,
        "notes": "\n".join(line for line in (lease.notes or "").splitlines() if line.strip() != IGNORE_LEASE_MARK),
        "active": lease.active,
        "ignored": lease_ignored(session, lease.id),
        "automation": get_lease_automation(session, lease.id) if session else {},
    }


def parse_receipt_details(receipt: PaymentReceipt) -> dict[str, Any]:
    raw = receipt.recipient_details or ""
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def receipt_meta(parsed: dict[str, Any]) -> dict[str, Any]:
    meta = parsed.get("_meta") if isinstance(parsed, dict) else None
    return meta if isinstance(meta, dict) else {}


def payment_source_label(receipt: PaymentReceipt) -> str:
    if receipt.source == UTILITY_ADVANCE_SOURCE:
        base = "зачёт аванса"
    elif receipt.channel == UTILITY_ADVANCE_CHANNEL:
        base = "аванс коммуналки"
    elif receipt.source == "manual":
        base = "ручной"
    elif receipt.channel == "ip":
        base = "ИП"
    elif receipt.channel == "personal":
        base = "перевод"
    elif receipt.channel == "expense_fund":
        base = "на расходы"
    elif receipt.channel == "utilities":
        base = "коммуналка"
    else:
        base = receipt.source or "платёж"
    return f"{base} от {receipt.created_at:%d.%m.%Y}"


def receipt_sender_name(parsed: dict[str, Any]) -> str:
    value = (parsed.get("payer_name") or "").strip()
    return value or "не указан"


def serialize_payment_receipt(receipt: PaymentReceipt, session: Session) -> dict[str, Any]:
    parsed = parse_receipt_details(receipt)
    charge = receipt.rent_charge
    utility_line = session.get(UtilityBillLine, receipt.utility_line_id) if receipt.utility_line_id else None
    lease = charge.lease if charge else utility_line.lease if utility_line else session.get(Lease, receipt.lease_id) if receipt.lease_id else None
    apartment = lease.apartment.name if lease else ""
    tenant = lease.tenant.full_name if lease else ""
    target_label = ""
    target_month = ""
    if charge:
        target_label = f"аренда {format_date(charge.due_date)}"
        target_month = charge.due_date.isoformat()
    elif utility_line:
        target_label = (
            f"аванс коммуналки {format_date(utility_line.bill.period_start)}"
            if utility_line_is_advance(utility_line)
            else f"коммуналка {format_date(utility_line.bill.period_end)}"
        )
        target_month = utility_line.bill.period_end.isoformat()
    elif receipt.channel == UTILITY_ADVANCE_CHANNEL:
        target_label = "аванс коммуналки"
        target_month = receipt.paid_at.date().isoformat()
    target_year = charge.due_date.year if charge else utility_line.bill.period_end.year if utility_line else receipt.paid_at.year if receipt.channel == UTILITY_ADVANCE_CHANNEL else date.today().year
    target_month_number = charge.due_date.month if charge else utility_line.bill.period_end.month if utility_line else receipt.paid_at.month if receipt.channel == UTILITY_ADVANCE_CHANNEL else 0
    return {
        "id": receipt.id,
        "lease_id": receipt.lease_id,
        "rent_charge_id": receipt.rent_charge_id,
        "utility_line_id": receipt.utility_line_id,
        "apartment": apartment,
        "tenant": tenant,
        "amount": money(receipt.amount),
        "channel": receipt.channel,
        "channel_label": receipt_channel_label(receipt.channel),
        "status": receipt.status,
        "paid_at": receipt.paid_at.isoformat(),
        "source": receipt.source,
        "source_label": payment_source_label(receipt),
        "created_at": receipt.created_at.isoformat(),
        "recipient_name": receipt.recipient_name,
        "sender_name": receipt_sender_name(parsed),
        "notes": receipt.notes or "",
        "file_path": receipt.file_path or "",
        "target_label": target_label,
        "target_month": target_month,
        "target_month_number": target_month_number,
        "target_year": target_year,
        "parsed": parsed,
    }


def lease_payment_target_options(session: Session, lease: Lease) -> dict[str, list[dict[str, Any]]]:
    rent_charges = session.scalars(
        select(RentCharge)
        .where(RentCharge.lease_id == lease.id, RentCharge.due_date >= RENT_GENERATION_START)
        .order_by(RentCharge.due_date.desc(), RentCharge.id.desc())
    ).all()
    utility_lines = session.scalars(
        select(UtilityBillLine)
        .where(UtilityBillLine.lease_id == lease.id)
        .order_by(UtilityBillLine.id.desc())
    ).all()
    return {
        "rent": [
            {
                "id": charge.id,
                "label": f"аренда {format_date(charge.due_date)}",
                "month": charge.due_date.month,
                "year": charge.due_date.year,
                "due_date": charge.due_date.isoformat(),
                "debt": money(max(0.0, charge.ip_due - charge.ip_paid) + max(0.0, charge.personal_due - charge.personal_paid)),
            }
            for charge in rent_charges
        ],
        "utility": [
            {
                "id": line.id,
                "label": (
                    f"аванс коммуналки {line.bill.period_start:%d.%m.%Y} -> {line.bill.period_end:%d.%m.%Y}"
                    if utility_line_is_advance(line)
                    else f"ком. услуги {line.bill.period_start:%d.%m.%Y} -> {line.bill.period_end:%d.%m.%Y}"
                ),
                "period_start": line.bill.period_start.isoformat(),
                "period_end": line.bill.period_end.isoformat(),
                "debt": money(max(0.0, line.total_amount - line.paid_amount)),
                "status": line.status,
                "service": utility_bill_service_label(line.bill),
                "line_type": utility_line_type(line),
            }
            for line in utility_lines
        ],
    }


def serialize_payment_brief(receipt: PaymentReceipt) -> dict[str, Any]:
    return {
        "id": receipt.id,
        "channel": receipt.channel,
        "channel_label": receipt_channel_label(receipt.channel),
        "amount": money(receipt.amount),
        "status": receipt.status,
        "paid_at": receipt.paid_at.isoformat(),
    }


def serialize_rent_charge(
    charge: RentCharge,
    session: Session | None = None,
    *,
    include_payments: bool = True,
    include_reminder: bool = True,
) -> dict[str, Any]:
    update_rent_charge_status(charge)
    reminder_template_key = "message_rent_due" if charge.status == "pending" and charge.due_date == date.today() else "message_rent_overdue"
    deferral_days_left = None
    if charge.deferral_until:
        deferral_days_left = max((charge.deferral_until - date.today()).days, 0)
    return {
        "id": charge.id,
        "lease_id": charge.lease_id,
        "object": charge.lease.apartment.object.name,
        "apartment": charge.lease.apartment.name,
        "tenant": charge.lease.tenant.full_name,
        "phone": charge.lease.tenant.phone,
        "telegram": charge.lease.tenant.telegram,
        "whatsapp": charge.lease.tenant.whatsapp,
        "period_start": charge.period_start.isoformat(),
        "period_end": charge.period_end.isoformat(),
        "due_date": charge.due_date.isoformat(),
        "ip_due": charge.ip_due,
        "personal_due": charge.personal_due,
        "ip_paid": charge.ip_paid,
        "personal_paid": charge.personal_paid,
        "total_due": money(charge.ip_due + charge.personal_due),
        "total_paid": money(charge.ip_paid + charge.personal_paid),
        "debt": money(max(0, charge.ip_due - charge.ip_paid) + max(0, charge.personal_due - charge.personal_paid)),
        "status": charge.status,
        "ip_status": status_for_amount(charge.ip_due, charge.ip_paid),
        "personal_status": status_for_amount(charge.personal_due, charge.personal_paid),
        "payments": [
            serialize_payment_brief(receipt)
            for receipt in sorted(charge.receipts, key=lambda item: (item.paid_at, item.id))
            if receipt.status == "accepted"
        ] if include_payments else [],
        "deferral_until": charge.deferral_until.isoformat() if charge.deferral_until else None,
        "deferral_days_left": deferral_days_left,
        "deferral_note": charge.deferral_note,
        "reminder": reminder_meta(
            session,
            charge.lease,
            charge.due_date,
            template_key=reminder_template_key,
            rent_charge_id=charge.id,
        ) if include_reminder else None,
    }


def serialize_service(service: UtilityService) -> dict[str, Any]:
    return {
        "id": service.id,
        "object_id": service.object_id,
        "object": service.object.name if service.object else "",
        "kind": service.kind,
        "name": service.name,
        "provider_reading_due_day": service.provider_reading_due_day,
        "provider_due_day": service.provider_due_day,
        "resident_due_days": service.resident_due_days,
        "active": service.active,
    }


def serialize_meter(meter: Meter) -> dict[str, Any]:
    latest = max(meter.readings, key=lambda item: item.reading_date) if meter.readings else None
    return {
        "id": meter.id,
        "service_id": meter.service_id,
        "service": meter.service.name,
        "object_id": meter.object_id,
        "object": meter.service.object.name,
        "apartment_id": meter.apartment_id,
        "apartment": meter.apartment.name if meter.apartment else "",
        "scope": meter.scope,
        "name": meter.name,
        "latest_date": latest.reading_date.isoformat() if latest else None,
        "latest_value": latest.value if latest else None,
        "active": meter.active,
    }


def utility_bill_type(bill: UtilityBill | None) -> str:
    return (getattr(bill, "bill_type", "") or UTILITY_BILL_USAGE).strip() or UTILITY_BILL_USAGE


def utility_line_type(line: UtilityBillLine | None) -> str:
    return (getattr(line, "line_type", "") or UTILITY_LINE_USAGE).strip() or UTILITY_LINE_USAGE


def utility_bill_is_advance(bill: UtilityBill | None) -> bool:
    return utility_bill_type(bill) == UTILITY_BILL_ADVANCE


def utility_line_is_advance(line: UtilityBillLine | None) -> bool:
    return utility_line_type(line) == UTILITY_LINE_ADVANCE


def utility_line_metadata(line: UtilityBillLine | None) -> dict[str, Any]:
    raw = getattr(line, "metadata_json", "") if line else ""
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def set_utility_line_metadata(line: UtilityBillLine, payload: dict[str, Any]) -> None:
    line.metadata_json = json.dumps(payload, ensure_ascii=False, sort_keys=True)


def utility_bill_service_label(bill: UtilityBill) -> str:
    return "Аванс коммуналки" if utility_bill_is_advance(bill) else bill.service.name


def utility_line_debt_amount(line: UtilityBillLine) -> float:
    return money(max(0.0, float(line.total_amount or 0) - float(line.paid_amount or 0)))


def utility_advance_applied_to_line(session: Session, line_id: int) -> float:
    return money(
        abs(
            float(
                session.scalar(
                    select(func.coalesce(func.sum(UtilityAdvanceLedger.amount), 0)).where(
                        UtilityAdvanceLedger.utility_line_id == line_id,
                        UtilityAdvanceLedger.amount < 0,
                    )
                )
                or 0
            )
        )
    )


def utility_advance_ledger_balance(session: Session, apartment_id: int) -> float:
    return money(
        float(
            session.scalar(
                select(func.coalesce(func.sum(UtilityAdvanceLedger.amount), 0)).where(
                    UtilityAdvanceLedger.apartment_id == apartment_id
                )
            )
            or 0
        )
    )


def legacy_unlinked_utility_advance_balance(session: Session, apartment_id: int) -> float:
    return money(
        float(
            session.scalar(
                select(func.coalesce(func.sum(PaymentReceipt.amount), 0)).where(
                    PaymentReceipt.apartment_id == apartment_id,
                    PaymentReceipt.status == "accepted",
                    PaymentReceipt.channel == UTILITY_ADVANCE_CHANNEL,
                    PaymentReceipt.utility_line_id.is_(None),
                    PaymentReceipt.rent_charge_id.is_(None),
                )
            )
            or 0
        )
    )


def utility_advance_balance(session: Session, apartment_id: int) -> float:
    return money(
        utility_advance_ledger_balance(session, apartment_id)
        + legacy_unlinked_utility_advance_balance(session, apartment_id)
    )


def average_money(values: list[float]) -> float:
    values = [float(value or 0) for value in values if float(value or 0) > EPS]
    return money(sum(values) / len(values)) if values else 0.0


def utility_advance_current_cap_multiplier(target_start: date) -> float:
    if target_start.month in {11, 12, 1, 2, 3}:
        return UTILITY_ADVANCE_CURRENT_CAP_WINTER
    if target_start.month in {4, 10}:
        return UTILITY_ADVANCE_CURRENT_CAP_SHOULDER
    return UTILITY_ADVANCE_CURRENT_CAP_DEFAULT


def capped_utility_advance_forecast(
    forecast: float,
    current_amount: float,
    target_start: date,
) -> tuple[float, dict[str, Any]]:
    uncapped = money(forecast)
    current = money(current_amount)
    multiplier = utility_advance_current_cap_multiplier(target_start)
    metadata = {
        "current_amount": current,
        "uncapped_forecast": uncapped,
        "forecast_cap_multiplier": multiplier,
        "forecast_cap": 0.0,
        "forecast_capped": False,
    }
    if current <= EPS or uncapped <= current + EPS:
        return uncapped, metadata

    forecast_cap = money(max(current * multiplier, current + UTILITY_ADVANCE_CURRENT_CAP_BUFFER))
    metadata["forecast_cap"] = forecast_cap
    if uncapped > forecast_cap + EPS:
        metadata["forecast_capped"] = True
        return forecast_cap, metadata
    return uncapped, metadata


def utility_advance_setting_for_apartment(session: Session, apartment_id: int) -> UtilityAdvanceSetting | None:
    return session.scalar(
        select(UtilityAdvanceSetting)
        .where(UtilityAdvanceSetting.apartment_id == apartment_id)
        .limit(1)
    )


def estimated_apartment_utility_amount(
    session: Session,
    apartment_id: int,
    target_start: date,
    *,
    current_amount: float = 0.0,
) -> tuple[float, dict[str, Any]]:
    setting = utility_advance_setting_for_apartment(session, apartment_id)
    if setting and setting.amount_override is not None:
        return money(setting.amount_override), {
            "source": "manual_override",
            "override": money(setting.amount_override),
        }

    historical_lines = session.scalars(
        select(UtilityBillLine)
        .options(joinedload(UtilityBillLine.bill))
        .join(UtilityBill)
        .where(
            UtilityBillLine.apartment_id == apartment_id,
            UtilityBillLine.line_type != UTILITY_LINE_ADVANCE,
            UtilityBill.status != "draft",
            UtilityBill.period_start < target_start,
        )
        .order_by(UtilityBill.period_start.desc(), UtilityBillLine.id.desc())
    ).all()
    monthly: dict[tuple[int, int], float] = {}
    for line in historical_lines:
        if not line.bill:
            continue
        key = (line.bill.period_start.year, line.bill.period_start.month)
        monthly[key] = money(monthly.get(key, 0.0) + float(line.total_amount or 0))

    ordered = sorted(monthly.items(), reverse=True)
    last_12 = ordered[:12]
    recent = [money(current_amount)] if current_amount > EPS else []
    recent.extend(value for _key, value in last_12[:3] if value > EPS)
    seasonal = [value for (_year, month), value in last_12 if month == target_start.month and value > EPS]
    annual = [value for _key, value in last_12 if value > EPS]

    recent_average = average_money(recent[:3])
    seasonal_average = average_money(seasonal)
    annual_average = average_money(annual)
    if seasonal_average > EPS and recent_average > EPS:
        forecast = money(seasonal_average * 0.65 + recent_average * 0.35)
        source = "seasonal_recent_blend"
    elif seasonal_average > EPS:
        forecast = seasonal_average
        source = "seasonal"
    elif recent_average > EPS:
        forecast = recent_average
        source = "recent"
    else:
        forecast = annual_average
        source = "annual_average" if annual_average > EPS else "empty_history"
    capped_forecast, cap_meta = capped_utility_advance_forecast(forecast, current_amount, target_start)
    return capped_forecast, {
        "source": source,
        "recent_average": recent_average,
        "seasonal_average": seasonal_average,
        "annual_average": annual_average,
        "history_months": len(last_12),
        **cap_meta,
    }


def add_calendar_month(value: date) -> date:
    year = value.year + (1 if value.month == 12 else 0)
    month = 1 if value.month == 12 else value.month + 1
    _, last_day = calendar.monthrange(year, month)
    return date(year, month, min(value.day, last_day))


def next_utility_advance_period(period_start: date, period_end: date) -> tuple[date, date]:
    return period_end, add_calendar_month(period_end)


def lease_should_skip_utility_advance(lease: Lease, advance_start: date, advance_end: date) -> bool:
    if lease.end_date is None:
        return False
    return lease.end_date <= advance_start


def lease_utility_advance_period(lease: Lease, advance_start: date, advance_end: date) -> tuple[date, date] | None:
    if lease_should_skip_utility_advance(lease, advance_start, advance_end):
        return None
    line_end = min(advance_end, lease.end_date) if lease.end_date else advance_end
    if line_end <= advance_start:
        return None
    return advance_start, line_end


def sync_utility_advance_credit_for_receipt(session: Session, receipt: PaymentReceipt) -> None:
    existing = session.scalar(
        select(UtilityAdvanceLedger)
        .where(
            UtilityAdvanceLedger.payment_receipt_id == receipt.id,
            UtilityAdvanceLedger.kind == "advance_payment",
        )
        .limit(1)
    )
    line = session.get(UtilityBillLine, receipt.utility_line_id) if receipt.utility_line_id else None
    should_have_credit = (
        receipt.status == "accepted"
        and receipt.channel == UTILITY_ADVANCE_CHANNEL
        and line is not None
        and utility_line_is_advance(line)
    )
    if not should_have_credit:
        if existing:
            session.delete(existing)
        return

    if not existing:
        existing = UtilityAdvanceLedger(
            payment_receipt_id=receipt.id,
            kind="advance_payment",
        )
        session.add(existing)
    existing.apartment_id = line.apartment_id
    existing.lease_id = line.lease_id
    existing.utility_line_id = line.id
    existing.period_start = line.bill.period_start if line.bill else None
    existing.period_end = line.bill.period_end if line.bill else None
    existing.amount = money(receipt.amount)
    existing.note = receipt.notes or "оплата аванса коммуналки"


def sync_utility_advance_credits_for_receipts(session: Session, receipts: list[PaymentReceipt]) -> None:
    for receipt in receipts:
        sync_utility_advance_credit_for_receipt(session, receipt)


def serialize_bill_line(
    line: UtilityBillLine,
    session: Session | None = None,
    *,
    include_payments: bool = True,
    include_reminder: bool = True,
    include_advance: bool = True,
) -> dict[str, Any]:
    update_utility_line_status(line)
    payments = []
    if session and include_payments:
        payments = [
            serialize_payment_brief(receipt)
            for receipt in session.scalars(
                select(PaymentReceipt)
                .where(PaymentReceipt.utility_line_id == line.id, PaymentReceipt.status == "accepted")
                .order_by(PaymentReceipt.paid_at, PaymentReceipt.id)
            ).all()
        ]
    period_label = utility_line_period_label(line)
    metadata = utility_line_metadata(line)
    advance_applied = utility_advance_applied_to_line(session, line.id) if include_advance and session and line.id else 0.0
    advance_balance_available = utility_advance_balance(session, line.apartment_id) if include_advance and session and line.apartment_id else 0.0
    return {
        "id": line.id,
        "bill_id": line.bill_id,
        "line_type": utility_line_type(line),
        "apartment_id": line.apartment_id,
        "apartment": line.apartment.name,
        "object": line.bill.service.object.name,
        "service": utility_bill_service_label(line.bill),
        "lease_id": line.lease_id,
        "tenant": line.lease.tenant.full_name if line.lease else "",
        "personal_consumption": line.personal_consumption,
        "odn_consumption": line.odn_consumption,
        "total_amount": line.total_amount,
        "paid_amount": line.paid_amount,
        "debt": utility_line_debt_amount(line),
        "advance_applied_amount": advance_applied,
        "advance_balance_available": advance_balance_available,
        "advance_balance_before": metadata.get("advance_balance_before", 0),
        "advance_recommended_amount": metadata.get("advance_recommended_amount", 0),
        "status": line.status,
        "due_date": line.due_date.isoformat() if line.due_date else None,
        "note": line.note,
        "metadata": metadata,
        "period_label": period_label,
        "payments": payments,
        "bill_period_start": line.bill.period_start.isoformat(),
        "bill_period_end": line.bill.period_end.isoformat(),
        "bill_period_label": f"{line.bill.period_start:%d.%m.%Y} -> {line.bill.period_end:%d.%m.%Y} ({(line.bill.period_end - line.bill.period_start).days} дн.)",
        "reminder": reminder_meta(
            session,
            line.lease,
            line.due_date,
            template_key="message_utility_bill",
            utility_line_id=line.id,
        ) if include_reminder and line.lease else None,
    }


def serialize_bill(
    bill: UtilityBill,
    session: Session | None = None,
    *,
    include_line_payments: bool = True,
    include_line_reminders: bool = True,
    include_line_advance: bool = True,
) -> dict[str, Any]:
    resident_total_amount = money(sum(line.total_amount for line in bill.lines))
    resident_paid_amount = money(sum(line.paid_amount for line in bill.lines))
    return {
        "id": bill.id,
        "service_id": bill.service_id,
        "bill_type": utility_bill_type(bill),
        "service": utility_bill_service_label(bill),
        "object": bill.service.object.name,
        "period_start": bill.period_start.isoformat(),
        "period_end": bill.period_end.isoformat(),
        "period_label": f"{bill.period_start:%d.%m.%Y} -> {bill.period_end:%d.%m.%Y} ({(bill.period_end - bill.period_start).days} дн.)",
        "days": (bill.period_end - bill.period_start).days,
        "status": bill.status,
        "total_consumption": bill.total_consumption,
        "apartment_consumption": bill.apartment_consumption,
        "odn_consumption": bill.odn_consumption,
        "total_cost": bill.total_cost,
        "resident_total_amount": resident_total_amount,
        "resident_paid_amount": resident_paid_amount,
        "average_unit_price": bill.average_unit_price,
        "due_date": bill.due_date.isoformat() if bill.due_date else None,
        "is_forecast": bill.is_forecast,
        "provider_paid": bill.provider_paid,
        "provider_paid_at": bill.provider_paid_at.isoformat() if bill.provider_paid_at else None,
        "notes": bill.notes,
        "lines": [
            serialize_bill_line(
                line,
                session,
                include_payments=include_line_payments,
                include_reminder=include_line_reminders,
                include_advance=include_line_advance,
            )
            for line in bill.lines
        ],
    }


def utility_line_period_label(line: UtilityBillLine) -> str:
    label = (line.note or "").strip()
    if label:
        return label
    if utility_line_is_advance(line):
        return f"аванс за {line.bill.period_start:%d.%m.%Y} -> {line.bill.period_end:%d.%m.%Y}"
    return f"{line.bill.period_start:%d.%m.%Y} -> {line.bill.period_end:%d.%m.%Y} ({(line.bill.period_end - line.bill.period_start).days} дн.)"


def serialize_expense(expense: Expense) -> dict[str, Any]:
    return {
        "id": expense.id,
        "expense_date": expense.expense_date.isoformat(),
        "object_id": expense.object_id,
        "object": expense.object.name if expense.object else "",
        "apartment_id": expense.apartment_id,
        "apartment": expense.apartment.name if expense.apartment else "",
        "category": expense.category,
        "amount": expense.amount,
        "source_funds": expense.source_funds,
        "payment_method": expense.payment_method,
        "description": expense.description,
        "compensation_status": expense.compensation_status,
        "compensated_at": expense.compensated_at.isoformat() if expense.compensated_at else None,
        "file_path": expense.file_path,
        "notes": expense.notes,
    }


def expense_fund_receipt_marker(receipt_id: int) -> str:
    return f"{EXPENSE_FUND_RECEIPT_MARK}{receipt_id}"


def linked_expense_fund_record(session: Session, receipt_id: int) -> Expense | None:
    marker = expense_fund_receipt_marker(receipt_id)
    return session.scalar(select(Expense).where(Expense.notes.contains(marker)).limit(1))


def sync_expense_fund_receipt(session: Session, receipt: PaymentReceipt) -> None:
    existing = linked_expense_fund_record(session, receipt.id)
    eligible = (
        receipt.status == "accepted"
        and receipt.channel == "expense_fund"
        and receipt.rent_charge_id is not None
        and receipt.lease_id is not None
    )
    if not eligible:
        if existing:
            session.delete(existing)
        return

    lease = receipt.rent_charge.lease if receipt.rent_charge else session.get(Lease, receipt.lease_id)
    if not lease:
        if existing:
            session.delete(existing)
        return

    marker = expense_fund_receipt_marker(receipt.id)
    expense = existing or Expense(notes=marker)
    expense.expense_date = receipt.paid_at.date()
    expense.object_id = lease.apartment.object_id
    expense.apartment_id = lease.apartment_id
    expense.category = "Пополнение на расходы"
    expense.amount = money(receipt.amount)
    expense.source_funds = "expense_fund_income"
    expense.payment_method = receipt.source or "manual"
    expense.description = f"Пополнение на расходы от {lease.tenant.full_name}"
    expense.compensation_status = "not_required"
    expense.file_path = receipt.file_path or ""
    if marker not in (expense.notes or ""):
        expense.notes = "; ".join(part for part in [expense.notes, marker] if part)
    session.add(expense)


def sync_expense_fund_receipts(session: Session, receipts: list[PaymentReceipt]) -> None:
    for receipt in receipts:
        sync_expense_fund_receipt(session, receipt)


def build_object_summary(session: Session) -> dict[str, Any]:
    apartments = session.scalars(select(Apartment).where(Apartment.active.is_(True))).all()
    occupied = [apartment for apartment in apartments if current_active_lease(apartment)]
    by_object = []
    for rental_object in session.scalars(select(RentalObject).order_by(RentalObject.name)).all():
        active_apartments = [apartment for apartment in rental_object.apartments if apartment.active]
        occupied_apartments = [apartment for apartment in active_apartments if current_active_lease(apartment)]
        by_object.append(
            {
                "object": rental_object.name,
                "occupied": len(occupied_apartments),
                "total": len(active_apartments),
            }
        )
    return {
        "occupied": len(occupied),
        "total": len(apartments),
        "by_object": by_object,
    }


@app.get("/api/auth/status")
def auth_status(request: Request, session: Session = Depends(get_session)) -> dict[str, Any]:
    role = panel_role_from_request(request, session)
    return {
        "authenticated": bool(role),
        "role": role,
    }


@app.post("/api/auth/pin")
def auth_with_pin(
    payload: dict[str, Any],
    request: Request,
    response: Response,
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    client_host = request.client.host if request.client else "unknown"
    fingerprint = login_fingerprint(client_host, request.headers.get("user-agent", ""))
    retry_after = login_retry_after(session, fingerprint)
    if retry_after:
        raise HTTPException(429, "Слишком много попыток входа. Повторите позже.", headers={"Retry-After": str(retry_after)})
    role = panel_role_for_pin(session, str(payload.get("pin_code") or ""))
    if not role:
        delay = record_login_failure(session, fingerprint)
        raise HTTPException(401, "Неверный PIN-код.", headers={"Retry-After": str(delay)})
    clear_login_failures(session, fingerprint)
    issued = issue_session(session, role, request.headers.get("user-agent", ""))
    remember_device = payload.get("remember_device", True)
    max_age = PANEL_AUTH_MAX_AGE_SECONDS if setting_bool_value(remember_device) else None
    secure_cookie = production_mode()
    response.set_cookie(
        PANEL_AUTH_COOKIE,
        issued.token,
        httponly=True,
        samesite="lax",
        max_age=max_age,
        expires=max_age,
        secure=secure_cookie,
        path="/",
    )
    response.set_cookie(
        PANEL_CSRF_COOKIE,
        issued.csrf_token,
        httponly=False,
        samesite="lax",
        max_age=max_age,
        expires=max_age,
        secure=secure_cookie,
        path="/",
    )
    return {
        "authenticated": True,
        "role": role,
        "csrf_token": issued.csrf_token,
    }


@app.post("/api/auth/logout")
def logout_panel(request: Request, response: Response, session: Session = Depends(get_session)) -> dict[str, bool]:
    token = str(request.cookies.get(PANEL_AUTH_COOKIE) or "")
    revoke_panel_session(session, token)
    response.delete_cookie(PANEL_AUTH_COOKIE, path="/", secure=production_mode(), samesite="lax")
    response.delete_cookie(PANEL_CSRF_COOKIE, path="/", secure=production_mode(), samesite="lax")
    return {"ok": True}


def build_registry_payload(session: Session, role: str | None) -> dict[str, Any]:
    if role != "owner":
        return {"objects": [], "leases": [], "meters": [], "services": []}
    return {
        "objects": [
            serialize_object(obj)
            for obj in session.scalars(
                select(RentalObject)
                .options(
                    selectinload(RentalObject.apartments).selectinload(Apartment.leases).joinedload(Lease.tenant),
                    selectinload(RentalObject.services),
                )
                .order_by(RentalObject.name)
            ).all()
        ],
        "leases": [
            serialize_lease(lease, session)
            for lease in session.scalars(
                select(Lease)
                .options(
                    joinedload(Lease.apartment).joinedload(Apartment.object),
                    joinedload(Lease.tenant),
                )
                .order_by(Lease.active.desc(), Lease.start_date.desc())
            ).all()
        ],
        "meters": [
            serialize_meter(meter)
            for meter in session.scalars(
                select(Meter)
                .options(
                    joinedload(Meter.service).joinedload(UtilityService.object),
                    joinedload(Meter.apartment),
                    selectinload(Meter.readings),
                )
                .order_by(Meter.object_id, Meter.scope, Meter.name)
            ).all()
        ],
        "services": [
            serialize_service(service)
            for service in session.scalars(
                select(UtilityService)
                .options(joinedload(UtilityService.object))
                .order_by(UtilityService.object_id, UtilityService.kind)
            ).all()
        ],
    }


def empty_registry_payload() -> dict[str, list[Any]]:
    return {"objects": [], "leases": [], "meters": [], "services": []}


def build_bootstrap_payload(
    request: Request,
    session: Session,
    *,
    include_registry: bool = False,
) -> dict[str, Any]:
    role = getattr(request.state, "panel_role", None) or panel_role_from_request(request, session)
    settings_payload = get_settings(session) if role == "owner" else public_settings(session)
    dashboard_payload = build_dashboard(session)
    if role != "owner":
        dashboard_payload = {
            "object_summary": dashboard_payload["object_summary"],
            "month_summary": dashboard_payload["month_summary"],
            "income_trend": dashboard_payload["income_trend"],
            "monthly_reports": dashboard_payload["monthly_reports"],
            "summary_counts": {
                "rent_overdue": len(dashboard_payload["rent_overdue"]),
                "rent_partial": len(dashboard_payload["rent_partial"]),
                "utility_overdue": len(dashboard_payload["utility_overdue"]),
                "utility_issued": len(dashboard_payload.get("utility_issued", [])),
                "provider_debts": len(dashboard_payload["provider_debts"]),
                "provider_reading_due": len(dashboard_payload.get("provider_reading_due", [])),
                "stale_readings": len(dashboard_payload["stale_readings"]),
                "suspicious_receipts": len(dashboard_payload["suspicious_receipts"]),
            },
        }
    registry_payload = build_registry_payload(session, role) if include_registry else empty_registry_payload()
    hermes_summary = {}
    if role == "owner":
        open_case_count = int(
            session.scalar(select(func.count(OperationalCase.id)).where(OperationalCase.status.in_(OPEN_CASE_STATUSES))) or 0
        )
        hermes_summary = {
            "enabled": setting_bool_value(get_setting_value(session, "ai_enabled")),
            "active_cases": open_case_count,
            "waiting_owner": int(
                session.scalar(select(func.count(OperationalCase.id)).where(OperationalCase.status == "waiting_owner")) or 0
            ),
            "pending_proposals": int(
                session.scalar(select(func.count(AgentActionProposal.id)).where(AgentActionProposal.status == "pending")) or 0
            ),
            "active_commitments": int(
                session.scalar(select(func.count(OwnerCommitment.id)).where(OwnerCommitment.status.in_({"active", "overdue"}))) or 0
            ),
        }
    return {
        "today": date.today().isoformat(),
        "auth": {"role": role},
        **registry_payload,
        "settings": settings_payload,
        "dashboard": dashboard_payload,
        "hermes_summary": hermes_summary,
    }


@app.get("/api/bootstrap")
def bootstrap(request: Request, session: Session = Depends(get_session)) -> dict[str, Any]:
    return build_bootstrap_payload(request, session, include_registry=True)


APP_STATE_SECTION_ORDER = (
    "bootstrap",
    "registry",
    "rent_charges",
    "utility_bills",
    "expenses",
    "tariffs",
    "utility_timeline",
    "message_targets",
    "suspicious_receipts",
)
APP_STATE_SECTION_SET = set(APP_STATE_SECTION_ORDER)


def parse_app_state_sections(value: str | None) -> list[str]:
    if not value:
        return list(APP_STATE_SECTION_ORDER)
    selected: list[str] = []
    for raw_item in re.split(r"[,;\s]+", value):
        item = raw_item.strip()
        if not item:
            continue
        if item == "all":
            return list(APP_STATE_SECTION_ORDER)
        if item not in APP_STATE_SECTION_SET:
            raise HTTPException(400, f"Неизвестная секция app-state: {item}")
        if item not in selected:
            selected.append(item)
    return selected or ["bootstrap"]


@app.get("/api/app-state")
def app_state(
    request: Request,
    sections: str | None = None,
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    selected = parse_app_state_sections(sections)
    payload: dict[str, Any] = {}
    timings: dict[str, int] = {}
    started = time_module.perf_counter()
    role = getattr(request.state, "panel_role", None) or panel_role_from_request(request, session)

    def load_section(key: str, loader) -> None:
        section_started = time_module.perf_counter()
        payload[key] = loader()
        timings[key] = int((time_module.perf_counter() - section_started) * 1000)

    if role != "owner":
        for key in selected:
            if key == "bootstrap":
                load_section("bootstrap", lambda: build_bootstrap_payload(request, session))
            elif key == "registry":
                load_section("registry", lambda: build_registry_payload(session, role))
            else:
                payload[key] = []
                timings[key] = 0
        record_perf_section(
            "app-state",
            ",".join(selected),
            int((time_module.perf_counter() - started) * 1000),
            {"role": role or "anonymous", "sections": timings},
        )
        return payload

    loaders = {
        "bootstrap": lambda: build_bootstrap_payload(request, session),
        "registry": lambda: build_registry_payload(session, role),
        "rent_charges": lambda: rent_charges_payload(
            session=session,
            include_payments=False,
            include_reminder=False,
            generate_missing=False,
        ),
        "utility_bills": lambda: utility_bills_payload(
            session,
            include_line_payments=False,
            include_line_reminders=False,
        ),
        "expenses": lambda: expenses_payload(session),
        "tariffs": lambda: tariffs_payload(session),
        "utility_timeline": lambda: utility_timeline_payload(session),
        "message_targets": lambda: message_targets_payload(session),
        "suspicious_receipts": lambda: suspicious_receipts_payload(session),
    }
    for key in selected:
        load_section(key, loaders[key])
    record_perf_section(
        "app-state",
        ",".join(selected),
        int((time_module.perf_counter() - started) * 1000),
        {"role": role or "anonymous", "sections": timings},
    )
    return payload


@app.get("/api/settings")
def api_get_settings(session: Session = Depends(get_session)) -> dict[str, str | bool]:
    return get_settings(session)


@app.get("/api/performance")
def api_performance(session: Session = Depends(get_session)) -> dict[str, Any]:
    snapshot = performance_snapshot()
    usage_since = date.today() - timedelta(days=13)
    usage_rows = session.scalars(
        select(AiUsageDaily)
        .where(AiUsageDaily.usage_date >= usage_since)
        .order_by(AiUsageDaily.usage_date.desc(), AiUsageDaily.provider, AiUsageDaily.model)
    ).all()
    today_usage = [row for row in usage_rows if row.usage_date == date.today()]
    snapshot["ai_usage"] = [
        {
            "date": row.usage_date.isoformat(),
            "provider": row.provider,
            "model": row.model,
            "calls": int(row.calls or 0),
            "total_tokens": int(row.total_tokens or 0),
            "cost_rub": money(float(row.cost_rub or 0)),
        }
        for row in usage_rows
    ]
    snapshot["ai_summary"] = {
        "enabled": setting_bool_value(get_setting_value(session, "ai_enabled")),
        "tenant_free_text_enabled": setting_bool_value(get_setting_value(session, "ai_tenant_free_text_enabled")),
        "supervisor_enabled": setting_bool_value(get_setting_value(session, "ai_supervisor_enabled")),
        "supervisor_cadence": get_setting_value(session, "ai_supervisor_cadence"),
        "supervisor_weekday": int(get_setting_value(session, "ai_supervisor_weekday") or 0),
        "supervisor_time": get_setting_value(session, "ai_supervisor_time"),
        "supervisor_model": get_setting_value(session, "ai_supervisor_model"),
        "today_calls": sum(int(row.calls or 0) for row in today_usage),
        "today_cost_rub": money(sum(float(row.cost_rub or 0) for row in today_usage)),
        "daily_call_limit": ai_daily_call_limit(session),
        "monthly_spent_rub": ai_monthly_spent_rub(session),
        "monthly_budget_rub": ai_monthly_budget_rub(session),
        "open_agent_tasks": int(session.scalar(select(func.count(AgentTask.id)).where(AgentTask.status == "open")) or 0),
        "pending_action_proposals": int(
            session.scalar(select(func.count(AgentActionProposal.id)).where(AgentActionProposal.status == "pending")) or 0
        ),
    }
    snapshot["data_counts"] = {
        "rent_charges": int(session.scalar(select(func.count(RentCharge.id))) or 0),
        "utility_bill_lines": int(session.scalar(select(func.count(UtilityBillLine.id))) or 0),
        "payment_receipts": int(session.scalar(select(func.count(PaymentReceipt.id))) or 0),
        "message_logs": int(session.scalar(select(func.count(MessageLog.id))) or 0),
    }
    return snapshot


@app.post("/api/settings")
def api_save_settings(request: Request, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, str | bool]:
    current_token = str(request.cookies.get(PANEL_AUTH_COOKIE) or "")
    return save_settings(session, payload, preserve_panel_token=current_token)


@app.post("/api/ai/test")
def api_test_ai_connection(session: Session = Depends(get_session)) -> dict[str, Any]:
    if ai_budget_exceeded(session):
        raise HTTPException(status_code=409, detail=AI_BUDGET_EXCEEDED_TEXT)
    if ai_daily_call_limit_exceeded(session):
        raise HTTPException(status_code=409, detail=AI_DAILY_LIMIT_EXCEEDED_TEXT)
    if ai_feature_daily_limit_exceeded(session, "key_test"):
        raise HTTPException(status_code=409, detail=AI_DAILY_LIMIT_EXCEEDED_TEXT)
    model = resolve_ai_model(get_setting_value(session, "deepseek_model"))
    manifest = ContextManifest(
        feature="key_test",
        model=model,
        selected_case_ids=[],
        included_entities={},
        included_messages_count=0,
        approximate_input_size=8,
        reason_for_llm_usage="explicit owner API key test",
        excluded_context=["all business data", "chat history", "database access"],
        output_limit=FEATURE_OUTPUT_LIMITS["key_test"],
        selected_tools=[],
        state_hash=hashlib.sha256(f"key-test:{date.today()}".encode("utf-8")).hexdigest(),
    )
    run = create_agent_run(session, manifest=manifest, trigger="api_key_test")
    try:
        runtime = build_provider_runtime(
            AiProvider.DEEPSEEK,
            requested_model=get_setting_value(session, "deepseek_model"),
            deepseek_api_key=get_setting_value(session, "deepseek_api_key"),
        )
        result = runtime.client.chat_completions(
            model=runtime.model,
            messages=[{"role": "user", "content": "Ответь только словом OK."}],
            temperature=0,
            max_tokens=16,
        )
    except AiProviderConfigError as exc:
        complete_agent_run(run, error=str(exc))
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except DeepSeekClientError as exc:
        complete_agent_run(run, error=str(exc))
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    normalized = DeepSeekResult(
        content=result.content,
        model=result.model,
        prompt_tokens=result.prompt_tokens or estimate_tokens("Ответь только словом OK."),
        completion_tokens=result.completion_tokens or estimate_tokens(result.content),
        raw=result.raw,
        provider=result.provider,
    )
    cost = ai_estimated_cost_rub(
        normalized.model,
        normalized.prompt_tokens,
        normalized.completion_tokens,
        ai_usd_rub_rate(session),
    )
    add_ai_usage(session, normalized, cost)
    record_feature_usage(
        session,
        feature="key_test",
        provider=normalized.provider,
        model=normalized.model,
        prompt_tokens=normalized.prompt_tokens,
        completion_tokens=normalized.completion_tokens,
        cost_rub=cost,
    )
    complete_agent_run(
        run,
        input_tokens=normalized.prompt_tokens,
        output_tokens=normalized.completion_tokens,
        estimated_cost=cost,
    )
    session.commit()
    return {
        "ok": True,
        "provider": normalized.provider,
        "model": normalized.model,
        "cost_rub": cost,
    }


HERMES_SETTING_KEYS = {
    "ai_enabled",
    "hermes_briefing_enabled",
    "hermes_auto_level_one_enabled",
    "hermes_auto_template_reminders",
    "hermes_mass_action_pin_threshold",
    "ai_tenant_mode",
    "ai_monthly_budget_rub",
    "ai_daily_call_limit",
    "ai_supervisor_time",
    "ai_feature_owner_chat_daily_limit",
    "ai_feature_tenant_chat_daily_limit",
    "ai_feature_case_details_daily_limit",
    "ai_feature_skill_builder_daily_limit",
    "ai_feature_deep_audit_daily_limit",
    "ai_feature_key_test_daily_limit",
    "ai_output_chars_daily_briefing",
    "ai_output_chars_owner_chat",
    "ai_output_chars_tenant_chat",
    "ai_output_chars_case_details",
    "ai_output_chars_skill_proposal",
}


def hermes_settings_payload(session: Session) -> dict[str, Any]:
    settings = get_settings(session)
    return {key: settings.get(key) for key in HERMES_SETTING_KEYS}


def hermes_control_center_payload(session: Session) -> dict[str, Any]:
    result = control_center_snapshot(
        session,
        enabled=setting_bool_value(get_setting_value(session, "ai_enabled")),
        monthly_budget_rub=ai_monthly_budget_rub(session),
    )
    result["settings"] = hermes_settings_payload(session)
    result["briefing_preview"] = briefing_preview(
        session,
        max_chars=ai_feature_output_limit(session, "daily_briefing"),
    )
    session.commit()
    return result


@app.get("/api/hermes")
@app.get("/api/android/hermes/summary")
def api_hermes_control_center(session: Session = Depends(get_session)) -> dict[str, Any]:
    return hermes_control_center_payload(session)


@app.get("/api/hermes/overview")
def api_hermes_overview(session: Session = Depends(get_session)) -> dict[str, Any]:
    result = hermes_overview(
        session,
        enabled=setting_bool_value(get_setting_value(session, "ai_enabled")),
        monthly_budget_rub=ai_monthly_budget_rub(session),
    )
    session.commit()
    return result


@app.get("/api/hermes/cases")
@app.get("/api/android/hermes/cases")
def api_hermes_cases(
    status: str = "",
    severity: str = "",
    property_id: int | None = None,
    session: Session = Depends(get_session),
) -> list[dict[str, Any]]:
    result = hermes_list_cases(
        session,
        status=status,
        severity=severity,
        property_id=property_id,
    )
    session.commit()
    return result


@app.get("/api/hermes/cases/{case_id}")
@app.get("/api/android/hermes/cases/{case_id}")
def api_hermes_case_details(case_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    reconcile_operational_cases(session)
    result = hermes_case_details(session, case_id)
    if result is None:
        raise HTTPException(404, "Кейс не найден")
    session.commit()
    return result


@app.post("/api/hermes/cases/{case_id}/close")
def api_hermes_close_case(
    case_id: int,
    payload: dict[str, Any] | None = None,
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    try:
        item = hermes_close_case(session, case_id, str((payload or {}).get("reason") or "Закрыто владельцем"))
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc
    session.commit()
    return case_snapshot(session, item)


@app.post("/api/hermes/cases/{case_id}/snooze")
def api_hermes_snooze_case(
    case_id: int,
    payload: dict[str, Any] | None = None,
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    raw_until = str((payload or {}).get("until") or "").strip()
    try:
        until = datetime.fromisoformat(raw_until) if raw_until else utc_now() + timedelta(days=1)
        item = hermes_snooze_case(session, case_id, until)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    session.commit()
    return case_snapshot(session, item)


@app.get("/api/hermes/commitments")
@app.get("/api/android/hermes/commitments")
def api_hermes_commitments(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    reconcile_owner_commitments(session)
    rows = session.scalars(
        select(OwnerCommitment).order_by(OwnerCommitment.status, OwnerCommitment.due_at, OwnerCommitment.id)
    ).all()
    result = [serialize_commitment(item) for item in rows]
    session.commit()
    return result


@app.post("/api/hermes/commitments/{commitment_id}/complete")
def api_hermes_complete_commitment(
    commitment_id: int,
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    try:
        item = complete_owner_commitment(session, commitment_id)
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc
    session.commit()
    return serialize_commitment(item)


@app.post("/api/hermes/commitments/{commitment_id}/postpone")
def api_hermes_postpone_commitment(
    commitment_id: int,
    payload: dict[str, Any] | None = None,
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    try:
        days = min(30, max(1, int((payload or {}).get("days") or 1)))
        item = postpone_owner_commitment(session, commitment_id, days=days)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    session.commit()
    return serialize_commitment(item)


@app.get("/api/hermes/preferences")
@app.get("/api/android/hermes/preferences")
def api_hermes_preferences(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    rows = session.scalars(
        select(OwnerPreference).order_by(OwnerPreference.enabled.desc(), OwnerPreference.scope, OwnerPreference.id)
    ).all()
    return [serialize_preference(item) for item in rows]


@app.post("/api/hermes/preferences")
def api_hermes_create_preference(
    payload: dict[str, Any],
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    scope = str(payload.get("scope") or "global").strip()[:120]
    key = str(payload.get("key") or "").strip()[:120]
    mode = str(payload.get("mode") or "persistent").strip()
    if not key or mode not in {"persistent", "once", "until_date"}:
        raise HTTPException(400, "Укажите ключ и корректный режим предпочтения")
    existing = session.scalar(
        select(OwnerPreference).where(
            OwnerPreference.scope == scope,
            OwnerPreference.key == key,
            OwnerPreference.mode == mode,
            OwnerPreference.enabled.is_(True),
        )
    )
    item = existing or OwnerPreference(scope=scope, key=key, mode=mode, source="owner_control_center")
    if not existing:
        session.add(item)
    value = payload.get("value") if isinstance(payload.get("value"), dict) else {"value": payload.get("value")}
    item.value_json = json.dumps(value, ensure_ascii=False, sort_keys=True)
    item.enabled = True
    item.consumed_at = None
    raw_until = str(payload.get("valid_until") or "").strip()
    item.valid_until = datetime.fromisoformat(raw_until) if raw_until else None
    session.commit()
    return serialize_preference(item)


@app.patch("/api/hermes/preferences/{preference_id}")
def api_hermes_update_preference(
    preference_id: int,
    payload: dict[str, Any],
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    item = session.get(OwnerPreference, preference_id)
    if not item:
        raise HTTPException(404, "Предпочтение не найдено")
    if "value" in payload:
        value = payload["value"] if isinstance(payload["value"], dict) else {"value": payload["value"]}
        item.value_json = json.dumps(value, ensure_ascii=False, sort_keys=True)
    if "enabled" in payload:
        item.enabled = setting_bool_value(payload["enabled"])
    if "mode" in payload and str(payload["mode"]) in {"persistent", "once", "until_date"}:
        item.mode = str(payload["mode"])
    item.updated_at = utc_now()
    session.commit()
    return serialize_preference(item)


@app.delete("/api/hermes/preferences/{preference_id}")
def api_hermes_disable_preference(
    preference_id: int,
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    item = session.get(OwnerPreference, preference_id)
    if not item:
        raise HTTPException(404, "Предпочтение не найдено")
    item.enabled = False
    item.updated_at = utc_now()
    session.commit()
    return serialize_preference(item)


@app.get("/api/hermes/skills")
def api_hermes_skills(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    rows = session.scalars(select(AiSkill).order_by(AiSkill.name, AiSkill.version.desc())).all()
    return [serialize_skill(item) for item in rows]


def require_hermes_skill(session: Session, skill_id: int) -> AiSkill:
    item = session.get(AiSkill, skill_id)
    if not item:
        raise HTTPException(404, "Навык не найден")
    return item


@app.post("/api/hermes/skills/{skill_id}/dry-run")
def api_hermes_skill_dry_run(
    skill_id: int,
    payload: dict[str, Any] | None = None,
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    skill = require_hermes_skill(session, skill_id)
    case_id = int((payload or {}).get("case_id") or 0)
    case = session.get(OperationalCase, case_id) if case_id else None
    return dry_run_skill(skill, case=case)


@app.post("/api/hermes/skills/{skill_id}/activate")
def api_hermes_activate_skill(skill_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    try:
        item = activate_skill(session, skill_id)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    session.commit()
    return serialize_skill(item)


@app.post("/api/hermes/skills/{skill_id}/disable")
def api_hermes_disable_skill(skill_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    try:
        item = disable_skill(session, skill_id)
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc
    session.commit()
    return serialize_skill(item)


@app.post("/api/hermes/skills/{skill_id}/rollback")
def api_hermes_rollback_skill(skill_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    try:
        item = rollback_skill(session, skill_id)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    session.commit()
    return serialize_skill(item)


@app.post("/api/hermes/skills/{skill_id}/version")
def api_hermes_version_skill(
    skill_id: int,
    payload: dict[str, Any],
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    changes: dict[str, Any] = {}
    json_fields = {
        "trigger_config_json",
        "preconditions_json",
        "steps_json",
        "allowed_tools_json",
        "autonomous_tools_json",
        "confirmation_required_tools_json",
        "success_metric_json",
    }
    for key, value in payload.items():
        changes[key] = json.dumps(value, ensure_ascii=False, sort_keys=True) if key in json_fields and not isinstance(value, str) else value
    try:
        item = version_skill(session, skill_id, changes)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    session.commit()
    return serialize_skill(item)


@app.get("/api/hermes/strategies")
def api_hermes_strategies(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    rows = session.scalars(select(TenantStrategyProfile).order_by(TenantStrategyProfile.contract_id)).all()
    return [serialize_strategy(item) for item in rows]


@app.patch("/api/hermes/strategies/{strategy_id}")
def api_hermes_update_strategy(
    strategy_id: int,
    payload: dict[str, Any],
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    item = session.get(TenantStrategyProfile, strategy_id)
    if not item:
        raise HTTPException(404, "Стратегия не найдена")
    allowed = {
        "preferred_channel",
        "preferred_contact_window",
        "response_speed_bucket",
        "reminder_stage",
        "soft_reminder_effectiveness",
        "broken_promises_count",
        "active_payment_situation",
        "escalation_threshold",
        "strategy_source",
    }
    unknown = set(payload) - allowed - {"owner_only_topics"}
    if unknown:
        raise HTTPException(400, "Неизвестные поля стратегии: " + ", ".join(sorted(unknown)))
    if "preferred_channel" in payload and str(payload["preferred_channel"]) not in {"telegram", "whatsapp", "phone"}:
        raise HTTPException(400, "Канал должен быть telegram, whatsapp или phone")
    if "preferred_contact_window" in payload and not re.fullmatch(
        r"(?:[01]\d|2[0-3]):[0-5]\d-(?:[01]\d|2[0-3]):[0-5]\d",
        str(payload["preferred_contact_window"]),
    ):
        raise HTTPException(400, "Окно связи должно иметь формат 10:00-20:00")
    if "response_speed_bucket" in payload and str(payload["response_speed_bucket"]) not in {
        "unknown",
        "fast",
        "same_day",
        "slow",
    }:
        raise HTTPException(400, "Неизвестная скорость ответа")
    if "escalation_threshold" in payload:
        try:
            threshold = int(payload["escalation_threshold"])
        except (TypeError, ValueError) as exc:
            raise HTTPException(400, "Порог эскалации должен быть целым числом") from exc
        if not 1 <= threshold <= 20:
            raise HTTPException(400, "Порог эскалации должен быть от 1 до 20")
        payload["escalation_threshold"] = threshold
    if "soft_reminder_effectiveness" in payload:
        try:
            effectiveness = float(payload["soft_reminder_effectiveness"])
        except (TypeError, ValueError) as exc:
            raise HTTPException(400, "Эффективность напоминаний должна быть числом") from exc
        if not 0 <= effectiveness <= 1:
            raise HTTPException(400, "Эффективность напоминаний должна быть от 0 до 1")
        payload["soft_reminder_effectiveness"] = effectiveness
    if "broken_promises_count" in payload:
        try:
            broken_promises = int(payload["broken_promises_count"])
        except (TypeError, ValueError) as exc:
            raise HTTPException(400, "Количество нарушенных обещаний должно быть целым числом") from exc
        if broken_promises < 0:
            raise HTTPException(400, "Количество нарушенных обещаний не может быть отрицательным")
        payload["broken_promises_count"] = broken_promises
    for key, value in payload.items():
        if key in allowed:
            setattr(item, key, value)
    if "owner_only_topics" in payload and isinstance(payload["owner_only_topics"], list):
        item.owner_only_topics_json = json.dumps(payload["owner_only_topics"], ensure_ascii=False)
    item.last_strategy_change = utc_now()
    session.commit()
    return serialize_strategy(item)


@app.post("/api/hermes/strategies/{strategy_id}/reset")
def api_hermes_reset_strategy(strategy_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    item = session.get(TenantStrategyProfile, strategy_id)
    if not item:
        raise HTTPException(404, "Стратегия не найдена")
    item.preferred_channel = "telegram"
    item.preferred_contact_window = "10:00-20:00"
    item.response_speed_bucket = "unknown"
    item.reminder_stage = "pre_due"
    item.soft_reminder_effectiveness = 0
    item.broken_promises_count = 0
    item.active_payment_situation = ""
    item.escalation_threshold = 3
    item.owner_only_topics_json = "[]"
    item.strategy_source = "owner_reset"
    item.last_strategy_change = utc_now()
    session.commit()
    return serialize_strategy(item)


@app.get("/api/hermes/proposals")
@app.get("/api/android/hermes/proposals")
def api_hermes_proposals(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    rows = session.scalars(select(AgentActionProposal).order_by(AgentActionProposal.id.desc()).limit(200)).all()
    return [serialize_proposal(item) for item in rows]


def decide_hermes_proposal(
    session: Session,
    proposal_id: int,
    *,
    decision: str,
    actor: str,
    confirmation_pin: str = "",
) -> AgentActionProposal:
    proposal = session.scalar(
        select(AgentActionProposal).where(AgentActionProposal.id == proposal_id).with_for_update()
    )
    if not proposal:
        raise HTTPException(404, "Предложение не найдено")
    if proposal.status != "pending":
        return proposal
    if proposal.expires_at and proposal.expires_at < utc_now():
        proposal.status = "expired"
        proposal.error_text = "Истёк срок подтверждения"
        session.commit()
        return proposal
    if decision == "reject":
        proposal.status = "rejected"
        proposal.result_text = "Отклонено владельцем"
        proposal.confirmed_by = actor[:80]
        session.add(
            AiActionLog(
                conversation_id=proposal.conversation_id,
                lease_id=proposal.lease_id,
                actor_role="owner",
                action_type=proposal.action_type,
                status="rejected",
                payload_json=proposal.payload_json,
                note=f"proposal rejected from {actor}",
            )
        )
        session.commit()
        return proposal
    if int(proposal.safety_level or 0) >= 3:
        owner_hash = configured_environment_pin_hash("owner") or get_setting_value(session, PIN_SETTING_KEYS["owner"])
        if not confirmation_pin or not verify_pin(owner_hash, confirmation_pin):
            raise HTTPException(403, "Для критического или массового действия повторно введите PIN владельца")
    proposal.status = "approved"
    proposal.confirmed_at = utc_now()
    proposal.confirmed_by = actor[:80]
    session.flush()
    try:
        result_text = execute_agent_action(session, proposal)
    except HTTPException as exc:
        proposal.status = "failed"
        proposal.result_text = str(exc.detail)
        proposal.error_text = str(exc.detail)
        session.commit()
        raise HTTPException(409, str(exc.detail)) from exc
    proposal.status = "executed"
    proposal.executed_at = utc_now()
    proposal.result_text = result_text
    proposal.error_text = ""
    reconcile_operational_cases(session)
    session.add(
        AiActionLog(
            conversation_id=proposal.conversation_id,
            lease_id=proposal.lease_id,
            actor_role="owner",
            action_type=proposal.action_type,
            status="executed",
            payload_json=proposal.payload_json,
            note=f"{actor}: {result_text}"[:2000],
        )
    )
    session.commit()
    return proposal


@app.post("/api/hermes/proposals/{proposal_id}/confirm")
@app.post("/api/android/hermes/proposals/{proposal_id}/confirm")
def api_hermes_confirm_proposal(
    proposal_id: int,
    request: Request,
    payload: dict[str, Any] | None = None,
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    actor = "android" if request.url.path.startswith("/api/android/") else "web"
    item = decide_hermes_proposal(
        session,
        proposal_id,
        decision="confirm",
        actor=actor,
        confirmation_pin=str((payload or {}).get("confirmation_pin") or ""),
    )
    return serialize_proposal(item)


@app.post("/api/hermes/proposals/{proposal_id}/reject")
@app.post("/api/android/hermes/proposals/{proposal_id}/reject")
def api_hermes_reject_proposal(
    proposal_id: int,
    request: Request,
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    actor = "android" if request.url.path.startswith("/api/android/") else "web"
    return serialize_proposal(
        decide_hermes_proposal(session, proposal_id, decision="reject", actor=actor)
    )


@app.get("/api/hermes/usage")
@app.get("/api/android/hermes/usage")
def api_hermes_usage(session: Session = Depends(get_session)) -> dict[str, Any]:
    return hermes_usage_summary(session)


@app.get("/api/hermes/runs")
def api_hermes_runs(limit: int = 100, session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    rows = session.scalars(
        select(HermesAgentRun).order_by(HermesAgentRun.id.desc()).limit(min(500, max(1, limit)))
    ).all()
    return [serialize_run(item) for item in rows]


@app.get("/api/hermes/runs/{run_id}")
def api_hermes_run_debug(run_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    item = session.get(HermesAgentRun, run_id)
    if not item:
        raise HTTPException(404, "Запуск не найден")
    return serialize_run(item, include_debug=True)


@app.get("/api/hermes/briefing/preview")
def api_hermes_briefing_preview(session: Session = Depends(get_session)) -> dict[str, Any]:
    result = briefing_preview(session, max_chars=ai_feature_output_limit(session, "daily_briefing"))
    session.rollback()
    return result


@app.get("/api/hermes/settings")
@app.get("/api/android/hermes/settings")
def api_hermes_settings(session: Session = Depends(get_session)) -> dict[str, Any]:
    return hermes_settings_payload(session)


@app.post("/api/hermes/settings")
@app.post("/api/android/hermes/settings")
def api_save_hermes_settings(
    payload: dict[str, Any],
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    filtered = {key: value for key, value in payload.items() if key in HERMES_SETTING_KEYS}
    save_settings(session, filtered)
    return hermes_settings_payload(session)


@app.get("/api/month-progress")
def api_month_progress(year: int, month: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    if month < 1 or month > 12:
        raise HTTPException(400, "Месяц должен быть от 1 до 12")
    start, end = month_range(year, month)
    today = date.today()
    generate_rent_charges(session)
    session.commit()
    rent_charges = session.scalars(
        select(RentCharge)
        .join(Lease)
        .join(Apartment)
        .where(
            Lease.active.is_(True),
            Apartment.active.is_(True),
            RentCharge.due_date >= start,
            RentCharge.due_date <= end,
        )
        .order_by(RentCharge.due_date, RentCharge.id)
    ).all()
    rent_charges = [charge for charge in rent_charges if not lease_ignored(session, charge.lease_id)]
    for charge in rent_charges:
        update_rent_charge_status(charge, today)

    utility_bills = session.scalars(
        select(UtilityBill)
        .where(
            UtilityBill.period_start >= start,
            UtilityBill.period_start <= end,
        )
        .order_by(UtilityBill.period_start, UtilityBill.service_id, UtilityBill.id)
    ).all()

    return {
        "year": year,
        "month": month,
        "period_start": start.isoformat(),
        "period_end": end.isoformat(),
        "summary": month_dashboard_summary(session, year, month, today),
        "rent_charges": [serialize_rent_charge(charge, session) for charge in rent_charges],
        "utility_bills": [serialize_bill(bill, session) for bill in utility_bills],
        "provider_readings": provider_reading_statuses_for_month(session, year, month, today),
    }


@app.get("/api/admin/database-export")
def export_database(session: Session = Depends(get_session)) -> Response:
    snapshot = current_database_snapshot(session)
    raw = json.dumps(snapshot, ensure_ascii=False, indent=2).encode("utf-8")
    filename = f"rental-manager-db-export-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
    return Response(
        content=raw,
        media_type="application/json; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(raw)),
            "X-Content-Type-Options": "nosniff",
        },
    )


@app.post("/api/admin/database-backup")
def export_protected_database_backup(
    payload: dict[str, Any],
    session: Session = Depends(get_session),
) -> Response:
    if str(payload.get("confirmation_text") or "").strip().upper() != DB_BACKUP_CONFIRM_TEXT:
        raise HTTPException(400, f'Для защищённой копии введите "{DB_BACKUP_CONFIRM_TEXT}".')
    snapshot = current_database_snapshot(session, include_secrets=True)
    raw = json.dumps(snapshot, ensure_ascii=False, indent=2).encode("utf-8")
    filename = f"rental-manager-protected-backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
    return Response(
        content=raw,
        media_type="application/json; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(raw)),
            "Cache-Control": "no-store",
            "X-Content-Type-Options": "nosniff",
        },
    )


@app.post("/api/admin/database-import/inspect")
async def inspect_database_import(file: UploadFile = File(...)) -> dict[str, Any]:
    payload = parse_database_import_bytes(await file.read())
    return inspect_database_import_payload(payload)


@app.post("/api/admin/database-import")
async def import_database(
    file: UploadFile = File(...),
    confirmation_text: str = Form(""),
    confirm_replace: bool = Form(False),
    create_backup: bool = Form(True),
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    if not confirm_replace:
        raise HTTPException(400, "Сначала подтвердите, что текущая база будет полностью заменена.")
    if confirmation_text.strip().upper() != DB_IMPORT_CONFIRM_TEXT:
        raise HTTPException(400, f'Для импорта введите подтверждение "{DB_IMPORT_CONFIRM_TEXT}".')
    payload = parse_database_import_bytes(await file.read())
    inspection = inspect_database_import_payload(payload)
    backup_path = ""
    if create_backup:
        backup_path = str(save_database_backup(session))
    try:
        result = apply_database_import_payload(session, payload)
        session.commit()
    except Exception:
        session.rollback()
        raise
    return {
        "ok": True,
        "inspection": inspection,
        "imported": result,
        "backup_path": backup_path,
    }


@app.get("/api/leases/{lease_id}/cadence")
def api_get_lease_cadence(lease_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    lease = session.get(Lease, lease_id)
    if not lease:
        raise HTTPException(404, "Договор не найден")
    return {"automation": get_lease_automation(session, lease_id)}


@app.post("/api/leases/{lease_id}/cadence")
def api_set_lease_cadence(lease_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    lease = session.get(Lease, lease_id)
    if not lease:
        raise HTTPException(404, "Договор не найден")
    automation = save_lease_automation(session, lease_id, payload)
    session.commit()
    return {"automation": automation}


@app.delete("/api/leases/{lease_id}/cadence")
def api_clear_lease_cadence(lease_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    lease = session.get(Lease, lease_id)
    if not lease:
        raise HTTPException(404, "Договор не найден")
    for template_key in LEASE_AUTOMATION_TEMPLATES:
        clear_lease_cadence(session, lease_id, template_key)
    session.commit()
    return {"automation": get_lease_automation(session, lease_id)}


@app.patch("/api/leases/{lease_id}/automation")
def api_update_lease_automation(lease_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    lease = session.get(Lease, lease_id)
    if not lease:
        raise HTTPException(404, "Договор не найден")
    automation = save_lease_automation(session, lease_id, payload)
    session.commit()
    return {"automation": automation}


@app.patch("/api/leases/{lease_id}/ignore")
def api_update_lease_ignore(lease_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    lease = session.get(Lease, lease_id)
    if not lease:
        raise HTTPException(404, "Договор не найден")
    set_lease_ignored(session, lease_id, setting_bool_value(payload.get("ignored")))
    session.commit()
    return serialize_lease(lease, session)


def import_release_baseline(session: Session) -> dict[str, int]:
    for model in [
        AgentActionProposal,
        AgentMemory,
        AgentTask,
        AgentTenantState,
        PaymentSituation,
        AiActionLog,
        AiMessage,
        AiConversation,
        AiUsageDaily,
        MessageLog,
        PaymentReceipt,
        ManualDebt,
        UtilityBillLine,
        UtilityBill,
        RentCharge,
        Lease,
        Tenant,
        Expense,
        MeterReading,
        Tariff,
    ]:
        session.execute(delete(model))
    session.flush()
    seed_if_empty(session)
    scripts_dir = ROOT_DIR / "scripts"
    import sys

    if str(scripts_dir) not in sys.path:
        sys.path.insert(0, str(scripts_dir))
    import seed_from_screenshots as legacy_seed  # type: ignore

    legacy_seed.seed_tariffs(session)
    leases = legacy_seed.seed_leases(session)
    legacy_seed.generate_rent_charges(session, until=date(2026, 5, 31))
    legacy_seed.mark_old_rent_paid(session)
    legacy_seed.seed_rent_payments(session, leases)
    legacy_seed.seed_utility_payments(session, leases)
    legacy_seed.seed_meter_readings(session)
    legacy_seed.seed_expenses(session)
    session.flush()
    return {
        "tenants": session.scalar(select(func.count(Tenant.id))) or 0,
        "leases": session.scalar(select(func.count(Lease.id))) or 0,
        "receipts": session.scalar(select(func.count(PaymentReceipt.id))) or 0,
        "readings": session.scalar(select(func.count(MeterReading.id))) or 0,
    }


@app.post("/api/admin/import-release-baseline")
def api_import_release_baseline(session: Session = Depends(get_session)) -> dict[str, int]:
    result = import_release_baseline(session)
    session.commit()
    return result


def add_report_issue(issues: list[dict[str, Any]], severity: str, title: str, count: int = 1, detail: str = "") -> None:
    if count <= 0:
        return
    issues.append(
        {
            "severity": severity,
            "title": title,
            "count": count,
            "detail": detail,
        }
    )


def utility_bill_for_month(session: Session, service: UtilityService, start: date, end: date) -> UtilityBill | None:
    next_day = end + timedelta(days=1)
    return session.scalar(
        select(UtilityBill)
        .where(
            UtilityBill.service_id == service.id,
            UtilityBill.period_start >= start,
            UtilityBill.period_start < next_day,
        )
        .order_by(UtilityBill.period_start.desc())
        .limit(1)
    )


def object_reading_in_month(session: Session, service: UtilityService, start: date, end: date) -> MeterReading | None:
    meter = session.scalar(
        select(Meter)
        .where(Meter.service_id == service.id, Meter.scope == "object")
        .limit(1)
    )
    if not meter:
        return None
    return session.scalar(
        select(MeterReading)
        .where(
            MeterReading.meter_id == meter.id,
            MeterReading.reading_date >= start,
            MeterReading.reading_date <= end,
        )
        .order_by(MeterReading.reading_date.desc())
        .limit(1)
    )


def service_due_date(year: int, month: int, day: int) -> date:
    _, last_day = calendar.monthrange(year, month)
    return date(year, month, min(max(int(day or 1), 1), last_day))


def provider_reading_statuses_for_month(
    session: Session,
    year: int,
    month: int,
    today: date | None = None,
) -> list[dict[str, Any]]:
    today = today or date.today()
    start, end = month_range(year, month)
    statuses: list[dict[str, Any]] = []
    services = session.scalars(
        select(UtilityService)
        .options(joinedload(UtilityService.object))
        .where(UtilityService.active.is_(True))
        .order_by(UtilityService.object_id, UtilityService.kind, UtilityService.id)
    ).all()
    service_ids = [service.id for service in services]
    readings_by_service: dict[int, date] = {}
    if service_ids:
        rows = session.execute(
            select(Meter.service_id, func.max(MeterReading.reading_date))
            .select_from(Meter)
            .join(MeterReading, MeterReading.meter_id == Meter.id)
            .where(
                Meter.service_id.in_(service_ids),
                Meter.scope == "object",
                MeterReading.reading_date >= start,
                MeterReading.reading_date <= end,
            )
            .group_by(Meter.service_id)
        ).all()
        readings_by_service = {int(service_id): reading_date for service_id, reading_date in rows if reading_date}
    for service in services:
        due = service_due_date(year, month, service.provider_reading_due_day)
        reading_date = readings_by_service.get(int(service.id))
        alert_from = due - timedelta(days=3)
        if reading_date:
            status = "paid"
            alert = False
        elif today < alert_from:
            status = "upcoming"
            alert = False
        elif today <= due:
            status = "pending"
            alert = True
        else:
            status = "overdue"
            alert = True
        statuses.append(
            {
                "id": service.id,
                "service_id": service.id,
                "object_id": service.object_id,
                "object": service.object.name if service.object else "",
                "service": service.name,
                "kind": service.kind,
                "year": year,
                "month": month,
                "period_start": start.isoformat(),
                "period_end": end.isoformat(),
                "due_date": due.isoformat(),
                "reading_date": reading_date.isoformat() if reading_date else None,
                "status": status,
                "alert": alert,
                "days_left": (due - today).days,
            }
        )
    return statuses


def monthly_report_status(session: Session, year: int, month: int, today: date | None = None, kind: str = "full") -> dict[str, Any]:
    today = today or date.today()
    start, end = month_range(year, month)
    kind = "preliminary" if kind == "preliminary" else "full"
    issues: list[dict[str, Any]] = []

    rent_charges = session.scalars(
        select(RentCharge)
        .join(Lease)
        .join(Apartment)
        .where(Apartment.active.is_(True))
        .where(RentCharge.due_date >= start, RentCharge.due_date <= end)
        .order_by(RentCharge.due_date)
    ).all()
    for charge in rent_charges:
        update_rent_charge_status(charge, today)
    critical_rent = [charge for charge in rent_charges if charge.status in {"overdue", "pending", "partial"}]
    deferred_rent = [charge for charge in rent_charges if charge.status == "deferred"]
    add_report_issue(issues, "danger", "не закрыта аренда", len(critical_rent), "Есть платежи аренды без полного закрытия.")
    add_report_issue(issues, "warn", "есть отсрочки по аренде", len(deferred_rent), "Проверьте актуальность сроков отсрочки.")

    utility_lines = session.scalars(
        select(UtilityBillLine)
        .join(UtilityBill)
        .join(Apartment, UtilityBillLine.apartment_id == Apartment.id)
        .where(
            Apartment.active.is_(True),
            UtilityBillLine.due_date >= start,
            UtilityBillLine.due_date <= end,
        )
    ).all()
    for line in utility_lines:
        update_utility_line_status(line, today)
    tenant_utility_debts = [line for line in utility_lines if line.status in {"overdue", "partial", "issued", "pending"}]
    add_report_issue(issues, "warn", "долги жильцов за коммуналку", len(tenant_utility_debts), "Коммуналка жильцов не закрыта полностью.")

    missing_bills = 0
    unpaid_provider_bills = 0
    missing_readings = 0
    for service in session.scalars(select(UtilityService).where(UtilityService.active.is_(True))).all():
        bill = utility_bill_for_month(session, service, start, end)
        if not bill:
            missing_bills += 1
        elif not bill.provider_paid:
            unpaid_provider_bills += 1
        if not object_reading_in_month(session, service, start, end):
            missing_readings += 1

    add_report_issue(issues, "danger", "нет коммунального счёта", missing_bills, "По услуге не создан месячный счёт.")
    add_report_issue(issues, "danger", "поставщик не оплачен", unpaid_provider_bills, "Деньги поставщику не отмечены как оплаченные.")
    add_report_issue(issues, "danger", "нет общих показаний", missing_readings, "За месяц не найдены общедомовые показания.")

    suspicious_receipts = session.scalar(
        select(PaymentReceipt.id)
        .where(PaymentReceipt.status == "suspicious", PaymentReceipt.created_at >= datetime.combine(start, datetime.min.time()))
        .limit(1)
    )
    add_report_issue(issues, "danger", "есть подозрительные чеки", 1 if suspicious_receipts else 0, "Проверить получателя и назначение платежа.")

    warning_count = sum(issue["count"] for issue in issues if issue["severity"] == "warn")
    danger_count = sum(issue["count"] for issue in issues if issue["severity"] == "danger")
    total_issues = warning_count + danger_count
    if danger_count >= 5 or total_issues >= 10:
        severity = "critical"
        label = "много проблем"
    elif danger_count:
        severity = "danger"
        label = "критические проблемы"
    elif warning_count:
        severity = "warn"
        label = "есть вопросы"
    else:
        severity = "ok"
        label = "закрыт"

    title_prefix = "Предварительный отчёт за" if kind == "preliminary" else "Готов отчёт за"
    key = monthly_report_key(year, month, kind)
    return {
        "key": key,
        "kind": kind,
        "year": year,
        "month": month,
        "month_name": MONTH_NAMES[month - 1],
        "title": f"{title_prefix} {MONTH_NAMES[month - 1]} {year}",
        "period_start": start.isoformat(),
        "period_end": end.isoformat(),
        "severity": severity,
        "label": label,
        "issue_count": total_issues,
        "warning_count": warning_count,
        "danger_count": danger_count,
        "issues": issues,
        "accepted": key in accepted_monthly_report_keys(session),
        "download_url": f"/api/reports/monthly.xlsx?year={year}&month={month}",
    }


def build_monthly_reports(session: Session, today: date | None = None) -> list[dict[str, Any]]:
    reports = [
        monthly_report_status(session, entry["year"], entry["month"], today, entry["kind"])
        for entry in iter_dashboard_report_entries(today)
    ]
    return [report for report in reports if not report["accepted"]]


def build_monthly_report_shells(session: Session, today: date | None = None) -> list[dict[str, Any]]:
    today = today or date.today()
    accepted = accepted_monthly_report_keys(session)
    reports: list[dict[str, Any]] = []
    for entry in iter_dashboard_report_entries(today):
        year = int(entry["year"])
        month = int(entry["month"])
        kind = "preliminary" if entry.get("kind") == "preliminary" else "full"
        key = monthly_report_key(year, month, kind)
        if key in accepted:
            continue
        start, end = month_range(year, month)
        title_prefix = "Предварительный отчёт за" if kind == "preliminary" else "Готов отчёт за"
        reports.append(
            {
                "key": key,
                "kind": kind,
                "year": year,
                "month": month,
                "month_name": MONTH_NAMES[month - 1],
                "title": f"{title_prefix} {MONTH_NAMES[month - 1]} {year}",
                "period_start": start.isoformat(),
                "period_end": end.isoformat(),
                "severity": "warn",
                "label": "ждёт проверки",
                "issue_count": 0,
                "warning_count": 0,
                "danger_count": 0,
                "issues": [],
                "accepted": False,
                "download_url": f"/api/reports/monthly.xlsx?year={year}&month={month}",
                "quick": True,
            }
        )
    return reports


def notify_available_monthly_reports(session: Session, today: date | None = None) -> int:
    today = today or date.today()
    owner_id = telegram_owner_chat_id(session)
    if not owner_id or not telegram_token(session):
        return 0
    notified = notified_monthly_report_keys(session)
    sent = 0
    last_complete = today.replace(day=1) - timedelta(days=1)
    fresh_reports = []
    if today.day >= 25:
        fresh_reports.append(monthly_report_status(session, today.year, today.month, today, "preliminary"))
    if today.day == 1 and last_complete >= REPORT_START_MONTH:
        fresh_reports.append(monthly_report_status(session, last_complete.year, last_complete.month, today, "full"))
    for report in fresh_reports:
        if report.get("accepted"):
            continue
        key = report["key"]
        if key in notified:
            continue
        try:
            send_telegram_text(
                session,
                owner_id,
                f"{report['title']}\nСтатус: {report['label']}. Проблем: {report['issue_count']}.\nОткрой дашборд и проверь отчёт.",
                app_keyboard(app_base_url(session)),
            )
        except HTTPException:
            continue
        notified.add(key)
        sent += 1
    if sent:
        set_internal_json(session, "notified_monthly_reports", sorted(notified))
    return sent


def active_leases_for_month(session: Session, start: date, end: date) -> list[Lease]:
    leases = session.scalars(
        select(Lease)
        .options(joinedload(Lease.apartment).joinedload(Apartment.object), joinedload(Lease.tenant))
        .join(Apartment)
        .where(Lease.active.is_(True), Apartment.active.is_(True))
        .order_by(Lease.start_date, Lease.id)
    ).all()
    return [
        lease
        for lease in leases
        if not lease_ignored(session, lease.id)
        and lease.start_date <= end
        and (lease.end_date is None or lease.end_date >= start)
    ]


def utility_advance_receipt_applies(receipt: PaymentReceipt) -> bool:
    return receipt.channel == UTILITY_ADVANCE_CHANNEL or receipt.source == UTILITY_ADVANCE_SOURCE


def estimated_utility_advance_due(
    session: Session,
    leases: list[Lease],
    start: date,
    current_lines_by_lease: dict[int, float],
) -> float:
    lease_ids = [lease.id for lease in leases]
    if not lease_ids:
        return 0.0

    history_by_lease: dict[int, dict[date, float]] = {lease.id: {} for lease in leases}
    historical_lines = session.scalars(
        select(UtilityBillLine)
        .options(joinedload(UtilityBillLine.bill))
        .join(UtilityBill)
        .where(
            UtilityBillLine.lease_id.in_(lease_ids),
            UtilityBill.period_end < start,
            UtilityBill.status != "draft",
        )
        .order_by(UtilityBill.period_start.desc(), UtilityBillLine.id.desc())
    ).all()
    for line in historical_lines:
        if not line.lease_id or not line.bill:
            continue
        month_start = date(line.bill.period_start.year, line.bill.period_start.month, 1)
        lease_history = history_by_lease.setdefault(int(line.lease_id), {})
        lease_history[month_start] = money(lease_history.get(month_start, 0.0) + float(line.total_amount or 0))

    total = 0.0
    for lease in leases:
        current_total = money(current_lines_by_lease.get(lease.id, 0.0))
        if current_total > EPS:
            total += current_total
            continue
        recent = [
            value
            for _month, value in sorted(history_by_lease.get(lease.id, {}).items(), reverse=True)[:3]
            if value > EPS
        ]
        if recent:
            total += money(sum(recent) / len(recent))
    return money(total)


def month_dashboard_summary(
    session: Session,
    year: int,
    month: int,
    today: date | None = None,
) -> dict[str, Any]:
    today = today or date.today()
    start, end = month_range(year, month)
    active_leases = active_leases_for_month(session, start, end)
    active_lease_ids = {lease.id for lease in active_leases}

    rent_charges = session.scalars(
        select(RentCharge)
        .options(
            joinedload(RentCharge.lease).joinedload(Lease.apartment).joinedload(Apartment.object),
            joinedload(RentCharge.lease).joinedload(Lease.tenant),
        )
        .join(Lease)
        .join(Apartment)
        .where(
            Lease.active.is_(True),
            Apartment.active.is_(True),
            RentCharge.due_date >= start,
            RentCharge.due_date <= end,
        )
        .order_by(RentCharge.due_date, RentCharge.id)
    ).all()
    rent_charges = [
        charge
        for charge in rent_charges
        if charge.lease_id in active_lease_ids and not lease_ignored(session, charge.lease_id)
    ]
    salary_due = 0.0
    salary_paid = 0.0
    paid_count = 0
    pending_count = 0
    overdue_count = 0
    for charge in rent_charges:
        update_rent_charge_status(charge, today)
        salary_due = money(salary_due + float(charge.personal_due or 0))
        salary_paid = money(salary_paid + float(charge.personal_paid or 0))
        rent_debt = money(max(0.0, float(charge.ip_due or 0) - float(charge.ip_paid or 0)) + max(0.0, float(charge.personal_due or 0) - float(charge.personal_paid or 0)))
        if charge.status in {"paid", "paid_ahead"} or rent_debt <= EPS:
            paid_count += 1
        elif charge.status in {"overdue", "partial"} and charge.due_date <= today:
            overdue_count += 1
        else:
            pending_count += 1

    utility_bills = session.scalars(
        select(UtilityBill)
        .options(
            joinedload(UtilityBill.service).joinedload(UtilityService.object),
            selectinload(UtilityBill.lines).joinedload(UtilityBillLine.lease),
        )
        .where(UtilityBill.period_start >= start, UtilityBill.period_start <= end)
        .order_by(UtilityBill.period_start, UtilityBill.service_id, UtilityBill.id)
    ).all()
    issued_bills = [bill for bill in utility_bills if bill.status != "draft"]
    current_lines_by_lease: dict[int, float] = {}
    bill_payment_due = 0.0
    bill_payment_paid = 0.0
    for bill in utility_bills:
        for line in bill.lines:
            if not line.lease_id or line.lease_id not in active_lease_ids or lease_ignored(session, line.lease_id):
                continue
            update_utility_line_status(line, today)
            if utility_line_is_advance(line):
                continue
            current_lines_by_lease[int(line.lease_id)] = money(current_lines_by_lease.get(int(line.lease_id), 0.0) + float(line.total_amount or 0))
            if bill.status == "draft":
                continue
            bill_payment_due = money(bill_payment_due + float(line.total_amount or 0))
            bill_payment_paid = money(bill_payment_paid + float(line.paid_amount or 0))

    active_services = session.scalars(select(UtilityService).where(UtilityService.active.is_(True))).all()
    active_service_ids = {service.id for service in active_services}
    issued_service_ids = {bill.service_id for bill in issued_bills}
    utility_bills_issued = not active_service_ids or active_service_ids.issubset(issued_service_ids)
    utility_provider_paid = bool(issued_bills) and all(bill.provider_paid for bill in issued_bills)

    start_dt = datetime.combine(start, time.min)
    end_dt = datetime.combine(end + timedelta(days=1), time.min)
    month_receipts = session.scalars(
        select(PaymentReceipt)
        .where(
            PaymentReceipt.status == "accepted",
            PaymentReceipt.paid_at >= start_dt,
            PaymentReceipt.paid_at < end_dt,
        )
        .order_by(PaymentReceipt.paid_at, PaymentReceipt.id)
    ).all()
    advance_paid = money(
        sum(
            float(receipt.amount or 0)
            for receipt in month_receipts
            if receipt.lease_id in active_lease_ids
            and utility_advance_receipt_applies(receipt)
            and not lease_ignored(session, receipt.lease_id)
        )
    )
    advance_balance = money(
        float(session.scalar(select(func.coalesce(func.sum(UtilityAdvanceLedger.amount), 0))) or 0)
        + float(
            session.scalar(
                select(func.coalesce(func.sum(PaymentReceipt.amount), 0)).where(
                    PaymentReceipt.status == "accepted",
                    PaymentReceipt.channel == UTILITY_ADVANCE_CHANNEL,
                    PaymentReceipt.utility_line_id.is_(None),
                    PaymentReceipt.rent_charge_id.is_(None),
                )
            )
            or 0
        )
    )
    total_apartments = int(session.scalar(select(func.count(Apartment.id)).where(Apartment.active.is_(True))) or 0)
    occupied_apartments = len({lease.apartment_id for lease in active_leases})
    advance_due = estimated_utility_advance_due(session, active_leases, start, current_lines_by_lease)

    return {
        "year": year,
        "month": month,
        "period_start": start.isoformat(),
        "period_end": end.isoformat(),
        "salary_paid": salary_paid,
        "salary_due": salary_due,
        "utility_bills_issued": utility_bills_issued,
        "utility_provider_paid": utility_provider_paid,
        "bill_payment_paid": bill_payment_paid,
        "bill_payment_due": bill_payment_due,
        "advance_paid": advance_paid,
        "advance_due": advance_due,
        "advance_balance": advance_balance,
        "occupied": occupied_apartments,
        "total_apartments": total_apartments,
        "paid_count": paid_count,
        "pending_count": pending_count,
        "overdue_count": overdue_count,
    }


def dashboard_income_trend(session: Session, today: date | None = None, months: int = 6) -> list[dict[str, Any]]:
    today = today or date.today()
    cursor_year = today.year
    cursor_month = today.month
    periods: list[tuple[int, int]] = []
    for _ in range(max(1, months)):
        periods.append((cursor_year, cursor_month))
        cursor_month -= 1
        if cursor_month == 0:
            cursor_month = 12
            cursor_year -= 1
    periods.reverse()

    start_year, start_month = periods[0]
    start_at = datetime(start_year, start_month, 1)
    receipts = session.scalars(
        select(PaymentReceipt)
        .where(PaymentReceipt.status == "accepted", PaymentReceipt.paid_at >= start_at)
        .order_by(PaymentReceipt.paid_at, PaymentReceipt.id)
    ).all()
    totals = {(year, month): 0.0 for year, month in periods}
    for receipt in receipts:
        if not receipt.paid_at:
            continue
        key = (receipt.paid_at.year, receipt.paid_at.month)
        if key in totals:
            totals[key] = money(totals[key] + float(receipt.amount or 0))
    return [
        {"period": f"{year:04d}-{month:02d}", "amount": money(totals[(year, month)])}
        for year, month in periods
    ]


def build_dashboard(session: Session) -> dict[str, Any]:
    today = date.today()
    cutoff = configured_notification_cutoff_date(session)
    ignored_ids = ignored_lease_ids(session)

    charge_query = (
        select(RentCharge)
        .options(
            joinedload(RentCharge.lease).joinedload(Lease.apartment).joinedload(Apartment.object),
            joinedload(RentCharge.lease).joinedload(Lease.tenant),
        )
        .join(Lease)
        .join(Apartment)
        .where(Lease.active.is_(True), Apartment.active.is_(True))
        .order_by(RentCharge.due_date)
    )
    if cutoff:
        charge_query = charge_query.where(RentCharge.due_date >= cutoff)
    if ignored_ids:
        charge_query = charge_query.where(Lease.id.not_in(ignored_ids))
    charges = session.scalars(charge_query).all()
    charges = [charge for charge in charges if IGNORE_LEASE_MARK not in (charge.lease.notes or "")]
    for charge in charges:
        update_rent_charge_status(charge, today)

    utility_query = (
        select(UtilityBillLine)
        .options(
            joinedload(UtilityBillLine.apartment),
            joinedload(UtilityBillLine.lease).joinedload(Lease.tenant),
            joinedload(UtilityBillLine.bill).joinedload(UtilityBill.service).joinedload(UtilityService.object),
        )
        .join(Apartment)
        .where(Apartment.active.is_(True))
    )
    if cutoff:
        utility_query = utility_query.where(or_(UtilityBillLine.due_date.is_(None), UtilityBillLine.due_date >= cutoff))
    if ignored_ids:
        utility_query = utility_query.where(or_(UtilityBillLine.lease_id.is_(None), UtilityBillLine.lease_id.not_in(ignored_ids)))
    utility_lines = session.scalars(utility_query).all()
    utility_lines = [
        line
        for line in utility_lines
        if line.lease_id
        and line.lease
        and IGNORE_LEASE_MARK not in (line.lease.notes or "")
    ]
    for line in utility_lines:
        update_utility_line_status(line, today)

    def rent_debt(charge: RentCharge) -> float:
        return money(max(0, float(charge.ip_due or 0) - float(charge.ip_paid or 0)) + max(0, float(charge.personal_due or 0) - float(charge.personal_paid or 0)))

    def utility_debt(line: UtilityBillLine) -> float:
        return money(max(0, float(line.total_amount or 0) - float(line.paid_amount or 0)))

    overdue_rent = [
        serialize_rent_charge(charge, session, include_payments=False, include_reminder=False)
        for charge in charges
        if charge.status == "overdue" and rent_debt(charge) > 0 and debt_visible_by_cutoff(charge.due_date, cutoff)
    ]
    partial_rent = [
        serialize_rent_charge(charge, session, include_payments=False, include_reminder=False)
        for charge in charges
        if charge.status == "partial" and rent_debt(charge) > 0 and debt_visible_by_cutoff(charge.due_date, cutoff)
    ]
    today_rent = [
        serialize_rent_charge(charge, session, include_payments=False, include_reminder=False)
        for charge in charges
        if charge.due_date == today and charge.status not in {"paid", "paid_ahead"} and rent_debt(charge) > 0 and debt_visible_by_cutoff(charge.due_date, cutoff)
    ]
    deferred = [
        serialize_rent_charge(charge, session, include_payments=False, include_reminder=False)
        for charge in charges
        if charge.status == "deferred"
        and charge.deferral_until
        and rent_debt(charge) > 0
        and debt_visible_by_cutoff(charge.due_date, cutoff)
    ]

    overdue_utilities = [
        serialize_bill_line(line, session, include_payments=False, include_reminder=False, include_advance=False)
        for line in utility_lines
        if line.status == "overdue" and utility_debt(line) > 0 and debt_visible_by_cutoff(line.due_date, cutoff)
    ]
    partial_utilities = [
        serialize_bill_line(line, session, include_payments=False, include_reminder=False, include_advance=False)
        for line in utility_lines
        if line.status == "partial" and utility_debt(line) > 0 and debt_visible_by_cutoff(line.due_date, cutoff)
    ]
    issued_utilities = [
        serialize_bill_line(line, session, include_payments=False, include_reminder=False, include_advance=False)
        for line in utility_lines
        if line.status in {"issued", "overdue", "partial"} and utility_debt(line) > 0 and debt_visible_by_cutoff(line.due_date, cutoff)
    ]
    manual_debts = [
        serialize_manual_debt(debt, session, include_reminder=False)
        for debt in outstanding_manual_debts(session, cutoff=cutoff)
        if debt.lease_id
        and int(debt.lease_id) not in ignored_ids
        and debt.lease
        and IGNORE_LEASE_MARK not in (debt.lease.notes or "")
    ]

    pending_expenses = session.scalars(
        select(Expense).where(Expense.source_funds == "personal", Expense.compensation_status != "compensated")
    ).all()
    expense_fund_received = money(
        session.scalar(select(func.coalesce(func.sum(Expense.amount), 0)).where(Expense.source_funds == "expense_fund_income")) or 0
    )
    expense_fund_spent = money(
        session.scalar(
            select(func.coalesce(func.sum(Expense.amount), 0)).where(
                Expense.source_funds != "expense_fund_income",
                Expense.compensation_status != "compensated",
            )
        )
        or 0
    )
    expense_fund_balance = money(expense_fund_received - expense_fund_spent)

    services = session.scalars(
        select(UtilityService).options(joinedload(UtilityService.object)).where(UtilityService.active.is_(True))
    ).all()
    service_ids = [service.id for service in services]
    last_reading_by_service: dict[int, date] = {}
    if service_ids:
        rows = session.execute(
            select(Meter.service_id, func.max(MeterReading.reading_date))
            .select_from(Meter)
            .outerjoin(MeterReading, MeterReading.meter_id == Meter.id)
            .where(Meter.service_id.in_(service_ids), Meter.scope == "object")
            .group_by(Meter.service_id)
        ).all()
        last_reading_by_service = {int(service_id): reading_date for service_id, reading_date in rows if reading_date}

    stale_readings = []
    for service in services:
        last_date = last_reading_by_service.get(int(service.id))
        if not last_date or (today - last_date).days >= 30:
            stale_readings.append(
                {
                    "service_id": service.id,
                    "object": service.object.name,
                    "service": service.name,
                    "last_date": last_date.isoformat() if last_date else None,
                    "days": (today - last_date).days if last_date else None,
                }
            )
    provider_reading_due = [
        item
        for item in provider_reading_statuses_for_month(session, today.year, today.month, today)
        if item["alert"]
    ]

    provider_debts = [
        serialize_bill(
            bill,
            include_line_payments=False,
            include_line_reminders=False,
            include_line_advance=False,
        )
        for bill in session.scalars(
            select(UtilityBill)
            .options(joinedload(UtilityBill.service).joinedload(UtilityService.object), selectinload(UtilityBill.lines))
            .where(UtilityBill.provider_paid.is_(False))
        ).all()
        if bill.status in {"issued", "draft"}
        and debt_visible_by_cutoff(bill.due_date or bill.period_end, cutoff)
        and any(line.status != "paid" for line in bill.lines)
    ]

    suspicious_receipts = session.scalars(select(PaymentReceipt).where(PaymentReceipt.status == "suspicious")).all()

    return {
        "object_summary": build_object_summary(session),
        "month_summary": month_dashboard_summary(session, today.year, today.month, today),
        "income_trend": dashboard_income_trend(session, today),
        "monthly_reports": build_monthly_report_shells(session, today),
        "rent_overdue": overdue_rent,
        "rent_partial": partial_rent,
        "rent_today": today_rent,
        "rent_deferred": deferred,
        "utility_overdue": overdue_utilities,
        "utility_partial": partial_utilities,
        "utility_issued": issued_utilities,
        "manual_debts": manual_debts,
        "pending_personal_expenses": [serialize_expense(expense) for expense in pending_expenses],
        "expense_fund": {
            "received": expense_fund_received,
            "spent": expense_fund_spent,
            "balance": expense_fund_balance,
            "has_mismatch": expense_fund_balance > 0.009,
        },
        "stale_readings": stale_readings,
        "provider_reading_due": provider_reading_due,
        "provider_debts": provider_debts,
        "suspicious_receipts": [serialize_payment_receipt(receipt, session) for receipt in suspicious_receipts],
    }


def build_status_dashboard_fast(session: Session, today: date | None = None) -> dict[str, Any]:
    today = today or date.today()
    cutoff = configured_notification_cutoff_date(session)
    ignored_ids = ignored_lease_ids(session)

    active_filters = [Lease.active.is_(True), Apartment.active.is_(True)]
    if ignored_ids:
        active_filters.append(Lease.id.not_in(ignored_ids))

    rent_base = (
        select(func.count(RentCharge.id))
        .select_from(RentCharge)
        .join(Lease, RentCharge.lease_id == Lease.id)
        .join(Apartment, Lease.apartment_id == Apartment.id)
        .where(*active_filters)
    )
    utility_base = (
        select(func.count(UtilityBillLine.id))
        .select_from(UtilityBillLine)
        .join(Lease, UtilityBillLine.lease_id == Lease.id)
        .join(Apartment, UtilityBillLine.apartment_id == Apartment.id)
        .where(*active_filters)
    )
    if cutoff:
        rent_base = rent_base.where(RentCharge.due_date >= cutoff)
        utility_base = utility_base.where(
            or_(UtilityBillLine.due_date.is_(None), UtilityBillLine.due_date >= cutoff)
        )

    latest_reading = (
        select(
            Meter.service_id.label("service_id"),
            func.max(MeterReading.reading_date).label("last_date"),
        )
        .select_from(Meter)
        .outerjoin(MeterReading, MeterReading.meter_id == Meter.id)
        .where(Meter.active.is_(True))
        .group_by(Meter.service_id)
        .subquery()
    )
    stale_threshold = today - timedelta(days=30)
    stale_readings_count = int(
        session.scalar(
            select(func.count(UtilityService.id))
            .select_from(UtilityService)
            .outerjoin(latest_reading, latest_reading.c.service_id == UtilityService.id)
            .where(
                UtilityService.active.is_(True),
                or_(latest_reading.c.last_date.is_(None), latest_reading.c.last_date <= stale_threshold),
            )
        )
        or 0
    )

    accepted = accepted_monthly_report_keys(session)
    open_report_count = sum(
        1
        for entry in iter_dashboard_report_entries(today)
        if monthly_report_key(entry["year"], entry["month"], entry["kind"]) not in accepted
    )
    counts = {
        "monthly_reports": open_report_count,
        "rent_overdue": int(session.scalar(rent_base.where(RentCharge.status == "overdue")) or 0),
        "rent_partial": int(session.scalar(rent_base.where(RentCharge.status == "partial")) or 0),
        "utility_overdue": int(session.scalar(utility_base.where(UtilityBillLine.status == "overdue")) or 0),
        "utility_issued": int(
            session.scalar(utility_base.where(UtilityBillLine.status.in_(["issued", "overdue", "partial"]))) or 0
        ),
        "stale_readings": stale_readings_count,
        "pending_personal_expenses": int(
            session.scalar(
                select(func.count(Expense.id)).where(
                    Expense.source_funds == "personal",
                    Expense.compensation_status != "compensated",
                )
            )
            or 0
        ),
        "suspicious_receipts": int(
            session.scalar(select(func.count(PaymentReceipt.id)).where(PaymentReceipt.status == "suspicious")) or 0
        ),
    }
    return {key: [None] * value for key, value in counts.items()}


def money_text(value: float) -> str:
    return f"{money(value):,.2f}".replace(",", " ").replace(".", ",") + " ₽"


def update_manual_debt_status(debt: ManualDebt, today: date | None = None) -> str:
    today = today or date.today()
    remaining = money(max(0.0, float(debt.amount or 0) - float(debt.paid_amount or 0)))
    if not debt.active:
        debt.status = "cancelled"
    elif remaining <= EPS:
        debt.status = "paid"
    elif debt.paid_amount > EPS:
        debt.status = "partial"
    elif debt.due_date and debt.due_date < today:
        debt.status = "overdue"
    else:
        debt.status = "issued"
    return debt.status


def manual_debt_kind_label(kind: str) -> str:
    return {
        "rent": "аренда",
        "utility": "коммуналка",
        "other": "другое",
    }.get(kind, kind or "другое")


def manual_debt_channel_label(channel: str) -> str:
    return {
        "ip": "ИП",
        "personal": "по номеру",
        "expense_fund": "мне на расходы",
    }.get(channel, channel or "")


def manual_debt_debt(debt: ManualDebt) -> float:
    return money(max(0.0, float(debt.amount or 0) - float(debt.paid_amount or 0)))


def manual_debt_reference_date(debt: ManualDebt) -> date | None:
    return debt.due_date or debt.period_end or debt.period_start or debt.created_at.date()


def serialize_manual_debt(
    debt: ManualDebt,
    session: Session | None = None,
    *,
    include_reminder: bool = True,
) -> dict[str, Any]:
    update_manual_debt_status(debt)
    period_label = ""
    if debt.period_start and debt.period_end:
        period_label = f"{debt.period_start:%d.%m.%Y} -> {debt.period_end:%d.%m.%Y}"
    elif debt.period_start:
        period_label = f"{debt.period_start:%d.%m.%Y}"
    return {
        "id": debt.id,
        "lease_id": debt.lease_id,
        "apartment_id": debt.apartment_id,
        "object": debt.lease.apartment.object.name if debt.lease else "",
        "apartment": debt.lease.apartment.name if debt.lease else "",
        "tenant": debt.lease.tenant.full_name if debt.lease else "",
        "kind": debt.kind,
        "kind_label": manual_debt_kind_label(debt.kind),
        "channel": debt.channel,
        "channel_label": manual_debt_channel_label(debt.channel),
        "title": debt.title or manual_debt_kind_label(debt.kind),
        "period_start": debt.period_start.isoformat() if debt.period_start else None,
        "period_end": debt.period_end.isoformat() if debt.period_end else None,
        "period_label": period_label,
        "due_date": debt.due_date.isoformat() if debt.due_date else None,
        "amount": money(debt.amount),
        "paid_amount": money(debt.paid_amount),
        "debt": manual_debt_debt(debt),
        "status": debt.status,
        "notes": debt.notes or "",
        "active": debt.active,
        "reminder": reminder_meta(
            session,
            debt.lease,
            manual_debt_reference_date(debt),
            template_key="message_rent_overdue" if debt.kind == "rent" else "message_utility_bill",
        ) if include_reminder and session and debt.lease else None,
    }


def outstanding_manual_debts(session: Session, lease: Lease | None = None, cutoff: date | None = None) -> list[ManualDebt]:
    query = select(ManualDebt).where(ManualDebt.active.is_(True))
    if lease:
        query = query.where(ManualDebt.lease_id == lease.id)
    debts = session.scalars(query.order_by(ManualDebt.due_date, ManualDebt.created_at, ManualDebt.id)).all()
    result = []
    for debt in debts:
        update_manual_debt_status(debt)
        ref_date = manual_debt_reference_date(debt)
        if not debt_visible_by_cutoff(ref_date, cutoff):
            continue
        if manual_debt_debt(debt) > EPS:
            result.append(debt)
    return result


def manual_debt_bullet(debt: ManualDebt) -> str:
    remaining = manual_debt_debt(debt)
    base = debt.title or manual_debt_kind_label(debt.kind)
    channel = f" ({manual_debt_channel_label(debt.channel)})" if debt.kind == "rent" and debt.channel else ""
    paid = money(float(debt.paid_amount or 0))
    if paid > EPS:
        return f"- {base}{channel}: начислено {money_text(debt.amount)}, оплачено {money_text(paid)}. Остаток {money_text(remaining)}."
    return f"- {base}{channel}: долг {money_text(remaining)}."


def first_open_rent_charge(lease: Lease, cutoff: date | None = None) -> RentCharge | None:
    charges = sorted(lease.rent_charges, key=lambda item: item.due_date)
    for charge in charges:
        update_rent_charge_status(charge)
        if not debt_visible_by_cutoff(charge.due_date, cutoff):
            continue
        if charge.status not in {"paid", "paid_ahead"}:
            return charge
    return None


def first_open_utility_line(session: Session, lease: Lease, cutoff: date | None = None) -> UtilityBillLine | None:
    lines = session.scalars(select(UtilityBillLine).where(UtilityBillLine.lease_id == lease.id)).all()
    lines = sorted(lines, key=lambda item: (item.due_date or date.max, item.id))
    for line in lines:
        update_utility_line_status(line)
        if not debt_visible_by_cutoff(line.due_date, cutoff):
            continue
        if line.status not in {"paid", "paid_ahead"}:
            return line
    return None


def month_title(value: date | None) -> str:
    if not value:
        return ""
    return f"{MONTH_NAMES[value.month - 1]} {value.year}"


def outstanding_rent_charges(lease: Lease, today: date | None = None, cutoff: date | None = None) -> list[RentCharge]:
    today = today or date.today()
    result: list[RentCharge] = []
    for charge in sorted(lease.rent_charges, key=lambda item: (item.due_date, item.id)):
        update_rent_charge_status(charge, today)
        debt = max(0.0, charge.ip_due - charge.ip_paid) + max(0.0, charge.personal_due - charge.personal_paid)
        if debt > 0.009 and charge.due_date <= today and debt_visible_by_cutoff(charge.due_date, cutoff):
            result.append(charge)
    return result


def outstanding_utility_lines(
    session: Session,
    lease: Lease,
    today: date | None = None,
    cutoff: date | None = None,
) -> list[UtilityBillLine]:
    today = today or date.today()
    result: list[UtilityBillLine] = []
    lines = session.scalars(select(UtilityBillLine).where(UtilityBillLine.lease_id == lease.id)).all()
    for line in sorted(lines, key=lambda item: (item.due_date or date.max, item.id)):
        update_utility_line_status(line, today)
        debt = max(0.0, line.total_amount - line.paid_amount)
        if debt > 0.009 and line.due_date and line.due_date <= today and debt_visible_by_cutoff(line.due_date, cutoff):
            result.append(line)
    return result


def issued_utility_lines(
    session: Session,
    lease: Lease,
    cutoff: date | None = None,
) -> list[UtilityBillLine]:
    result: list[UtilityBillLine] = []
    lines = session.scalars(select(UtilityBillLine).where(UtilityBillLine.lease_id == lease.id)).all()
    for line in sorted(lines, key=lambda item: (item.due_date or date.max, item.id)):
        update_utility_line_status(line)
        debt = max(0.0, line.total_amount - line.paid_amount)
        if debt > 0.009 and line.status == "issued" and debt_visible_by_cutoff(line.due_date, cutoff):
            result.append(line)
    return result


def debt_month_heading(value: date | None) -> str:
    if not value:
        return ""
    current_year = date.today().year
    month_name = MONTH_NAMES[value.month - 1].capitalize()
    return month_name if value.year == current_year else f"{month_name} {value.year}"


def rent_debt_bullet(charge: RentCharge) -> str:
    ip_left = money(max(0.0, charge.ip_due - charge.ip_paid))
    personal_left = money(max(0.0, charge.personal_due - charge.personal_paid))
    ip_paid = money(charge.ip_paid) > 0.009
    personal_paid = money(charge.personal_paid) > 0.009

    if ip_left <= 0.009 and personal_left <= 0.009:
        return "аренда закрыта."
    if ip_left > 0.009 and personal_left > 0.009:
        if not ip_paid and not personal_paid:
            return "аренда не оплачена. ИП не оплачено, перевод по номеру телефона не поступил."
        if ip_paid and not personal_paid:
            return f"неполная оплата аренды. По ИП есть оплата, не хватает {money_text(ip_left)}. Перевод по номеру телефона не поступил."
        if not ip_paid and personal_paid:
            return f"неполная оплата аренды. ИП не оплачено, по номеру телефона не хватает {money_text(personal_left)}."
        return f"неполная оплата аренды. По ИП не хватает {money_text(ip_left)}, по номеру телефона не хватает {money_text(personal_left)}."
    if ip_left > 0.009:
        if ip_paid:
            return f"неполная оплата аренды. По ИП не хватает {money_text(ip_left)}, перевод по номеру телефона оплачен полностью."
        return "неполная оплата аренды. ИП не оплачено, перевод по номеру телефона оплачен полностью."
    if personal_paid:
        return f"неполная оплата аренды. ИП оплачено, по номеру телефона не хватает {money_text(personal_left)}."
    return "неполная оплата аренды. ИП оплачено, перевод по номеру телефона не поступил."


def rent_channel_summary_text(charge: RentCharge) -> str:
    ip_done = money(charge.ip_paid) + 0.009 >= money(charge.ip_due)
    personal_done = money(charge.personal_paid) + 0.009 >= money(charge.personal_due)
    if ip_done and personal_done:
        return "всё оплачено"
    if ip_done and not personal_done:
        return "ИП оплачен, перевод нет"
    if not ip_done and personal_done:
        return "ИП нет, перевод оплачен"
    return "ИП нет, перевода нет"


def utility_debt_bullet(lines: list[UtilityBillLine]) -> str:
    total = money(sum(line.total_amount for line in lines))
    paid = money(sum(line.paid_amount for line in lines))
    debt = money(max(0.0, total - paid))
    if paid <= 0.009:
        return f"неоплачены счета по коммунальным платежам. Всего {money_text(total)}."
    return f"неоплачены счета по коммунальным платежам. Всего {money_text(total)} — оплачено {money_text(paid)}. Остаток {money_text(debt)}."


def utility_message_line(line: UtilityBillLine, debt_override: float | None = None) -> str:
    debt = money(debt_override if debt_override is not None else max(0.0, line.total_amount - line.paid_amount))
    if utility_line_is_advance(line):
        period = utility_line_period_label(line)
        return (
            f"{line.bill.service.object.name}, аванс коммуналки на следующий период ({period}): {money_text(debt)}. "
            "Это предоплата: при следующем расчёте она вычтется из коммунального счёта."
        )
    service_name = (line.bill.service.name or "коммуналка").strip().lower()
    period = utility_line_period_label(line)
    return f"{line.bill.service.object.name}, {service_name} за период {period}: {money_text(debt)}"


def selected_utility_lines(
    session: Session,
    lease: Lease,
    line: UtilityBillLine | None = None,
    utility_lines_override: list[UtilityBillLine] | None = None,
    today: date | None = None,
) -> list[UtilityBillLine]:
    cutoff = configured_notification_cutoff_date(session)
    if utility_lines_override is not None:
        lines = [item for item in utility_lines_override if debt_visible_by_cutoff(item.due_date, cutoff)]
    else:
        lines = outstanding_utility_lines(session, lease, today, cutoff)
        if line and debt_visible_by_cutoff(line.due_date, cutoff) and all(existing.id != line.id for existing in lines):
            lines = [*lines, line]
    unique: dict[int, UtilityBillLine] = {}
    for item in sorted(lines, key=lambda current: (current.bill.period_end, current.id)):
        if item.status == "cancelled":
            continue
        if max(0.0, item.total_amount - item.paid_amount) <= 0.009:
            continue
        unique[item.id] = item
    return list(unique.values())


def build_all_debts_breakdown(session: Session, lease: Lease, today: date | None = None) -> str:
    today = today or date.today()
    cutoff = configured_notification_cutoff_date(session)
    rent_debts = outstanding_rent_charges(lease, today, cutoff)
    utility_debts = outstanding_utility_lines(session, lease, today, cutoff)
    manual_debts = outstanding_manual_debts(session, lease, cutoff)

    by_month: dict[tuple[int, int], list[str]] = {}
    for charge in rent_debts:
        key = (charge.due_date.year, charge.due_date.month)
        by_month.setdefault(key, []).append(f"- {rent_debt_bullet(charge)}")

    utility_groups: dict[tuple[int, int], list[UtilityBillLine]] = {}
    for line in utility_debts:
        key = (line.bill.period_end.year, line.bill.period_end.month)
        utility_groups.setdefault(key, []).append(line)

    for key, lines in utility_groups.items():
        by_month.setdefault(key, []).append(f"- {utility_debt_bullet(lines)}")

    for debt in manual_debts:
        ref = manual_debt_reference_date(debt) or today
        by_month.setdefault((ref.year, ref.month), []).append(manual_debt_bullet(debt))

    if not by_month:
        return "Задолженностей на сегодня нет."

    parts: list[str] = []
    for year, month in sorted(by_month):
        parts.append(f"{debt_month_heading(date(year, month, 1))}:")
        parts.extend(by_month[(year, month)])
    return "\n".join(parts)


def build_message_context(
    session: Session,
    lease: Lease,
    charge: RentCharge | None = None,
    line: UtilityBillLine | None = None,
    custom_text: str = "",
    utility_lines_override: list[UtilityBillLine] | None = None,
    utility_due_date_override: date | None = None,
) -> dict[str, str]:
    today = date.today()
    cutoff = configured_notification_cutoff_date(session)
    charge = charge or first_open_rent_charge(lease, cutoff)
    line = line or first_open_utility_line(session, lease, cutoff)
    current_rent_debt = 0.0
    ip_due = 0.0
    personal_due = 0.0
    rent_due_date = ""
    rent_period = ""
    if charge:
        update_rent_charge_status(charge)
        ip_due = max(0.0, charge.ip_due - charge.ip_paid)
        personal_due = max(0.0, charge.personal_due - charge.personal_paid)
        current_rent_debt = ip_due + personal_due
        rent_due_date = format_date(charge.due_date)
        rent_period = f"{format_date(charge.period_start)} - {format_date(charge.period_end)}"

    utility_lines = selected_utility_lines(session, lease, line, utility_lines_override, today)
    current_utility_total = sum(max(0.0, item.total_amount - item.paid_amount) for item in utility_lines)
    utility_due_date_value = utility_due_date_override or next((item.due_date for item in utility_lines if item.due_date), line.due_date if line else None)
    utility_due_date = format_full_date(utility_due_date_value) if utility_due_date_value else ""
    utility_period = ""
    utility_service = ""
    if utility_lines:
        first_line = utility_lines[0]
        if len(utility_lines) == 1:
            utility_period = f"{format_date(first_line.bill.period_start)} - {format_date(first_line.bill.period_end)}"
            utility_service = first_line.bill.service.name
        else:
            utility_service = "коммунальные платежи"
            utility_period = "; ".join(
                f"{format_date(item.bill.period_start)} - {format_date(item.bill.period_end)}"
                for item in utility_lines
            )

    rent_debts = outstanding_rent_charges(lease, today, cutoff)
    utility_debts = outstanding_utility_lines(session, lease, today, cutoff)
    manual_debts = outstanding_manual_debts(session, lease, cutoff)
    total_rent_debt = sum(max(0.0, item.ip_due - item.ip_paid) + max(0.0, item.personal_due - item.personal_paid) for item in rent_debts)
    total_utility_debt = sum(max(0.0, item.total_amount - item.paid_amount) for item in utility_debts)
    total_manual_debt = sum(manual_debt_debt(item) for item in manual_debts)

    rent_months = [month_title(item.due_date) for item in rent_debts]
    utility_items = [utility_message_line(item) for item in utility_lines] or [
        f"{item.bill.service.name} за {month_title(item.bill.period_end)} — {money_text(max(0.0, item.total_amount - item.paid_amount))}"
        for item in utility_debts
    ]
    rent_details = [
        f"{month_title(item.due_date)} — {money_text(max(0.0, item.ip_due - item.ip_paid) + max(0.0, item.personal_due - item.personal_paid))}"
        for item in rent_debts
    ]

    rent_debt = total_rent_debt if total_rent_debt > 0 else current_rent_debt
    utility_total = total_utility_debt if total_utility_debt > 0 else current_utility_total

    return {
        "tenant_name": lease.tenant.full_name,
        "apartment": lease.apartment.name,
        "object": lease.apartment.object.name,
        "phone": lease.tenant.phone or "",
        "telegram": lease.tenant.telegram or "",
        "payment_day": str(lease.payment_day),
        "ip_due": money_text(ip_due),
        "personal_due": money_text(personal_due),
        "total_due": money_text(rent_debt),
        "debt": money_text(rent_debt),
        "rent_due_date": rent_due_date,
        "rent_period": rent_period,
        "utility_total": money_text(utility_total),
        "utility_due_date": utility_due_date,
        "period": utility_period or rent_period,
        "utility_service": utility_service,
        "rent_debt_total": money_text(total_rent_debt),
        "rent_debt_months_count": str(len(rent_months)),
        "rent_debt_months": ", ".join(rent_months),
        "rent_debt_months_details": "; ".join(rent_details),
        "utility_debt_total": money_text(total_utility_debt),
        "utility_debt_count": str(len(utility_items)),
        "utility_debt_details": "\n".join(utility_items),
        "all_debt_total": money_text(total_rent_debt + total_utility_debt + total_manual_debt),
        "all_debts_breakdown": build_all_debts_breakdown(session, lease, today),
        "custom_text": custom_text,
        "ip_recipient_name_text": get_settings(session).get("ip_recipient_name") or "ИП не указан",
        "ip_recipient_account_text": get_settings(session).get("ip_recipient_account") or "счёт не указан",
        "personal_recipient_phone_text": get_settings(session).get("personal_recipient_phone") or "номер не указан",
        "personal_recipient_bank_text": get_settings(session).get("personal_recipient_bank") or "банк не указан",
    }


def render_message_text(
    session: Session,
    template_key: str,
    lease: Lease,
    charge: RentCharge | None = None,
    line: UtilityBillLine | None = None,
    custom_text: str = "",
    utility_lines_override: list[UtilityBillLine] | None = None,
    utility_due_date_override: date | None = None,
    context_overrides: dict[str, str] | None = None,
) -> str:
    if template_key == "custom":
        return custom_text.strip()
    context = build_message_context(
        session,
        lease,
        charge,
        line,
        custom_text,
        utility_lines_override=utility_lines_override,
        utility_due_date_override=utility_due_date_override,
    )
    if context_overrides:
        context.update(context_overrides)
    template = message_template(session, template_key)
    if template_key == "message_rent_overdue" and "{rent_debt_months" not in template:
        template = (
            template.rstrip()
            + "\nДолги по аренде: {rent_debt_months_count} мес. ({rent_debt_months})."
            + "\nПо месяцам: {rent_debt_months_details}."
            + "\nКоммуналка: {utility_debt_total}. Общий долг: {all_debt_total}."
        )
    if template_key == "message_all_debts" and "{all_debts_breakdown}" not in template:
        template = template.rstrip() + "\n{all_debts_breakdown}"
    if template_key == "message_utility_bill" and "{utility_debt_details}" not in template:
        template = DEFAULT_SETTINGS["message_utility_bill"]
    return fill_template(template, context)


def serialize_message_target(session: Session, lease: Lease) -> dict[str, Any]:
    cutoff = configured_notification_cutoff_date(session)
    today = date.today()
    charge = first_open_rent_charge(lease, cutoff)
    line = first_open_utility_line(session, lease, cutoff)
    rent_issues = outstanding_rent_charges(lease, today, cutoff)
    utility_issues = [*outstanding_utility_lines(session, lease, today, cutoff), *issued_utility_lines(session, lease, cutoff)]
    manual_issues = outstanding_manual_debts(session, lease, cutoff)
    rent_debt = money(max(0.0, (charge.ip_due - charge.ip_paid) if charge else 0.0) + max(0.0, (charge.personal_due - charge.personal_paid) if charge else 0.0))
    utility_debt = money(max(0.0, (line.total_amount - line.paid_amount) if line else 0.0))
    manual_debt_total = money(sum(manual_debt_debt(item) for item in manual_issues))
    return {
        "lease_id": lease.id,
        "tenant_id": lease.tenant_id,
        "tenant": lease.tenant.full_name,
        "object": lease.apartment.object.name,
        "apartment": lease.apartment.name,
        "telegram": lease.tenant.telegram,
        "linked": tenant_chat_linked(session, lease),
        "automation": get_lease_automation(session, lease.id),
        "rent_charge_id": charge.id if charge else None,
        "rent_status": charge.status if charge else "",
        "rent_debt": rent_debt,
        "rent_reminder": reminder_meta(
            session,
            lease,
            charge.due_date,
            template_key="message_rent_overdue",
            rent_charge_id=charge.id,
        ) if charge else None,
        "utility_line_id": line.id if line else None,
        "utility_status": line.status if line else "",
        "utility_debt": utility_debt,
        "utility_reminder": reminder_meta(
            session,
            lease,
            line.due_date,
            template_key="message_utility_bill",
            utility_line_id=line.id,
        ) if line else None,
        "rent_items": [
            {
                "id": item.id,
                "status": item.status,
                "due_date": item.due_date.isoformat(),
                "label": debt_month_heading(item.due_date),
                "debt": money(max(0.0, item.ip_due - item.ip_paid) + max(0.0, item.personal_due - item.personal_paid)),
                "channel_summary": rent_channel_summary_text(item),
            }
            for item in rent_issues
        ],
        "utility_items": [
            {
                "id": item.id,
                "status": item.status,
                "due_date": item.due_date.isoformat() if item.due_date else "",
                "label": f"ком. услуги {utility_line_period_label(item)}",
                "debt": money(max(0.0, item.total_amount - item.paid_amount)),
                "service": item.bill.service.name,
                "period_label": utility_line_period_label(item),
            }
            for item in utility_issues
        ],
        "manual_debt_items": [serialize_manual_debt(item, session) for item in manual_issues],
        "all_debt_total": money(rent_debt + utility_debt + manual_debt_total),
    }


def resolve_message_request(
    session: Session,
    payload: dict[str, Any],
) -> tuple[Lease, str, RentCharge | None, UtilityBillLine | None, str]:
    lease = session.get(Lease, int(payload["lease_id"]))
    if not lease or not lease.active:
        raise HTTPException(404, "Активный договор не найден")

    template_key = (payload.get("template_key") or "").strip()
    if template_key not in {"message_rent_due", "message_rent_overdue", "message_utility_bill", "message_all_debts", "custom"}:
        raise HTTPException(400, "Неизвестный шаблон сообщения")

    charge = session.get(RentCharge, int(payload["charge_id"])) if payload.get("charge_id") else None
    line = session.get(UtilityBillLine, int(payload["utility_line_id"])) if payload.get("utility_line_id") else None
    if charge and charge.lease_id != lease.id:
        raise HTTPException(400, "Этот арендный долг относится к другому жильцу")
    if line and line.lease_id != lease.id:
        raise HTTPException(400, "Этот коммунальный счёт относится к другому жильцу")
    custom_text = (payload.get("custom_text") or "").strip()
    return lease, template_key, charge, line, custom_text


def send_tenant_message(
    session: Session,
    lease: Lease,
    template_key: str,
    charge: RentCharge | None = None,
    line: UtilityBillLine | None = None,
    custom_text: str = "",
    note: str = "manual",
    utility_lines_override: list[UtilityBillLine] | None = None,
    utility_due_date_override: date | None = None,
    keyboard: dict[str, Any] | None = None,
) -> dict[str, Any]:
    chat_id = lease_chat_id(session, lease)
    if not chat_id:
        raise HTTPException(400, f"{lease.tenant.full_name} ещё не привязан к боту. Пусть сначала напишет /start.")
    text = render_message_text(
        session,
        template_key,
        lease,
        charge,
        line,
        custom_text,
        utility_lines_override=utility_lines_override,
        utility_due_date_override=utility_due_date_override,
    )
    if not text.strip():
        raise HTTPException(400, "Текст сообщения пустой")
    send_telegram_text(session, chat_id, text, keyboard)
    session.add(
        MessageLog(
            lease_id=lease.id,
            rent_charge_id=charge.id if charge else None,
            utility_line_id=line.id if line else None,
            channel="telegram",
            template_key=template_key,
            status="sent",
            recipient_chat_id=str(chat_id),
            text=text,
            note=note,
        )
    )
    session.flush()
    return {
        "ok": True,
        "tenant": lease.tenant.full_name,
        "lease_id": lease.id,
        "template_key": template_key,
        "text": text,
    }


BOT_DIALOG_LIST_LIMIT = 1200
BOT_DIALOG_HISTORY_LIMIT = 220


def bot_dialog_id_for_lease(lease_id: int) -> str:
    return f"lease:{lease_id}"


def bot_dialog_id_for_chat(chat_id: int | str, *, owner: bool = False) -> str:
    return f"{'owner' if owner else 'chat'}:{str(chat_id)}"


def bot_dialog_created_at(value: datetime | None) -> str:
    return value.isoformat() if value else ""


def bot_dialog_message_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())[:180]


def lease_for_dialog(session: Session, lease_id: int) -> Lease | None:
    return session.scalar(
        select(Lease)
        .options(
            joinedload(Lease.apartment).joinedload(Apartment.object),
            joinedload(Lease.tenant),
        )
        .where(Lease.id == lease_id)
        .limit(1)
    )


def latest_dialog_chat_id(session: Session, lease: Lease) -> str:
    linked = lease_chat_id(session, lease)
    if linked:
        return linked
    recent = session.scalar(
        select(MessageLog.recipient_chat_id)
        .where(MessageLog.lease_id == lease.id, MessageLog.recipient_chat_id != "")
        .order_by(MessageLog.created_at.desc(), MessageLog.id.desc())
        .limit(1)
    )
    return str(recent or "")


def ensure_bot_dialog_for_lease(
    session: Session,
    dialogs: dict[str, dict[str, Any]],
    lease: Lease,
    *,
    chat_id: str | None = None,
) -> dict[str, Any]:
    dialog_id = bot_dialog_id_for_lease(lease.id)
    linked_chat_id = str(chat_id or latest_dialog_chat_id(session, lease) or "")
    dialog = dialogs.get(dialog_id)
    if not dialog:
        dialog = {
            "id": dialog_id,
            "kind": "tenant",
            "lease_id": lease.id,
            "tenant_id": lease.tenant_id,
            "tenant": lease.tenant.full_name,
            "object": lease.apartment.object.name if lease.apartment and lease.apartment.object else "",
            "apartment": lease.apartment.name if lease.apartment else "",
            "title": lease.tenant.full_name,
            "subtitle": f"{lease.apartment.object.name}, {lease.apartment.name}" if lease.apartment and lease.apartment.object else "",
            "chat_id": linked_chat_id,
            "linked": bool(linked_chat_id),
            "status": "active" if lease.active else "archived",
            "last_at": "",
            "last_text": "",
            "last_direction": "",
            "message_count": 0,
        }
        dialogs[dialog_id] = dialog
    elif linked_chat_id and not dialog.get("chat_id"):
        dialog["chat_id"] = linked_chat_id
        dialog["linked"] = True
    return dialog


def ensure_bot_dialog_for_chat(
    session: Session,
    dialogs: dict[str, dict[str, Any]],
    chat_id: int | str,
    *,
    owner: bool | None = None,
) -> dict[str, Any]:
    chat_ref = str(chat_id or "").strip()
    is_owner = bool(owner) if owner is not None else chat_ref in telegram_owner_ids(session)
    dialog_id = bot_dialog_id_for_chat(chat_ref, owner=is_owner)
    dialog = dialogs.get(dialog_id)
    if not dialog:
        dialog = {
            "id": dialog_id,
            "kind": "owner" if is_owner else "chat",
            "lease_id": None,
            "tenant_id": None,
            "tenant": "",
            "object": "",
            "apartment": "",
            "title": "Владелец" if is_owner else f"Telegram {chat_ref}",
            "subtitle": "owner-чат бота" if is_owner else "чат без активной привязки к жильцу",
            "chat_id": chat_ref,
            "linked": bool(chat_ref),
            "status": "active",
            "last_at": "",
            "last_text": "",
            "last_direction": "",
            "message_count": 0,
        }
        dialogs[dialog_id] = dialog
    return dialog


def update_bot_dialog_last(dialog: dict[str, Any], text: str, created_at: datetime | None, direction: str) -> None:
    dialog["message_count"] = int(dialog.get("message_count") or 0) + 1
    created_text = bot_dialog_created_at(created_at)
    if created_text and created_text >= str(dialog.get("last_at") or ""):
        dialog["last_at"] = created_text
        dialog["last_text"] = bot_dialog_message_text(text)
        dialog["last_direction"] = direction


def bot_dialog_for_message_log(
    session: Session,
    dialogs: dict[str, dict[str, Any]],
    log: MessageLog,
) -> dict[str, Any] | None:
    if log.lease_id:
        lease = lease_for_dialog(session, int(log.lease_id))
        if lease:
            return ensure_bot_dialog_for_lease(session, dialogs, lease, chat_id=log.recipient_chat_id)
    if log.recipient_chat_id:
        return ensure_bot_dialog_for_chat(session, dialogs, log.recipient_chat_id)
    return None


def bot_dialog_for_ai_message(
    session: Session,
    dialogs: dict[str, dict[str, Any]],
    message: AiMessage,
) -> dict[str, Any] | None:
    if message.lease_id:
        lease = lease_for_dialog(session, int(message.lease_id))
        if lease:
            return ensure_bot_dialog_for_lease(session, dialogs, lease, chat_id=message.conversation.chat_id if message.conversation else "")
    conversation = message.conversation
    if conversation and conversation.chat_id:
        return ensure_bot_dialog_for_chat(
            session,
            dialogs,
            conversation.chat_id,
            owner=conversation.role == "owner",
        )
    return None


def serialize_bot_dialog_log_message(session: Session, log: MessageLog) -> dict[str, Any]:
    incoming = log.status == "incoming"
    owner_incoming = incoming and (
        "incoming:owner" in (log.note or "") or str(log.recipient_chat_id or "") in telegram_owner_ids(session)
    )
    return {
        "id": f"log:{log.id}",
        "source": "message_log",
        "direction": "incoming" if incoming else "outgoing",
        "author": "Владелец" if owner_incoming else "Жилец" if incoming else "Бот",
        "role": "user" if incoming else "assistant",
        "text": log.text or "",
        "created_at": bot_dialog_created_at(log.created_at),
        "status": log.status,
        "template_key": log.template_key or "",
        "note": log.note or "",
    }


def serialize_bot_dialog_ai_message(message: AiMessage) -> dict[str, Any]:
    conversation_role = message.conversation.role if message.conversation else ""
    incoming = message.role == "user"
    author = (
        "Владелец"
        if incoming and conversation_role == "owner"
        else "Жилец"
        if incoming
        else "Бот"
        if message.role == "assistant"
        else message.role
    )
    return {
        "id": f"ai:{message.id}",
        "source": "ai",
        "direction": "incoming" if incoming else "outgoing" if message.role == "assistant" else "system",
        "author": author,
        "role": message.role,
        "text": message.text or "",
        "created_at": bot_dialog_created_at(message.created_at),
        "status": "sent" if message.role == "assistant" else "received",
        "template_key": "",
        "note": message.model or "",
    }


def bot_dialog_message_matches_log(message: AiMessage, logs: list[MessageLog]) -> bool:
    text = (message.text or "").strip()
    if not text:
        return False
    for log in logs:
        if (log.text or "").strip() != text:
            continue
        if not log.created_at or not message.created_at:
            return True
        if abs((log.created_at - message.created_at).total_seconds()) <= 300:
            return True
    return False


def bot_dialogs_payload(session: Session) -> list[dict[str, Any]]:
    dialogs: dict[str, dict[str, Any]] = {}
    leases = session.scalars(
        select(Lease)
        .options(
            joinedload(Lease.apartment).joinedload(Apartment.object),
            joinedload(Lease.tenant),
        )
        .where(Lease.active.is_(True))
        .order_by(Lease.start_date.desc(), Lease.id.desc())
    ).all()
    for lease in leases:
        if lease.apartment and lease.apartment.active and not lease_ignored(session, lease.id):
            ensure_bot_dialog_for_lease(session, dialogs, lease)

    logs = session.scalars(
        select(MessageLog)
        .where(MessageLog.channel == "telegram")
        .order_by(MessageLog.created_at.desc(), MessageLog.id.desc())
        .limit(BOT_DIALOG_LIST_LIMIT)
    ).all()
    for log in logs:
        dialog = bot_dialog_for_message_log(session, dialogs, log)
        if dialog:
            update_bot_dialog_last(
                dialog,
                log.text,
                log.created_at,
                "incoming" if log.status == "incoming" else "outgoing",
            )

    ai_messages = session.scalars(
        select(AiMessage)
        .options(joinedload(AiMessage.conversation))
        .where(AiMessage.channel == "telegram")
        .order_by(AiMessage.created_at.desc(), AiMessage.id.desc())
        .limit(BOT_DIALOG_LIST_LIMIT)
    ).all()
    for message in ai_messages:
        dialog = bot_dialog_for_ai_message(session, dialogs, message)
        if dialog:
            update_bot_dialog_last(
                dialog,
                message.text,
                message.created_at,
                "incoming" if message.role == "user" else "outgoing",
            )

    return sorted(
        dialogs.values(),
        key=lambda item: (
            bool(item.get("last_at")),
            str(item.get("last_at") or ""),
            bool(item.get("linked")),
            str(item.get("title") or ""),
        ),
        reverse=True,
    )


def resolve_bot_dialog_target(session: Session, dialog_id: str) -> dict[str, Any]:
    raw = str(dialog_id or "").strip()
    if raw.startswith("lease:"):
        try:
            lease_id = int(raw.split(":", 1)[1])
        except ValueError as exc:
            raise HTTPException(400, "Некорректный id диалога") from exc
        lease = lease_for_dialog(session, lease_id)
        if not lease:
            raise HTTPException(404, "Диалог не найден")
        return {"dialog_id": bot_dialog_id_for_lease(lease.id), "lease": lease, "chat_id": latest_dialog_chat_id(session, lease)}
    if raw.startswith("owner:") or raw.startswith("chat:"):
        _kind, chat_id = raw.split(":", 1)
        chat_id = chat_id.strip()
        if not chat_id:
            raise HTTPException(400, "В диалоге нет chat id")
        return {"dialog_id": raw, "lease": None, "chat_id": chat_id}
    raise HTTPException(400, "Некорректный id диалога")


def bot_dialog_messages_payload(session: Session, dialog_id: str, limit: int = BOT_DIALOG_HISTORY_LIMIT) -> dict[str, Any]:
    target = resolve_bot_dialog_target(session, dialog_id)
    lease: Lease | None = target["lease"]
    chat_id = str(target["chat_id"] or "")
    conditions = []
    if lease:
        conditions.append(MessageLog.lease_id == lease.id)
    if chat_id:
        conditions.append(MessageLog.recipient_chat_id == chat_id)
    if not conditions:
        raise HTTPException(400, "У диалога ещё нет Telegram chat id")

    logs = session.scalars(
        select(MessageLog)
        .where(MessageLog.channel == "telegram", or_(*conditions))
        .order_by(MessageLog.created_at.desc(), MessageLog.id.desc())
        .limit(max(limit * 2, limit))
    ).all()

    if lease:
        ai_query = (
            select(AiMessage)
            .options(joinedload(AiMessage.conversation))
            .where(AiMessage.channel == "telegram", AiMessage.lease_id == lease.id)
        )
    else:
        ai_query = (
            select(AiMessage)
            .join(AiConversation, AiMessage.conversation_id == AiConversation.id)
            .options(joinedload(AiMessage.conversation))
            .where(AiMessage.channel == "telegram", AiConversation.chat_id == chat_id)
        )
    ai_messages = session.scalars(
        ai_query.order_by(AiMessage.created_at.desc(), AiMessage.id.desc()).limit(max(limit * 2, limit))
    ).all()

    incoming_logs = [log for log in logs if log.status == "incoming"]
    messages = [serialize_bot_dialog_log_message(session, log) for log in logs]
    for message in ai_messages:
        if message.role == "user" and bot_dialog_message_matches_log(message, incoming_logs):
            continue
        messages.append(serialize_bot_dialog_ai_message(message))

    messages = sorted(messages, key=lambda item: (item["created_at"], item["id"]))[-limit:]
    return {
        "dialog": target["dialog_id"],
        "chat_id": chat_id,
        "linked": bool(chat_id),
        "messages": messages,
    }


def telegram_history_text(message: dict[str, Any]) -> str:
    parts: list[str] = []
    text = (message.get("text") or message.get("caption") or "").strip()
    if text:
        parts.append(text)
    contact = message.get("contact") or {}
    if contact:
        phone = str(contact.get("phone_number") or "").strip()
        name = " ".join(str(contact.get(key) or "").strip() for key in ("first_name", "last_name")).strip()
        parts.append(f"[контакт] {name or phone}".strip())
    document = message.get("document") or {}
    if document:
        parts.append(f"[документ] {document.get('file_name') or document.get('file_id') or 'файл'}")
    if message.get("photo"):
        parts.append("[фото]")
    if message.get("voice"):
        parts.append("[голосовое сообщение]")
    if message.get("sticker"):
        parts.append("[стикер]")
    return "\n".join(part for part in parts if part).strip() or "[служебное сообщение Telegram]"


def log_telegram_incoming_message(
    session: Session,
    message: dict[str, Any],
    lease: Lease | None,
    *,
    is_owner: bool,
) -> None:
    chat = message.get("chat") or {}
    chat_id = chat.get("id")
    if not chat_id:
        return
    message_id = message.get("message_id")
    note = f"incoming:{'owner' if is_owner else 'tenant'}:{message_id or ''}"
    if message_id:
        existing = session.scalar(
            select(MessageLog.id)
            .where(
                MessageLog.channel == "telegram",
                MessageLog.status == "incoming",
                MessageLog.recipient_chat_id == str(chat_id),
                MessageLog.note.like(f"{note}%"),
            )
            .limit(1)
        )
        if existing:
            return
    sender = message_sender_label(message)
    session.add(
        MessageLog(
            lease_id=lease.id if lease else None,
            channel="telegram",
            template_key="telegram_incoming",
            status="incoming",
            recipient_chat_id=str(chat_id),
            text=telegram_history_text(message),
            note=f"{note}; sender={sender}",
        )
    )
    session.flush()


def int_selection(values: Any) -> set[int]:
    if values is None or values == "":
        return set()
    if isinstance(values, str):
        values = [part.strip() for part in values.split(",")]
    if not isinstance(values, list | tuple | set):
        values = [values]
    result: set[int] = set()
    for value in values:
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            continue
        if parsed > 0:
            result.add(parsed)
    return result


def broadcast_candidate_leases(session: Session, payload: dict[str, Any]) -> list[Lease]:
    all_selected = setting_bool_value(payload.get("all")) or str(payload.get("scope") or "").strip() == "all"
    lease_ids = int_selection(payload.get("lease_ids"))
    apartment_ids = int_selection(payload.get("apartment_ids"))
    object_ids = int_selection(payload.get("object_ids"))
    if not all_selected and not lease_ids and not apartment_ids and not object_ids:
        raise HTTPException(400, "Выберите хотя бы одного получателя")

    leases = session.scalars(
        select(Lease)
        .options(
            joinedload(Lease.apartment).joinedload(Apartment.object),
            joinedload(Lease.tenant),
        )
        .join(Apartment)
        .where(Lease.active.is_(True), Apartment.active.is_(True))
        .order_by(Apartment.object_id, Apartment.sort_order, Apartment.name, Lease.start_date.desc())
    ).all()
    selected = []
    for lease in leases:
        if lease_ignored(session, lease.id):
            continue
        if all_selected or lease.id in lease_ids or lease.apartment_id in apartment_ids or lease.apartment.object_id in object_ids:
            selected.append(lease)
    return selected


def serialize_broadcast_target(lease: Lease, status: str, detail: str = "") -> dict[str, Any]:
    return {
        "lease_id": lease.id,
        "tenant": lease.tenant.full_name,
        "object": lease.apartment.object.name,
        "apartment": lease.apartment.name,
        "status": status,
        "detail": detail,
    }


def resolve_broadcast_recipients(session: Session, payload: dict[str, Any]) -> dict[str, Any]:
    recipients: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    seen_chats: set[str] = set()
    for lease in broadcast_candidate_leases(session, payload):
        chat_id = lease_chat_id(session, lease)
        if not chat_id:
            skipped.append(serialize_broadcast_target(lease, "unlinked", "ждём /start от жильца"))
            continue
        if chat_id in seen_chats:
            skipped.append(serialize_broadcast_target(lease, "duplicate_chat", "этот Telegram-чат уже выбран"))
            continue
        seen_chats.add(chat_id)
        recipients.append({"lease": lease, "chat_id": chat_id})
    return {"recipients": recipients, "skipped": skipped}


@app.post("/api/messages/broadcast")
def broadcast_message_to_tenants(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    text = str(payload.get("message") or "").strip()
    if not text:
        raise HTTPException(400, "Введите текст рассылки")
    if len(text) > 4000:
        raise HTTPException(400, "Telegram не любит простыни длиннее 4000 символов. Разбей сообщение на части.")
    if not telegram_token(session):
        raise HTTPException(400, "Сначала сохрани Telegram bot token")

    resolved = resolve_broadcast_recipients(session, payload)
    sent: list[dict[str, Any]] = []
    failed: list[dict[str, Any]] = []
    for item in resolved["recipients"]:
        lease = item["lease"]
        chat_id = item["chat_id"]
        try:
            send_telegram_text(session, chat_id, text)
            session.add(
                MessageLog(
                    lease_id=lease.id,
                    channel="telegram",
                    template_key="broadcast",
                    status="sent",
                    recipient_chat_id=str(chat_id),
                    text=text,
                    note="mass-broadcast",
                )
            )
            sent.append(serialize_broadcast_target(lease, "sent"))
        except HTTPException as exc:
            detail = str(exc.detail)
            session.add(
                MessageLog(
                    lease_id=lease.id,
                    channel="telegram",
                    template_key="broadcast",
                    status="failed",
                    recipient_chat_id=str(chat_id),
                    text=text,
                    note=f"mass-broadcast: {detail}",
                )
            )
            failed.append(serialize_broadcast_target(lease, "failed", detail))
    session.commit()
    skipped = resolved["skipped"]
    return {
        "ok": True,
        "sent": len(sent),
        "failed": len(failed),
        "skipped_unlinked": len([item for item in skipped if item["status"] == "unlinked"]),
        "skipped_duplicate": len([item for item in skipped if item["status"] == "duplicate_chat"]),
        "recipients": sent,
        "failed_recipients": failed,
        "skipped_recipients": skipped,
    }


def utility_issue_targets(session: Session, bill: UtilityBill) -> list[dict[str, Any]]:
    return utility_issue_targets_for_bills(session, utility_issue_group_bills(session, bill))


def utility_issue_group_bills(session: Session, bill: UtilityBill) -> list[UtilityBill]:
    object_id = bill.service.object_id
    bills: list[UtilityBill] = []
    if utility_bill_is_advance(bill):
        usage_bills = session.scalars(
            select(UtilityBill)
            .options(selectinload(UtilityBill.lines))
            .join(UtilityService)
            .where(
                UtilityService.object_id == object_id,
                UtilityBill.bill_type != UTILITY_BILL_ADVANCE,
                UtilityBill.status == "draft",
                UtilityBill.period_end == bill.period_start,
            )
            .order_by(UtilityBill.period_start, UtilityBill.id)
        ).all()
        bills.extend(usage_bills)
        bills.append(bill)
    else:
        usage_bills = session.scalars(
            select(UtilityBill)
            .options(selectinload(UtilityBill.lines))
            .join(UtilityService)
            .where(
                UtilityService.object_id == object_id,
                UtilityBill.bill_type != UTILITY_BILL_ADVANCE,
                UtilityBill.status == "draft",
                UtilityBill.period_start == bill.period_start,
                UtilityBill.period_end == bill.period_end,
            )
            .order_by(UtilityBill.period_start, UtilityBill.id)
        ).all()
        bills.extend(usage_bills)
        advance_start, advance_end = next_utility_advance_period(bill.period_start, bill.period_end)
        advance_bill = advance_bill_for_object_period(session, object_id, advance_start, advance_end)
        if advance_bill and advance_bill.status == "draft":
            bills.append(advance_bill)
    if not bills:
        bills = [bill]
    unique: dict[int, UtilityBill] = {}
    for item in bills:
        unique[item.id] = item
    return sorted(unique.values(), key=lambda item: (1 if utility_bill_is_advance(item) else 0, item.period_start, item.id))


def resident_due_date_for_bills(bills: list[UtilityBill]) -> date:
    usage_bills = [bill for bill in bills if not utility_bill_is_advance(bill)]
    candidates = [resident_due_date(bill.service) for bill in usage_bills or bills]
    return min(candidates) if candidates else date.today()


def projected_utility_debts_for_issue_preview(
    session: Session,
    selected_lines: list[UtilityBillLine],
    combined_lines: list[UtilityBillLine],
) -> dict[int, float]:
    selected_ids = {line.id for line in selected_lines}
    balance_by_apartment: dict[int, float] = {}
    projected: dict[int, float] = {}
    for line in combined_lines:
        debt = utility_line_debt_amount(line)
        if (
            line.id in selected_ids
            and not utility_line_is_advance(line)
            and debt > EPS
        ):
            balance = balance_by_apartment.setdefault(line.apartment_id, utility_advance_balance(session, line.apartment_id))
            portion = money(min(balance, debt))
            if portion > EPS:
                debt = money(debt - portion)
                balance_by_apartment[line.apartment_id] = money(balance - portion)
        projected[line.id] = debt
    return projected


def utility_message_projection_context(
    session: Session,
    selected_lines: list[UtilityBillLine],
    combined_lines: list[UtilityBillLine],
) -> dict[str, str]:
    projected_debts = projected_utility_debts_for_issue_preview(session, selected_lines, combined_lines)
    visible_lines = [line for line in combined_lines if projected_debts.get(line.id, 0.0) > EPS]
    return {
        "utility_debt_details": "\n".join(
            utility_message_line(line, projected_debts.get(line.id))
            for line in visible_lines
        ),
        "utility_total": money_text(sum(projected_debts.get(line.id, 0.0) for line in visible_lines)),
        "utility_debt_count": str(len(visible_lines)),
    }


def utility_issue_targets_for_bills(session: Session, bills: list[UtilityBill]) -> list[dict[str, Any]]:
    if not bills:
        return []
    due = resident_due_date_for_bills(bills)
    by_lease: dict[int, list[UtilityBillLine]] = {}
    for bill in bills:
        for line in bill.lines:
            if not line.lease_id:
                continue
            debt = max(0.0, line.total_amount - line.paid_amount)
            if debt <= 0.009 or line.status == "cancelled":
                continue
            by_lease.setdefault(line.lease_id, []).append(line)

    previews: list[dict[str, Any]] = []
    for lease_id, selected_lines in sorted(by_lease.items()):
        lease = session.get(Lease, lease_id)
        if not lease or lease_ignored(session, lease.id):
            continue
        existing_lines = [
            item
            for item in outstanding_utility_lines(session, lease)
            if item.id not in {selected.id for selected in selected_lines}
        ]
        combined_lines = sorted(
            [*existing_lines, *selected_lines],
            key=lambda item: (item.bill.period_end, item.bill.service.name, item.id),
        )
        text = render_message_text(
            session,
            "message_utility_bill",
            lease,
            line=selected_lines[0],
            utility_lines_override=combined_lines,
            utility_due_date_override=due,
            context_overrides=utility_message_projection_context(session, selected_lines, combined_lines),
        )
        previews.append(
            {
                "lease_id": lease.id,
                "tenant": lease.tenant.full_name,
                "apartment": lease.apartment.name,
                "object": lease.apartment.object.name,
                "linked": tenant_chat_linked(session, lease),
                "text": text,
                "line_ids": [line.id for line in selected_lines],
                "all_line_ids": [line.id for line in combined_lines],
                "total": money(sum(max(0.0, item.total_amount - item.paid_amount) for item in selected_lines)),
            }
        )
    return previews


def move_out_process_key(lease_id: int, end_date: date) -> str:
    return f"{lease_id}:{end_date.isoformat()}"


def processed_move_out_keys(session: Session) -> set[str]:
    raw = get_internal_json(session, "processed_move_out_notifications", [])
    if not isinstance(raw, list):
        return set()
    return {str(item) for item in raw if isinstance(item, str) and item}


def save_processed_move_out_keys(session: Session, values: set[str]) -> None:
    set_internal_json(session, "processed_move_out_notifications", sorted(values))


def last_lease_service_period_end(session: Session, lease: Lease, service_id: int) -> date:
    ends = session.scalars(
        select(UtilityBill.period_end)
        .join(UtilityBillLine, UtilityBillLine.bill_id == UtilityBill.id)
        .where(
            UtilityBill.service_id == service_id,
            UtilityBillLine.lease_id == lease.id,
            UtilityBill.status != "draft",
            UtilityBillLine.status != "draft",
        )
        .order_by(UtilityBill.period_end.desc())
    ).all()
    return ends[0] if ends else lease.start_date


def existing_lease_service_line(
    session: Session,
    lease: Lease,
    service_id: int,
    period_start: date,
    period_end: date,
) -> UtilityBillLine | None:
    return session.scalar(
        select(UtilityBillLine)
        .join(UtilityBill, UtilityBillLine.bill_id == UtilityBill.id)
        .where(
            UtilityBill.service_id == service_id,
            UtilityBill.period_start == period_start,
            UtilityBill.period_end == period_end,
            UtilityBillLine.lease_id == lease.id,
        )
        .order_by(UtilityBillLine.id.desc())
        .limit(1)
    )


def create_move_out_utility_lines(session: Session, lease: Lease) -> tuple[list[UtilityBillLine], list[str]]:
    end_date = lease.end_date
    if not end_date:
        return [], []
    created_lines: list[UtilityBillLine] = []
    warnings: list[str] = []
    services = session.scalars(
        select(UtilityService)
        .where(UtilityService.object_id == lease.apartment.object_id, UtilityService.active.is_(True))
        .order_by(UtilityService.id)
    ).all()
    for service in services:
        period_start = last_lease_service_period_end(session, lease, service.id)
        if end_date <= period_start:
            continue
        existing = existing_lease_service_line(session, lease, service.id, period_start, end_date)
        if existing:
            if existing.bill.status == "draft":
                due = resident_due_date(service, end_date)
                existing.bill.status = "issued"
                existing.bill.due_date = due
                existing.bill.is_forecast = True
                existing.bill.provider_paid = True
                existing.bill.provider_paid_at = utc_now()
                existing.status = "issued"
                existing.issued_at = utc_now()
                existing.due_date = due
            created_lines.append(existing)
            continue
        try:
            preview_bill, preview_warnings = calculate_utility_bill(
                session=session,
                service_id=service.id,
                period_start=period_start,
                period_end=end_date,
                allow_estimate=True,
            )
        except ValueError as exc:
            warnings.append(f"{service.object.name}, {service.name}: {exc}")
            continue
        selected_lines = [
            line
            for line in preview_bill.lines
            if line.lease_id == lease.id and line.apartment_id == lease.apartment_id and money(line.total_amount) > EPS
        ]
        if not selected_lines:
            continue
        due = resident_due_date(service, end_date)
        note_lines = list(dict.fromkeys([*(preview_warnings or []), *(preview_bill.notes or "").splitlines()]))
        note_lines = [item for item in note_lines if item]
        note_lines.append(f"Авторасчёт при выезде {lease.tenant.full_name} на {end_date:%d.%m.%Y}.")
        bill = UtilityBill(
            service_id=service.id,
            period_start=period_start,
            period_end=end_date,
            status="issued",
            total_consumption=money(sum(float(line.personal_consumption or 0) + float(line.odn_consumption or 0) for line in selected_lines)),
            apartment_consumption=money(sum(float(line.personal_consumption or 0) for line in selected_lines)),
            odn_consumption=money(sum(float(line.odn_consumption or 0) for line in selected_lines)),
            total_cost=money(sum(float(line.total_amount or 0) for line in selected_lines)),
            average_unit_price=preview_bill.average_unit_price,
            due_date=due,
            is_forecast=True,
            provider_paid=True,
            provider_paid_at=utc_now(),
            notes="\n".join(note_lines),
        )
        issued_at = utc_now()
        for preview_line in selected_lines:
            bill.lines.append(
                UtilityBillLine(
                    apartment_id=preview_line.apartment_id,
                    lease_id=preview_line.lease_id,
                    personal_consumption=money(preview_line.personal_consumption),
                    odn_consumption=money(preview_line.odn_consumption),
                    total_amount=money(preview_line.total_amount),
                    paid_amount=0,
                    status="issued",
                    issued_at=issued_at,
                    due_date=due,
                    note=(preview_line.note or "").strip(),
                )
            )
        session.add(bill)
        session.flush()
        created_lines.extend(bill.lines)
        warnings.extend([warning for warning in preview_warnings if warning])
    return created_lines, list(dict.fromkeys(warnings))


def send_move_out_utility_message(session: Session, lease: Lease, lines: list[UtilityBillLine]) -> bool:
    if not lines or not lease_chat_id(session, lease):
        return False
    sorted_lines = sorted(lines, key=lambda item: (item.bill.period_end, item.bill.service.name, item.id))
    payload = send_tenant_message(
        session,
        lease,
        "message_utility_bill",
        line=sorted_lines[0],
        note="move-out-auto",
        utility_lines_override=sorted_lines,
        utility_due_date_override=sorted_lines[0].due_date,
    )
    for extra_line in sorted_lines[1:]:
        session.add(
            MessageLog(
                lease_id=lease.id,
                utility_line_id=extra_line.id,
                channel="telegram",
                template_key="message_utility_bill",
                status="sent",
                recipient_chat_id=str(lease_chat_id(session, lease)),
                text=payload["text"],
                note="move-out-auto",
            )
        )
    session.flush()
    return True


def notify_owner_about_move_out(
    session: Session,
    lease: Lease,
    lines: list[UtilityBillLine],
    warnings: list[str],
    tenant_notified: bool,
) -> bool:
    owner_id = telegram_owner_chat_id(session)
    if not owner_id or not telegram_token(session):
        return False
    total = money(sum(max(0.0, float(line.total_amount or 0) - float(line.paid_amount or 0)) for line in lines))
    if lines:
        utility_rows = "\n".join(f"- {utility_message_line(line)}" for line in lines)
    else:
        utility_rows = "- коммунальные начисления не созданы: сумма равна нулю или недостаточно данных."
    warning_text = ""
    if warnings:
        warning_text = "\nПредупреждения:\n" + "\n".join(f"- {item}" for item in warnings)
    tenant_state = "отправлено жильцу" if tenant_notified else "жильцу не отправлено: чат не привязан или нечего слать"
    text = (
        f"Сегодня выезд жильца.\n"
        f"{lease.apartment.object.name}, {lease.apartment.name}, {lease.tenant.full_name}\n"
        f"Дата выезда: {lease.end_date:%d.%m.%Y}\n"
        f"Расчётная коммуналка создана на {money_text(total)}.\n"
        f"{utility_rows}\n"
        f"Статус рассылки: {tenant_state}.{warning_text}"
    )
    send_telegram_text(session, owner_id, text, app_keyboard(app_base_url(session)))
    return True


def process_move_out_notifications(session: Session, today: date | None = None) -> dict[str, int]:
    today = today or date.today()
    processed = processed_move_out_keys(session)
    leases = session.scalars(
        select(Lease)
        .join(Apartment)
        .join(Tenant)
        .where(Lease.end_date == today, Apartment.active.is_(True))
        .order_by(Apartment.object_id, Apartment.sort_order, Lease.id)
    ).all()
    summary = {"processed": 0, "created_lines": 0, "tenant_sent": 0, "owner_sent": 0}
    for lease in leases:
        if lease_ignored(session, lease.id):
            continue
        key = move_out_process_key(lease.id, today)
        if key in processed:
            continue
        lines, warnings = create_move_out_utility_lines(session, lease)
        tenant_sent = send_move_out_utility_message(session, lease, lines)
        owner_sent = notify_owner_about_move_out(session, lease, lines, warnings, tenant_sent)
        processed.add(key)
        summary["processed"] += 1
        summary["created_lines"] += len(lines)
        summary["tenant_sent"] += 1 if tenant_sent else 0
        summary["owner_sent"] += 1 if owner_sent else 0
    save_processed_move_out_keys(session, processed)
    session.flush()
    return summary


def reminder_sent_today(
    session: Session,
    template_key: str,
    *,
    rent_charge_id: int | None = None,
    utility_line_id: int | None = None,
    today: date | None = None,
) -> bool:
    today = today or date.today()
    query = select(MessageLog.id).where(
        MessageLog.template_key == template_key,
        func.date(MessageLog.created_at) == today,
    )
    if rent_charge_id is not None:
        query = query.where(MessageLog.rent_charge_id == rent_charge_id)
    if utility_line_id is not None:
        query = query.where(MessageLog.utility_line_id == utility_line_id)
    return bool(session.scalar(query.limit(1)))


PAYMENT_SITUATION_FINAL_STATUSES = {"paid", "cancelled"}
PAYMENT_SITUATION_PAUSED_STATUSES = {
    "awaiting_receipt",
    "awaiting_payment_date",
    "receipt_pending_review",
    "promised",
    "short_deferral",
    "question",
    "escalated",
    "manual",
    "ignored",
}
MAX_AUTOMATIC_SITUATION_REMINDERS = 3


def situation_metadata(situation: PaymentSituation) -> dict[str, Any]:
    try:
        parsed = json.loads(situation.metadata_json or "{}")
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def save_situation_metadata(situation: PaymentSituation, metadata: dict[str, Any]) -> None:
    situation.metadata_json = json.dumps(metadata, ensure_ascii=False, sort_keys=True)


def payment_situation(
    session: Session,
    kind: str,
    reference_id: int,
    lease_id: int,
) -> PaymentSituation:
    situation = session.scalar(
        select(PaymentSituation)
        .where(
            PaymentSituation.kind == kind,
            PaymentSituation.reference_id == int(reference_id),
        )
        .limit(1)
    )
    if situation:
        return situation
    situation = PaymentSituation(
        lease_id=lease_id,
        kind=kind,
        reference_id=int(reference_id),
        status="awaiting_payment",
        mode="normal",
    )
    session.add(situation)
    session.flush()
    return situation


def payment_situation_target(
    session: Session,
    situation: PaymentSituation,
) -> RentCharge | UtilityBillLine | None:
    if situation.kind == "rent":
        return session.get(RentCharge, situation.reference_id)
    if situation.kind == "utility":
        return session.get(UtilityBillLine, situation.reference_id)
    return None


def situation_due_date(target: RentCharge | UtilityBillLine | None) -> date | None:
    return target.due_date if target else None


def situation_debt(target: RentCharge | UtilityBillLine | None) -> float:
    if isinstance(target, RentCharge):
        return money(
            max(0.0, float(target.ip_due or 0) - float(target.ip_paid or 0))
            + max(0.0, float(target.personal_due or 0) - float(target.personal_paid or 0))
        )
    if isinstance(target, UtilityBillLine):
        return money(max(0.0, float(target.total_amount or 0) - float(target.paid_amount or 0)))
    return 0.0


def situation_lease(target: RentCharge | UtilityBillLine | None) -> Lease | None:
    return target.lease if target else None


def sync_payment_situation(
    session: Session,
    situation: PaymentSituation,
    today: date | None = None,
) -> PaymentSituation:
    today = today or date.today()
    target = payment_situation_target(session, situation)
    if isinstance(target, RentCharge):
        update_rent_charge_status(target, today)
    elif isinstance(target, UtilityBillLine):
        update_utility_line_status(target, today)
    if target is None:
        situation.status = "cancelled"
        return situation
    if situation_debt(target) <= EPS:
        situation.status = "paid"
        situation.paused_until = None
        situation.promise_date = None
        return situation
    if situation.status in {"promised", "short_deferral"} and situation.paused_until:
        if today < situation.paused_until:
            return situation
        situation.status = "awaiting_payment"
        situation.paused_until = None
    if situation.mode == "manual":
        situation.status = "manual"
    return situation


def situation_payment_breakdown(target: RentCharge | UtilityBillLine) -> list[str]:
    if isinstance(target, RentCharge):
        ip_left = money(max(0.0, float(target.ip_due or 0) - float(target.ip_paid or 0)))
        personal_left = money(max(0.0, float(target.personal_due or 0) - float(target.personal_paid or 0)))
        return [
            f"на расчётный счёт ИП: {money_text(ip_left)}",
            f"переводом по номеру телефона: {money_text(personal_left)}",
        ]
    service = target.bill.service.name if target.bill and target.bill.service else "коммунальные услуги"
    return [f"{service}: {money_text(situation_debt(target))} — перевод по номеру телефона"]


def tenant_situation_keyboard(situation: PaymentSituation, *, overdue: bool = False) -> dict[str, Any]:
    later_text = "📅 Указать дату оплаты" if overdue else "🕓 Оплачу позже"
    return {
        "inline_keyboard": [
            [
                {"text": "✅ Оплатил(а)", "callback_data": f"tenantpay:{situation.id}:paid"},
                {"text": "📎 Отправить чек", "callback_data": f"tenantpay:{situation.id}:receipt"},
            ],
            [
                {"text": later_text, "callback_data": f"tenantpay:{situation.id}:later"},
                {"text": "💬 Есть вопрос", "callback_data": f"tenantpay:{situation.id}:question"},
            ],
        ]
    }


def owner_situation_keyboard(situation: PaymentSituation) -> dict[str, Any]:
    return {
        "inline_keyboard": [
            [
                {"text": "💬 Подготовить сообщение", "callback_data": f"ownersit:{situation.id}:suggest"},
                {"text": "⏳ Подождать день", "callback_data": f"ownersit:{situation.id}:wait"},
            ],
            [
                {"text": "🔕 В ручной режим", "callback_data": f"ownersit:{situation.id}:manual"},
            ],
        ]
    }


def financial_notification_sent_today(session: Session, lease_id: int, today: date) -> bool:
    local_start = datetime.combine(today, time.min, tzinfo=LOCAL_TZ)
    utc_start = local_start.astimezone(timezone.utc).replace(tzinfo=None)
    utc_end = (local_start + timedelta(days=1)).astimezone(timezone.utc).replace(tzinfo=None)
    return bool(
        session.scalar(
            select(MessageLog.id)
            .where(
                MessageLog.lease_id == lease_id,
                MessageLog.channel == "telegram",
                MessageLog.note.like("auto-financial:%"),
                MessageLog.created_at >= utc_start,
                MessageLog.created_at < utc_end,
            )
            .limit(1)
        )
    )


def situation_stage_sent(session: Session, situation: PaymentSituation, stage: str) -> bool:
    note = f"auto-financial:{situation.kind}:{situation.id}:{stage}"
    return bool(session.scalar(select(MessageLog.id).where(MessageLog.note == note).limit(1)))


def send_payment_situation_message(
    session: Session,
    situation: PaymentSituation,
    template_key: str,
    stage: str,
    today: date,
) -> bool:
    target = payment_situation_target(session, situation)
    lease = situation_lease(target)
    if target is None or lease is None:
        return False
    if financial_notification_sent_today(session, lease.id, today):
        return False
    if situation.notification_count >= MAX_AUTOMATIC_SITUATION_REMINDERS:
        return False
    text = render_message_text(
        session,
        template_key,
        lease,
        charge=target if isinstance(target, RentCharge) else None,
        line=target if isinstance(target, UtilityBillLine) else None,
    )
    chat_id = lease_chat_id(session, lease)
    if not chat_id:
        return False
    send_telegram_text(
        session,
        chat_id,
        text,
        tenant_situation_keyboard(situation, overdue=bool(situation_due_date(target) and situation_due_date(target) < today)),
    )
    session.add(
        MessageLog(
            lease_id=lease.id,
            rent_charge_id=target.id if isinstance(target, RentCharge) else None,
            utility_line_id=target.id if isinstance(target, UtilityBillLine) else None,
            channel="telegram",
            template_key=template_key,
            status="sent",
            recipient_chat_id=str(chat_id),
            text=text,
            note=f"auto-financial:{situation.kind}:{situation.id}:{stage}",
        )
    )
    situation.notification_count = int(situation.notification_count or 0) + 1
    situation.last_notification_at = utc_now()
    situation.status = {
        "upcoming": "first_reminder_sent",
        "due": "due_today",
        "overdue_1": "overdue",
        "overdue_3": "overdue",
    }.get(stage, situation.status)
    session.flush()
    return True


def situation_owner_event_sent(situation: PaymentSituation, event: str) -> bool:
    events = situation_metadata(situation).get("owner_events") or []
    return event in events


def mark_situation_owner_event(situation: PaymentSituation, event: str) -> None:
    metadata = situation_metadata(situation)
    events = [str(item) for item in metadata.get("owner_events") or []]
    if event not in events:
        events.append(event)
    metadata["owner_events"] = events[-30:]
    save_situation_metadata(situation, metadata)
    situation.owner_last_notified_at = utc_now()


def owner_situation_text(
    situation: PaymentSituation,
    target: RentCharge | UtilityBillLine,
    event: str,
    today: date,
) -> str:
    lease = target.lease
    due = situation_due_date(target) or today
    overdue_days = max(0, (today - due).days)
    kind_label = "аренды" if situation.kind == "rent" else "коммунальных платежей"
    event_title = {
        "due_unpaid": "Сегодня был день оплаты, но платёж пока не поступил.",
        "overdue_1": f"Оплата {kind_label} просрочена на 1 день.",
        "overdue_3": f"Оплата {kind_label} просрочена на 3 дня.",
        "overdue_5": f"Оплата {kind_label} просрочена на 5 дней.",
        "ignored": "Жилец не отвечает на финансовые уведомления.",
        "question": f"Жилец задал вопрос по оплате {kind_label}.",
        "promise": f"Жилец сообщил новую дату оплаты {kind_label}.",
        "short_deferral": "Система автоматически зафиксировала короткую отсрочку.",
        "receipt": f"Жилец отправил чек по оплате {kind_label}.",
        "manual": "Ситуация переведена в ручной режим.",
    }.get(event, f"Обновилась ситуация по оплате {kind_label}.")
    lines = [
        event_title,
        "",
        f"Объект: {lease.apartment.object.name}",
        f"Квартира: {lease.apartment.name}",
        f"Жилец: {lease.tenant.full_name}",
        f"Сумма к оплате: {money_text(situation_debt(target))}",
        f"Срок: {format_full_date(due)}",
        f"Просрочка: {overdue_days} дн.",
        f"Автоматических уведомлений: {int(situation.notification_count or 0)}",
        f"Статус: {situation.status}",
    ]
    if isinstance(target, RentCharge):
        lines.extend(["", "Назначения платежей:", *[f"• {item}" for item in situation_payment_breakdown(target)]])
    elif target.bill and target.bill.service:
        lines.append(f"Услуга: {target.bill.service.name}; оплата переводом по номеру телефона.")
    if situation.promise_date:
        lines.append(f"Новая дата ожидания: {format_full_date(situation.promise_date)}")
    if event == "ignored":
        lines.extend(["", "Автосистема остановила сообщения по этой ситуации, чтобы не спамить."])
    return "\n".join(lines)


def notify_owner_situation_event(
    session: Session,
    situation: PaymentSituation,
    event: str,
    today: date,
) -> bool:
    if situation_owner_event_sent(situation, event):
        return False
    owner_id = telegram_owner_chat_id(session)
    target = payment_situation_target(session, situation)
    if not owner_id or not telegram_token(session) or target is None:
        return False
    send_telegram_text(
        session,
        owner_id,
        owner_situation_text(situation, target, event, today),
        owner_situation_keyboard(situation),
    )
    mark_situation_owner_event(situation, event)
    return True


def due_day_notice_allows_send(session: Session, charge: RentCharge, now: datetime) -> bool:
    slot = datetime.combine(charge.due_date, time(hour=10), tzinfo=LOCAL_TZ)
    if now < slot:
        return False
    return not reminder_sent_today(session, "message_rent_due", rent_charge_id=charge.id, today=charge.due_date)


def run_owner_due_payment_digest(session: Session, today: date, now: datetime) -> int:
    if now < datetime.combine(today, time(hour=20), tzinfo=LOCAL_TZ):
        return 0
    owner_id = telegram_owner_chat_id(session)
    if not owner_id or not telegram_token(session):
        return 0
    sent_dates = get_internal_json(session, "owner_due_digest_dates", [])
    sent_set = {str(item) for item in sent_dates if isinstance(item, str)} if isinstance(sent_dates, list) else set()
    today_key = today.isoformat()
    if today_key in sent_set:
        return 0
    charges = session.scalars(
        select(RentCharge)
        .join(Lease)
        .join(Apartment)
        .where(Lease.active.is_(True), Apartment.active.is_(True), RentCharge.due_date == today)
        .order_by(RentCharge.due_date, RentCharge.id)
    ).all()
    missed = []
    for charge in charges:
        if lease_ignored(session, charge.lease_id):
            continue
        update_rent_charge_status(charge, today)
        paid = money(float(charge.ip_paid or 0) + float(charge.personal_paid or 0))
        debt = money(max(0.0, float(charge.ip_due or 0) - float(charge.ip_paid or 0)) + max(0.0, float(charge.personal_due or 0) - float(charge.personal_paid or 0)))
        if paid <= EPS and debt > EPS:
            missed.append(f"{charge.lease.apartment.object.name}, {charge.lease.apartment.name}, {charge.lease.tenant.full_name}: {money_text(debt)}")
    if not missed:
        sent_set.add(today_key)
        set_internal_json(session, "owner_due_digest_dates", sorted(sent_set))
        return 0
    text = "Сегодня был день оплаты, но денег не поступило:\n" + "\n".join(f"- {line}" for line in missed)
    try:
        send_telegram_text(session, owner_id, text, app_keyboard(app_base_url(session)))
    except HTTPException:
        return 0
    sent_set.add(today_key)
    set_internal_json(session, "owner_due_digest_dates", sorted(sent_set))
    return 1


def run_owner_payment_digest(session: Session, today: date, now: datetime) -> int:
    if now < datetime.combine(today, time(hour=10), tzinfo=LOCAL_TZ):
        return 0
    owner_id = telegram_owner_chat_id(session)
    if not owner_id or not telegram_token(session):
        return 0
    sent_dates = get_internal_json(session, "owner_payment_digest_dates", [])
    sent_set = {str(item) for item in sent_dates if isinstance(item, str)} if isinstance(sent_dates, list) else set()
    today_key = today.isoformat()
    if today_key in sent_set:
        return 0

    for charge in session.scalars(
        select(RentCharge).join(Lease).where(Lease.active.is_(True))
    ).all():
        update_rent_charge_status(charge, today)
        if situation_debt(charge) > EPS:
            sync_payment_situation(
                session,
                payment_situation(session, "rent", charge.id, charge.lease_id),
                today,
            )
    for line in session.scalars(
        select(UtilityBillLine).join(Lease, UtilityBillLine.lease_id == Lease.id).where(Lease.active.is_(True))
    ).all():
        update_utility_line_status(line, today)
        if line.lease_id and situation_debt(line) > EPS:
            sync_payment_situation(
                session,
                payment_situation(session, "utility", line.id, int(line.lease_id)),
                today,
            )

    situations = session.scalars(
        select(PaymentSituation)
        .where(PaymentSituation.status.not_in(PAYMENT_SITUATION_FINAL_STATUSES))
        .order_by(PaymentSituation.kind, PaymentSituation.id)
    ).all()
    due_today: list[str] = []
    overdue: list[str] = []
    needs_review: list[str] = []
    automated: list[str] = []
    for situation in situations:
        sync_payment_situation(session, situation, today)
        target = payment_situation_target(session, situation)
        if target is None or situation.status in PAYMENT_SITUATION_FINAL_STATUSES:
            continue
        lease = target.lease
        label = f"{lease.tenant.full_name} — {lease.apartment.object.name}, {lease.apartment.name}"
        debt = money_text(situation_debt(target))
        due = situation_due_date(target)
        if situation.status in {"receipt_pending_review", "question", "escalated", "manual", "ignored"}:
            needs_review.append(f"{label}: {situation.status}, {debt}")
        elif situation.status == "short_deferral":
            automated.append(f"{label}: пауза до {format_full_date(situation.promise_date)}, {debt}")
        elif due == today:
            due_today.append(f"{label}: {debt}")
        elif due and due < today:
            overdue.append(f"{label}: {(today - due).days} дн., {debt}, статус {situation.status}")

    if not any([due_today, overdue, needs_review, automated]):
        sent_set.add(today_key)
        set_internal_json(session, "owner_payment_digest_dates", sorted(sent_set))
        return 0
    lines = ["Сводка по платежам на сегодня:"]
    sections = [
        ("Сегодня должны оплатить", due_today),
        ("Просрочки", overdue),
        ("Нужна проверка владельца", needs_review),
        ("Автоматически обработано", automated),
    ]
    for title, items in sections:
        if not items:
            continue
        lines.extend(["", f"{title}:"])
        lines.extend(f"{index}. {item}" for index, item in enumerate(items[:8], start=1))
        if len(items) > 8:
            lines.append(f"…ещё {len(items) - 8}")
    send_telegram_text(session, owner_id, "\n".join(lines))
    sent_set.add(today_key)
    set_internal_json(session, "owner_payment_digest_dates", sorted(sent_set))
    return 1


def tenant_recent_dialogue_text(session: Session, lease: Lease, limit: int = 10) -> str:
    conversation = session.scalar(
        select(AiConversation)
        .where(
            AiConversation.lease_id == lease.id,
            AiConversation.role == "tenant",
            AiConversation.status == "active",
        )
        .order_by(AiConversation.updated_at.desc(), AiConversation.id.desc())
        .limit(1)
    )
    if not conversation:
        return "Недавний диалог: сообщений ещё нет."
    rows = session.scalars(
        select(AiMessage)
        .where(AiMessage.conversation_id == conversation.id)
        .order_by(AiMessage.created_at.desc(), AiMessage.id.desc())
        .limit(limit)
    ).all()
    if not rows:
        return "Недавний диалог: сообщений ещё нет."
    labels = {"user": "жилец", "assistant": "агент"}
    return "Недавний диалог:\n" + "\n".join(
        f"- {labels.get(row.role, row.role)}: {(row.text or '')[:700]}" for row in reversed(rows)
    )


def fallback_debt_followup_text(state: AgentTenantState, debt: float) -> str:
    level = int(state.escalation_level or 0)
    if level <= 0:
        return f"Добрый день. Вчера ожидалась оплата, но пока она не поступила. Подскажите, сегодня сможете оплатить {money_text(debt)}?"
    if level == 1:
        return f"Добрый день. Оплата {money_text(debt)} всё ещё не поступила. Назовите, пожалуйста, конкретную дату оплаты."
    if level == 2:
        return (
            f"Оплата {money_text(debt)} по-прежнему не поступила, а сроки уже переносились. "
            f"Нужна конкретная и выполнимая дата оплаты."
        )
    return (
        f"Задолженность {money_text(debt)} не погашена, предыдущие обещания не выполнены. "
        f"Ситуация передана владельцу. Сообщите окончательный план оплаты."
    )


def run_tenant_rent_dialogue(
    session: Session,
    charge: RentCharge,
    today: date,
) -> str:
    lease = charge.lease
    state = agent_tenant_state(session, lease)
    debt = money(
        max(0.0, float(charge.ip_due or 0) - float(charge.ip_paid or 0))
        + max(0.0, float(charge.personal_due or 0) - float(charge.personal_paid or 0))
    )
    if debt <= EPS:
        state.status = "normal"
        state.escalation_level = 0
        state.consecutive_excuses = 0
        state.promise_date = None
        state.promise_text = ""
        state.next_contact_on = None
        return "paid"

    overdue_days = max(1, (today - charge.due_date).days)
    if state.promise_date and today <= state.promise_date:
        return "waiting_promise"
    if state.next_contact_on and state.next_contact_on > today:
        return "waiting_schedule"
    if reminder_sent_today(session, "agent_rent_followup", rent_charge_id=charge.id, today=today):
        return "duplicate"

    if state.promise_date and today > state.promise_date:
        state.consecutive_excuses = int(state.consecutive_excuses or 0) + 1
        state.escalation_level = min(3, max(int(state.escalation_level or 0) + 1, 1))
        state.status = "promise_broken"
    elif overdue_days >= 7:
        state.escalation_level = max(int(state.escalation_level or 0), 2)
    elif overdue_days >= 3:
        state.escalation_level = max(int(state.escalation_level or 0), 1)

    context = "\n\n".join(
        [
            tenant_context_text(session, lease, get_settings(session), today),
            tenant_state_context(state),
            tenant_recent_dialogue_text(session, lease),
            agent_memory_context(
                session,
                scope_type="lease",
                scope_id=str(lease.id),
                lease_id=lease.id,
            ),
            f"Просрочка: {overdue_days} дн. Текущий долг: {money_text(debt)}.",
        ]
    )
    answer = call_deepseek_ai(
        session,
        chat_id=lease_chat_id(session, lease) or f"lease:{lease.id}",
        actor_role="tenant_automation",
        lease=lease,
        system_prompt=append_ai_instructions(
            TENANT_DEBT_FOLLOWUP_SYSTEM_PROMPT,
            get_setting_value(session, "ai_tenant_instructions"),
            "автоматических сообщений арендаторам",
        ),
        context=context,
        user_text="Сформулируй следующее сообщение жильцу по текущей просрочке.",
        model=resolve_ai_model(get_setting_value(session, "deepseek_model")),
        max_tokens=350,
    )
    if answer in {
        AI_UNAVAILABLE_TEXT,
        AI_DISABLED_TEXT,
        AI_BUDGET_EXCEEDED_TEXT,
        AI_DAILY_LIMIT_EXCEEDED_TEXT,
    }:
        answer = fallback_debt_followup_text(state, debt)
    send_tenant_message(
        session,
        lease,
        "custom",
        charge=charge,
        custom_text=answer,
        note="agent-auto-followup",
    )
    session.add(
        MessageLog(
            lease_id=lease.id,
            rent_charge_id=charge.id,
            channel="internal",
            template_key="agent_rent_followup",
            status="sent",
            recipient_chat_id=str(lease_chat_id(session, lease) or ""),
            text=answer,
            note=f"level={int(state.escalation_level or 0)} overdue_days={overdue_days}",
        )
    )
    state.last_agent_message_at = utc_now()
    state.next_contact_on = today + timedelta(days=2 if int(state.escalation_level or 0) < 3 else 1)
    state.status = "contacted"

    owner_id = telegram_owner_chat_id(session)
    if owner_id and telegram_token(session):
        send_telegram_text(
            session,
            owner_id,
            (
                f"Агент написал должнику\n"
                f"{lease.apartment.object.name}, {lease.apartment.name}, {lease.tenant.full_name}\n"
                f"Долг: {money_text(debt)}, просрочка: {overdue_days} дн., уровень: {int(state.escalation_level or 0)}\n"
                f"Текст: {answer}"
            ),
            app_keyboard(app_base_url(session)),
        )
        state.last_owner_update_at = utc_now()
    return "sent"


def upsert_agent_task(
    session: Session,
    *,
    issue_key: str,
    category: str,
    severity: str,
    title: str,
    details: str,
    lease_id: int | None = None,
    due_date: date | None = None,
) -> AgentTask:
    task = session.scalar(select(AgentTask).where(AgentTask.issue_key == issue_key).limit(1))
    now = utc_now()
    if not task:
        task = AgentTask(
            issue_key=issue_key,
            lease_id=lease_id,
            category=category,
            first_seen_at=now,
        )
        session.add(task)
    task.status = "open"
    task.severity = severity
    task.title = title[:240]
    task.details = details[:2000]
    task.due_date = due_date
    task.last_seen_at = now
    task.resolved_at = None
    return task


def sync_agent_tasks(session: Session, dashboard: dict[str, Any] | None = None) -> list[AgentTask]:
    if dashboard is None:
        dashboard = build_dashboard(session)
    seen: set[str] = set()
    today = date.today()

    for item in [*dashboard.get("rent_overdue", []), *dashboard.get("rent_partial", [])]:
        charge_id = int(item.get("id") or 0)
        if not charge_id:
            continue
        key = f"rent:{charge_id}"
        seen.add(key)
        due = date.fromisoformat(item["due_date"]) if item.get("due_date") else None
        days = (today - due).days if due else 0
        upsert_agent_task(
            session,
            issue_key=key,
            category="rent",
            severity="critical" if days >= 7 else "high",
            title=f"Просроченная аренда: {item.get('apartment') or ''} {item.get('tenant') or ''}".strip(),
            details=f"Долг {money_text(float(item.get('debt') or 0))}, просрочка {max(days, 0)} дн.",
            lease_id=int(item.get("lease_id") or 0) or None,
            due_date=due,
        )

    for item in dashboard.get("utility_overdue", []):
        line_id = int(item.get("id") or 0)
        if not line_id:
            continue
        key = f"utility:{line_id}"
        seen.add(key)
        due = date.fromisoformat(item["due_date"]) if item.get("due_date") else None
        upsert_agent_task(
            session,
            issue_key=key,
            category="utility",
            severity="high",
            title=f"Просрочена коммуналка: {item.get('apartment') or ''} {item.get('tenant') or ''}".strip(),
            details=f"Долг {money_text(float(item.get('debt') or 0))}.",
            lease_id=int(item.get("lease_id") or 0) or None,
            due_date=due,
        )

    for item in dashboard.get("suspicious_receipts", []):
        receipt_id = int(item.get("id") or 0)
        if not receipt_id:
            continue
        key = f"receipt:{receipt_id}"
        seen.add(key)
        upsert_agent_task(
            session,
            issue_key=key,
            category="receipt",
            severity="high",
            title=f"Проверить подозрительный чек #{receipt_id}",
            details=f"Сумма {money_text(float(item.get('amount') or 0))}, статус требует ручного решения.",
            lease_id=int(item.get("lease_id") or 0) or None,
        )

    for item in dashboard.get("stale_readings", []):
        service_id = int(item.get("service_id") or 0)
        key = f"reading:{service_id}"
        seen.add(key)
        upsert_agent_task(
            session,
            issue_key=key,
            category="reading",
            severity="medium",
            title=f"Обновить показания: {item.get('object') or ''} / {item.get('service') or ''}",
            details=f"Последняя дата: {item.get('last_date') or 'нет данных'}.",
        )

    for item in dashboard.get("manual_debts", []):
        debt_id = int(item.get("id") or 0)
        if not debt_id:
            continue
        key = f"manual-debt:{debt_id}"
        seen.add(key)
        upsert_agent_task(
            session,
            issue_key=key,
            category="manual_debt",
            severity="high" if item.get("status") == "overdue" else "medium",
            title=f"Ручной долг: {item.get('apartment') or ''} {item.get('tenant') or ''}".strip(),
            details=f"{item.get('title') or 'долг'} — {money_text(float(item.get('debt') or 0))}.",
            lease_id=int(item.get("lease_id") or 0) or None,
            due_date=date.fromisoformat(item["due_date"]) if item.get("due_date") else None,
        )

    for item in dashboard.get("pending_personal_expenses", []):
        expense_id = int(item.get("id") or 0)
        if not expense_id:
            continue
        key = f"expense:{expense_id}"
        seen.add(key)
        upsert_agent_task(
            session,
            issue_key=key,
            category="expense",
            severity="medium",
            title=f"Вернуть личные расходы: {item.get('category') or 'расход'}",
            details=f"{money_text(float(item.get('amount') or 0))}. {item.get('description') or ''}".strip(),
        )

    for item in dashboard.get("provider_debts", []):
        bill_id = int(item.get("id") or 0)
        if not bill_id:
            continue
        key = f"provider:{bill_id}"
        seen.add(key)
        upsert_agent_task(
            session,
            issue_key=key,
            category="provider",
            severity="high",
            title=f"Оплатить поставщика: {item.get('object') or ''} / {item.get('service') or ''}",
            details=f"Счёт на {money_text(float(item.get('total_cost') or 0))}, период до {item.get('period_end') or ''}.",
            due_date=date.fromisoformat(item["due_date"]) if item.get("due_date") else None,
        )

    for item in dashboard.get("monthly_reports", []):
        year = int(item.get("year") or 0)
        month = int(item.get("month") or 0)
        if not year or not month:
            continue
        key = f"report:{year:04d}-{month:02d}"
        seen.add(key)
        upsert_agent_task(
            session,
            issue_key=key,
            category="report",
            severity="medium",
            title=f"Закрыть месячный отчёт: {item.get('month_name') or month} {year}",
            details=f"{int(item.get('issue_count') or 0)} незакрытых проблем.",
        )

    managed_categories = {"rent", "utility", "receipt", "reading", "manual_debt", "expense", "provider", "report"}
    open_tasks = session.scalars(select(AgentTask).where(AgentTask.status == "open")).all()
    for task in open_tasks:
        if task.category in managed_categories and task.issue_key not in seen:
            task.status = "resolved"
            task.resolved_at = utc_now()
    session.flush()
    return session.scalars(
        select(AgentTask)
        .where(AgentTask.status == "open")
        .order_by(AgentTask.severity.desc(), AgentTask.first_seen_at, AgentTask.id)
    ).all()


def run_owner_supervisor_digest(
    session: Session,
    today: date,
    now: datetime,
    dashboard: dict[str, Any] | None = None,
) -> int:
    if not ai_enabled(session) or not setting_bool_value(get_setting_value(session, "ai_supervisor_enabled")):
        return 0
    if not setting_bool_value(get_setting_value(session, "hermes_briefing_enabled")):
        return 0
    if not ai_supervisor_schedule_due(session, today, now):
        return 0
    owner_id = telegram_owner_chat_id(session)
    if not owner_id or not telegram_token(session):
        return 0
    build = build_owner_briefing(
        session,
        today=today,
        max_chars=ai_feature_output_limit(session, "daily_briefing"),
    )
    if not build.briefing or not build.text:
        return 0
    response = send_telegram_text(session, owner_id, build.text, briefing_keyboard(session, build))
    message_id = int(((response.get("result") or {}) if isinstance(response, dict) else {}).get("message_id") or 0) or None
    mark_briefing_sent(build, telegram_message_id=message_id)
    return 1


def run_due_reminders(session: Session, today: date | None = None) -> dict[str, Any]:
    today = today or date.today()
    now = local_now()
    tenant_window_open = 10 <= now.hour < 20
    cutoff = notification_cutoff_date(session)
    if not notifications_enabled(session):
        dashboard = build_dashboard(session)
        supervisor_sent = run_owner_supervisor_digest(session, today, now, dashboard)
        payment_digest_sent = (
            0
            if setting_bool_value(get_setting_value(session, "hermes_briefing_enabled"))
            else run_owner_payment_digest(session, today, now)
        )
        return {
            "enabled": False,
            "cutoff_date": cutoff.isoformat(),
            "sent": 0,
            "skipped_disabled": 1,
            "skipped_legacy": 0,
            "skipped_duplicate": 0,
            "skipped_unlinked": 0,
            "failed": 0,
            "supervisor_digest_sent": supervisor_sent,
            "owner_payment_digest_sent": payment_digest_sent,
        }

    summary = {
        "enabled": True,
        "cutoff_date": cutoff.isoformat(),
        "sent": 0,
        "skipped_disabled": 0,
        "skipped_legacy": 0,
        "skipped_duplicate": 0,
        "skipped_unlinked": 0,
        "failed": 0,
    }

    charges = session.scalars(
        select(RentCharge)
        .join(Lease)
        .join(Apartment)
        .where(Lease.active.is_(True), Apartment.active.is_(True))
        .order_by(RentCharge.due_date, RentCharge.id)
    ).all()
    for charge in charges:
        update_rent_charge_status(charge, today)
        if lease_ignored(session, charge.lease_id):
            summary["skipped_disabled"] += 1
            continue
        if charge.due_date < cutoff:
            summary["skipped_legacy"] += 1
            continue
        if not lease_chat_id(session, charge.lease):
            summary["skipped_unlinked"] += 1
            continue
        situation = sync_payment_situation(
            session,
            payment_situation(session, "rent", charge.id, charge.lease_id),
            today,
        )
        rent_cadence_key = "message_rent_due" if charge.due_date >= today else "message_rent_overdue"
        if get_lease_cadence(session, charge.lease_id, rent_cadence_key) == "never":
            situation.mode = "manual"
            situation.status = "manual"
        if situation.status in PAYMENT_SITUATION_FINAL_STATUSES:
            continue
        if situation.mode == "manual" or situation.status in PAYMENT_SITUATION_PAUSED_STATUSES:
            summary["skipped_disabled"] += 1
            continue
        due_delta = (today - charge.due_date).days
        stage = ""
        template_key = ""
        if due_delta == -3 and tenant_window_open:
            stage, template_key = "upcoming", "message_rent_upcoming"
        elif due_delta == 0 and tenant_window_open:
            stage, template_key = "due", "message_rent_due"
        elif due_delta == 1 and tenant_window_open:
            stage, template_key = "overdue_1", "message_rent_overdue"
        elif due_delta >= 3 and tenant_window_open and situation.notification_count < MAX_AUTOMATIC_SITUATION_REMINDERS:
            if situation.promise_date and today >= situation.promise_date and situation.paused_until is None:
                stage = f"promise_followup_{situation.promise_date.isoformat()}"
            else:
                stage = "overdue_3"
            template_key = "message_rent_status_request"

        if stage and situation_stage_sent(session, situation, stage):
            summary["skipped_duplicate"] += 1
        elif stage:
            allowed, reason = hermes_reminder_allowed(session, situation=situation, now=now)
            if not allowed:
                summary["skipped_duplicate" if reason == "already_sent_today" else "skipped_disabled"] += 1
                stage = ""
        if stage:
            try:
                if send_payment_situation_message(session, situation, template_key, stage, today):
                    summary["sent"] += 1
                    record_reminder_outcome(
                        session,
                        lease_id=situation.lease_id,
                        stage=stage,
                        template_key=template_key,
                        payment_situation_id=situation.id,
                    )
                else:
                    summary["skipped_duplicate"] += 1
            except HTTPException as exc:
                summary["failed"] += 1
                emit_domain_event(
                    session,
                    "automation_failed",
                    entity_type="PaymentSituation",
                    entity_id=situation.id,
                    contract_id=situation.lease_id,
                    source="rent_reminder",
                    payload={"error": str(exc.detail), "stage": stage, "template_key": template_key},
                    idempotency_key=f"automation:rent-reminder:{situation.id}:{stage}:{today.isoformat()}",
                )

        if due_delta == 0 and now >= datetime.combine(today, time(hour=20), tzinfo=LOCAL_TZ):
            notify_owner_situation_event(session, situation, "due_unpaid", today)
        if due_delta in {1, 3, 5}:
            notify_owner_situation_event(session, situation, f"overdue_{due_delta}", today)
        if due_delta >= 3 and situation.notification_count >= MAX_AUTOMATIC_SITUATION_REMINDERS:
            situation.status = "ignored"
            situation.escalated_at = situation.escalated_at or utc_now()
            notify_owner_situation_event(session, situation, "ignored", today)

    lines = session.scalars(
        select(UtilityBillLine)
        .join(Lease, UtilityBillLine.lease_id == Lease.id)
        .join(Apartment, UtilityBillLine.apartment_id == Apartment.id)
        .where(Lease.active.is_(True), Apartment.active.is_(True))
        .order_by(UtilityBillLine.id)
    ).all()
    for line in lines:
        update_utility_line_status(line, today)
        if lease_ignored(session, line.lease_id):
            summary["skipped_disabled"] += 1
            continue
        if not line.due_date:
            continue
        if line.due_date < cutoff:
            summary["skipped_legacy"] += 1
            continue
        if not line.lease or not lease_chat_id(session, line.lease):
            summary["skipped_unlinked"] += 1
            continue
        if line.status not in {"issued", "overdue", "partial"}:
            continue
        situation = sync_payment_situation(
            session,
            payment_situation(session, "utility", line.id, int(line.lease_id or 0)),
            today,
        )
        if get_lease_cadence(session, int(line.lease_id or 0), "message_utility_bill") == "never":
            situation.mode = "manual"
            situation.status = "manual"
        if situation.status in PAYMENT_SITUATION_FINAL_STATUSES:
            continue
        if situation.mode == "manual" or situation.status in PAYMENT_SITUATION_PAUSED_STATUSES:
            summary["skipped_disabled"] += 1
            continue
        due_delta = (today - line.due_date).days
        stage = ""
        if due_delta == 0 and tenant_window_open:
            stage = "due"
        elif due_delta == 1 and tenant_window_open:
            stage = "overdue_1"
        elif due_delta >= 3 and tenant_window_open and situation.notification_count < MAX_AUTOMATIC_SITUATION_REMINDERS:
            if situation.promise_date and today >= situation.promise_date and situation.paused_until is None:
                stage = f"promise_followup_{situation.promise_date.isoformat()}"
            else:
                stage = "overdue_3"
        template_key = "message_utility_bill" if due_delta == 0 else "message_utility_overdue"

        if stage and situation_stage_sent(session, situation, stage):
            summary["skipped_duplicate"] += 1
        elif stage:
            allowed, reason = hermes_reminder_allowed(session, situation=situation, now=now)
            if not allowed:
                summary["skipped_duplicate" if reason == "already_sent_today" else "skipped_disabled"] += 1
                stage = ""
        if stage:
            try:
                if send_payment_situation_message(session, situation, template_key, stage, today):
                    summary["sent"] += 1
                    record_reminder_outcome(
                        session,
                        lease_id=situation.lease_id,
                        stage=stage,
                        template_key=template_key,
                        payment_situation_id=situation.id,
                    )
                else:
                    summary["skipped_duplicate"] += 1
            except HTTPException as exc:
                summary["failed"] += 1
                emit_domain_event(
                    session,
                    "automation_failed",
                    entity_type="PaymentSituation",
                    entity_id=situation.id,
                    contract_id=situation.lease_id,
                    source="utility_reminder",
                    payload={"error": str(exc.detail), "stage": stage, "template_key": template_key},
                    idempotency_key=f"automation:utility-reminder:{situation.id}:{stage}:{today.isoformat()}",
                )

        if due_delta == 0 and now >= datetime.combine(today, time(hour=20), tzinfo=LOCAL_TZ):
            notify_owner_situation_event(session, situation, "due_unpaid", today)
        if due_delta in {1, 3, 5}:
            notify_owner_situation_event(session, situation, f"overdue_{due_delta}", today)
        if due_delta >= 3 and situation.notification_count >= MAX_AUTOMATIC_SITUATION_REMINDERS:
            situation.status = "ignored"
            situation.escalated_at = situation.escalated_at or utc_now()
            notify_owner_situation_event(session, situation, "ignored", today)

    dashboard = build_dashboard(session)
    summary["owner_due_digest_sent"] = 0
    summary["owner_payment_digest_sent"] = (
        0
        if setting_bool_value(get_setting_value(session, "hermes_briefing_enabled"))
        else run_owner_payment_digest(session, today, now)
    )
    summary["supervisor_digest_sent"] = run_owner_supervisor_digest(session, today, now, dashboard)
    summary["move_out_processed"] = process_move_out_notifications(session, today)["processed"]
    return summary


def format_date(value: date | None) -> str:
    if not value:
        return ""
    month = MONTH_NAMES[value.month - 1]
    today = date.today()
    if value.year != today.year:
        return f"{value.day} {month} {value.year}"
    if value.month != today.month:
        return f"{value.day} {month}"
    return f"{value.day} число"


def format_full_date(value: date | None) -> str:
    if not value:
        return ""
    return value.strftime("%d.%m.%Y")


def message_sender_label(message: dict[str, Any]) -> str:
    sender = message.get("from") or {}
    full_name = " ".join(part for part in [sender.get("first_name"), sender.get("last_name")] if part).strip()
    username = normalize_telegram_handle(sender.get("username") or "")
    if full_name and username:
        return f"{full_name} (@{username})"
    if full_name:
        return full_name
    if username:
        return f"@{username}"
    return "неизвестный отправитель"


def receipt_channel_label(value: str) -> str:
    return {
        "ip": "ИП",
        "personal": "По номеру",
        "expense_fund": "Мне на расходы",
        "utilities": "коммуналка",
        "utility_advance": "аванс коммуналки",
        "unknown": "неизвестно",
    }.get(value, value or "неизвестно")


def receipt_status_label(value: str) -> str:
    return {
        "accepted": "принят",
        "suspicious": "нужна проверка",
        "duplicate": "дубликат",
        "rejected": "отклонён",
        "ignored": "скрыт",
    }.get(value, value or "неизвестно")


def receipt_storage_dir() -> Path:
    path = ROOT_DIR / "data" / "telegram_receipts"
    path.mkdir(parents=True, exist_ok=True)
    return path


def telegram_media_descriptor(message: dict[str, Any]) -> dict[str, str]:
    document = message.get("document")
    if document:
        return {
            "file_id": document.get("file_id") or "",
            "file_name": Path(document.get("file_name") or "telegram-document.bin").name,
            "mime_type": document.get("mime_type") or "",
            "file_unique_id": document.get("file_unique_id") or "",
        }
    photos = message.get("photo") or []
    if photos:
        largest = photos[-1]
        return {
            "file_id": largest.get("file_id") or "",
            "file_name": f"telegram-photo-{largest.get('file_unique_id') or utc_now().timestamp()}.jpg",
            "mime_type": "image/jpeg",
            "file_unique_id": largest.get("file_unique_id") or "",
        }
    return {"file_id": "", "file_name": "", "mime_type": "", "file_unique_id": ""}


def save_telegram_media(session: Session, message: dict[str, Any]) -> Path | None:
    token = telegram_token(session)
    media = telegram_media_descriptor(message)
    file_id = media["file_id"]
    name = media["file_name"]
    if not token or not file_id:
        return None
    info = telegram_file_info(token, file_id)
    file_path = info.get("file_path") or ""
    if not file_path:
        return None
    extension = Path(name).suffix or Path(file_path).suffix or ".bin"
    stamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    target = receipt_storage_dir() / f"{stamp}-{Path(name).stem[:40]}{extension}"
    return download_telegram_file(token, file_path, target)


def receipt_candidate_lease(session: Session, message: dict[str, Any], parsed: dict[str, Any] | None, linked_lease: Lease | None) -> Lease | None:
    if linked_lease:
        return linked_lease
    if parsed and parsed.get("purpose"):
        return find_lease_by_receipt_purpose(session, parsed.get("purpose") or "")
    return None


def receipt_operation_fingerprint(parsed: dict[str, Any] | None) -> str:
    if not parsed:
        return ""
    parts = [
        str(parsed.get("source_bank") or "").strip().lower(),
        str(parsed.get("receipt_number") or "").strip().lower(),
        str(parsed.get("paid_at") or "").strip(),
        f"{float(parsed.get('amount') or 0):.2f}",
        str(parsed.get("recipient_name") or "").strip().lower(),
        str(parsed.get("recipient_account") or "").strip(),
        str(parsed.get("recipient_phone") or "").strip(),
        str(parsed.get("payer_name") or "").strip().lower(),
    ]
    if not any(parts[1:4]):
        return ""
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


def receipt_document_sha256(saved_path: Path) -> str:
    return hashlib.sha256(saved_path.read_bytes()).hexdigest()


def build_receipt_details_payload(message: dict[str, Any], parsed: dict[str, Any] | None, saved_path: Path) -> dict[str, Any]:
    media = telegram_media_descriptor(message)
    payload = dict(parsed or {})
    payload["_meta"] = {
        "received_at": utc_now().isoformat(),
        "document_sha256": receipt_document_sha256(saved_path),
        "operation_fingerprint": receipt_operation_fingerprint(parsed),
        "telegram_chat_id": str((message.get("chat") or {}).get("id") or ""),
        "telegram_message_id": str(message.get("message_id") or ""),
        "telegram_file_id": media.get("file_id") or "",
        "telegram_file_unique_id": media.get("file_unique_id") or "",
        "file_name": saved_path.name,
    }
    return payload


def duplicate_receipt(session: Session, document_sha256: str, operation_fingerprint: str) -> PaymentReceipt | None:
    if not document_sha256 and not operation_fingerprint:
        return None
    receipts = session.scalars(select(PaymentReceipt).where(PaymentReceipt.source == "telegram").order_by(PaymentReceipt.id.desc())).all()
    for receipt in receipts:
        parsed = parse_receipt_details(receipt)
        meta = receipt_meta(parsed)
        if document_sha256 and meta.get("document_sha256") == document_sha256:
            return receipt
        if operation_fingerprint and meta.get("operation_fingerprint") == operation_fingerprint:
            return receipt
    return None


def recipient_mismatch_issue(channel: str, issues: list[str]) -> bool:
    if channel == "ip":
        return any("получатель ип" in issue.lower() or "счёт получателя ип" in issue.lower() for issue in issues)
    if channel == "personal":
        return any(
            marker in issue.lower()
            for marker in [
                "получатель перевода",
                "номер получателя перевода",
            ]
            for issue in issues
        )
    return False


def personal_rent_candidates_for_payment(session: Session, lease: Lease, paid_at: datetime, cutoff: date) -> list[RentCharge]:
    paid_day = paid_at.date()
    generate_rent_charges(session, until=paid_day + timedelta(days=40))
    current_charge: RentCharge | None = None
    candidates: list[RentCharge] = []
    charges = session.scalars(
        select(RentCharge)
        .where(RentCharge.lease_id == lease.id, RentCharge.due_date >= cutoff)
        .order_by(RentCharge.due_date.desc(), RentCharge.id.desc())
    ).all()
    for charge in charges:
        personal_debt = money(max(0.0, charge.personal_due - charge.personal_paid))
        if personal_debt <= EPS:
            continue
        if charge.due_date.year == paid_day.year and charge.due_date.month == paid_day.month:
            current_charge = charge
        elif charge.due_date <= paid_day:
            candidates.append(charge)
    if current_charge:
        candidates.insert(0, current_charge)
    return candidates


def utility_transfer_parts(lines: list[UtilityBillLine], amount: float) -> tuple[list[tuple[str, UtilityBillLine, float]], float]:
    remaining = money(amount)
    parts: list[tuple[str, UtilityBillLine, float]] = []
    for line in lines:
        debt = money(max(0.0, line.total_amount - line.paid_amount))
        portion = min(remaining, debt)
        if portion <= EPS:
            continue
        parts.append(("utility", line, money(portion)))
        remaining = money(remaining - portion)
        if remaining <= EPS:
            break
    return parts, money(remaining)


def build_personal_transfer_plan(
    session: Session,
    lease: Lease,
    amount: float,
    *,
    paid_at: datetime,
) -> tuple[list[tuple[str, RentCharge | UtilityBillLine, float]], float, str]:
    amount = money(amount)
    cutoff = allocation_cutoff_date(session)
    utility_candidates = [line for line in utility_line_candidates(session, lease.id) if debt_visible_by_cutoff(line.due_date, cutoff)]
    rent_candidates = personal_rent_candidates_for_payment(session, lease, paid_at, cutoff)

    parts: list[tuple[str, RentCharge | UtilityBillLine, float]] = []
    remaining = amount
    if rent_candidates:
        charge = rent_candidates[0]
        personal_debt = money(max(0.0, charge.personal_due - charge.personal_paid))
        rent_portion = min(remaining, personal_debt)
        if rent_portion > EPS:
            parts.append(("rent", charge, money(rent_portion)))
            remaining = money(remaining - rent_portion)

    if remaining > EPS:
        utility_parts, utility_remaining = utility_transfer_parts(utility_candidates, remaining)
        parts.extend(utility_parts)
        if utility_parts and utility_remaining > EPS:
            kind, line, portion = parts[-1]
            parts[-1] = (kind, line, money(portion + utility_remaining))
            remaining = 0.0
        else:
            remaining = utility_remaining

    if not parts:
        return [], amount, ""
    kinds = {kind for kind, _item, _portion in parts}
    if kinds == {"utility"}:
        comment = "зачёт перевода: вся сумма ушла в коммуналку"
    elif kinds == {"rent"}:
        comment = "зачёт перевода: личная часть аренды"
    else:
        comment = "зачёт перевода: личная часть аренды, всё сверх неё ушло в коммуналку"
    return parts, money(remaining), comment


def create_personal_priority_receipts(
    session: Session,
    lease: Lease,
    amount: float,
    *,
    paid_at: datetime,
    source: str,
    status: str,
    recipient_name: str = "",
    recipient_details: str = "",
    notes: str = "",
    file_path: str = "",
) -> list[PaymentReceipt]:
    parts, remaining, _comment = build_personal_transfer_plan(session, lease, amount, paid_at=paid_at)
    if not parts or remaining > EPS:
        raise ValueError("по переводу не найдено долга для зачёта")

    receipts: list[PaymentReceipt] = []
    for kind, item, portion in parts:
        if kind == "utility":
            line = item
            receipts.append(
                PaymentReceipt(
                    lease_id=lease.id,
                    utility_line_id=line.id,
                    apartment_id=lease.apartment_id,
                    amount=portion,
                    channel="utilities",
                    paid_at=paid_at,
                    source=source,
                    status=status,
                    recipient_name=recipient_name,
                    recipient_details=recipient_details,
                    file_path=file_path,
                    notes=notes,
                )
            )
        else:
            charge = item
            receipts.append(
                PaymentReceipt(
                    lease_id=lease.id,
                    rent_charge_id=charge.id,
                    apartment_id=lease.apartment_id,
                    amount=portion,
                    channel="personal",
                    paid_at=paid_at,
                    source=source,
                    status=status,
                    recipient_name=recipient_name,
                    recipient_details=recipient_details,
                    file_path=file_path,
                    notes=notes,
                )
            )

    session.add_all(receipts)
    session.flush()
    recalculate_lease_balances(session, lease.id)
    return receipts


def apply_receipt_match(session: Session, lease: Lease, parsed: dict[str, Any]) -> tuple[str, str, int | None, list[str], str]:
    amount = float(parsed.get("amount") or 0)
    channel = detect_receipt_channel(parsed)
    issues = receipt_validation_issues(parsed, get_settings(session), channel)
    allocation_comment = ""
    paid_at = datetime.fromisoformat(parsed["paid_at"]) if parsed.get("paid_at") else None
    cutoff = allocation_cutoff_date(session)
    if amount <= 0:
        issues.append("в чеке не удалось определить положительную сумму")

    if channel == "ip":
        plan, remaining, decision = build_rent_plan(
            session,
            lease,
            "ip",
            amount,
            exact_only=False,
            paid_at=paid_at,
            prefer_document_month=True,
            cutoff=cutoff,
        )
        allocation_comment = describe_rent_allocation_decision(decision)
        if not plan or remaining > EPS:
            issues.append("не удалось честно разложить платёж по аренде")
        if issues:
            return "suspicious", "review", None, issues, allocation_comment
        return "accepted", "rent", plan[0][0].id, [], allocation_comment

    if channel == "personal":
        if recipient_mismatch_issue(channel, issues):
            return "rejected", "review", None, issues, ""
        if issues:
            return "suspicious", "review", None, issues, ""
        plan_paid_at = paid_at if isinstance(paid_at, datetime) else utc_now()
        parts, remaining, allocation_comment = build_personal_transfer_plan(session, lease, amount, paid_at=plan_paid_at)
        if parts and remaining <= EPS:
            kinds = {kind for kind, _item, _portion in parts}
            first_kind, first_item, _portion = parts[0]
            match_type = "mixed" if len(kinds) > 1 else "utility" if first_kind == "utility" else "rent"
            return "accepted", match_type, first_item.id, [], allocation_comment
        return "suspicious", "review", None, ["по переводу не найдено долга для зачёта"], ""

    if recipient_mismatch_issue(channel, issues):
        return "rejected", "review", None, issues, ""
    return "suspicious", "review", None, ["тип перевода пока не удалось определить"], ""


def owner_receipt_alert_text(
    session: Session,
    lease: Lease | None,
    parsed: dict[str, Any] | None,
    receipt_status: str,
    issues: list[str],
    sender_label: str,
    allocation_comment: str = "",
) -> str:
    summary_bits = [sender_label]
    if parsed:
        if parsed.get("recipient_name"):
            summary_bits.append(f"получатель: {parsed['recipient_name']}")
        if parsed.get("purpose"):
            summary_bits.append(f"назначение: {parsed['purpose']}")
    if allocation_comment:
        summary_bits.append(f"зачёт: {allocation_comment}")
    if issues:
        summary_bits.append("вопросы: " + "; ".join(issues))
    context = {
        "tenant_name": lease.tenant.full_name if lease else sender_label,
        "apartment": lease.apartment.name if lease else "не определена",
        "amount": money_text(float(parsed.get("amount") or 0)) if parsed else "0,00 ₽",
        "channel": receipt_channel_label(detect_receipt_channel(parsed or {})),
        "receipt_status": receipt_status_label(receipt_status),
        "receipt_summary": ". ".join(summary_bits),
    }
    return fill_template(message_template(session, "message_owner_receipt_alert"), context)


def handle_tenant_receipt_message(session: Session, message: dict[str, Any], linked_lease: Lease | None) -> None:
    chat_id = (message.get("chat") or {}).get("id")
    if not chat_id:
        return
    owner_id = telegram_owner_chat_id(session)
    saved_path = save_telegram_media(session, message)
    if not saved_path:
        if owner_id and message.get("message_id"):
            try:
                copy_message(
                    telegram_token(session),
                    to_chat_id=owner_id,
                    from_chat_id=chat_id,
                    message_id=int(message["message_id"]),
                    reply_markup=app_keyboard(app_base_url(session)),
                )
            except RuntimeError:
                pass
        send_telegram_text(session, chat_id, message_template(session, "message_receipt_review"), tenant_keyboard())
        return

    parsed: dict[str, Any] | None = None
    if saved_path.suffix.lower() == ".pdf":
        try:
            parsed = parse_receipt_file(saved_path)
        except Exception:
            parsed = None

    receipt_payload = build_receipt_details_payload(message, parsed, saved_path)
    meta = receipt_meta(receipt_payload)
    receipt_details = json.dumps(receipt_payload, ensure_ascii=False)
    lease = receipt_candidate_lease(session, message, parsed, linked_lease)
    status = "suspicious"
    match_type = "review"
    linked_id: int | None = None
    allocation_comment = ""
    issues = ["чек сохранён, но не удалось ничего распознать"]
    duplicate = duplicate_receipt(
        session,
        str(meta.get("document_sha256") or ""),
        str(meta.get("operation_fingerprint") or ""),
    )
    if duplicate:
        status = "duplicate"
        issues = [f"этот чек уже есть в системе с {duplicate.created_at:%d.%m.%Y}, повторно не зачитываю"]
    elif parsed and lease:
        status, match_type, linked_id, issues, allocation_comment = apply_receipt_match(session, lease, parsed)
    elif parsed and not lease:
        issues = ["чек разобран, но арендатор по нему пока не определён"]

    channel = detect_receipt_channel(parsed or {})
    paid_at = datetime.fromisoformat(parsed["paid_at"]) if parsed and parsed.get("paid_at") else utc_now()
    stored_path = str(saved_path.relative_to(ROOT_DIR))
    if status == "accepted" and lease and parsed:
        try:
            if channel == "personal":
                create_personal_priority_receipts(
                    session,
                    lease,
                    float(parsed.get("amount") or 0),
                    paid_at=paid_at,
                    source="telegram",
                    status="accepted",
                    recipient_name=parsed.get("recipient_name") or "",
                    recipient_details=receipt_details,
                    notes="автоматически принято из Telegram",
                    file_path=stored_path,
                )
            elif match_type == "rent":
                create_rent_receipts(
                    session,
                    lease,
                    "ip" if channel == "ip" else "personal",
                    float(parsed.get("amount") or 0),
                    paid_at=paid_at,
                    source="telegram",
                    status="accepted",
                    recipient_name=parsed.get("recipient_name") or "",
                    recipient_details=receipt_details,
                    notes="автоматически принято из Telegram",
                    file_path=stored_path,
                    exact_only=channel == "personal",
                    prefer_document_month=True,
                    cutoff=allocation_cutoff_date(session),
                )
            elif match_type == "utility":
                receipts = create_utility_receipts(
                    session,
                    lease,
                    float(parsed.get("amount") or 0),
                    paid_at=paid_at,
                    source="telegram",
                    status="accepted",
                    recipient_name=parsed.get("recipient_name") or "",
                    recipient_details=receipt_details,
                    notes="автоматически принято из Telegram",
                    file_path=stored_path,
                    exact_only=True,
                )
                sync_utility_advance_credits_for_receipts(session, receipts)
            else:
                status = "suspicious"
                issues = ["бот не понял, куда зачесть чек"]
        except ValueError as exc:
            status = "suspicious"
            match_type = "review"
            issues = [str(exc)]

    if status != "accepted":
        session.add(
            PaymentReceipt(
                lease_id=lease.id if lease else None,
                apartment_id=lease.apartment_id if lease else None,
                amount=float((parsed or {}).get("amount") or 0),
                channel=channel or "unknown",
                paid_at=paid_at,
                source="telegram",
                status="duplicate" if status == "duplicate" else "rejected" if status == "rejected" else "suspicious",
                recipient_name=(parsed or {}).get("recipient_name") or "",
                recipient_details=receipt_details,
                file_path=stored_path,
                notes="; ".join(issues),
            )
        )

    if lease:
        situations = session.scalars(
            select(PaymentSituation)
            .where(
                PaymentSituation.lease_id == lease.id,
                PaymentSituation.status.not_in(PAYMENT_SITUATION_FINAL_STATUSES),
            )
            .order_by(PaymentSituation.updated_at.desc(), PaymentSituation.id.desc())
        ).all()
        preferred_kind = "utility" if match_type == "utility" or channel == "utilities" else "rent" if match_type == "rent" or channel in {"ip", "personal"} else ""
        affected = [item for item in situations if not preferred_kind or item.kind == preferred_kind]
        if not affected and situations:
            affected = [situations[0]]
        for situation in affected:
            situation.last_tenant_response_at = utc_now()
            situation.tenant_response_count = int(situation.tenant_response_count or 0) + 1
            if status == "accepted":
                sync_payment_situation(session, situation)
                if situation.status != "paid":
                    situation.status = "awaiting_payment"
            elif status != "duplicate":
                situation.status = "receipt_pending_review"
                situation.escalated_at = utc_now()
            notify_owner_situation_event(session, situation, "receipt", date.today())

    if status == "accepted":
        send_telegram_text(session, chat_id, message_template(session, "message_receipt_received"), tenant_keyboard())
    elif status == "duplicate":
        send_telegram_text(session, chat_id, message_template(session, "message_receipt_duplicate"), tenant_keyboard())
    elif status == "rejected":
        send_telegram_text(session, chat_id, "Некорректный получатель, отправляю владельцу для рассмотрения.", tenant_keyboard())
    else:
        send_telegram_text(session, chat_id, message_template(session, "message_receipt_review"), tenant_keyboard())

    if owner_id:
        send_telegram_text(
            session,
            owner_id,
            owner_receipt_alert_text(session, lease, parsed, status, issues, message_sender_label(message), allocation_comment),
        )
        if message.get("message_id"):
            try:
                copy_message(
                    telegram_token(session),
                    to_chat_id=owner_id,
                    from_chat_id=chat_id,
                    message_id=int(message["message_id"]),
                )
            except RuntimeError:
                pass


def send_telegram_text(session: Session, chat_id: int | str, text: str, keyboard: dict[str, Any] | None = None) -> dict[str, Any]:
    started = time_module.perf_counter()
    token = telegram_token(session)
    runtime_log("TELEGRAM", f"send attempt chat_id={chat_id} token_configured={bool(token)} text_len={len(text or '')}")
    if not token:
        runtime_log("TELEGRAM", f"send skipped chat_id={chat_id} reason=missing_token")
        raise HTTPException(400, "Не задан токен Telegram-бота")
    try:
        result = send_message(token, chat_id, text, reply_markup=keyboard)
        message_id = ((result.get("result") or {}) if isinstance(result, dict) else {}).get("message_id")
        elapsed_ms = int((time_module.perf_counter() - started) * 1000)
        runtime_log("TELEGRAM", f"send ok chat_id={chat_id} message_id={message_id} duration_ms={elapsed_ms}")
        return result
    except TelegramApiError as exc:
        elapsed_ms = int((time_module.perf_counter() - started) * 1000)
        runtime_log("TELEGRAM", f"send failed chat_id={chat_id} duration_ms={elapsed_ms} error={exc}")
        raise HTTPException(400, str(exc)) from exc


def ai_enabled(session: Session) -> bool:
    return setting_bool_value(get_setting_value(session, "ai_enabled"))


def ai_tenant_free_text_enabled(session: Session) -> bool:
    return setting_bool_value(get_setting_value(session, "ai_tenant_free_text_enabled"))


def bounded_ai_int_setting(session: Session, key: str, fallback: int) -> int:
    minimum, maximum, _label = AI_INTEGER_SETTING_LIMITS[key]
    try:
        return min(maximum, max(minimum, int(get_setting_value(session, key) or fallback)))
    except ValueError:
        return fallback


def ai_daily_call_limit(session: Session) -> int:
    return bounded_ai_int_setting(session, "ai_daily_call_limit", 100)


def ai_max_output_tokens(session: Session) -> int:
    return bounded_ai_int_setting(session, "ai_max_output_tokens", 1500)


def ai_supervisor_max_tokens(session: Session) -> int:
    return bounded_ai_int_setting(session, "ai_supervisor_max_tokens", 1200)


def ai_usd_rub_rate(session: Session) -> float:
    try:
        raw_rate = str(get_setting_value(session, "ai_usd_rub_rate") or 100).replace(",", ".")
        return min(1000.0, max(1.0, float(raw_rate)))
    except ValueError:
        return AI_DEFAULT_USD_RUB_RATE


def ai_daily_calls(session: Session, today: date | None = None) -> int:
    usage_date = today or date.today()
    return int(
        session.scalar(
            select(func.coalesce(func.sum(AiUsageDaily.calls), 0)).where(AiUsageDaily.usage_date == usage_date)
        )
        or 0
    )


def ai_daily_call_limit_exceeded(session: Session, today: date | None = None) -> bool:
    limit = ai_daily_call_limit(session)
    return bool(limit > 0 and ai_daily_calls(session, today) >= limit)


AI_FEATURE_LIMIT_SETTINGS = {
    "owner_chat": "ai_feature_owner_chat_daily_limit",
    "tenant_chat": "ai_feature_tenant_chat_daily_limit",
    "case_details": "ai_feature_case_details_daily_limit",
    "skill_builder": "ai_feature_skill_builder_daily_limit",
    "deep_audit": "ai_feature_deep_audit_daily_limit",
    "key_test": "ai_feature_key_test_daily_limit",
}

AI_OUTPUT_LIMIT_SETTINGS = {
    "daily_briefing": "ai_output_chars_daily_briefing",
    "case_details": "ai_output_chars_case_details",
    "owner_chat": "ai_output_chars_owner_chat",
    "tenant_chat": "ai_output_chars_tenant_chat",
    "skill_builder": "ai_output_chars_skill_proposal",
}


def ai_feature_output_limit(session: Session, feature: str, *, detailed: bool = False) -> int:
    fallback = FEATURE_OUTPUT_LIMITS.get(feature, 800)
    setting_key = AI_OUTPUT_LIMIT_SETTINGS.get(feature)
    if not setting_key:
        return fallback
    value = bounded_ai_int_setting(session, setting_key, fallback)
    if feature == "daily_briefing":
        return min(800, value)
    if detailed and feature in {"owner_chat", "case_details"}:
        return min(AI_INTEGER_SETTING_LIMITS[setting_key][1], max(value, value * 2))
    return value


def ai_feature_daily_limit_exceeded(session: Session, feature: str, today: date | None = None) -> bool:
    setting_key = AI_FEATURE_LIMIT_SETTINGS.get(feature)
    if not setting_key:
        return False
    limit = bounded_ai_int_setting(session, setting_key, 0)
    if limit <= 0:
        return False
    usage_date = today or date.today()
    calls = int(
        session.scalar(
            select(func.coalesce(func.sum(AiFeatureUsageDaily.calls), 0)).where(
                AiFeatureUsageDaily.usage_date == usage_date,
                AiFeatureUsageDaily.feature == feature,
            )
        )
        or 0
    )
    return calls >= limit


def ai_supervisor_schedule_due(session: Session, today: date, now: datetime) -> bool:
    cadence = get_setting_value(session, "ai_supervisor_cadence") or "daily"
    if cadence == "weekdays" and today.weekday() >= 5:
        return False
    if cadence == "weekly":
        weekday = bounded_ai_int_setting(session, "ai_supervisor_weekday", 0)
        if today.weekday() != weekday:
            return False
    configured_time = get_setting_value(session, "ai_supervisor_time") or "10:00"
    try:
        hour, minute = (int(part) for part in configured_time.split(":", 1))
        scheduled_time = time(hour=hour, minute=minute)
    except (TypeError, ValueError):
        scheduled_time = time(hour=10)
    return now >= datetime.combine(today, scheduled_time, tzinfo=LOCAL_TZ)


def ai_monthly_budget_rub(session: Session) -> float:
    try:
        return max(0.0, float(str(get_setting_value(session, "ai_monthly_budget_rub") or "0").replace(",", ".")))
    except ValueError:
        return 0.0


def ai_monthly_spent_rub(session: Session, today: date | None = None) -> float:
    value = today or date.today()
    month_start = value.replace(day=1)
    spent = session.scalar(
        select(func.coalesce(func.sum(AiUsageDaily.cost_rub), 0.0)).where(AiUsageDaily.usage_date >= month_start)
    )
    return money(float(spent or 0.0))


def ai_budget_exceeded(session: Session, today: date | None = None) -> bool:
    budget = ai_monthly_budget_rub(session)
    return bool(budget > 0 and ai_monthly_spent_rub(session, today) >= budget)


def resolve_ai_model(model: str) -> str:
    value = (model or "").strip() or str(DEFAULT_SETTINGS["deepseek_model"])
    try:
        return normalize_deepseek_model(value)
    except AiProviderConfigError:
        return str(DEFAULT_SETTINGS["deepseek_model"])


def ai_model_price_key(model: str) -> str:
    normalized = (model or "").strip().lower()
    if "deepseek-v4-flash" in normalized:
        return "deepseek-v4-flash"
    if "deepseek-v4-pro" in normalized:
        return "deepseek-v4-pro"
    return "deepseek-v4-flash"


def ai_estimated_cost_rub(
    model: str,
    prompt_tokens: int,
    completion_tokens: int,
    usd_rub_rate: float = AI_DEFAULT_USD_RUB_RATE,
) -> float:
    price_key = ai_model_price_key(model)
    input_usd, output_usd = AI_MODEL_PRICES_USD_PER_M.get(
        price_key,
        AI_MODEL_PRICES_USD_PER_M["deepseek-v4-flash"],
    )
    cost_usd = (prompt_tokens / 1_000_000) * input_usd + (completion_tokens / 1_000_000) * output_usd
    return money(cost_usd * usd_rub_rate)


def ai_conversation(session: Session, chat_id: int | str, role: str, lease: Lease | None = None) -> AiConversation:
    chat_ref = str(chat_id)
    query = select(AiConversation).where(
        AiConversation.chat_id == chat_ref,
        AiConversation.role == role,
        AiConversation.status == "active",
    )
    if lease:
        query = query.where(AiConversation.lease_id == lease.id)
    conversation = session.scalar(query.order_by(AiConversation.updated_at.desc(), AiConversation.id.desc()).limit(1))
    if conversation:
        return conversation
    conversation = AiConversation(
        chat_id=chat_ref,
        role=role,
        lease_id=lease.id if lease else None,
        tenant_id=lease.tenant_id if lease else None,
        title=f"{role}:{chat_ref}",
    )
    session.add(conversation)
    session.flush()
    return conversation


def log_ai_message(
    session: Session,
    conversation: AiConversation,
    role: str,
    text: str,
    *,
    model: str = "",
    prompt_tokens: int = 0,
    completion_tokens: int = 0,
    cost_rub: float = 0.0,
) -> AiMessage:
    message = AiMessage(
        conversation_id=conversation.id,
        lease_id=conversation.lease_id,
        role=role,
        channel="telegram",
        text=text,
        model=model,
        prompt_tokens=prompt_tokens,
        completion_tokens=completion_tokens,
        cost_rub=cost_rub,
    )
    conversation.updated_at = utc_now()
    session.add(message)
    session.flush()
    return message


def recent_ai_history(session: Session, conversation: AiConversation, limit: int = 12) -> list[dict[str, str]]:
    rows = session.scalars(
        select(AiMessage)
        .where(AiMessage.conversation_id == conversation.id, AiMessage.role.in_(["user", "assistant"]))
        .order_by(AiMessage.created_at.desc(), AiMessage.id.desc())
        .limit(max(1, limit))
    ).all()
    return [
        {"role": row.role, "content": (row.text or "")[:2400]}
        for row in reversed(rows)
        if (row.text or "").strip()
    ]


def agent_memory_context(
    session: Session,
    *,
    scope_type: str,
    scope_id: str,
    lease_id: int | None = None,
    limit: int = 20,
) -> str:
    query = select(AgentMemory).where(
        AgentMemory.scope_type == scope_type,
        AgentMemory.scope_id == scope_id,
        AgentMemory.status == "active",
    )
    if lease_id is not None:
        query = query.where(AgentMemory.lease_id == lease_id)
    memories = session.scalars(
        query.order_by(AgentMemory.importance.desc(), AgentMemory.updated_at.desc(), AgentMemory.id.desc()).limit(limit)
    ).all()
    if not memories:
        return "Долговременная память и навыки: пока пусто."
    return "Долговременная память и навыки:\n" + "\n".join(
        f"- [{memory.kind}] {memory.content}" for memory in memories
    )


def store_agent_memories(
    session: Session,
    conversation: AiConversation,
    memories: list[dict[str, Any]],
    *,
    scope_type: str,
    scope_id: str,
    lease_id: int | None = None,
) -> None:
    for item in memories[:5]:
        content = str(item.get("content") or "").strip()[:500]
        if not content:
            continue
        existing = session.scalar(
            select(AgentMemory)
            .where(
                AgentMemory.scope_type == scope_type,
                AgentMemory.scope_id == scope_id,
                AgentMemory.kind == str(item.get("kind") or "fact"),
                func.lower(AgentMemory.content) == content.lower(),
                AgentMemory.status == "active",
            )
            .limit(1)
        )
        if existing:
            existing.importance = max(int(existing.importance or 1), int(item.get("importance") or 1))
            existing.updated_at = utc_now()
            continue
        session.add(
            AgentMemory(
                scope_type=scope_type,
                scope_id=scope_id,
                conversation_id=conversation.id,
                lease_id=lease_id,
                kind=str(item.get("kind") or "fact"),
                content=content,
                importance=min(3, max(1, int(item.get("importance") or 1))),
                source="agent",
            )
        )


def upsert_owner_preference(session: Session, content: str, importance: int = 3) -> AgentMemory:
    normalized = re.sub(r"\s+", " ", content or "").strip()[:500]
    existing = session.scalar(
        select(AgentMemory)
        .where(
            AgentMemory.scope_type == "owner",
            AgentMemory.scope_id == "owner",
            AgentMemory.kind == "preference",
            func.lower(AgentMemory.content) == normalized.lower(),
            AgentMemory.status == "active",
        )
        .limit(1)
    )
    if existing:
        existing.importance = max(int(existing.importance or 1), importance)
        existing.updated_at = utc_now()
        return existing
    memory = AgentMemory(
        scope_type="owner",
        scope_id="owner",
        kind="preference",
        content=normalized,
        importance=importance,
        source="owner_explicit",
    )
    session.add(memory)
    session.flush()
    return memory


def capture_owner_style_preferences(session: Session, text: str) -> list[str]:
    lowered = (text or "").lower()
    captured: list[str] = []
    if re.search(r"(не\s+(?:надо|нужно|используй|пиши|обращайся)|без).{0,50}(им[её]н|фамил)", lowered):
        captured.append("В сообщениях жильцам не обращаться по имени или фамилии; начинать нейтрально: «Здравствуйте.»")
    if "фамильяр" in lowered or re.search(r"не\s+надо.{0,40}(сюсюк|на\s+ты|по-дружеск)", lowered):
        captured.append("Избегать фамильярного тона; писать жильцам нейтрально, уважительно и по-деловому.")
    for preference in captured:
        upsert_owner_preference(session, preference)
    return captured


def backfill_owner_style_preferences(session: Session, limit: int = 50) -> int:
    messages = session.scalars(
        select(AiMessage)
        .join(AiConversation, AiMessage.conversation_id == AiConversation.id)
        .where(AiConversation.role == "owner", AiMessage.role == "user")
        .order_by(AiMessage.created_at.desc(), AiMessage.id.desc())
        .limit(limit)
    ).all()
    captured: set[str] = set()
    for message in messages:
        captured.update(capture_owner_style_preferences(session, message.text or ""))
    return len(captured)


def owner_request_is_debt_details(text: str) -> bool:
    lowered = (text or "").lower()
    debt_topic = bool(re.search(r"\b(долг|должен|должны|должник|задолж|просроч|коммунал)", lowered))
    detail_request = bool(
        re.search(r"\b(кто|сколько|какие|покажи|подроб|список|распиши|сводк|состояни|как\s+там|что\s+(?:там\s+)?по)", lowered)
    )
    return debt_topic and detail_request


def owner_request_allows_actions(text: str) -> bool:
    lowered = (text or "").lower()
    explicit_action = re.search(
        r"\b(отправ|уведом|напиш|создай|оформ|выдай|дай|предостав|отсроч|установ|перенес|"
        r"добав|начисл|подготов|сделай|выполни|запусти|предложи)\w*",
        lowered,
    )
    recommendation = re.search(r"\b(что\s+делать|как\s+решить|предложи\s+решени|кого\s+пнуть)\b", lowered)
    return bool(explicit_action or recommendation)


def owner_request_explicitly_requests_action(text: str) -> bool:
    lowered = (text or "").lower()
    return bool(
        re.search(
            r"\b(отправ|уведом|напиш|создай|оформ|выдай|дай|предостав|отсроч|установ|перенес|добав|начисл|подготов|сделай|"
            r"выполни|запусти|удали|измени|обнови|зачти|прими|отметь|сохрани|рассчитай|"
            r"выстав|компенсир|засели|переведи|скрой|игнорир)\w*",
            lowered,
        )
    )


def owner_action_unavailable_message(user_text: str, proposal_errors: list[str]) -> str:
    lowered = (user_text or "").lower()
    lines = [
        "Кнопки подтверждения не созданы: не удалось собрать безопасное действие с проверяемыми параметрами.",
        "Ничего не изменено и никому ничего не отправлено.",
    ]

    cleaned_errors: list[str] = []
    for error in proposal_errors:
        value = clean_ai_response(str(error), max_chars=240).replace("\n", " ").strip()
        if value and value not in cleaned_errors:
            cleaned_errors.append(value)
    if cleaned_errors:
        lines.append("Почему: " + "; ".join(cleaned_errors[:3]) + ".")

    suggestions: list[str] = []
    if re.search(r"\b(долг|долж|задолж|начисл|ручн)", lowered):
        suggestions.append(
            "Создать ручной долг: нужен конкретный договор/квартира/жилец, сумма, срок оплаты и назначение "
            "(аренда, коммуналка, на ИП, перевод, прочее)."
        )
    if re.search(r"\b(отсроч|перенес|срок|дат)", lowered):
        suggestions.append("Дать отсрочку: нужен конкретный долг/начисление или платёжная ситуация и новая дата.")
    if re.search(r"\b(отправ|уведом|напиш|сообщ)", lowered):
        suggestions.append(
            "Подготовить сообщение жильцу: отправка возможна только если Telegram-чат жильца привязан; "
            "если нет — дам текст для ручной отправки."
        )
    if re.search(r"\b(напомин|пнут|пни|запусти)", lowered):
        suggestions.append("Запустить напоминания: могу подготовить run_reminders на подтверждение кнопками.")
    if re.search(r"\b(чек|плат[её]ж|оплат|зач[её]т|разнес)", lowered):
        suggestions.append("Разнести платёж или чек: нужен чек/id платежа, сумма, назначение и куда зачесть.")
    if not suggestions:
        suggestions.append(
            "Могу подготовить только операции, которые есть в owner_operations backend. "
            "Если операции нет, предложу ближайший безопасный ручной путь."
        )

    lines.append("")
    lines.append("Что могу сделать из доступного:")
    lines.extend(f"- {item}" for item in suggestions[:4])
    lines.append("")
    lines.append("Напиши недостающие данные одной строкой — подготовлю предложение с кнопками подтверждения.")
    return "\n".join(lines)


def owner_text_confirmation_decision(text: str) -> str:
    normalized = re.sub(r"[.!?,\s]+", " ", (text or "").strip().lower()).strip()
    if normalized in {"да", "ок", "окей", "подтверждаю", "подтвердить", "согласен", "выполняй", "делай"}:
        return "confirm"
    if normalized in {"нет", "отмена", "отклоняю", "отклонить", "не надо", "не делай"}:
        return "reject"
    return ""


def tenant_message_without_name(lease: Lease, text: str) -> str:
    value = (text or "").strip()
    if not value:
        return value
    full_name = re.sub(r"\s+", " ", lease.tenant.full_name or "").strip()
    variants = [full_name]
    if full_name:
        variants.extend(part for part in full_name.split() if len(part) > 1)
    variants = sorted({item for item in variants if item}, key=len, reverse=True)
    for name in variants:
        escaped = re.escape(name)
        patterns = [
            rf"^\s*уважаем(?:ый|ая)\s+{escaped}\s*[,!.:-]*\s*",
            rf"^\s*здравствуйте\s*,?\s*{escaped}\s*[,!.:-]*\s*",
            rf"^\s*добрый\s+(?:день|вечер)\s*,?\s*{escaped}\s*[,!.:-]*\s*",
            rf"^\s*{escaped}\s*[,!.:-]+\s*",
        ]
        for pattern in patterns:
            if re.search(pattern, value, flags=re.IGNORECASE):
                value = re.sub(pattern, "", value, count=1, flags=re.IGNORECASE).lstrip()
                return "Здравствуйте." + (f" {value[:1].upper()}{value[1:]}" if value else "")
    return value


def increment_ai_usage_row(usage: AiUsageDaily, result: DeepSeekResult, cost_rub: float) -> None:
    prompt_tokens = int(result.prompt_tokens or 0)
    completion_tokens = int(result.completion_tokens or 0)
    usage.prompt_tokens = int(usage.prompt_tokens or 0) + prompt_tokens
    usage.completion_tokens = int(usage.completion_tokens or 0) + completion_tokens
    usage.total_tokens = int(usage.total_tokens or 0) + prompt_tokens + completion_tokens
    usage.cost_rub = money(float(usage.cost_rub or 0) + cost_rub)
    usage.calls = int(usage.calls or 0) + 1


def add_ai_usage(session: Session, result: DeepSeekResult, cost_rub: float, today: date | None = None) -> None:
    usage_date = today or date.today()
    provider = (result.provider or "deepseek").strip()
    usage = session.scalar(
        select(AiUsageDaily).where(
            AiUsageDaily.usage_date == usage_date,
            AiUsageDaily.provider == provider,
            AiUsageDaily.model == result.model,
        )
    )
    if not usage:
        usage = AiUsageDaily(usage_date=usage_date, provider=provider, model=result.model)
        session.add(usage)
    increment_ai_usage_row(usage, result, cost_rub)


def deepseek_client_for_settings(session: Session) -> DeepSeekClient:
    runtime = build_provider_runtime(
        AiProvider.DEEPSEEK,
        requested_model=get_setting_value(session, "deepseek_model"),
        deepseek_api_key=get_setting_value(session, "deepseek_api_key"),
    )
    return cast(DeepSeekClient, runtime.client)


def invoke_ai_completion(
    session: Session,
    *,
    conversation: AiConversation,
    actor_role: str,
    lease: Lease | None,
    messages: list[dict[str, str]],
    model: str,
    max_tokens: int,
    temperature: float = 0.2,
    feature: str = "owner_chat",
    hermes_run: HermesAgentRun | None = None,
) -> DeepSeekResult | None:
    max_tokens = min(max_tokens, ai_max_output_tokens(session))
    if hermes_run is None:
        raw_size = sum(len(str(item.get("content") or "")) for item in messages)
        manifest = ContextManifest(
            feature=feature,
            model=model,
            selected_case_ids=[],
            included_entities={},
            included_messages_count=max(0, len(messages) - 2),
            approximate_input_size=max(1, raw_size // 4),
            reason_for_llm_usage=f"legacy {actor_role} completion routed through Hermes audit",
            excluded_context=["database access", "SQL execution"],
            output_limit=FEATURE_OUTPUT_LIMITS.get(feature, 800),
            selected_tools=[],
            state_hash=hashlib.sha256("\n".join(str(item.get("content") or "") for item in messages).encode("utf-8")).hexdigest(),
        )
        hermes_run = create_agent_run(session, manifest=manifest, trigger=actor_role)
    result = None
    try:
        configured_providers = provider_chain()
    except AiProviderConfigError as exc:
        runtime_log("AI", f"provider configuration failed error={exc}")
        session.add(
            AiActionLog(
                conversation_id=conversation.id,
                lease_id=lease.id if lease else None,
                actor_role=actor_role,
                action_type="agent_call",
                status="failed",
                note=str(exc)[:2000],
            )
        )
        complete_agent_run(hermes_run, error=str(exc))
        return None
    for provider in configured_providers:
        provider_started = time_module.perf_counter()
        runtime_model = model
        try:
            runtime = build_provider_runtime(
                provider,
                requested_model=model,
                deepseek_api_key=get_setting_value(session, "deepseek_api_key"),
            )
            runtime_model = runtime.model
            runtime_log(
                "AI",
                f"call provider={provider.value} role={actor_role} model={runtime.model}",
            )
            result = runtime.client.chat_completions(
                model=runtime.model,
                messages=messages,
                temperature=temperature,
                max_tokens=max_tokens,
                session_id=f"rental-manager-{actor_role}-{conversation.id}",
            )
            duration_ms = int((time_module.perf_counter() - provider_started) * 1000)
            runtime_log(
                "AI",
                (
                    f"call ok provider={provider.value} role={actor_role} model={runtime.model} "
                    f"duration_ms={duration_ms}"
                ),
            )
            record_background_event(
                "ai_call",
                duration_ms,
                status="ok",
                detail={"provider": provider.value, "role": actor_role, "model": runtime.model},
            )
            break
        except (AiProviderConfigError, DeepSeekClientError) as exc:
            duration_ms = int((time_module.perf_counter() - provider_started) * 1000)
            runtime_log(
                "AI",
                (
                    f"call failed provider={provider.value} role={actor_role} model={runtime_model} "
                    f"duration_ms={duration_ms} error={exc}"
                ),
            )
            record_background_event(
                "ai_call",
                duration_ms,
                status="failed",
                detail={"provider": provider.value, "role": actor_role, "model": runtime_model, "error": str(exc)[:240]},
            )
            session.add(
                AiActionLog(
                    conversation_id=conversation.id,
                    lease_id=lease.id if lease else None,
                    actor_role=actor_role,
                    action_type="agent_call",
                    status="failed",
                    payload_json=json.dumps(
                        {"provider": provider.value, "requested_model": model},
                        ensure_ascii=False,
                    ),
                    note=str(exc)[:2000],
                )
            )
    if result is None:
        complete_agent_run(hermes_run, error="AI provider unavailable")
        return None

    prompt_tokens = result.prompt_tokens or estimate_tokens("\n".join(item["content"] for item in messages))
    completion_tokens = result.completion_tokens or estimate_tokens(result.content)
    normalized = DeepSeekResult(
        content=result.content,
        model=result.model,
        prompt_tokens=prompt_tokens,
        completion_tokens=completion_tokens,
        raw=result.raw,
        provider=result.provider,
    )
    cost = ai_estimated_cost_rub(
        normalized.model,
        normalized.prompt_tokens,
        normalized.completion_tokens,
        ai_usd_rub_rate(session),
    )
    add_ai_usage(session, normalized, cost)
    record_feature_usage(
        session,
        feature=feature,
        provider=normalized.provider,
        model=normalized.model,
        prompt_tokens=normalized.prompt_tokens,
        completion_tokens=normalized.completion_tokens,
        cost_rub=cost,
    )
    complete_agent_run(
        hermes_run,
        input_tokens=normalized.prompt_tokens,
        output_tokens=normalized.completion_tokens,
        estimated_cost=cost,
    )
    return normalized


def append_ai_instructions(base_prompt: str, instructions: str, audience: str) -> str:
    custom = str(instructions or "").strip()
    if not custom:
        return base_prompt
    return (
        f"{base_prompt.rstrip()}\n\n"
        f"Дополнительные инструкции владельца для {audience}:\n{custom}\n\n"
        "Эти инструкции не отменяют формат ответа, ограничения безопасности, подтверждение действий и backend-валидацию."
    )


def configured_owner_agent_system_prompt(session: Session) -> str:
    configured_path = os.environ.get("AI_SYSTEM_PROMPT_PATH", "").strip()
    if not configured_path:
        prompt = OWNER_AGENT_SYSTEM_PROMPT
    else:
        prompt_path = Path(configured_path)
        if not prompt_path.is_absolute():
            prompt_path = ROOT_DIR / prompt_path
        try:
            prompt = prompt_path.read_text(encoding="utf-8").strip()
        except OSError as exc:
            runtime_log("AI", f"system prompt path unavailable path={prompt_path} error={exc}")
            prompt = OWNER_AGENT_SYSTEM_PROMPT
        if not prompt:
            runtime_log("AI", f"system prompt path is empty path={prompt_path}")
            prompt = OWNER_AGENT_SYSTEM_PROMPT
    return append_ai_instructions(
        prompt[:20_000],
        get_setting_value(session, "ai_owner_instructions"),
        "диалогов владельца",
    )


def configured_tenant_agent_system_prompt(session: Session) -> str:
    return append_ai_instructions(
        TENANT_AGENT_SYSTEM_PROMPT,
        get_setting_value(session, "ai_tenant_instructions"),
        "диалогов с арендаторами",
    )


def configured_audit_system_prompt(session: Session) -> str:
    base_prompt = configured_owner_agent_system_prompt(session) + "\n\n" + SUPERVISOR_SYSTEM_PROMPT
    return append_ai_instructions(
        base_prompt,
        get_setting_value(session, "ai_audit_instructions"),
        "операционного аудита",
    )


def call_deepseek_ai(
    session: Session,
    *,
    chat_id: int | str,
    actor_role: str,
    lease: Lease | None,
    system_prompt: str,
    context: str,
    user_text: str,
    model: str,
    max_tokens: int = 700,
    feature: str = "owner_chat",
    manifest: ContextManifest | None = None,
) -> str:
    if not ai_enabled(session):
        return AI_DISABLED_TEXT
    if ai_budget_exceeded(session):
        return AI_BUDGET_EXCEEDED_TEXT
    if ai_daily_call_limit_exceeded(session):
        return AI_DAILY_LIMIT_EXCEEDED_TEXT
    if ai_feature_daily_limit_exceeded(session, feature):
        return AI_DAILY_LIMIT_EXCEEDED_TEXT

    conversation = ai_conversation(session, chat_id, actor_role, lease)
    history = recent_ai_history(session, conversation)
    log_ai_message(session, conversation, "user", user_text)
    messages = [
        {"role": "system", "content": system_prompt.strip()},
        {"role": "system", "content": context.strip()},
        *history,
        {"role": "user", "content": user_text.strip()},
    ]
    hermes_run = create_agent_run(session, manifest=manifest, trigger=actor_role) if manifest else None
    normalized = invoke_ai_completion(
        session,
        conversation=conversation,
        actor_role=actor_role,
        lease=lease,
        messages=messages,
        model=model,
        max_tokens=max_tokens,
        feature=feature,
        hermes_run=hermes_run,
    )
    if normalized is None:
        return AI_UNAVAILABLE_TEXT

    answer = clean_ai_response(normalized.content) or AI_UNAVAILABLE_TEXT
    cost = ai_estimated_cost_rub(
        normalized.model,
        normalized.prompt_tokens,
        normalized.completion_tokens,
        ai_usd_rub_rate(session),
    )
    log_ai_message(
        session,
        conversation,
        "assistant",
        answer,
        model=normalized.model,
        prompt_tokens=normalized.prompt_tokens,
        completion_tokens=normalized.completion_tokens,
        cost_rub=cost,
    )
    return answer


def call_agent_envelope(
    session: Session,
    *,
    chat_id: int | str,
    actor_role: str,
    lease: Lease | None,
    system_prompt: str,
    context: str,
    user_text: str,
    model: str,
    max_tokens: int = 1000,
    feature: str = "owner_chat",
    manifest: ContextManifest | None = None,
) -> tuple[AiConversation, AgentEnvelope]:
    conversation = ai_conversation(session, chat_id, actor_role, lease)
    if not ai_enabled(session):
        return conversation, AgentEnvelope(reply=AI_DISABLED_TEXT)
    if ai_budget_exceeded(session):
        return conversation, AgentEnvelope(reply=AI_BUDGET_EXCEEDED_TEXT)
    if ai_daily_call_limit_exceeded(session):
        return conversation, AgentEnvelope(reply=AI_DAILY_LIMIT_EXCEEDED_TEXT)
    if ai_feature_daily_limit_exceeded(session, feature):
        return conversation, AgentEnvelope(reply=AI_DAILY_LIMIT_EXCEEDED_TEXT)

    history = recent_ai_history(session, conversation)
    log_ai_message(session, conversation, "user", user_text)
    messages = [
        {"role": "system", "content": system_prompt.strip()},
        {"role": "system", "content": context.strip()},
        *history,
        {"role": "user", "content": user_text.strip()},
    ]
    hermes_run = create_agent_run(session, manifest=manifest, trigger=actor_role) if manifest else None
    normalized = invoke_ai_completion(
        session,
        conversation=conversation,
        actor_role=actor_role,
        lease=lease,
        messages=messages,
        model=model,
        max_tokens=max_tokens,
        temperature=0.15,
        feature=feature,
        hermes_run=hermes_run,
    )
    if normalized is None:
        return conversation, AgentEnvelope(reply=AI_UNAVAILABLE_TEXT)
    envelope = parse_agent_envelope(normalized.content)
    reply_limit = manifest.output_limit if manifest else FEATURE_OUTPUT_LIMITS.get(feature, 800)
    answer = clean_ai_response(envelope.reply, max_chars=reply_limit) or AI_UNAVAILABLE_TEXT
    envelope = AgentEnvelope(
        reply=answer,
        requested_reads=envelope.requested_reads,
        proposed_actions=envelope.proposed_actions,
        memory_changes=envelope.memory_changes,
        commitment_changes=envelope.commitment_changes,
        skill_proposals=envelope.skill_proposals,
        confidence=envelope.confidence,
        need_disambiguation=envelope.need_disambiguation,
        actions=envelope.actions,
        memories=envelope.memories,
        intent=envelope.intent,
        promise_date=envelope.promise_date,
        needs_owner=envelope.needs_owner,
        owner_summary=envelope.owner_summary,
    )
    cost = ai_estimated_cost_rub(
        normalized.model,
        normalized.prompt_tokens,
        normalized.completion_tokens,
        ai_usd_rub_rate(session),
    )
    log_ai_message(
        session,
        conversation,
        "assistant",
        answer,
        model=normalized.model,
        prompt_tokens=normalized.prompt_tokens,
        completion_tokens=normalized.completion_tokens,
        cost_rub=cost,
    )
    if hermes_run:
        complete_agent_run(
            hermes_run,
            input_tokens=normalized.prompt_tokens,
            output_tokens=normalized.completion_tokens,
            estimated_cost=cost,
            envelope=envelope.model_dump(),
        )
    return conversation, envelope


def agent_tenant_state(session: Session, lease: Lease) -> AgentTenantState:
    state = session.scalar(select(AgentTenantState).where(AgentTenantState.lease_id == lease.id).limit(1))
    if state:
        return state
    state = AgentTenantState(lease_id=lease.id)
    session.add(state)
    session.flush()
    return state


def tenant_state_context(state: AgentTenantState) -> str:
    return "\n".join(
        [
            "Состояние диалога:",
            f"- уровень строгости: {int(state.escalation_level or 0)}",
            f"- последовательных отписок/сорванных обещаний: {int(state.consecutive_excuses or 0)}",
            f"- обещанная дата оплаты: {state.promise_date.isoformat() if state.promise_date else 'нет'}",
            f"- формулировка обещания: {state.promise_text or 'нет'}",
            f"- следующая проверка: {state.next_contact_on.isoformat() if state.next_contact_on else 'не назначена'}",
        ]
    )


def parse_agent_date(value: Any, field_label: str) -> date:
    raw = str(value or "").strip()
    if not raw:
        raise HTTPException(400, f"Не указана дата: {field_label}")
    try:
        return date.fromisoformat(raw)
    except ValueError as exc:
        raise HTTPException(400, f"Дата «{field_label}» должна быть в формате ГГГГ-ММ-ДД") from exc


def require_agent_lease(session: Session, payload: dict[str, Any]) -> Lease:
    try:
        lease_id = int(payload.get("lease_id") or 0)
    except (TypeError, ValueError) as exc:
        raise HTTPException(400, "Агент не указал корректный lease_id") from exc
    lease = session.get(Lease, lease_id)
    if not lease:
        raise HTTPException(404, "Договор из предложения агента не найден")
    return lease


def owner_operation_target(
    session: Session,
    arguments: dict[str, Any],
) -> tuple[Lease | None, str]:
    entity_fields: list[tuple[str, type[Any], str]] = [
        ("lease_id", Lease, "договор"),
        ("charge_id", RentCharge, "начисление"),
        ("apartment_id", Apartment, "квартира"),
        ("object_id", RentalObject, "объект"),
        ("service_id", UtilityService, "коммунальная услуга"),
        ("meter_id", Meter, "счётчик"),
        ("bill_id", UtilityBill, "коммунальный счёт"),
        ("line_id", UtilityBillLine, "строка коммунального счёта"),
        ("debt_id", ManualDebt, "ручной долг"),
        ("receipt_id", PaymentReceipt, "платёж"),
        ("expense_id", Expense, "расход"),
        ("situation_id", PaymentSituation, "платёжная ситуация"),
    ]
    resolved: dict[str, Any] = {}
    for field, model, label in entity_fields:
        raw_id = arguments.get(field)
        if raw_id in {None, ""}:
            continue
        try:
            entity_id = int(raw_id)
        except (TypeError, ValueError) as exc:
            raise HTTPException(400, f"Некорректный идентификатор: {field}") from exc
        entity = session.get(model, entity_id)
        if entity is None:
            raise HTTPException(404, f"Не найден {label} с id={entity_id}")
        arguments[field] = entity_id
        resolved[field] = entity

    lease = resolved.get("lease_id")
    charge = resolved.get("charge_id")
    line = resolved.get("line_id")
    debt = resolved.get("debt_id")
    receipt = resolved.get("receipt_id")
    if lease is None and charge is not None:
        lease = charge.lease
    if lease is None and line is not None and line.lease_id:
        lease = session.get(Lease, line.lease_id)
    if lease is None and debt is not None:
        lease = debt.lease
    if lease is None and receipt is not None and receipt.lease_id:
        lease = session.get(Lease, receipt.lease_id)

    if lease is not None:
        return lease, f"{lease.apartment.object.name}, {lease.apartment.name}, {lease.tenant.full_name} (lease_id={lease.id})"
    apartment = resolved.get("apartment_id")
    if apartment is not None:
        return None, f"{apartment.object.name}, {apartment.name} (apartment_id={apartment.id})"
    rental_object = resolved.get("object_id")
    if rental_object is not None:
        return None, f"{rental_object.name} (object_id={rental_object.id})"
    service = resolved.get("service_id")
    if service is not None:
        return None, f"{service.object.name}, {service.name} (service_id={service.id})"
    meter = resolved.get("meter_id")
    if meter is not None:
        return None, f"{meter.service.object.name}, {meter.service.name}, {meter.name} (meter_id={meter.id})"
    bill = resolved.get("bill_id")
    if bill is not None:
        return None, (
            f"{bill.service.object.name}, {bill.service.name}, "
            f"{bill.period_start:%d.%m.%Y}–{bill.period_end:%d.%m.%Y} (bill_id={bill.id})"
        )
    expense = resolved.get("expense_id")
    if expense is not None:
        return None, f"расход {expense.category} на {money_text(expense.amount)} (expense_id={expense.id})"
    payment_state = resolved.get("situation_id")
    if payment_state is not None:
        target = payment_situation_target(session, payment_state)
        lease = situation_lease(target)
        if lease:
            return lease, (
                f"{lease.apartment.object.name}, {lease.apartment.name}, {lease.tenant.full_name}; "
                f"{payment_state.kind}, situation_id={payment_state.id}"
            )
    return None, ""


def tenant_message_unverified_change(text: str) -> str:
    lowered = re.sub(r"\s+", " ", (text or "").strip().lower())
    if not lowered:
        return ""

    if re.search(r"\bотсроч", lowered):
        safe_deferral_context = (
            re.search(r"\b(?:запрос|просьб|вопрос)\w*.{0,60}\bотсроч", lowered)
            or re.search(r"\b(?:рассмотр|обсуд|уточн|возможност)\w*.{0,60}\bотсроч", lowered)
            or re.search(r"\bотсроч\w*.{0,40}\bне\s+(?:предоставлен|подтвержден|одобрен|оформлен)", lowered)
        )
        if not safe_deferral_context:
            return "отсрочка"

    mutation_patterns = (
        ("списание или закрытие долга", r"\b(?:долг|задолженн)\w*.{0,50}\b(?:списан|удал[её]н|закрыт|аннулирован)"),
        ("зачёт платежа", r"\b(?:плат[её]ж|оплат)\w*.{0,50}\b(?:зачт[её]н|подтвержд[её]н|провед[её]н)"),
        ("изменение договора", r"\bдоговор\w*.{0,50}\b(?:измен[её]н|продл[её]н|расторгнут|закрыт)"),
        ("оформление выезда", r"\bвыезд\w*.{0,50}\b(?:оформлен|подтвержд[её]н|зафиксирован)"),
        ("возврат залога", r"\bзалог\w*.{0,50}\b(?:возвращ[её]н|перечислен|выплачен)"),
        ("предоставление скидки", r"\bскидк\w*.{0,50}\b(?:предоставлен|одобрен|примен[её]н)"),
    )
    for label, pattern in mutation_patterns:
        if re.search(pattern, lowered):
            return label
    return ""


def reject_tenant_message_claiming_unverified_change(text: str) -> None:
    change = tenant_message_unverified_change(text)
    if not change:
        return
    raise HTTPException(
        400,
        f"Сообщение заявляет, что уже выполнено изменение «{change}», но отдельная отправка сообщения не меняет данные. "
        "Сначала подготовьте соответствующую операцию с данными; backend уведомит жильца только после её подтверждения и успешного сохранения.",
    )


def normalize_agent_action(
    session: Session,
    action: dict[str, Any],
) -> tuple[Lease | None, str, dict[str, Any], str]:
    action_type = str(action.get("type") or "").strip()
    payload = dict(action.get("payload") or {})
    if action_type == "owner_operation":
        operation = str(payload.get("operation") or "").strip()
        raw_arguments = payload.get("arguments")
        if not isinstance(raw_arguments, dict):
            raise HTTPException(400, "Для owner_operation нужен объект arguments")
        try:
            payload = validate_owner_operation(operation, raw_arguments)
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc
        action_type = operation
    elif action_type in {"defer_rent", "move_out", "create_manual_debt", "send_tenant_message"}:
        try:
            payload = validate_owner_operation(action_type, payload)
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc
    else:
        raise HTTPException(400, "Агент предложил неподдерживаемое действие")

    if action_type not in OWNER_OPERATION_SPECS:
        raise HTTPException(400, "Агент предложил неподдерживаемую операцию")

    if action_type not in {"defer_rent", "move_out", "create_manual_debt", "send_tenant_message"}:
        lease, target = owner_operation_target(session, payload)
        if action_type == "update_lease":
            if lease is None:
                raise HTTPException(404, "Договор для изменения не найден")
            payload = {
                "lease_id": lease.id,
                "apartment_id": payload.get("apartment_id", lease.apartment_id),
                "start_date": payload.get("start_date", lease.start_date.isoformat()),
                "payment_day": payload.get("payment_day", lease.payment_day),
                "ip_amount": payload.get("ip_amount", lease.ip_amount),
                "personal_amount": payload.get("personal_amount", lease.personal_amount),
                "deposit_amount": payload.get("deposit_amount", lease.deposit_amount),
                "deposit_location": payload.get("deposit_location", lease.deposit_location),
                "deposit_terms": payload.get("deposit_terms", lease.deposit_terms),
                "notes": payload.get("notes", lease.notes),
                "full_name": payload.get("full_name", lease.tenant.full_name),
                "phone": payload.get("phone", lease.tenant.phone),
                "telegram": payload.get("telegram", lease.tenant.telegram),
                "whatsapp": payload.get("whatsapp", lease.tenant.whatsapp),
                **({"end_date": payload["end_date"]} if "end_date" in payload else {}),
                **({"tenant_notes": payload["tenant_notes"]} if "tenant_notes" in payload else {}),
                **({"ignored": payload["ignored"]} if "ignored" in payload else {}),
            }
        preview = owner_operation_preview(action_type, payload, target)
        return lease, action_type, payload, preview

    lease = require_agent_lease(session, payload)
    place = f"{lease.apartment.object.name}, {lease.apartment.name}, {lease.tenant.full_name}"

    if action_type == "defer_rent":
        try:
            charge_id = int(payload.get("charge_id") or 0)
        except (TypeError, ValueError) as exc:
            raise HTTPException(400, "Для отсрочки не указан корректный charge_id") from exc
        charge = session.get(RentCharge, charge_id)
        if not charge or charge.lease_id != lease.id:
            raise HTTPException(400, "Начисление для отсрочки не относится к указанному договору")
        update_rent_charge_status(charge)
        debt = money(
            max(0.0, float(charge.ip_due or 0) - float(charge.ip_paid or 0))
            + max(0.0, float(charge.personal_due or 0) - float(charge.personal_paid or 0))
        )
        if debt <= EPS:
            raise HTTPException(400, "По этому начислению уже нет долга")
        until = parse_agent_date(payload.get("deferral_until"), "дата отсрочки")
        if until <= date.today() or until > date.today() + timedelta(days=180):
            raise HTTPException(400, "Отсрочка должна быть в пределах следующих 180 дней")
        notify_tenant = setting_bool_value(payload.get("notify_tenant", True))
        if notify_tenant and not lease_chat_id(session, lease):
            raise HTTPException(
                400,
                f"{lease.tenant.full_name} ещё не выполнил /start, поэтому уведомление об отсрочке не дойдёт",
            )
        normalized = {
            "lease_id": lease.id,
            "charge_id": charge.id,
            "deferral_until": until.isoformat(),
            "note": str(payload.get("note") or action.get("reason") or "").strip()[:1000],
            "notify_tenant": notify_tenant,
            "tenant_message": tenant_message_without_name(
                lease,
                str(payload.get("tenant_message") or "").strip()[:3500],
            ),
        }
        preview = (
            f"Предлагаю выдать отсрочку\n"
            f"Жилец: {place}\n"
            f"Долг: {money_text(debt)}\n"
            f"Новый срок: {until:%d.%m.%Y}\n"
            f"Причина: {normalized['note'] or 'не указана'}"
        )
        return lease, action_type, normalized, preview

    if action_type == "move_out":
        end = parse_agent_date(payload.get("end_date"), "дата выезда")
        if end < date.today() or end < lease.start_date:
            raise HTTPException(400, "Дата выезда не может быть в прошлом или раньше начала договора")
        notify_tenant = setting_bool_value(payload.get("notify_tenant", False))
        if notify_tenant and not lease_chat_id(session, lease):
            raise HTTPException(
                400,
                f"{lease.tenant.full_name} ещё не выполнил /start, поэтому уведомление о выезде не дойдёт",
            )
        normalized = {
            "lease_id": lease.id,
            "end_date": end.isoformat(),
            "notes": str(payload.get("notes") or action.get("reason") or "").strip()[:1500],
            "notify_tenant": notify_tenant,
            "tenant_message": tenant_message_without_name(
                lease,
                str(payload.get("tenant_message") or "").strip()[:3500],
            ),
        }
        preview = (
            f"Предлагаю оформить выезд\n"
            f"Жилец: {place}\n"
            f"Дата выезда: {end:%d.%m.%Y}\n"
            f"Примечание: {normalized['notes'] or 'не указано'}"
        )
        return lease, action_type, normalized, preview

    if action_type == "create_manual_debt":
        try:
            amount = money(float(str(payload.get("amount") or "0").replace(" ", "").replace(",", ".")))
        except ValueError as exc:
            raise HTTPException(400, "Сумма ручного долга должна быть числом") from exc
        if amount <= 0 or amount > 10_000_000:
            raise HTTPException(400, "Сумма ручного долга должна быть больше нуля и меньше 10 млн")
        due = parse_agent_date(payload.get("due_date"), "срок ручного долга")
        title = str(payload.get("title") or "Ручной долг").strip()[:180]
        normalized = {
            "lease_id": lease.id,
            "kind": str(payload.get("kind") or "other"),
            "channel": str(payload.get("channel") or ""),
            "title": title,
            "amount": amount,
            "due_date": due.isoformat(),
            "notes": str(payload.get("note") or action.get("reason") or "").strip()[:1500],
        }
        for field in ("period_start", "period_end"):
            if payload.get(field):
                normalized[field] = parse_agent_date(payload.get(field), field).isoformat()
        preview = (
            f"Предлагаю создать ручной долг\n"
            f"Жилец: {place}\n"
            f"Назначение: {title}\n"
            f"Сумма: {money_text(amount)}\n"
            f"Срок: {due:%d.%m.%Y}"
        )
        return lease, action_type, normalized, preview

    if action_type == "send_tenant_message":
        if not lease_chat_id(session, lease):
            raise HTTPException(
                400,
                f"{lease.tenant.full_name} ещё не выполнил /start, сообщение отправить невозможно",
            )
        message_text = tenant_message_without_name(lease, str(payload.get("text") or "").strip())
        if not message_text:
            raise HTTPException(400, "Агент предложил пустое сообщение")
        if len(message_text) > 3500:
            raise HTTPException(400, "Сообщение жильцу длиннее 3500 символов")
        reject_tenant_message_claiming_unverified_change(message_text)
        normalized = {"lease_id": lease.id, "text": message_text}
        preview = f"Предлагаю отправить сообщение\nЖилец: {place}\n\n{message_text}"
        return lease, action_type, normalized, preview

    raise HTTPException(400, "Агент предложил неподдерживаемое действие")


def create_agent_action_proposal(
    session: Session,
    conversation: AiConversation,
    owner_chat_id: int | str,
    action: dict[str, Any],
) -> AgentActionProposal:
    lease, action_type, payload, preview = normalize_agent_action(session, action)
    payload_json = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    idempotency_key = "agent-proposal:" + hashlib.sha256(
        f"{owner_chat_id}:{action_type}:{payload_json}".encode("utf-8")
    ).hexdigest()
    pending = session.scalars(
        select(AgentActionProposal)
        .where(
            AgentActionProposal.idempotency_key == idempotency_key,
            AgentActionProposal.status == "pending",
        )
        .order_by(AgentActionProposal.id.desc())
        .limit(20)
    ).all()
    for existing in pending:
        try:
            existing_payload = json.dumps(
                json.loads(existing.payload_json or "{}"),
                ensure_ascii=False,
                sort_keys=True,
            )
        except json.JSONDecodeError:
            continue
        if existing_payload == payload_json:
            runtime_log("AGENT", f"duplicate proposal skipped existing_id={existing.id} action={action_type}")
            return existing
    try:
        ttl_hours = min(168, max(1, int(get_setting_value(session, "ai_action_confirmation_ttl_hours") or 48)))
    except ValueError:
        ttl_hours = 48
    decision = ACTION_SAFETY_REGISTRY.classify(
        action_type,
        payload,
        owner_level_one_enabled=setting_bool_value(get_setting_value(session, "hermes_auto_level_one_enabled")),
    )
    related_case = None
    if lease:
        related_case = session.scalar(
            select(OperationalCase)
            .where(
                OperationalCase.contract_id == lease.id,
                OperationalCase.status.in_(OPEN_CASE_STATUSES),
            )
            .order_by(OperationalCase.priority_score.desc(), OperationalCase.id)
            .limit(1)
        )
    proposal = AgentActionProposal(
        conversation_id=conversation.id,
        lease_id=lease.id if lease else None,
        action_type=action_type,
        status="pending",
        payload_json=payload_json,
        preview_text=preview,
        requested_by="agent",
        owner_chat_id=str(owner_chat_id),
        expires_at=utc_now() + timedelta(hours=ttl_hours),
        case_id=related_case.id if related_case else None,
        safety_level=decision.level,
        idempotency_key=idempotency_key,
        validation_hash=hashlib.sha256(payload_json.encode("utf-8")).hexdigest(),
    )
    session.add(proposal)
    session.flush()
    session.add(
        AiActionLog(
            conversation_id=conversation.id,
            lease_id=lease.id if lease else None,
            actor_role="agent",
            action_type=action_type,
            status="pending",
            payload_json=payload_json,
            note=preview[:2000],
        )
    )
    response = send_telegram_text(
        session,
        owner_chat_id,
        preview + "\n\nНикаких изменений пока не внесено.",
        action_confirmation_keyboard(proposal.id),
    )
    proposal.owner_message_id = int(((response.get("result") or {}) if isinstance(response, dict) else {}).get("message_id") or 0) or None
    return proposal


def owner_agent_context(session: Session, chat_id: int | str, dashboard: dict[str, Any]) -> str:
    leases = session.scalars(
        select(Lease)
        .options(
            joinedload(Lease.apartment).joinedload(Apartment.object),
            joinedload(Lease.tenant),
        )
        .where(Lease.active.is_(True))
        .order_by(Lease.apartment_id, Lease.id)
    ).all()
    lease_lines = ["Активные договоры и идентификаторы для возможных действий:"]
    for lease in leases[:50]:
        chat_status = "привязан" if lease_chat_id(session, lease) else "не привязан, сначала нужен /start"
        lease_lines.append(
            f"- lease_id={lease.id}; {lease.apartment.object.name}; {lease.apartment.name}; "
            f"{lease.tenant.full_name}; дата начала={lease.start_date.isoformat()}; "
            f"залог={money_text(lease.deposit_amount)}; Telegram-чат: {chat_status}"
        )
        charges = session.scalars(
            select(RentCharge)
            .where(RentCharge.lease_id == lease.id)
            .order_by(RentCharge.due_date.desc(), RentCharge.id.desc())
            .limit(4)
        ).all()
        for charge in reversed(charges):
            update_rent_charge_status(charge)
            debt = money(
                max(0.0, float(charge.ip_due or 0) - float(charge.ip_paid or 0))
                + max(0.0, float(charge.personal_due or 0) - float(charge.personal_paid or 0))
            )
            if debt > EPS or charge.due_date >= date.today():
                lease_lines.append(
                    f"  charge_id={charge.id}; срок={charge.due_date.isoformat()}; "
                    f"статус={charge.status}; долг={debt}; "
                    f"отсрочка={charge.deferral_until.isoformat() if charge.deferral_until else 'нет'}"
                )

    tasks = session.scalars(
        select(AgentTask)
        .where(AgentTask.status == "open")
        .order_by(AgentTask.severity.desc(), AgentTask.first_seen_at, AgentTask.id)
        .limit(20)
    ).all()
    task_text = "Открытые задачи супервизора:\n" + (
        "\n".join(f"- [{task.severity}] {task.title}: {task.details}" for task in tasks)
        if tasks
        else "- нет"
    )
    pending_count = int(
        session.scalar(select(func.count(AgentActionProposal.id)).where(AgentActionProposal.status == "pending")) or 0
    )
    return "\n\n".join(
        [
            owner_context_text(session, dashboard),
            build_owner_read_tools_context(session, dashboard, chat_id),
            "\n".join(lease_lines),
            task_text,
            f"Предложений действий, ожидающих подтверждения: {pending_count}",
            agent_memory_context(session, scope_type="owner", scope_id="owner"),
        ]
    )


def owner_debt_details_text(session: Session, dashboard: dict[str, Any], query: str = "") -> str:
    lowered = (query or "").lower()
    utility_only = "коммунал" in lowered and not re.search(r"\bаренд", lowered)
    rent_only = bool(re.search(r"\bаренд", lowered)) and "коммунал" not in lowered
    sections: list[str] = []
    grand_total = 0.0
    linked_tenant_ids = {
        int(raw_tenant_id)
        for raw_tenant_id, chat_id in get_tenant_links(session).items()
        if str(raw_tenant_id).isdigit() and str(chat_id).strip()
    }
    linked_lease_ids = {
        lease.id
        for lease in session.scalars(
            select(Lease).where(Lease.active.is_(True), Lease.tenant_id.in_(linked_tenant_ids))
        ).all()
    }

    def unique_items(keys: list[str]) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        seen: set[tuple[str, int]] = set()
        for key in keys:
            for item in dashboard.get(key, []):
                item_id = int(item.get("id") or 0)
                identity = (key.split("_", 1)[0], item_id)
                if item_id and identity in seen:
                    continue
                if item_id:
                    seen.add(identity)
                result.append(item)
        return result

    def add_section(title: str, items: list[dict[str, Any]]) -> None:
        nonlocal grand_total
        lines: list[str] = []
        subtotal = 0.0
        for item in items:
            debt = money(float(item.get("debt") or item.get("amount") or 0))
            if debt <= EPS:
                continue
            subtotal += debt
            place = ", ".join(
                part for part in [item.get("object"), item.get("apartment")] if str(part or "").strip()
            )
            tenant = str(item.get("tenant") or "").strip()
            due = str(item.get("due_date") or "").strip()
            due_text = f", срок {due}" if due else ""
            lease_id = int(item.get("lease_id") or 0)
            delivery = (
                "Telegram привязан"
                if lease_id and lease_id in linked_lease_ids
                else "Telegram не привязан: сначала нужен /start"
            )
            lines.append(
                f"• {place or 'объект не указан'} — {tenant or 'жилец не указан'}: "
                f"{money_text(debt)}{due_text}; {delivery}"
            )
        if not lines:
            return
        grand_total += subtotal
        sections.append(f"{title} — {money_text(subtotal)}:\n" + "\n".join(lines))

    if not utility_only:
        add_section(
            "Аренда",
            unique_items(["rent_overdue", "rent_partial", "rent_today", "rent_deferred"]),
        )
    if not rent_only:
        add_section("Коммунальные платежи", unique_items(["utility_issued"]))
    if not utility_only and not rent_only:
        add_section("Прочие начисления", unique_items(["manual_debts"]))

    if not sections:
        topic = "по коммунальным платежам" if utility_only else "по аренде" if rent_only else ""
        return f"Открытых долгов {topic} сейчас не вижу.".replace("  ", " ")
    return "\n\n".join(["Сейчас в системе:", *sections, f"Итого: {money_text(grand_total)}."])


def update_tenant_state_from_envelope(
    state: AgentTenantState,
    envelope: AgentEnvelope,
    user_text: str,
    today: date | None = None,
) -> None:
    today = today or date.today()
    state.last_tenant_message_at = utc_now()
    if envelope.intent == "payment_promise":
        promised = None
        try:
            promised = date.fromisoformat(envelope.promise_date) if envelope.promise_date else None
        except ValueError:
            promised = None
        if promised and today <= promised <= today + timedelta(days=60):
            state.promise_date = promised
            state.promise_text = user_text[:1000]
            state.next_contact_on = promised + timedelta(days=1)
            state.status = "promised"
        else:
            state.status = "awaiting_date"
    elif envelope.intent == "payment_sent":
        state.status = "awaiting_receipt"
        state.consecutive_excuses = 0
        state.next_contact_on = today + timedelta(days=1)
    elif envelope.intent in {"excuse", "deferral_request"}:
        state.consecutive_excuses = int(state.consecutive_excuses or 0) + 1
        state.escalation_level = min(3, max(int(state.escalation_level or 0), state.consecutive_excuses - 1))
        state.status = "needs_attention"
        state.next_contact_on = today + timedelta(days=1)


def notify_owner_about_tenant_dialogue(
    session: Session,
    lease: Lease,
    text: str,
    summary: str,
) -> None:
    owner_id = telegram_owner_chat_id(session)
    if not owner_id or not telegram_token(session):
        return
    send_telegram_text(
        session,
        owner_id,
        (
            f"Жилец запросил решение владельца\n"
            f"{lease.apartment.object.name}, {lease.apartment.name}, {lease.tenant.full_name}\n"
            f"Сообщение: {text}\n"
            f"Суть: {summary or 'нужно посмотреть диалог'}\n\n"
            f"Агент ничего в системе не менял."
        ),
        app_keyboard(app_base_url(session)),
    )


def notify_owner_about_tenant_update(
    session: Session,
    lease: Lease,
    envelope: AgentEnvelope,
    state: AgentTenantState,
) -> None:
    if envelope.intent not in {"payment_promise", "payment_sent", "excuse"}:
        return
    owner_id = telegram_owner_chat_id(session)
    if not owner_id or not telegram_token(session):
        return
    if envelope.intent == "payment_promise":
        detail = f"обещал оплатить {state.promise_date:%d.%m.%Y}" if state.promise_date else "обещал оплатить, но точную дату не назвал"
    elif envelope.intent == "payment_sent":
        detail = "сообщил об оплате; агент попросил прислать чек"
    else:
        detail = f"очередная отписка, счётчик: {int(state.consecutive_excuses or 0)}"
    send_telegram_text(
        session,
        owner_id,
        f"Диалог по оплате: {lease.apartment.name}, {lease.tenant.full_name} — {detail}.",
        app_keyboard(app_base_url(session)),
    )
    state.last_owner_update_at = utc_now()


def handle_tenant_ai_message(session: Session, chat_id: int | str, lease: Lease, text: str) -> bool:
    if not text or text.startswith("/") or not ai_enabled(session) or not ai_tenant_free_text_enabled(session):
        return False
    intent = classify_tenant_intent(text)
    settings = get_settings(session)
    policy = tenant_response_policy(intent, settings, now=local_now())
    state = agent_tenant_state(session, lease)
    if intent == "payment_promise":
        update_latest_reminder_outcome(session, lease_id=lease.id, outcome="promised")
    else:
        update_latest_reminder_outcome(session, lease_id=lease.id, outcome="replied")
    if not policy["use_llm"]:
        night = not (8 <= local_now().hour < 21)
        answer = tenant_acknowledgement(intent, night=night)
        send_telegram_text(session, chat_id, answer, tenant_keyboard())
        if policy["escalate"]:
            summary = f"{intent}: {text[:700]}"
            notify_owner_about_tenant_dialogue(session, lease, text, summary)
            situation = active_payment_situation_for_lease(session, lease.id)
            if situation:
                situation.status = "question" if intent != "deferral_request" else "escalated"
                situation.escalated_at = utc_now()
            emit_domain_event(
                session,
                "tenant_message_escalated",
                entity_type="Lease",
                entity_id=lease.id,
                property_id=lease.apartment.object_id,
                apartment_id=lease.apartment_id,
                tenant_id=lease.tenant_id,
                contract_id=lease.id,
                payload={"intent": intent, "delayed": policy["delay_escalation"], "text": text[:700]},
            )
        return True
    model = resolve_ai_model(get_setting_value(session, "deepseek_model"))
    selected = build_hermes_tenant_context(
        session,
        lease=lease,
        user_text=text,
        model=model,
        reason="free tenant message needs scoped natural-language answer",
        output_limit=ai_feature_output_limit(session, "tenant_chat"),
    )
    conversation, envelope = call_agent_envelope(
        session,
        chat_id=chat_id,
        actor_role="tenant",
        lease=lease,
        system_prompt=configured_tenant_agent_system_prompt(session),
        context=selected.text,
        user_text=text,
        model=model,
        max_tokens=500,
        feature="tenant_chat",
        manifest=selected.manifest,
    )
    update_tenant_state_from_envelope(state, envelope, text)
    situation = active_payment_situation_for_lease(session, lease.id)
    if situation:
        situation.last_tenant_response_at = utc_now()
        situation.tenant_response_count = int(situation.tenant_response_count or 0) + 1
        if envelope.intent == "payment_sent":
            situation.status = "awaiting_receipt"
        elif envelope.intent == "payment_promise" and envelope.promise_date:
            promised = parse_tenant_payment_date(envelope.promise_date)
            target = payment_situation_target(session, situation)
            if promised and target:
                reason = f"жилец сообщил: {text[:600]}"
                if short_deferral_is_safe(situation, target, promised, date.today()):
                    apply_short_payment_deferral(session, situation, promised, reason)
                    return True
                propose_long_payment_deferral(session, situation, promised, reason)
                send_telegram_text(
                    session,
                    chat_id,
                    f"Запрос на перенос оплаты до {format_full_date(promised)} передан владельцу.",
                )
                return True
    notify_owner_about_tenant_update(session, lease, envelope, state)
    needs_owner = policy["mode"] == "draft" or envelope.needs_owner or tenant_question_needs_owner(text)
    if needs_owner:
        if situation:
            situation.status = "question"
            situation.escalated_at = utc_now()
            notify_owner_situation_event(session, situation, "question", date.today())
        session.add(
            AiActionLog(
                conversation_id=conversation.id,
                lease_id=lease.id,
                actor_role="tenant",
                action_type="owner_escalation",
                status="pending",
                payload_json=json.dumps({"text": text, "summary": envelope.owner_summary}, ensure_ascii=False),
                note="tenant dialogue requires owner decision",
            )
        )
        notify_owner_about_tenant_dialogue(session, lease, text, envelope.owner_summary)
    answer = envelope.reply[: selected.manifest.output_limit]
    if policy["mode"] == "draft":
        notify_owner_about_tenant_dialogue(session, lease, text, answer or envelope.owner_summary)
        answer = tenant_acknowledgement(intent)
    if needs_owner and answer in {"", AI_UNAVAILABLE_TEXT}:
        answer = TENANT_ESCALATION_TEXT
    send_telegram_text(session, chat_id, answer or AI_UNAVAILABLE_TEXT, tenant_keyboard())
    return True


def hermes_case_choice_keyboard(session: Session, cases: list[OperationalCase], action: str = "details") -> dict[str, Any]:
    return {
        "inline_keyboard": [
            [
                {
                    "text": case_label(session, item)[:50],
                    "callback_data": f"ai:case:{item.id}:{action}",
                }
            ]
            for item in cases[:8]
        ]
    }


def hermes_open_cases_text(session: Session, cases: list[OperationalCase] | None = None, *, limit: int = 8) -> str:
    rows = cases or reconcile_operational_cases(session)
    if not rows:
        return "Активных операционных кейсов сейчас нет."
    total = round(sum(float(item.amount_total or 0) for item in rows), 2)
    rendered = f"{total:,.2f}".replace(",", " ").replace(".00", "")
    lines = [f"Открыто {len(rows)} кейсов · сумма {rendered} ₽"]
    for item in rows[:limit]:
        lines.append(f"• {case_label(session, item)} — {item.compact_summary.rstrip('. ')}.")
    if len(rows) > limit:
        lines.append(f"Ещё {len(rows) - limit} — в Hermes Control Center.")
    return clean_ai_response("\n".join(lines), max_chars=ai_feature_output_limit(session, "owner_chat"))


def handle_owner_ai_message(
    session: Session,
    chat_id: int | str,
    text: str,
    *,
    audit_deep: bool = False,
    audit_requested: bool = False,
) -> bool:
    user_text = text.strip()
    if not user_text:
        return False
    reconcile_operational_cases(session)
    conversation = ai_conversation(session, chat_id, "owner", None)
    intent = classify_owner_intent(user_text)
    audit_mode = audit_requested or audit_deep
    debt_summary_request = not audit_mode and owner_request_is_debt_details(user_text) and not intent.wants_action
    matched_cases = [] if audit_mode or debt_summary_request else resolve_case_alias(session, user_text)
    selected_case = matched_cases[0] if len(matched_cases) == 1 else None
    structured_preference = capture_structured_owner_preference(session, user_text, case=selected_case)
    captured_style_preferences = capture_owner_style_preferences(session, user_text)
    if structured_preference or captured_style_preferences:
        parts = []
        if structured_preference:
            mode_text = "на один раз" if structured_preference.mode == "once" else "до срока" if structured_preference.mode == "until_date" else "постоянно"
            parts.append(f"Правило сохранено {mode_text}: {structured_preference.key}.")
        parts.extend(captured_style_preferences)
        answer = "Принял. " + " ".join(parts)
        log_ai_message(session, conversation, "user", user_text)
        log_ai_message(session, conversation, "assistant", answer)
        send_telegram_text(session, chat_id, answer)
        update_conversation_summary(session, conversation)
        return True
    if is_commitment_phrase(user_text):
        if len(matched_cases) > 1:
            send_telegram_text(
                session,
                chat_id,
                "К какому кейсу относится обязательство?",
                hermes_case_choice_keyboard(session, matched_cases, "owner_commitment"),
            )
            return True
        if not selected_case:
            open_cases = reconcile_operational_cases(session)
            if len(open_cases) == 1:
                selected_case = open_cases[0]
            elif open_cases:
                send_telegram_text(
                    session,
                    chat_id,
                    "Уточните кейс — обещание нельзя привязать наугад.",
                    hermes_case_choice_keyboard(session, open_cases, "owner_commitment"),
                )
                return True
        commitment = create_owner_commitment(
            session,
            case=selected_case,
            text=user_text,
            briefing_time=get_setting_value(session, "ai_supervisor_time") or "10:00",
        )
        label = case_label(session, selected_case) if selected_case else "Обязательство"
        send_telegram_text(session, chat_id, f"{label}: запомнил. Проверю результат {commitment.due_at:%d.%m в %H:%M}.")
        return True
    if intent.wants_skill:
        skill = create_skill_draft(session, user_text)
        send_telegram_text(
            session,
            chat_id,
            skill_card(skill),
            {
                "inline_keyboard": [
                    [
                        {"text": "Активировать", "callback_data": f"ai:skill:{skill.id}:activate"},
                        {"text": "Изменить", "callback_data": f"ai:skill:{skill.id}:edit"},
                    ],
                    [{"text": "Отмена", "callback_data": f"ai:skill:{skill.id}:cancel"}],
                ]
            },
        )
        return True
    grouped = plan_grouped_deferral(session, user_text)
    if grouped.matched:
        if grouped.ambiguous_lease_ids:
            cases = [item for item in reconcile_operational_cases(session) if item.contract_id in grouped.ambiguous_lease_ids]
            send_telegram_text(
                session,
                chat_id,
                "Нашёл несколько похожих целей. Выберите нужную:",
                hermes_case_choice_keyboard(session, cases),
            )
            return True
        if not grouped.items:
            send_telegram_text(session, chat_id, "Подходящих открытых платежей для пакетной отсрочки не найдено.")
            return True
        try:
            ttl = int(get_setting_value(session, "ai_action_confirmation_ttl_hours") or 48)
        except ValueError:
            ttl = 48
        proposal = create_grouped_deferral_proposal(
            session,
            conversation=conversation,
            owner_chat_id=chat_id,
            plan=grouped,
            ttl_hours=ttl,
            mass_action_threshold=int(get_setting_value(session, "hermes_mass_action_pin_threshold") or 20),
        )
        response = send_telegram_text(
            session,
            chat_id,
            proposal.preview_text + "\n\nДо подтверждения данные не изменятся.",
            action_confirmation_keyboard(proposal.id),
        )
        proposal.owner_message_id = int(((response.get("result") or {}) if isinstance(response, dict) else {}).get("message_id") or 0) or None
        return True
    if debt_summary_request:
        send_telegram_text(session, chat_id, owner_debt_details_text(session, build_dashboard(session), user_text))
        return True
    if len(matched_cases) > 1:
        send_telegram_text(
            session,
            chat_id,
            "Нашёл несколько похожих кейсов. Выберите нужный:",
            hermes_case_choice_keyboard(session, matched_cases),
        )
        return True
    allows_actions = owner_request_allows_actions(user_text)
    explicitly_requests_action = owner_request_explicitly_requests_action(user_text)
    model_key = "ai_supervisor_model" if audit_mode else "deepseek_model"
    model = resolve_ai_model(get_setting_value(session, model_key))
    feature = "deep_audit" if audit_mode else "owner_chat"
    selected = build_hermes_owner_context(
        session,
        chat_id=chat_id,
        user_text=user_text,
        model=model,
        feature=feature,
        output_limit=ai_feature_output_limit(session, feature, detailed=intent.wants_detail),
    )
    action_policy = (
        "Владелец разрешил подготовить предложения действий: actions допустимы."
        if allows_actions
        else "Это информационный разговор. actions обязан быть пустым массивом; только ответь владельцу."
    )
    conversation, envelope = call_agent_envelope(
        session,
        chat_id=chat_id,
        actor_role="owner",
        lease=None,
        system_prompt=(
            configured_audit_system_prompt(session)
            if audit_mode
            else configured_owner_agent_system_prompt(session)
        ),
        context=selected.text + "\n\nПолитика текущего запроса:\n" + action_policy,
        user_text=user_text,
        model=model,
        max_tokens=ai_supervisor_max_tokens(session) if audit_mode else 1100,
        feature=feature,
        manifest=selected.manifest,
    )
    actions = envelope.actions if allows_actions else []
    if envelope.actions and not allows_actions:
        runtime_log("AGENT", f"suppressed {len(envelope.actions)} unsolicited actions chat_id={chat_id}")
    proposals: list[AgentActionProposal] = []
    proposal_errors: list[str] = []
    for action in actions:
        try:
            proposals.append(create_agent_action_proposal(session, conversation, chat_id, action))
        except HTTPException as exc:
            proposal_errors.append(str(exc.detail))

    if not explicitly_requests_action:
        send_telegram_text(session, chat_id, envelope.reply or AI_UNAVAILABLE_TEXT)
    elif not proposals:
        send_telegram_text(
            session,
            chat_id,
            owner_action_unavailable_message(user_text, proposal_errors),
        )
    elif proposal_errors:
        send_telegram_text(
            session,
            chat_id,
            f"Часть предложений не создана: {'; '.join(proposal_errors)}. Остальные ждут подтверждения кнопками.",
        )
    update_conversation_summary(session, conversation)
    return True


def owner_operation_result_text(operation: str, result: Any) -> str:
    label = owner_operation_label(operation)
    if isinstance(result, dict):
        if operation == "run_reminders":
            return (
                f"Напоминания обработаны: отправлено {result.get('sent', 0)}, "
                f"ошибок {result.get('failed', 0)}, без Telegram-привязки {result.get('skipped_unlinked', 0)}."
            )
        if operation == "broadcast_message":
            return (
                f"Рассылка завершена: отправлено {len(result.get('sent') or [])}, "
                f"ошибок {len(result.get('failed') or [])}, пропущено {len(result.get('skipped') or [])}."
            )
        if operation == "generate_rent_charges":
            return f"Начисления сформированы: новых записей {result.get('created', 0)}."
        if operation == "send_tenant_message":
            return f"Сообщение отправлено жильцу {result.get('tenant') or ''}.".strip()
        if operation == "issue_utility_bill":
            return (
                f"Коммунальный счёт выставлен: отправлено {result.get('sent', 0)}, "
                f"без привязки {result.get('skipped_unlinked', 0)}, "
                f"зачтено авансов {money_text(float(result.get('applied_advances') or 0))}."
            )
        if operation == "save_meter_readings_batch":
            return f"Показания сохранены: {result.get('saved', 0)}."
    return f"Операция «{label}» выполнена."


def execute_owner_web_operation(session: Session, operation: str, payload: dict[str, Any]) -> str:
    if operation == "create_object":
        result = create_object(payload, session)
    elif operation == "create_apartment":
        result = create_apartment(payload, session)
    elif operation == "update_apartment":
        result = update_apartment(int(payload["apartment_id"]), payload, session)
    elif operation == "onboard_tenant":
        result = onboard_tenant(payload, session)
    elif operation == "update_lease":
        result = update_lease(int(payload["lease_id"]), payload, session)
    elif operation == "transfer_lease":
        result = transfer_lease(int(payload["lease_id"]), payload, session)
    elif operation == "delete_lease":
        result = delete_lease(int(payload["lease_id"]), session)
    elif operation == "set_lease_automation":
        result = api_update_lease_automation(int(payload["lease_id"]), payload, session)
    elif operation == "clear_lease_automation":
        result = api_clear_lease_cadence(int(payload["lease_id"]), session)
    elif operation == "set_lease_ignored":
        result = api_update_lease_ignore(int(payload["lease_id"]), payload, session)
    elif operation == "broadcast_message":
        result = broadcast_message_to_tenants(payload, session)
    elif operation == "run_reminders":
        result = run_reminders_now(session)
    elif operation == "generate_rent_charges":
        result = generate_charges(payload, session)
    elif operation == "defer_payment_situation":
        situation = session.get(PaymentSituation, int(payload["situation_id"]))
        if not situation:
            raise HTTPException(404, "Платёжная ситуация не найдена")
        target = payment_situation_target(session, situation)
        if target is None:
            raise HTTPException(404, "Начисление платёжной ситуации больше не существует")
        sync_payment_situation(session, situation)
        if situation_debt(target) <= EPS:
            raise HTTPException(400, "Платёж уже закрыт, перенос даты не требуется")
        promised = parse_agent_date(payload.get("promise_date"), "новая дата оплаты")
        if promised < date.today():
            raise HTTPException(400, "Новая дата оплаты уже в прошлом")
        notify_tenant = setting_bool_value(payload.get("notify_tenant", True))
        tenant_chat_id = lease_chat_id(session, target.lease)
        if notify_tenant and not tenant_chat_id:
            raise HTTPException(400, "Жилец не привязан к Telegram, уведомить его о переносе невозможно")
        situation.promise_date = promised
        situation.paused_until = promised
        situation.status = "promised"
        if isinstance(target, RentCharge):
            target.deferral_until = promised
            target.deferral_note = str(payload.get("reason") or "подтверждено владельцем")
            update_rent_charge_status(target)
        if notify_tenant:
            send_telegram_text(
                session,
                tenant_chat_id,
                (
                    f"Спасибо, зафиксировал дату оплаты: {format_full_date(promised)}.\n\n"
                    "До этой даты автоматические напоминания по этому платежу поставлены на паузу."
                ),
            )
        session.commit()
        result = {"ok": True}
    elif operation == "set_payment_situation_mode":
        situation = session.get(PaymentSituation, int(payload["situation_id"]))
        if not situation:
            raise HTTPException(404, "Платёжная ситуация не найдена")
        mode = str(payload.get("mode") or "").strip()
        if mode not in {"normal", "manual"}:
            raise HTTPException(400, "Режим должен быть normal или manual")
        situation.mode = mode
        situation.status = "manual" if mode == "manual" else "awaiting_payment"
        session.commit()
        result = {"ok": True}
    elif operation == "add_rent_payment":
        result = add_rent_payment(int(payload["charge_id"]), payload, session)
    elif operation == "create_manual_payment":
        result = create_manual_payment(payload, session)
    elif operation == "update_manual_debt":
        result = update_manual_debt(int(payload["debt_id"]), payload, session)
    elif operation == "add_manual_debt_payment":
        result = add_manual_debt_payment(int(payload["debt_id"]), payload, session)
    elif operation == "delete_manual_debt":
        result = delete_manual_debt(int(payload["debt_id"]), session)
    elif operation == "update_payment_receipt":
        result = update_payment_receipt(int(payload["receipt_id"]), payload, session)
    elif operation == "delete_payment_receipt":
        result = delete_payment_receipt(int(payload["receipt_id"]), session)
    elif operation == "ignore_payment_receipt":
        result = ignore_payment_receipt(int(payload["receipt_id"]), session)
    elif operation == "moderate_payment_receipt":
        result = moderate_payment_receipt(int(payload["receipt_id"]), payload, session)
    elif operation == "update_utility_service":
        result = update_utility_service(int(payload["service_id"]), payload, session)
    elif operation == "save_meter_reading":
        result = save_meter_reading(payload, session)
    elif operation == "save_meter_readings_batch":
        result = save_meter_readings_batch(payload, session)
    elif operation == "create_tariff":
        result = create_tariff(payload, session)
    elif operation == "calculate_utility_bill":
        result = create_utility_bill(payload, session)
    elif operation == "delete_utility_bill":
        result = delete_utility_bill(int(payload["bill_id"]), session)
    elif operation == "issue_utility_bill":
        result = issue_utility_bill(int(payload["bill_id"]), session)
    elif operation == "mark_provider_paid":
        result = mark_provider_paid(int(payload["bill_id"]), session)
    elif operation == "add_utility_payment":
        result = add_utility_payment(int(payload["line_id"]), payload, session)
    elif operation == "create_expense":
        result = create_expense(payload, session)
    elif operation == "compensate_expense":
        result = compensate_expense(int(payload["expense_id"]), session)
    elif operation == "accept_monthly_report":
        result = accept_monthly_report(int(payload["year"]), int(payload["month"]), payload, session)
    else:
        raise HTTPException(400, f"Операция «{operation}» не подключена к backend")
    return owner_operation_result_text(operation, result)


def execute_agent_action(session: Session, proposal: AgentActionProposal) -> str:
    try:
        payload = json.loads(proposal.payload_json or "{}")
    except json.JSONDecodeError as exc:
        raise HTTPException(400, "Повреждены параметры предложения") from exc

    if proposal.action_type == "grouped_deferral":
        try:
            return execute_grouped_deferral(session, proposal)
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc

    if proposal.action_type == "defer_rent":
        lease = session.get(Lease, int(payload.get("lease_id") or proposal.lease_id or 0))
        if not lease:
            raise HTTPException(404, "Договор для действия не найден")
        charge = session.get(RentCharge, int(payload.get("charge_id") or 0))
        if not charge or charge.lease_id != lease.id:
            raise HTTPException(400, "Начисление для отсрочки больше не найдено")
        update_rent_charge_status(charge)
        debt = money(
            max(0.0, float(charge.ip_due or 0) - float(charge.ip_paid or 0))
            + max(0.0, float(charge.personal_due or 0) - float(charge.personal_paid or 0))
        )
        if debt <= EPS:
            raise HTTPException(400, "Долг уже закрыт, отсрочка не нужна")
        until = parse_agent_date(payload.get("deferral_until"), "дата отсрочки")
        if until <= date.today():
            raise HTTPException(400, "Дата отсрочки уже наступила")
        charge.deferral_until = until
        charge.deferral_note = str(payload.get("note") or "").strip()
        update_rent_charge_status(charge)
        situation = payment_situation(session, "rent", charge.id, lease.id)
        situation.promise_date = until
        situation.paused_until = until
        situation.status = "promised"
        if setting_bool_value(payload.get("notify_tenant")):
            message_text = str(payload.get("tenant_message") or "").strip() or (
                f"Владелец подтвердил отсрочку оплаты до {until:%d.%m.%Y}. "
                f"Текущий долг: {money_text(debt)}."
            )
            send_tenant_message(
                session,
                lease,
                "custom",
                charge=charge,
                custom_text=message_text,
                note=f"agent-action:{proposal.id}",
            )
        return f"Отсрочка для {lease.tenant.full_name} установлена до {until:%d.%m.%Y}."

    if proposal.action_type == "move_out":
        lease = session.get(Lease, int(payload.get("lease_id") or proposal.lease_id or 0))
        if not lease:
            raise HTTPException(404, "Договор для действия не найден")
        end = parse_agent_date(payload.get("end_date"), "дата выезда")
        if end < date.today() or end < lease.start_date:
            raise HTTPException(400, "Дата выезда больше не допустима")
        generate_rent_charges(session, until=end)
        lease.end_date = end
        lease.active = end > date.today()
        if payload.get("notes"):
            lease.notes = (lease.notes + "\n" + str(payload["notes"])).strip()
        prune_unpaid_rent_charges_after(session, lease, end)
        other_active = session.scalar(
            select(Lease.id).where(
                Lease.tenant_id == lease.tenant_id,
                Lease.id != lease.id,
                Lease.active.is_(True),
            ).limit(1)
        )
        lease.tenant.active = bool(lease.active or other_active)
        if end == date.today():
            process_move_out_notifications(session, end)
        if setting_bool_value(payload.get("notify_tenant")):
            message_text = str(payload.get("tenant_message") or "").strip() or (
                f"Владелец подтвердил оформление выезда на {end:%d.%m.%Y}. "
                f"Итоговый расчёт будет сформирован по данным системы."
            )
            send_tenant_message(
                session,
                lease,
                "custom",
                custom_text=message_text,
                note=f"agent-action:{proposal.id}",
            )
        return f"Выезд {lease.tenant.full_name} оформлен на {end:%d.%m.%Y}."

    if proposal.action_type == "create_manual_debt":
        lease = session.get(Lease, int(payload.get("lease_id") or proposal.lease_id or 0))
        if not lease:
            raise HTTPException(404, "Договор для действия не найден")
        debt = apply_manual_debt_payload(ManualDebt(), lease, payload)
        session.add(debt)
        session.flush()
        return f"Создан ручной долг «{debt.title}» на {money_text(debt.amount)}."

    if proposal.action_type == "send_tenant_message":
        lease = session.get(Lease, int(payload.get("lease_id") or proposal.lease_id or 0))
        if not lease:
            raise HTTPException(404, "Договор для действия не найден")
        message_text = str(payload.get("text") or "").strip()
        reject_tenant_message_claiming_unverified_change(message_text)
        result = send_tenant_message(
            session,
            lease,
            "custom",
            custom_text=message_text,
            note=f"agent-action:{proposal.id}",
        )
        return f"Сообщение отправлено жильцу {result['tenant']}."

    return execute_owner_web_operation(session, proposal.action_type, payload)


def active_payment_situation_for_lease(session: Session, lease_id: int) -> PaymentSituation | None:
    situations = session.scalars(
        select(PaymentSituation)
        .where(
            PaymentSituation.lease_id == lease_id,
            PaymentSituation.status.not_in(PAYMENT_SITUATION_FINAL_STATUSES),
        )
        .order_by(PaymentSituation.updated_at.desc(), PaymentSituation.id.desc())
    ).all()
    priority = {
        "awaiting_payment_date": 0,
        "question": 1,
        "overdue": 2,
        "due_today": 3,
        "first_reminder_sent": 4,
        "awaiting_payment": 5,
    }
    return min(situations, key=lambda item: priority.get(item.status, 10), default=None)


def parse_tenant_payment_date(text: str, today: date | None = None) -> date | None:
    today = today or date.today()
    lowered = re.sub(r"\s+", " ", (text or "").strip().lower())
    if not lowered:
        return None
    if "сегодня" in lowered:
        return today
    if "завтра" in lowered:
        return today + timedelta(days=1)
    after_days = re.search(r"через\s+(\d{1,2})\s*(?:дн|дня|день|дней)", lowered)
    if after_days:
        return today + timedelta(days=int(after_days.group(1)))
    iso = re.search(r"\b(20\d{2}-\d{2}-\d{2})\b", lowered)
    if iso:
        try:
            return date.fromisoformat(iso.group(1))
        except ValueError:
            return None
    short = re.search(r"\b(\d{1,2})[./](\d{1,2})(?:[./](20\d{2}))?\b", lowered)
    if short:
        day, month = int(short.group(1)), int(short.group(2))
        year = int(short.group(3) or today.year)
        try:
            result = date(year, month, day)
        except ValueError:
            return None
        if not short.group(3) and result < today:
            try:
                result = date(today.year + 1, month, day)
            except ValueError:
                return None
        return result
    return None


def short_deferral_is_safe(
    situation: PaymentSituation,
    target: RentCharge | UtilityBillLine,
    promised: date,
    today: date,
) -> bool:
    due = situation_due_date(target) or today
    return (
        today <= promised <= today + timedelta(days=3)
        and max(0, (today - due).days) <= 3
        and situation.mode == "normal"
        and situation.status not in {"ignored", "escalated", "manual", "receipt_pending_review"}
        and int(situation.notification_count or 0) < MAX_AUTOMATIC_SITUATION_REMINDERS
    )


def apply_short_payment_deferral(
    session: Session,
    situation: PaymentSituation,
    promised: date,
    reason: str,
) -> None:
    target = payment_situation_target(session, situation)
    if target is None:
        raise HTTPException(404, "Платёжная ситуация больше не существует")
    situation.promise_date = promised
    situation.paused_until = promised
    situation.status = "short_deferral"
    situation.last_tenant_response_at = utc_now()
    situation.tenant_response_count = int(situation.tenant_response_count or 0) + 1
    metadata = situation_metadata(situation)
    metadata["deferral_reason"] = reason[:1000]
    metadata["automatic_short_deferral"] = True
    save_situation_metadata(situation, metadata)
    if isinstance(target, RentCharge):
        target.deferral_until = promised
        target.deferral_note = f"короткая автоотсрочка: {reason[:500]}"
        update_rent_charge_status(target)
    lease = target.lease
    send_telegram_text(
        session,
        lease_chat_id(session, lease),
        (
            f"Спасибо, зафиксировал дату оплаты: {format_full_date(promised)}.\n\n"
            "До этой даты напоминания по этому платежу поставлены на паузу. "
            "После оплаты отправьте чек PDF-документом."
        ),
    )
    notify_owner_situation_event(session, situation, "short_deferral", date.today())


def propose_long_payment_deferral(
    session: Session,
    situation: PaymentSituation,
    promised: date,
    reason: str,
) -> None:
    owner_id = telegram_owner_chat_id(session)
    if not owner_id:
        raise HTTPException(400, "Owner chat не настроен")
    conversation = ai_conversation(session, owner_id, "owner", None)
    create_agent_action_proposal(
        session,
        conversation,
        owner_id,
        {
            "type": "owner_operation",
            "payload": {
                "operation": "defer_payment_situation",
                "arguments": {
                    "situation_id": situation.id,
                    "promise_date": promised.isoformat(),
                    "reason": reason,
                    "notify_tenant": True,
                },
            },
            "reason": reason,
        },
    )
    situation.status = "escalated"
    situation.escalated_at = utc_now()


def handle_tenant_payment_text(session: Session, lease: Lease, chat_id: int | str, text: str) -> bool:
    situation = active_payment_situation_for_lease(session, lease.id)
    if not situation:
        return False
    promised = parse_tenant_payment_date(text)
    expects_date = situation.status == "awaiting_payment_date"
    mentions_payment_date = bool(
        re.search(r"\b(оплачу|переведу|смогу\s+оплатить|через\s+\d+\s*д|завтра|сегодня)\b", text.lower())
    )
    if not expects_date and not mentions_payment_date:
        return False
    if not promised:
        send_telegram_text(
            session,
            chat_id,
            "Не смог определить дату. Напишите её, например: «завтра», «через 2 дня» или «28.06.2026».",
        )
        return True
    target = payment_situation_target(session, situation)
    if target is None:
        return False
    reason = f"жилец написал: {text.strip()[:600]}"
    if short_deferral_is_safe(situation, target, promised, date.today()):
        apply_short_payment_deferral(session, situation, promised, reason)
    else:
        situation.promise_date = promised
        situation.last_tenant_response_at = utc_now()
        situation.tenant_response_count = int(situation.tenant_response_count or 0) + 1
        propose_long_payment_deferral(session, situation, promised, reason)
        send_telegram_text(
            session,
            chat_id,
            (
                f"Запрос на перенос оплаты до {format_full_date(promised)} передан владельцу. "
                "До его решения я не буду обещать изменение срока."
            ),
        )
    return True


def handle_payment_situation_callback_query(session: Session, callback: dict[str, Any]) -> bool:
    data = str(callback.get("data") or "")
    match = re.fullmatch(r"(tenantpay|ownersit):(\d+):([a-z_]+)", data)
    if not match:
        return False
    scope, situation_id_text, action = match.groups()
    callback_id = str(callback.get("id") or "")
    sender = callback.get("from") or {}
    message = callback.get("message") or {}
    chat_id = ((message.get("chat") or {}).get("id")) or sender.get("id")
    token = telegram_token(session)
    situation = session.get(PaymentSituation, int(situation_id_text))
    if not callback_id or not chat_id or not token or not situation:
        return True
    target = payment_situation_target(session, situation)
    if target is None:
        safe_answer_agent_callback(token, callback_id, "Платёж уже не найден.")
        return True
    lease = target.lease

    if scope == "tenantpay":
        if str(lease_chat_id(session, lease)) != str(chat_id):
            safe_answer_agent_callback(token, callback_id, "Эта кнопка относится к другому договору.")
            return True
        situation.last_tenant_response_at = utc_now()
        situation.tenant_response_count = int(situation.tenant_response_count or 0) + 1
        if action in {"paid", "receipt"}:
            situation.status = "awaiting_receipt"
            situation.paused_until = None
            safe_answer_agent_callback(token, callback_id, "Жду чек PDF-документом.")
            send_telegram_text(
                session,
                chat_id,
                (
                    "Спасибо. Пришлите, пожалуйста, чек PDF-документом для подтверждения. "
                    "Для аренды нужны отдельные чеки на перевод ИП и перевод по номеру, если обе части оплачены отдельно."
                ),
            )
        elif action == "later":
            situation.status = "awaiting_payment_date"
            safe_answer_agent_callback(token, callback_id, "Напишите точную дату оплаты.")
            send_telegram_text(
                session,
                chat_id,
                "Укажите дату, когда сможете оплатить: например, «завтра», «через 2 дня» или «28.06.2026».",
            )
        elif action == "question":
            situation.status = "question"
            situation.escalated_at = utc_now()
            safe_answer_agent_callback(token, callback_id, "Вопрос передан владельцу.")
            send_telegram_text(session, chat_id, "Спасибо, передал вопрос владельцу. До ответа напоминания поставлены на паузу.")
            notify_owner_situation_event(session, situation, "question", date.today())
        return True

    if not owner_chat_allowed(session, sender.get("id") or ""):
        safe_answer_agent_callback(token, callback_id, "Эти действия доступны только владельцу.")
        return True
    if action == "manual":
        conversation = ai_conversation(session, chat_id, "owner", None)
        create_agent_action_proposal(
            session,
            conversation,
            chat_id,
            {
                "type": "owner_operation",
                "payload": {
                    "operation": "set_payment_situation_mode",
                    "arguments": {"situation_id": situation.id, "mode": "manual"},
                },
                "reason": "владелец выбрал ручной режим в карточке платёжной ситуации",
            },
        )
        safe_answer_agent_callback(token, callback_id, "Нужно подтверждение владельца.")
    elif action == "wait":
        conversation = ai_conversation(session, chat_id, "owner", None)
        create_agent_action_proposal(
            session,
            conversation,
            chat_id,
            {
                "type": "owner_operation",
                "payload": {
                    "operation": "defer_payment_situation",
                    "arguments": {
                        "situation_id": situation.id,
                        "promise_date": (date.today() + timedelta(days=1)).isoformat(),
                        "reason": "владелец выбрал паузу на один день",
                        "notify_tenant": False,
                    },
                },
                "reason": "владелец выбрал паузу на один день",
            },
        )
        safe_answer_agent_callback(token, callback_id, "Нужно подтверждение владельца.")
    elif action == "suggest":
        safe_answer_agent_callback(token, callback_id, "Готовлю предложение сообщения.")
        handle_owner_ai_message(
            session,
            chat_id,
            (
                f"Подготовь индивидуальное сообщение жильцу по payment situation {situation.id}. "
                "Не отправляй без отдельного подтверждения."
            ),
        )
    return True


def safe_answer_agent_callback(token: str, callback_id: str, text: str) -> None:
    try:
        answer_callback_query(token, callback_id, text)
    except TelegramApiError as exc:
        runtime_log("TELEGRAM", f"callback answer failed error={exc}")


def hermes_case_details_text(session: Session, item: OperationalCase) -> str:
    snapshot = case_snapshot(session, item)
    lines = [
        f"{snapshot['label']} · {item.title}",
        "",
        item.compact_summary,
        f"Статус: {item.status}; важность: {item.severity}.",
    ]
    if item.waiting_for:
        lines.append(f"Сейчас ожидается: {item.waiting_for}.")
    if item.next_review_at:
        lines.append(f"Следующая проверка: {item.next_review_at:%d.%m.%Y %H:%M}.")
    metadata = snapshot.get("metadata") or {}
    if metadata.get("promise_date"):
        lines.append(f"Обещанная дата: {metadata['promise_date']}.")
    suggestion = {
        "broken_payment_promise": "Предлагаю проверить ситуацию и зафиксировать следующий шаг.",
        "rent_overdue": "Предлагаю отправить напоминание или согласовать отсрочку кнопкой.",
        "rent_partial": "Предлагаю уточнить остаток и ожидаемую дату оплаты.",
        "rent_due": "Предлагаю напомнить о сроке или согласовать отсрочку кнопкой.",
        "utility_overdue": "Предлагаю отправить напоминание о коммуналке.",
        "stale_meter_reading": "Внесите свежие показания на странице счётчиков.",
        "suspicious_payment": "Проверьте чек на странице сообщений.",
        "expense_compensation": "Проверьте компенсацию на странице расходов.",
        "automation_failure": "Проверьте журнал автоматизации и причину ошибки.",
    }.get(item.case_type, "Зафиксируйте следующий шаг или закройте кейс.")
    lines.extend(["", suggestion])
    return clean_ai_response("\n".join(lines), max_chars=ai_feature_output_limit(session, "case_details"))


def hermes_case_actions_keyboard(item: OperationalCase) -> dict[str, Any]:
    rows: list[list[dict[str, str]]] = []
    if item.case_type in {"rent_due", "rent_overdue", "rent_partial"}:
        rows.append(
            [
                {"text": "Отправить напоминание", "callback_data": f"ai:case:{item.id}:remind"},
                {"text": "Дать отсрочку", "callback_data": f"ai:case:{item.id}:defer"},
            ]
        )
    elif item.case_type == "utility_overdue":
        rows.append([{"text": "Отправить напоминание", "callback_data": f"ai:case:{item.id}:remind"}])

    commitment_label = {
        "stale_meter_reading": "Внесу показания",
        "suspicious_payment": "Проверю платёж",
        "expense_compensation": "Проверю расход",
        "automation_failure": "Проверю ошибку",
        "tenant_message_escalation": "Отвечу жильцу",
        "deferral_request": "Решу запрос",
    }.get(item.case_type, "Я разберусь")
    rows.extend(
        [
            [
                {"text": commitment_label, "callback_data": f"ai:case:{item.id}:owner_commitment"},
                {"text": "Закрыть", "callback_data": f"ai:case:{item.id}:close"},
            ],
            [{"text": "Отложить на день", "callback_data": f"ai:case:{item.id}:snooze"}],
        ]
    )
    return {"inline_keyboard": rows}


def send_case_template_reminder(session: Session, item: OperationalCase) -> str:
    try:
        metadata = json.loads(item.metadata_json or "{}")
    except json.JSONDecodeError:
        metadata = {}
    if metadata.get("rent_charge_id"):
        charge = session.get(RentCharge, int(metadata["rent_charge_id"]))
        if not charge:
            raise HTTPException(404, "Начисление больше не найдено")
        situation = payment_situation(session, "rent", charge.id, charge.lease_id)
        template_key = "message_rent_status_request" if charge.due_date < date.today() else "message_rent_due"
        if not send_payment_situation_message(session, situation, template_key, "owner_case_action", date.today()):
            return "Напоминание сегодня уже отправлялось."
        record_reminder_outcome(
            session,
            lease_id=charge.lease_id,
            stage="owner_case_action",
            template_key=template_key,
            payment_situation_id=situation.id,
        )
        return "Шаблонное напоминание отправлено."
    if metadata.get("utility_line_id"):
        line = session.get(UtilityBillLine, int(metadata["utility_line_id"]))
        if not line or not line.lease_id:
            raise HTTPException(404, "Коммунальный платёж больше не найден")
        situation = payment_situation(session, "utility", line.id, int(line.lease_id))
        if not send_payment_situation_message(
            session,
            situation,
            "message_utility_overdue",
            "owner_case_action",
            date.today(),
        ):
            return "Напоминание сегодня уже отправлялось."
        record_reminder_outcome(
            session,
            lease_id=int(line.lease_id),
            stage="owner_case_action",
            template_key="message_utility_overdue",
            payment_situation_id=situation.id,
        )
        return "Шаблонное напоминание отправлено."
    raise HTTPException(400, "Для этого кейса нет разрешённого шаблонного напоминания")


def handle_hermes_callback_query(
    session: Session,
    callback: dict[str, Any],
    *,
    callback_id: str,
    data: str,
    chat_id: int | str,
    token: str,
) -> bool:
    case_match = re.fullmatch(r"ai:case:(\d+):([a-z_]+)", data)
    if case_match:
        case_id_text, action = case_match.groups()
        item = session.get(OperationalCase, int(case_id_text))
        if not item:
            safe_answer_agent_callback(token, callback_id, "Кейс не найден.")
            return True
        if action in {"details", "actions"}:
            safe_answer_agent_callback(token, callback_id, "Открываю кейс.")
            send_telegram_text(
                session,
                chat_id,
                hermes_case_details_text(session, item),
                hermes_case_actions_keyboard(item),
            )
        elif action == "snooze":
            hermes_snooze_case(session, item.id, utc_now() + timedelta(days=1))
            safe_answer_agent_callback(token, callback_id, "Отложено на день.")
        elif action == "owner_commitment":
            commitment = create_owner_commitment(
                session,
                case=item,
                text="Я разберусь",
                briefing_time=get_setting_value(session, "ai_supervisor_time") or "10:00",
            )
            safe_answer_agent_callback(token, callback_id, "Обязательство сохранено.")
            send_telegram_text(
                session,
                chat_id,
                f"{case_label(session, item)}: спрошу о результате {commitment.due_at:%d.%m в %H:%M}.",
            )
        elif action == "close":
            hermes_close_case(session, item.id, "Владелец закрыл кейс через Telegram")
            safe_answer_agent_callback(token, callback_id, "Кейс закрыт.")
        elif action == "remind":
            try:
                result = send_case_template_reminder(session, item)
            except HTTPException as exc:
                safe_answer_agent_callback(token, callback_id, str(exc.detail)[:180])
            else:
                safe_answer_agent_callback(token, callback_id, result[:180])
        elif action == "defer":
            try:
                metadata = json.loads(item.metadata_json or "{}")
            except json.JSONDecodeError:
                metadata = {}
            conversation = ai_conversation(session, chat_id, "owner", None)
            if metadata.get("rent_charge_id") and item.contract_id:
                create_agent_action_proposal(
                    session,
                    conversation,
                    chat_id,
                    {
                        "type": "defer_rent",
                        "payload": {
                            "lease_id": item.contract_id,
                            "charge_id": metadata["rent_charge_id"],
                            "deferral_until": (date.today() + timedelta(days=7)).isoformat(),
                            "notify_tenant": True,
                        },
                        "reason": "владелец запросил отсрочку из карточки кейса",
                    },
                )
                safe_answer_agent_callback(token, callback_id, "Создано предложение на 7 дней.")
            else:
                safe_answer_agent_callback(token, callback_id, "Уточните срок отсрочки сообщением Hermes.")
        else:
            safe_answer_agent_callback(token, callback_id, "Неизвестное действие кейса.")
        return True

    briefing_match = re.fullmatch(r"ai:briefing:(\d+):(remaining|snooze)", data)
    if briefing_match:
        briefing_id_text, action = briefing_match.groups()
        briefing = session.get(OwnerBriefing, int(briefing_id_text))
        if not briefing:
            safe_answer_agent_callback(token, callback_id, "Сводка не найдена.")
            return True
        if action == "snooze":
            snooze_briefing(session, briefing.id)
            safe_answer_agent_callback(token, callback_id, "Сводка отложена на 3 часа.")
        else:
            included = set(
                session.scalars(
                    select(OwnerBriefingItem.case_id).where(OwnerBriefingItem.briefing_id == briefing.id)
                ).all()
            )
            remaining = [item for item in reconcile_operational_cases(session) if item.id not in included]
            safe_answer_agent_callback(token, callback_id, "Показываю остальные кейсы.")
            send_telegram_text(
                session,
                chat_id,
                hermes_open_cases_text(session, remaining, limit=8),
                hermes_case_choice_keyboard(session, remaining),
            )
        return True

    commitment_match = re.fullmatch(r"ai:commitment:(\d+):(done|postpone|details)", data)
    if commitment_match:
        commitment_id_text, action = commitment_match.groups()
        callback_commitment = session.get(OwnerCommitment, int(commitment_id_text))
        if not callback_commitment:
            safe_answer_agent_callback(token, callback_id, "Обязательство не найдено.")
            return True
        if action == "done":
            complete_owner_commitment(session, callback_commitment.id)
            if callback_commitment.case_id:
                hermes_close_case(session, callback_commitment.case_id, "Владелец подтвердил выполнение обязательства")
            safe_answer_agent_callback(token, callback_id, "Отмечено выполненным.")
        elif action == "postpone":
            postpone_owner_commitment(session, callback_commitment.id)
            safe_answer_agent_callback(token, callback_id, "Перенесено на завтра.")
        elif callback_commitment.case_id:
            item = session.get(OperationalCase, callback_commitment.case_id)
            if item:
                send_telegram_text(session, chat_id, hermes_case_details_text(session, item), hermes_case_actions_keyboard(item))
            safe_answer_agent_callback(token, callback_id, "Открываю подробности.")
        return True

    skill_match = re.fullmatch(r"ai:skill:(\d+):(activate|edit|cancel|disable|rollback)", data)
    if skill_match:
        skill_id_text, action = skill_match.groups()
        try:
            callback_skill: AiSkill | None
            if action == "activate":
                callback_skill = activate_skill(session, int(skill_id_text))
                answer = "Навык активирован."
            elif action in {"cancel", "disable"}:
                callback_skill = disable_skill(session, int(skill_id_text))
                answer = "Навык отключён."
            elif action == "rollback":
                callback_skill = rollback_skill(session, int(skill_id_text))
                answer = "Предыдущая версия активирована."
            else:
                callback_skill = session.get(AiSkill, int(skill_id_text))
                answer = "Опишите изменения обычным сообщением с названием навыка."
            safe_answer_agent_callback(token, callback_id, answer)
            if callback_skill:
                send_telegram_text(session, chat_id, skill_card(callback_skill))
        except ValueError as exc:
            safe_answer_agent_callback(token, callback_id, str(exc)[:180])
        return True
    return False


def handle_agent_callback_query(session: Session, callback: dict[str, Any]) -> None:
    callback_id = str(callback.get("id") or "")
    data = str(callback.get("data") or "")
    sender = callback.get("from") or {}
    message = callback.get("message") or {}
    chat = message.get("chat") or {}
    chat_id = chat.get("id") or sender.get("id")
    token = telegram_token(session)
    if not callback_id or not chat_id or not token:
        return

    if not owner_chat_allowed(session, sender.get("id") or ""):
        safe_answer_agent_callback(token, callback_id, "Подтверждать действия может только владелец.")
        return

    proposal_alias = re.fullmatch(r"ai:proposal:(\d+):(confirm|reject|edit)", data)
    if proposal_alias:
        proposal_id_text, alias_action = proposal_alias.groups()
        if alias_action == "edit":
            safe_answer_agent_callback(token, callback_id, "Опишите изменения сообщением — старое предложение не выполнится.")
            return
        data = f"agent:{alias_action}:{proposal_id_text}"

    if handle_hermes_callback_query(
        session,
        callback,
        callback_id=callback_id,
        data=data,
        chat_id=chat_id,
        token=token,
    ):
        return

    match = re.fullmatch(r"agent:(confirm|reject):(\d+)", data)
    if not match:
        safe_answer_agent_callback(token, callback_id, "Неизвестная кнопка.")
        return
    decision, proposal_id_text = match.groups()
    proposal = session.scalar(
        select(AgentActionProposal)
        .where(AgentActionProposal.id == int(proposal_id_text))
        .with_for_update()
    )
    if not proposal:
        safe_answer_agent_callback(token, callback_id, "Предложение не найдено.")
        return
    if proposal.status != "pending":
        safe_answer_agent_callback(token, callback_id, f"Уже обработано: {proposal.status}.")
        return
    if proposal.expires_at and proposal.expires_at < utc_now():
        proposal.status = "expired"
        proposal.error_text = "Истёк срок подтверждения"
        session.add(
            AiActionLog(
                conversation_id=proposal.conversation_id,
                lease_id=proposal.lease_id,
                actor_role="system",
                action_type=proposal.action_type,
                status="expired",
                payload_json=proposal.payload_json,
                note=proposal.error_text,
            )
        )
        safe_answer_agent_callback(token, callback_id, "Срок подтверждения истёк.")
        return
    if proposal.owner_chat_id and str(chat_id) != proposal.owner_chat_id:
        safe_answer_agent_callback(token, callback_id, "Подтверждение доступно только в исходном owner chat.")
        return
    proposal_owner_message_id = proposal.owner_message_id
    if decision == "confirm" and int(proposal.safety_level or 0) >= 3:
        safe_answer_agent_callback(token, callback_id, "Для уровня 3 нужен повторный PIN в Control Center.")
        send_telegram_text(
            session,
            chat_id,
            "Это критическое или массовое действие. Откройте Hermes Control Center в веб-панели или Android и подтвердите его повторным PIN владельца.",
        )
        return

    if decision == "reject":
        proposal.status = "rejected"
        proposal.result_text = "Отклонено владельцем"
        session.add(
            AiActionLog(
                conversation_id=proposal.conversation_id,
                lease_id=proposal.lease_id,
                actor_role="owner",
                action_type=proposal.action_type,
                status="rejected",
                payload_json=proposal.payload_json,
                note="owner rejected agent proposal",
            )
        )
        safe_answer_agent_callback(token, callback_id, "Отклонено. Данные не менялись.")
        send_telegram_text(session, chat_id, f"Отклонено: {proposal.preview_text.splitlines()[0]}.")
    else:
        proposal.status = "approved"
        proposal.confirmed_at = utc_now()
        proposal.confirmed_by = str(sender.get("id") or chat_id)
        session.add(
            AiActionLog(
                conversation_id=proposal.conversation_id,
                lease_id=proposal.lease_id,
                actor_role="owner",
                action_type=proposal.action_type,
                status="approved",
                payload_json=proposal.payload_json,
                note="owner approved agent proposal",
            )
        )
        session.flush()
        try:
            result_text = execute_agent_action(session, proposal)
        except HTTPException as exc:
            proposal.status = "failed"
            proposal.result_text = str(exc.detail)
            proposal.error_text = str(exc.detail)
            session.add(
                AiActionLog(
                    conversation_id=proposal.conversation_id,
                    lease_id=proposal.lease_id,
                    actor_role="owner",
                    action_type=proposal.action_type,
                    status="failed",
                    payload_json=proposal.payload_json,
                    note=str(exc.detail),
                )
            )
            safe_answer_agent_callback(token, callback_id, "Не выполнено: данные изменились.")
            send_telegram_text(session, chat_id, f"Действие не выполнено: {exc.detail}")
        except Exception as exc:
            runtime_log(
                "AGENT",
                f"owner operation failed proposal_id={proposal.id} action={proposal.action_type} "
                f"error={type(exc).__name__}: {exc}",
            )
            proposal_id = proposal.id
            session.rollback()
            proposal = session.get(AgentActionProposal, proposal_id)
            if proposal:
                proposal.status = "failed"
                proposal.result_text = "Внутренняя ошибка выполнения"
                proposal.error_text = f"{type(exc).__name__}: {exc}"[:2000]
                session.add(
                    AiActionLog(
                        conversation_id=proposal.conversation_id,
                        lease_id=proposal.lease_id,
                        actor_role="owner",
                        action_type=proposal.action_type,
                        status="failed",
                        payload_json=proposal.payload_json,
                        note=proposal.error_text,
                    )
                )
            safe_answer_agent_callback(token, callback_id, "Не выполнено: внутренняя ошибка.")
            send_telegram_text(
                session,
                chat_id,
                "Действие не выполнено из-за внутренней ошибки. Данные не считаются изменёнными; подробность записана в лог.",
            )
        else:
            proposal.status = "executed"
            proposal.executed_at = utc_now()
            proposal.result_text = result_text
            proposal.error_text = ""
            reconcile_operational_cases(session)
            session.add(
                AiActionLog(
                    conversation_id=proposal.conversation_id,
                    lease_id=proposal.lease_id,
                    actor_role="owner",
                    action_type=proposal.action_type,
                    status="executed",
                    payload_json=proposal.payload_json,
                    note=result_text,
                )
            )
            safe_answer_agent_callback(token, callback_id, "Подтверждено и выполнено.")
            send_telegram_text(session, chat_id, f"Готово. {result_text}")

    message_id = message.get("message_id") or proposal_owner_message_id
    if message_id:
        try:
            clear_inline_keyboard(token, chat_id, int(message_id))
        except TelegramApiError:
            pass


def tenant_requisites_text(session: Session) -> str:
    settings = get_settings(session)
    lines = [
        "Реквизиты для перевода на ИП:",
        f"Наименование: {settings.get('ip_recipient_name') or 'не задано'}",
        f"ИНН: {settings.get('ip_recipient_inn') or 'не задано'}",
        f"ОГРНИП: {settings.get('ip_recipient_ogrnip') or 'не задано'}",
        f"Расчётный счёт: {settings.get('ip_recipient_account') or 'не задано'}",
        f"Банк: {settings.get('ip_recipient_bank') or 'не задано'}",
        f"БИК банка: {settings.get('ip_recipient_bik') or 'не задано'}",
        f"Корр. счёт банка: {settings.get('ip_recipient_correspondent_account') or 'не задано'}",
        f"ИНН банка: {settings.get('ip_recipient_bank_inn') or 'не задано'}",
        f"КПП банка: {settings.get('ip_recipient_bank_kpp') or 'не задано'}",
        "",
        "Реквизиты для перевода по номеру телефона (дополнительная часть аренды и коммуналка):",
        (
            f"{settings.get('personal_recipient_phone') or 'номер не задан'} "
            f"{settings.get('personal_recipient_bank') or 'банк не задан'} "
            f"{settings.get('personal_recipient_name') or 'получатель не задан'}"
        ).strip(),
        "",
        "Назначение платежей:",
        "• часть аренды «ИП» — только на расчётный счёт ИП;",
        "• дополнительная часть аренды — переводом по номеру телефона;",
        "• коммунальные платежи — переводом по номеру телефона;",
        "• если переводов несколько, отправьте отдельный чек PDF по каждому.",
    ]
    return "\n".join(lines)


def tenant_help_text() -> str:
    return "\n".join(
        [
            "Что умеет бот для жильца:",
            "• принять чек PDF документом и попытаться зачесть его автоматически;",
            "• если что-то не понял — передать чек владельцу на ручную проверку;",
            "• показать реквизиты по кнопке «Реквизиты» или по команде /requisites;",
            "• показать все текущие долги по кнопке «Все долги» или по команде /debts;",
            "• подтвердить, что чек уже был получен, чтобы не крутить один и тот же документ по кругу.",
        ]
    )


def tenant_personal_data_consent_text() -> str:
    return "\n".join(
        [
            "Соглашение на обработку персональных данных:",
            "Продолжая пользоваться ботом, вы соглашаетесь на обработку ваших персональных данных по 152-ФЗ РФ.",
            "Обрабатываются: ФИО, номер телефона, Telegram ID/username, сведения по аренде, платежам, коммунальным начислениям и отправленные вами документы/чеки.",
            "Цель обработки: учёт аренды, проверка оплат, расчёт коммунальных платежей и связь по вопросам проживания.",
            "Если не согласны, не отправляйте боту документы и свяжитесь с владельцем напрямую.",
        ]
    )


def telegram_help_text() -> str:
    return "\n".join(
        [
            "🤖 Команды owner-бота:",
            "Обычный текст — разговор с агентом; команды ниже обрабатываются механически.",
            "/ping - проверить скорость ответа",
            "/start - приветствие",
            "/id - показать chat id",
            "/status - 📊 статус пульта",
            "/reports - 🗂 открытые месячные отчёты",
            "/requisites - 💳 показать реквизиты для оплаты",
            "/run_reminders - 🔔 прогнать напоминания",
            "/app - 🚪 открыть пульт",
            "/help - подсказка",
        ]
    )


def handle_telegram_message(session: Session, message: dict[str, Any]) -> None:
    started = time_module.perf_counter()
    chat = message.get("chat") or {}
    sender = message.get("from") or {}
    chat_id = chat.get("id")
    if not chat_id:
        return

    text = (message.get("text") or "").strip()
    command, _args = parse_command(text)
    owner_id = telegram_owner_chat_id(session)
    is_owner = owner_message_allowed(session, message)
    linked_lease = None if is_owner else maybe_link_tenant_chat(session, message)
    log_telegram_incoming_message(session, message, linked_lease, is_owner=is_owner)
    preparation_ms = int((time_module.perf_counter() - started) * 1000)
    runtime_log(
        "TELEGRAM",
        (
            f"message chat_id={chat_id} from_id={sender.get('id')} chat_type={chat.get('type') or '-'} "
            f"command={command or '-'} text_len={len(text)} "
            f"is_owner={is_owner} owner_configured={bool(owner_id)} "
            f"linked_lease_id={getattr(linked_lease, 'id', None)} preparation_ms={preparation_ms}"
        ),
    )

    if command == "/ping":
        elapsed_ms = int((time_module.perf_counter() - started) * 1000)
        send_telegram_text(session, chat_id, f"🏓 Pong. Подготовка ответа на сервере: {elapsed_ms} мс.")
        return

    if command == "/id":
        send_telegram_text(session, chat_id, f"Ваш chat id: {chat_id}")
        return

    base_url = app_base_url(session)

    if message.get("contact") and not is_owner:
        if linked_lease:
            send_telegram_text(
                session,
                chat_id,
                f"Готово, привязал этот чат к квартире {linked_lease.apartment.name}.\n\n{tenant_personal_data_consent_text()}",
                tenant_keyboard(),
            )
        else:
            send_telegram_text(session, chat_id, "Не нашёл активного жильца с таким телефоном в базе.")
        return

    if command == "/start":
        if is_owner:
            send_telegram_text(
                session,
                chat_id,
                "✅ Бот подключён.\n📊 Можно смотреть статус, отчёты и открывать пульт с телефона.",
                app_keyboard(base_url, include_app=True),
            )
            return
        if linked_lease:
            if not tenant_consent_sent(session, chat_id):
                mark_tenant_consent_sent(session, chat_id)
                send_telegram_text(session, chat_id, tenant_personal_data_consent_text(), tenant_keyboard())
                return
            send_telegram_text(
                session,
                chat_id,
                f"Привет. Я привязал этот чат к квартире {linked_lease.apartment.name}. Теперь сюда можно присылать чеки PDF документом, спрашивать реквизиты и получать ответы по платежам.",
                tenant_keyboard(),
            )
            return
        send_telegram_text(
            session,
            chat_id,
            "Если вы жилец без @username, нажмите «Привязать по телефону». Пульт владельца через бота не выдаётся.",
            tenant_keyboard(),
        )
        return

    if command == "/help":
        if is_owner:
            send_telegram_text(session, chat_id, telegram_help_text(), app_keyboard(base_url))
        else:
            send_telegram_text(session, chat_id, tenant_help_text(), tenant_keyboard())
        return

    if command == "/requisites" or text.lower() == "реквизиты":
        # Показываем реквизиты только владельцу или привязанным жильцам
        if is_owner or linked_lease:
            send_telegram_text(
                session,
                chat_id,
                tenant_requisites_text(session),
                tenant_keyboard() if not is_owner else app_keyboard(base_url),
            )
        else:
            send_telegram_text(session, chat_id, "Реквизиты доступны только для привязанных жильцов.")
        return

    if command == "/debts" or text.lower() == "все долги":
        if linked_lease:
            send_telegram_text(
                session,
                chat_id,
                render_message_text(session, "message_all_debts", linked_lease),
                tenant_keyboard(),
            )
        elif is_owner:
            send_telegram_text(session, chat_id, "Команда /debts доступна только в привязанном чате жильца. Сводка владельца доступна в пульте.")
        else:
            send_telegram_text(session, chat_id, "Сначала привяжи чат через @username в базе и /start.")
        return

    # Обработка файлов от жильцов - до проверки owner_id
    if not is_owner and (message.get("document") or message.get("photo")):
        if linked_lease:
            handle_tenant_receipt_message(session, message, linked_lease)
        return

    if not owner_id:
        send_telegram_text(
            session,
            chat_id,
            f"Owner chat id ещё не настроен. Отправь /id и сохрани это значение в приложении. Сейчас chat id: {chat_id}",
        )
        return

    if not is_owner and linked_lease and text:
        if handle_tenant_payment_text(session, linked_lease, chat_id, text):
            return
        if handle_tenant_ai_message(session, chat_id, linked_lease, text):
            return

    if not is_owner:
        send_telegram_text(
            session,
            chat_id,
            tenant_help_text(),
            tenant_keyboard(),
        )
        return

    if command == "/status":
        status_started = time_module.perf_counter()
        status_dashboard = build_status_dashboard_fast(session)
        runtime_log(
            "TELEGRAM",
            f"status snapshot chat_id={chat_id} duration_ms={int((time_module.perf_counter() - status_started) * 1000)}",
        )
        send_telegram_text(session, chat_id, build_status_message(status_dashboard), app_keyboard(base_url))
        return
    if command == "/reports":
        reports_started = time_module.perf_counter()
        reports = build_monthly_reports(session)
        runtime_log(
            "TELEGRAM",
            f"reports snapshot chat_id={chat_id} duration_ms={int((time_module.perf_counter() - reports_started) * 1000)}",
        )
        send_telegram_text(session, chat_id, build_reports_message(reports), app_keyboard(base_url, reports))
        return
    if command == "/ask":
        question = " ".join(_args).strip() or text[len("/ask") :].strip()
        if not question:
            send_telegram_text(session, chat_id, "Напиши вопрос после /ask.", app_keyboard(base_url))
            return
        handle_owner_ai_message(session, chat_id, question)
        return
    if command == "/audit":
        deep = "deep" in text.lower().split()[1:]
        handle_owner_ai_message(session, chat_id, AUDIT_USER_PROMPT, audit_deep=deep, audit_requested=True)
        return
    if command == "/run_reminders":
        summary = run_due_reminders(session)
        session.commit()
        send_telegram_text(
            session,
            chat_id,
            "\n".join(
                [
                    "🔔 Напоминания прогнаны.",
                    f"✅ Отправлено: {summary['sent']}",
                    f"🔁 Дубликаты за сегодня: {summary['skipped_duplicate']}",
                    f"⏳ Старые долги под молчанием: {summary['skipped_legacy']}",
                    f"🔗 Без привязки к боту: {summary['skipped_unlinked']}",
                    f"⚠️ Ошибки: {summary['failed']}",
                ]
            ),
            app_keyboard(base_url),
        )
        return
    if command == "/app":
        if is_owner:
            send_telegram_text(session, chat_id, "🚪 Открываю пульт.", app_keyboard(base_url, include_app=True))
        else:
            send_telegram_text(session, chat_id, "Доступ к пульту только у владельца.")
        return
    if command == "/requisites":
        send_telegram_text(session, chat_id, tenant_requisites_text(session), app_keyboard(base_url))
        return
    if message.get("document") or message.get("photo"):
        send_telegram_text(
            session,
            chat_id,
            "Файл получил. OCR и разбор чеков ещё не включены, но вход для этого уже подготовлен.",
        )
        return

    if text:
        confirmation_decision = owner_text_confirmation_decision(text)
        if confirmation_decision:
            pending_count = int(
                session.scalar(
                    select(func.count(AgentActionProposal.id)).where(
                        AgentActionProposal.status == "pending",
                        AgentActionProposal.owner_chat_id == str(chat_id),
                    )
                )
                or 0
            )
            if pending_count:
                send_telegram_text(
                    session,
                    chat_id,
                    "Подтверждение текстом не выполняет действия. Нажмите «Подтвердить» или «Отклонить» "
                    "под нужным предложением — там зафиксированы точные параметры.",
                )
            else:
                send_telegram_text(
                    session,
                    chat_id,
                    "Сейчас нет действия, ожидающего подтверждения. Сначала сформулируйте, что именно нужно сделать.",
                )
            return
        handle_owner_ai_message(session, chat_id, text)
        return
    send_telegram_text(session, chat_id, telegram_help_text(), app_keyboard(base_url))


def process_telegram_update_background(payload: dict[str, Any], received_at: float | None = None) -> None:
    update_id = payload.get("update_id")
    started = time_module.perf_counter()
    queue_ms = int((started - received_at) * 1000) if received_at is not None else 0
    status = "ok"
    perf_detail: dict[str, Any] = {"update_id": update_id, "queue_ms": queue_ms}
    try:
        callback = payload.get("callback_query")
        message = payload.get("message") or payload.get("edited_message") or payload.get("business_message")
        perf_detail["kind"] = "callback" if callback else "message" if message else "unknown"
        if not message and not callback:
            status = "ignored"
            perf_detail["reason"] = "no_supported_update"
            runtime_log("TELEGRAM", f"webhook ignored update_id={update_id} reason=no_supported_update")
            return
        with SessionLocal() as session:
            session_opened = time_module.perf_counter()
            with TELEGRAM_UPDATE_LOCK:
                if not mark_telegram_update_seen(session, update_id):
                    session.commit()
                    status = "duplicate"
                    runtime_log("TELEGRAM", f"webhook duplicate skipped update_id={update_id}")
                    return
                session.commit()
            dedupe_finished = time_module.perf_counter()
            if callback:
                if not handle_payment_situation_callback_query(session, callback):
                    handle_agent_callback_query(session, callback)
            else:
                handle_telegram_message(session, message)
            handler_finished = time_module.perf_counter()
            session.commit()
            finished = time_module.perf_counter()
            total_ms = int((finished - (received_at or started)) * 1000)
            perf_detail.update(
                {
                    "session_ms": int((session_opened - started) * 1000),
                    "dedupe_ms": int((dedupe_finished - session_opened) * 1000),
                    "handler_ms": int((handler_finished - dedupe_finished) * 1000),
                    "commit_ms": int((finished - handler_finished) * 1000),
                    "total_ms": total_ms,
                }
            )
            runtime_log(
                "TELEGRAM",
                (
                    f"timing update_id={update_id} queue_ms={queue_ms} "
                    f"session_ms={int((session_opened - started) * 1000)} "
                    f"dedupe_ms={int((dedupe_finished - session_opened) * 1000)} "
                    f"handler_ms={int((handler_finished - dedupe_finished) * 1000)} "
                    f"commit_ms={int((finished - handler_finished) * 1000)} "
                    f"total_ms={total_ms}"
                ),
            )
    except Exception as exc:
        status = "failed"
        elapsed_ms = int((time_module.perf_counter() - (received_at or started)) * 1000)
        perf_detail["error"] = repr(exc)[:240]
        runtime_log("TELEGRAM", f"background failed update_id={update_id} total_ms={elapsed_ms} error={exc!r}")
    finally:
        record_background_event(
            "telegram_update",
            int((time_module.perf_counter() - (received_at or started)) * 1000),
            status=status,
            detail=perf_detail,
        )


def telegram_worker_loop() -> None:
    while not TELEGRAM_WORKER_STOP.is_set():
        try:
            payload, received_at = TELEGRAM_UPDATE_QUEUE.get(timeout=0.5)
        except queue.Empty:
            continue
        try:
            process_telegram_update_background(payload, received_at)
        finally:
            TELEGRAM_UPDATE_QUEUE.task_done()


def start_telegram_workers() -> None:
    if TELEGRAM_WORKERS:
        return
    TELEGRAM_WORKER_STOP.clear()
    for index in range(TELEGRAM_WORKER_COUNT):
        worker = threading.Thread(target=telegram_worker_loop, name=f"telegram-worker-{index + 1}", daemon=True)
        worker.start()
        TELEGRAM_WORKERS.append(worker)


def stop_telegram_workers() -> None:
    TELEGRAM_WORKER_STOP.set()
    for worker in TELEGRAM_WORKERS:
        worker.join(timeout=2)
    TELEGRAM_WORKERS.clear()


def queue_telegram_update(payload: dict[str, Any], received_at: float | None = None) -> bool:
    try:
        TELEGRAM_UPDATE_QUEUE.put_nowait((dict(payload), received_at))
        return True
    except queue.Full:
        return False


def telegram_webhook_rate_allowed(client_host: str, *, limit: int = 60, window_seconds: int = 60) -> bool:
    now = time_module.monotonic()
    cutoff = now - window_seconds
    key = hashlib.sha256(str(client_host or "unknown").encode("utf-8")).hexdigest()
    with TELEGRAM_RATE_LOCK:
        bucket = TELEGRAM_RATE_BUCKETS.setdefault(key, deque())
        while bucket and bucket[0] < cutoff:
            bucket.popleft()
        if len(bucket) >= limit:
            return False
        bucket.append(now)
        if len(TELEGRAM_RATE_BUCKETS) > 2048:
            stale = [item_key for item_key, values in TELEGRAM_RATE_BUCKETS.items() if not values or values[-1] < cutoff]
            for item_key in stale:
                TELEGRAM_RATE_BUCKETS.pop(item_key, None)
        return True


@app.post("/api/integrations/telegram/webhook")
async def telegram_webhook(request: Request, payload: dict[str, Any]) -> dict[str, bool]:
    received_at = time_module.perf_counter()
    update_types = ",".join(sorted(key for key in payload if key != "update_id")) or "-"
    runtime_log("TELEGRAM", f"webhook update_id={payload.get('update_id')} types={update_types}")
    client_host = request.client.host if request.client else "unknown"
    if not telegram_webhook_rate_allowed(client_host):
        raise HTTPException(429, "Слишком много webhook-запросов.")
    with SessionLocal() as settings_session:
        secret = str(os.environ.get("TELEGRAM_WEBHOOK_SECRET") or telegram_secret(settings_session) or "").strip()
    if production_mode() and not secret:
        runtime_log("TELEGRAM", f"webhook rejected update_id={payload.get('update_id')} reason=secret_missing")
        raise HTTPException(503, "Telegram webhook secret не настроен")
    provided = request.headers.get("X-Telegram-Bot-Api-Secret-Token", "")
    if secret and not hmac.compare_digest(provided, secret):
        runtime_log("TELEGRAM", f"webhook rejected update_id={payload.get('update_id')} reason=bad_secret")
        raise HTTPException(403, "Неверный Telegram secret token")
    if not queue_telegram_update(payload, received_at):
        raise HTTPException(503, "Очередь Telegram временно заполнена")
    return {"ok": True}


def ensure_runtime_telegram_webhook(session: Session) -> None:
    token = telegram_token(session)
    base_url = app_base_url(session)
    if not token or not base_url.startswith("https://"):
        return
    payload: dict[str, Any] = {
        "url": f"{base_url}/api/integrations/telegram/webhook",
        "allowed_updates": ["message", "edited_message", "business_message", "callback_query"],
    }
    secret = telegram_secret(session)
    if secret:
        payload["secret_token"] = secret
    try:
        result = telegram_api_request(token, "setWebhook", payload)
        runtime_log("TELEGRAM", f"runtime webhook refresh ok={bool(result.get('ok'))}")
    except TelegramApiError as exc:
        runtime_log("TELEGRAM", f"runtime webhook refresh skipped error={exc}")
    try:
        commands_result = telegram_api_request(token, "setMyCommands", {"commands": owner_commands()})
        runtime_log("TELEGRAM", f"runtime commands refresh ok={bool(commands_result.get('ok'))}")
    except TelegramApiError as exc:
        runtime_log("TELEGRAM", f"runtime commands refresh skipped error={exc}")


def refresh_runtime_telegram_webhook_background() -> None:
    started = time_module.perf_counter()
    status = "ok"
    detail: dict[str, Any] = {}
    try:
        with SessionLocal() as session:
            token_configured = bool(telegram_token(session))
            base_url = app_base_url(session)
            detail = {
                "token_configured": token_configured,
                "base_url_configured": bool(base_url),
                "https_base_url": base_url.startswith("https://") if base_url else False,
            }
            ensure_runtime_telegram_webhook(session)
    except Exception as exc:
        status = "failed"
        detail["error"] = repr(exc)[:240]
        runtime_log("TELEGRAM", f"runtime webhook background failed error={exc!r}")
    finally:
        record_background_event(
            "telegram_webhook_refresh",
            int((time_module.perf_counter() - started) * 1000),
            status=status,
            detail=detail,
        )


def queue_runtime_telegram_webhook_refresh() -> None:
    thread = threading.Thread(
        target=refresh_runtime_telegram_webhook_background,
        name="telegram-webhook-refresh",
        daemon=True,
    )
    thread.start()


@app.post("/api/integrations/telegram/set-webhook")
def telegram_set_webhook(payload: dict[str, Any] | None = None, session: Session = Depends(get_session)) -> dict[str, Any]:
    token = telegram_token(session)
    base_url = save_app_base_url_if_changed(session, (payload or {}).get("app_base_url"))
    if not token:
        raise HTTPException(400, "Сначала сохрани Telegram bot token")
    if not base_url.startswith("https://"):
        raise HTTPException(400, "Для webhook нужен публичный HTTPS URL в app_base_url")
    payload: dict[str, Any] = {
        "url": f"{base_url}/api/integrations/telegram/webhook",
        "allowed_updates": ["message", "edited_message", "business_message", "callback_query"],
        "drop_pending_updates": True,
    }
    secret = telegram_secret(session)
    if secret:
        payload["secret_token"] = secret
    delete_warning = ""
    commands_warning = ""
    commands_result: dict[str, Any] = {"ok": False}
    try:
        try:
            telegram_api_request(token, "deleteWebhook", {"drop_pending_updates": True})
        except TelegramApiError as exc:
            delete_warning = str(exc)
            runtime_log("TELEGRAM", f"delete webhook skipped error={exc}")
        webhook_result = telegram_api_request(token, "setWebhook", payload)
        try:
            commands_result = telegram_api_request(token, "setMyCommands", {"commands": owner_commands()})
        except TelegramApiError as exc:
            commands_warning = str(exc)
            runtime_log("TELEGRAM", f"set commands skipped error={exc}")
        session.commit()
        return {
            "ok": bool(webhook_result.get("ok")),
            "description": webhook_result.get("description", ""),
            "commands_configured": bool(commands_result.get("ok")),
            "url": payload["url"],
            "delete_warning": delete_warning,
            "commands_warning": commands_warning,
        }
    except TelegramApiError as exc:
        runtime_log("TELEGRAM", f"set webhook failed error={exc}")
        raise HTTPException(400, str(exc)) from exc


@app.get("/api/integrations/telegram/webhook-info")
def telegram_webhook_info(session: Session = Depends(get_session)) -> dict[str, Any]:
    token = telegram_token(session)
    if not token:
        raise HTTPException(400, "Сначала сохрани Telegram bot token")
    try:
        response = telegram_api_request(token, "getWebhookInfo", {})
    except TelegramApiError as exc:
        runtime_log("TELEGRAM", f"webhook info failed error={exc}")
        return {
            "ok": False,
            "url": "",
            "expected_url": f"{app_base_url(session)}/api/integrations/telegram/webhook" if app_base_url(session) else "",
            "matches_expected": False,
            "pending_update_count": 0,
            "last_error_date": None,
            "last_error_message": str(exc),
            "max_connections": None,
            "allowed_updates": [],
        }
    info = response.get("result") or {}
    expected_url = ""
    base_url = app_base_url(session)
    if base_url:
        expected_url = f"{base_url}/api/integrations/telegram/webhook"
    actual_url = str(info.get("url") or "")
    return {
        "ok": bool(response.get("ok")),
        "url": actual_url,
        "expected_url": expected_url,
        "matches_expected": bool(expected_url and actual_url == expected_url),
        "pending_update_count": info.get("pending_update_count") or 0,
        "last_error_date": info.get("last_error_date"),
        "last_error_message": info.get("last_error_message") or "",
        "max_connections": info.get("max_connections"),
        "allowed_updates": info.get("allowed_updates") or [],
    }


@app.post("/api/integrations/telegram/send-test")
def telegram_send_test(session: Session = Depends(get_session)) -> dict[str, Any]:
    owner_id = telegram_owner_chat_id(session)
    if not owner_id:
        raise HTTPException(400, "Сначала сохрани Telegram owner chat id")
    dashboard = build_dashboard(session)
    send_telegram_text(
        session,
        owner_id,
        build_status_message(dashboard),
        app_keyboard(app_base_url(session), dashboard.get("monthly_reports")),
    )
    return {"ok": True}


def message_targets_payload(session: Session) -> list[dict[str, Any]]:
    leases = session.scalars(
        select(Lease).join(Apartment).where(Lease.active.is_(True), Apartment.active.is_(True)).order_by(Lease.start_date.desc())
    ).all()
    return [serialize_message_target(session, lease) for lease in leases if not lease_ignored(session, lease.id)]


@app.get("/api/messages/targets")
def message_targets(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    return message_targets_payload(session)


@app.post("/api/messages/preview")
def preview_message_to_tenant(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    lease, template_key, charge, line, custom_text = resolve_message_request(session, payload)
    text = render_message_text(session, template_key, lease, charge, line, custom_text)
    if not text.strip():
        raise HTTPException(400, "Текст сообщения пустой")
    return {
        "lease_id": lease.id,
        "tenant": lease.tenant.full_name,
        "apartment": lease.apartment.name,
        "object": lease.apartment.object.name,
        "template_key": template_key,
        "template_label": reminder_label(template_key),
        "linked": tenant_chat_linked(session, lease),
        "text": text,
        "payload": {
            "lease_id": lease.id,
            "template_key": template_key,
            "charge_id": charge.id if charge else None,
            "utility_line_id": line.id if line else None,
            "custom_text": custom_text,
        },
    }


@app.post("/api/messages/send")
def send_message_to_tenant(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    lease, template_key, charge, line, custom_text = resolve_message_request(session, payload)
    result = send_tenant_message(
        session,
        lease,
        template_key,
        charge=charge,
        line=line,
        custom_text=custom_text,
    )
    session.commit()
    return result


@app.get("/api/bot-dialogs")
def list_bot_dialogs(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    return bot_dialogs_payload(session)


@app.get("/api/bot-dialogs/{dialog_id}/messages")
def get_bot_dialog_messages(
    dialog_id: str,
    limit: int = BOT_DIALOG_HISTORY_LIMIT,
    session: Session = Depends(get_session),
) -> dict[str, Any]:
    safe_limit = max(20, min(int(limit or BOT_DIALOG_HISTORY_LIMIT), BOT_DIALOG_HISTORY_LIMIT))
    return bot_dialog_messages_payload(session, dialog_id, safe_limit)


@app.post("/api/bot-dialogs/{dialog_id}/send")
def send_bot_dialog_message(dialog_id: str, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    target = resolve_bot_dialog_target(session, dialog_id)
    lease: Lease | None = target["lease"]
    chat_id = str(target["chat_id"] or "").strip()
    text = str(payload.get("text") or "").strip()
    if not text:
        raise HTTPException(400, "Введите текст сообщения")
    if len(text) > 4000:
        raise HTTPException(400, "Telegram не любит сообщения длиннее 4000 символов. Тут без романа, пожалуйста.")
    if not chat_id:
        raise HTTPException(400, "Этот диалог ещё не привязан к Telegram-чату")

    log = MessageLog(
        lease_id=lease.id if lease else None,
        channel="telegram",
        template_key="direct",
        status="sent",
        recipient_chat_id=chat_id,
        text=text,
        note="web-dialog",
    )
    try:
        send_telegram_text(session, chat_id, text)
    except HTTPException as exc:
        log.status = "failed"
        log.note = f"web-dialog: {exc.detail}"
        session.add(log)
        session.commit()
        raise
    session.add(log)
    session.commit()
    return {"ok": True, "message": serialize_bot_dialog_log_message(session, log)}


@app.post("/api/reminders/run")
def run_reminders_now(session: Session = Depends(get_session)) -> dict[str, Any]:
    summary = run_due_reminders(session)
    session.commit()
    return summary


@app.get("/api/objects")
def list_objects(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    return [serialize_object(obj) for obj in session.scalars(select(RentalObject).order_by(RentalObject.name)).all()]


@app.post("/api/objects")
def create_object(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    obj = RentalObject(
        name=(payload.get("name") or "").strip(),
        short_code=(payload.get("short_code") or "").strip(),
        notes=payload.get("notes") or "",
    )
    if not obj.name:
        raise HTTPException(400, "Название объекта обязательно")
    session.add(obj)
    session.commit()
    session.refresh(obj)
    return serialize_object(obj)


@app.post("/api/apartments")
def create_apartment(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    apartment = Apartment(
        object_id=int(payload["object_id"]),
        name=(payload.get("name") or "").strip(),
        sort_order=int(payload.get("sort_order") or 0),
        odn_share_percent=float(payload.get("odn_share_percent") or 0),
        active=payload.get("active", True) not in {False, "false", "0", 0},
    )
    if not apartment.name:
        raise HTTPException(400, "Название квартиры обязательно")
    session.add(apartment)
    session.commit()
    session.refresh(apartment)
    return serialize_apartment(apartment)


@app.patch("/api/apartments/{apartment_id}")
def update_apartment(apartment_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    apartment = session.get(Apartment, apartment_id)
    if not apartment:
        raise HTTPException(404, "Квартира не найдена")
    if "name" in payload:
        apartment.name = (payload.get("name") or apartment.name).strip()
    if "sort_order" in payload:
        apartment.sort_order = int(payload.get("sort_order") or 0)
    if "odn_share_percent" in payload:
        apartment.odn_share_percent = float(payload.get("odn_share_percent") or 0)
    if "active" in payload:
        apartment.active = payload.get("active") not in {False, "false", "0", 0}
    session.flush()
    generate_rent_charges(session)
    session.commit()
    session.refresh(apartment)
    return serialize_apartment(apartment)


@app.patch("/api/apartments/{apartment_id}/utility-advance")
def update_apartment_utility_advance(apartment_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    apartment = session.get(Apartment, apartment_id)
    if not apartment:
        raise HTTPException(404, "Квартира не найдена")
    raw_amount = payload.get("amount_override")
    amount_override = None
    if raw_amount not in {None, ""}:
        amount_override = money(float(raw_amount))
        if amount_override < 0:
            raise HTTPException(400, "аванс не может быть отрицательным")
    note = (payload.get("note") or "").strip()
    setting = utility_advance_setting_for_apartment(session, apartment.id)
    old_amount = setting.amount_override if setting else None
    if not setting:
        setting = UtilityAdvanceSetting(apartment_id=apartment.id)
        session.add(setting)
    setting.amount_override = amount_override
    setting.note = note
    setting.updated_at = utc_now()
    session.add(
        UtilityAdvanceSettingHistory(
            apartment_id=apartment.id,
            old_amount=old_amount,
            new_amount=amount_override,
            actor="owner",
            note=note,
        )
    )
    session.commit()
    session.refresh(apartment)
    return serialize_apartment(apartment)


@app.get("/api/tenants")
def list_tenants(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    return [serialize_tenant(tenant) for tenant in session.scalars(select(Tenant).order_by(Tenant.active.desc(), Tenant.full_name)).all()]


@app.get("/api/leases")
def list_leases(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    return [serialize_lease(lease, session) for lease in session.scalars(select(Lease).order_by(Lease.active.desc(), Lease.start_date.desc())).all()]


@app.post("/api/leases/onboard")
def onboard_tenant(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    apartment_id = int(payload.get("apartment_id") or 0)
    if not apartment_id:
        raise HTTPException(400, "Выберите свободную квартиру")
    apartment = session.get(Apartment, apartment_id)
    if not apartment:
        raise HTTPException(404, "Квартира не найдена")
    active_lease = next(
        (
            lease
            for lease in session.scalars(select(Lease).where(Lease.apartment_id == apartment.id, Lease.active.is_(True))).all()
            if not lease_ignored(session, lease.id)
        ),
        None,
    )
    if active_lease:
        raise HTTPException(400, "В квартире уже есть активный жилец. Сначала оформите выезд.")

    start = parse_date(payload.get("start_date"), date.today())
    payment_day = int(payload.get("payment_day") or start.day)
    tenant = Tenant(
        full_name=(payload.get("full_name") or "Без имени").strip(),
        phone=payload.get("phone") or "",
        telegram=payload.get("telegram") or "",
        whatsapp=payload.get("whatsapp") or "",
        notes=payload.get("tenant_notes") or "",
    )
    session.add(tenant)
    session.flush()
    lease = Lease(
        apartment_id=apartment.id,
        tenant_id=tenant.id,
        start_date=start,
        payment_day=payment_day,
        ip_amount=float(payload.get("ip_amount") or 0),
        personal_amount=float(payload.get("personal_amount") or 0),
        deposit_amount=float(payload.get("deposit_amount") or 0),
        deposit_location=payload.get("deposit_location") or "",
        deposit_terms=payload.get("deposit_terms") or "",
        notes=payload.get("notes") or "",
    )
    session.add(lease)
    session.flush()
    if "ignored" in payload:
        set_lease_ignored(session, lease.id, setting_bool_value(payload.get("ignored")))
    generate_rent_charges(session)
    session.commit()
    session.refresh(lease)
    return serialize_lease(lease, session)


@app.patch("/api/leases/{lease_id}")
def update_lease(lease_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    lease = session.get(Lease, lease_id)
    if not lease:
        raise HTTPException(404, "Договор не найден")

    was_ignored = lease_ignored(session, lease.id)
    apartment_id = int(payload.get("apartment_id") or lease.apartment_id)
    apartment = session.get(Apartment, apartment_id)
    if not apartment:
        raise HTTPException(404, "Квартира не найдена")
    if apartment.id != lease.apartment_id:
        active_lease = session.scalar(select(Lease).where(Lease.apartment_id == apartment.id, Lease.active.is_(True), Lease.id != lease.id))
        if active_lease:
            raise HTTPException(400, "В квартире уже есть активный жилец. Сначала оформите выезд.")
        lease.apartment_id = apartment.id

    start = parse_date(payload.get("start_date"), lease.start_date)
    payment_day = int(payload.get("payment_day") or start.day)

    lease.start_date = start
    lease.payment_day = payment_day
    lease.ip_amount = float(payload.get("ip_amount") or 0)
    lease.personal_amount = float(payload.get("personal_amount") or 0)
    lease.deposit_amount = float(payload.get("deposit_amount") or 0)
    lease.deposit_location = payload.get("deposit_location") or ""
    lease.deposit_terms = payload.get("deposit_terms") or ""
    lease.notes = payload.get("notes") or ""
    if was_ignored and IGNORE_LEASE_MARK not in lease.notes:
        lease.notes = (lease.notes + "\n" + IGNORE_LEASE_MARK).strip()
    if "end_date" in payload:
        raw_end = str(payload.get("end_date") or "").strip()
        lease.end_date = parse_date(raw_end) if raw_end else None
        lease.active = not (lease.end_date and lease.end_date <= date.today())
        lease.tenant.active = lease.active
    if "ignored" in payload:
        set_lease_ignored(session, lease.id, setting_bool_value(payload.get("ignored")))

    lease.tenant.full_name = (payload.get("full_name") or "Без имени").strip()
    lease.tenant.phone = payload.get("phone") or ""
    lease.tenant.telegram = payload.get("telegram") or ""
    lease.tenant.whatsapp = payload.get("whatsapp") or ""
    tenant_notes = payload.get("tenant_notes")
    if tenant_notes is not None:
        lease.tenant.notes = tenant_notes

    session.flush()
    generate_rent_charges(session)
    session.commit()
    session.refresh(lease)
    return serialize_lease(lease, session)


def prune_unpaid_rent_charges_after(session: Session, lease: Lease, end_date: date) -> None:
    charges = session.scalars(
        select(RentCharge)
        .where(
            RentCharge.lease_id == lease.id,
            RentCharge.due_date > end_date,
        )
        .order_by(RentCharge.due_date, RentCharge.id)
    ).all()
    for charge in charges:
        accepted_receipts = [receipt for receipt in charge.receipts if receipt.status == "accepted"]
        if accepted_receipts:
            continue
        for receipt in charge.receipts:
            receipt.rent_charge_id = None
        for log in session.scalars(select(MessageLog).where(MessageLog.rent_charge_id == charge.id)).all():
            log.rent_charge_id = None
        session.delete(charge)


def copy_lease_automation(session: Session, source_lease_id: int, target_lease_id: int) -> None:
    for template_key in LEASE_AUTOMATION_TEMPLATES:
        cadence = get_lease_cadence(session, source_lease_id, template_key)
        if cadence:
            set_lease_cadence(session, target_lease_id, template_key, cadence)


def parse_transfer_apartment_id(payload: dict[str, Any]) -> int:
    raw = payload.get("apartment_id") or payload.get("target_apartment_id") or 0
    try:
        return int(raw)
    except (TypeError, ValueError) as exc:
        raise HTTPException(400, "Выберите квартиру для переезда") from exc


def parse_transfer_money(payload: dict[str, Any], key: str, fallback: float) -> float:
    raw = payload.get(key)
    if raw in {None, ""}:
        return money(fallback)
    normalized = str(raw).replace("\u00a0", "").replace(" ", "").replace(",", ".").strip()
    try:
        value = float(normalized)
    except (TypeError, ValueError) as exc:
        raise HTTPException(400, "Стоимость аренды должна быть числом") from exc
    if value < 0:
        raise HTTPException(400, "Стоимость аренды не может быть отрицательной")
    return money(value)


@app.post("/api/leases/{lease_id}/transfer")
def transfer_lease(lease_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    lease = session.get(Lease, lease_id)
    if not lease:
        raise HTTPException(404, "Договор не найден")
    if not lease.active:
        raise HTTPException(400, "Переезд можно оформить только для активного договора")

    target_apartment_id = parse_transfer_apartment_id(payload)
    if not target_apartment_id:
        raise HTTPException(400, "Выберите квартиру для переезда")
    target_apartment = session.get(Apartment, target_apartment_id)
    if not target_apartment:
        raise HTTPException(404, "Квартира для переезда не найдена")
    if not target_apartment.active:
        raise HTTPException(400, "Квартира выключена из учёта")
    if target_apartment.id == lease.apartment_id:
        raise HTTPException(400, "Новая квартира совпадает с текущей.")

    try:
        transfer_date = parse_date(payload.get("transfer_date") or payload.get("move_date"), date.today())
    except ValueError as exc:
        raise HTTPException(400, "Дата переезда должна быть в формате ГГГГ-ММ-ДД") from exc
    if transfer_date <= lease.start_date:
        raise HTTPException(400, "Дата переезда должна быть позже даты текущего заезда")
    old_end = transfer_date - timedelta(days=1)

    target_conflict = next(
        (
            item
            for item in session.scalars(
                select(Lease)
                .where(
                    Lease.apartment_id == target_apartment.id,
                    Lease.active.is_(True),
                    Lease.id != lease.id,
                    or_(Lease.end_date.is_(None), Lease.end_date >= transfer_date),
                )
                .order_by(Lease.start_date)
            ).all()
            if not lease_ignored(session, item.id)
        ),
        None,
    )
    if target_conflict:
        raise HTTPException(400, "В выбранной квартире есть активный жилец на дату переезда")

    existing_transfer = session.scalar(
        select(Lease)
        .where(
            Lease.tenant_id == lease.tenant_id,
            Lease.apartment_id == target_apartment.id,
            Lease.start_date == transfer_date,
        )
        .limit(1)
    )
    if existing_transfer:
        raise HTTPException(400, "Переезд на эту дату уже оформлен")

    ip_amount = parse_transfer_money(payload, "ip_amount", float(lease.ip_amount or 0))
    personal_amount = parse_transfer_money(payload, "personal_amount", float(lease.personal_amount or 0))

    generate_rent_charges(session, until=old_end)
    old_notes = [lease.notes or "", f"Переезд в {target_apartment.name} с {transfer_date:%d.%m.%Y}."]
    lease.end_date = old_end
    lease.active = old_end >= date.today()
    lease.notes = "\n".join(part for part in old_notes if part).strip()
    prune_unpaid_rent_charges_after(session, lease, old_end)

    new_lease = Lease(
        apartment_id=target_apartment.id,
        tenant_id=lease.tenant_id,
        start_date=transfer_date,
        payment_day=lease.payment_day,
        ip_amount=money(ip_amount),
        personal_amount=money(personal_amount),
        deposit_amount=lease.deposit_amount,
        deposit_location=lease.deposit_location,
        deposit_terms=lease.deposit_terms,
        notes=f"Переезд из {lease.apartment.name} с {transfer_date:%d.%m.%Y}.",
        active=True,
    )
    session.add(new_lease)
    session.flush()
    copy_lease_automation(session, lease.id, new_lease.id)
    if lease_ignored(session, lease.id):
        set_lease_ignored(session, new_lease.id, True)
    lease.tenant.active = True
    generate_rent_charges(session)
    session.commit()
    session.refresh(lease)
    session.refresh(new_lease)
    return {
        "old_lease": serialize_lease(lease, session),
        "new_lease": serialize_lease(new_lease, session),
        "summary": {
            "transfer_date": transfer_date.isoformat(),
            "old_end_date": old_end.isoformat(),
            "target_apartment": target_apartment.name,
            "ip_amount": money(ip_amount),
            "personal_amount": money(personal_amount),
        },
    }


@app.delete("/api/leases/{lease_id}")
def delete_lease(lease_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    lease = session.get(Lease, lease_id)
    if not lease:
        raise HTTPException(404, "Договор не найден")
    tenant = lease.tenant
    for state in session.scalars(select(AgentTenantState).where(AgentTenantState.lease_id == lease.id)).all():
        session.delete(state)
    for situation in session.scalars(select(PaymentSituation).where(PaymentSituation.lease_id == lease.id)).all():
        session.delete(situation)
    for memory in session.scalars(select(AgentMemory).where(AgentMemory.lease_id == lease.id)).all():
        session.delete(memory)
    for proposal in session.scalars(select(AgentActionProposal).where(AgentActionProposal.lease_id == lease.id)).all():
        proposal.lease_id = None
        if proposal.status == "pending":
            proposal.status = "cancelled"
            proposal.result_text = "Договор удалён до подтверждения действия"
    for task in session.scalars(select(AgentTask).where(AgentTask.lease_id == lease.id)).all():
        task.lease_id = None
        task.status = "resolved"
        task.resolved_at = utc_now()
    for ai_message in session.scalars(select(AiMessage).where(AiMessage.lease_id == lease.id)).all():
        ai_message.lease_id = None
    for action_log in session.scalars(select(AiActionLog).where(AiActionLog.lease_id == lease.id)).all():
        action_log.lease_id = None
    for conversation in session.scalars(
        select(AiConversation).where(
            or_(AiConversation.lease_id == lease.id, AiConversation.tenant_id == lease.tenant_id)
        )
    ).all():
        conversation.lease_id = None
        conversation.tenant_id = None
    session.execute(delete(MessageLog).where(MessageLog.lease_id == lease.id))
    for receipt in session.scalars(select(PaymentReceipt).where(PaymentReceipt.lease_id == lease.id)).all():
        linked_expense = linked_expense_fund_record(session, receipt.id)
        if linked_expense:
            session.delete(linked_expense)
        session.delete(receipt)
    for line in session.scalars(select(UtilityBillLine).where(UtilityBillLine.lease_id == lease.id)).all():
        line.lease_id = None
        line.status = "not_required"
    session.execute(delete(RentCharge).where(RentCharge.lease_id == lease.id))
    set_lease_ignored(session, lease.id, False)
    session.delete(lease)
    session.flush()
    if tenant and not session.scalar(select(Lease.id).where(Lease.tenant_id == tenant.id).limit(1)):
        session.delete(tenant)
    session.commit()
    return {"ok": True}


@app.post("/api/leases/{lease_id}/move-out")
def move_out(lease_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    lease = session.get(Lease, lease_id)
    if not lease:
        raise HTTPException(404, "Договор не найден")
    end = parse_date(payload.get("end_date"), date.today())
    lease.end_date = end
    lease.active = False
    lease.tenant.active = False
    if payload.get("notes"):
        lease.notes = (lease.notes + "\n" + payload["notes"]).strip()

    for charge in lease.rent_charges:
        update_rent_charge_status(charge)

    move_out_summary = {"processed": 0, "created_lines": 0, "tenant_sent": 0, "owner_sent": 0}
    if end == date.today():
        move_out_summary = process_move_out_notifications(session, end)

    cutoff = configured_notification_cutoff_date(session)
    rent_debt = sum(
        max(0, charge.ip_due - charge.ip_paid) + max(0, charge.personal_due - charge.personal_paid)
        for charge in lease.rent_charges
        if charge.period_start <= end
        and charge.status not in {"paid", "paid_ahead"}
        and debt_visible_by_cutoff(charge.due_date, cutoff)
    )
    utility_lines = session.scalars(select(UtilityBillLine).where(UtilityBillLine.lease_id == lease.id)).all()
    utility_debt = sum(max(0, line.total_amount - line.paid_amount) for line in utility_lines if line.status != "paid")
    paid_charges = [charge for charge in lease.rent_charges if charge.status in {"paid", "paid_ahead"}]
    last_paid_day = max((charge.period_end for charge in paid_charges), default=lease.start_date - timedelta(days=1))

    session.commit()
    return {
        "lease": serialize_lease(lease, session),
        "summary": {
            "full_months_lived": full_months_lived(lease, end),
            "last_paid_day": last_paid_day.isoformat(),
            "rent_debt": money(rent_debt),
            "utility_debt": money(utility_debt),
            "deposit_amount": lease.deposit_amount,
            "deposit_location": lease.deposit_location,
            "deposit_terms": lease.deposit_terms,
            "final_total_due": money(rent_debt + utility_debt),
            "move_out_utility_lines_created": move_out_summary["created_lines"],
        },
    }


def rent_charges_payload(
    start: str | None = None,
    end: str | None = None,
    session: Session | None = None,
    *,
    include_payments: bool = True,
    include_reminder: bool = True,
    generate_missing: bool = True,
    limit: int | None = None,
    offset: int = 0,
) -> list[dict[str, Any]]:
    if session is None:
        return []
    if generate_missing:
        generate_rent_charges(session)
    start_date, end_date = current_month_range()
    if start:
        start_date = parse_date(start)
    if end:
        end_date = parse_date(end)
    query = (
        select(RentCharge)
        .options(
            joinedload(RentCharge.lease).joinedload(Lease.apartment).joinedload(Apartment.object),
            joinedload(RentCharge.lease).joinedload(Lease.tenant),
        )
        .join(Lease)
        .join(Apartment)
        .where(Lease.active.is_(True), Apartment.active.is_(True))
        .where(RentCharge.due_date >= start_date, RentCharge.due_date <= end_date)
        .order_by(RentCharge.due_date, RentCharge.id)
    )
    if include_payments:
        query = query.options(selectinload(RentCharge.receipts))
    if offset:
        query = query.offset(max(0, int(offset)))
    if limit is not None:
        query = query.limit(max(1, min(int(limit), 500)))
    charges = session.scalars(query).all()
    charges = [charge for charge in charges if not lease_ignored(session, charge.lease_id)]
    if include_reminder:
        prefetch_latest_message_logs(session, rent_charge_ids=[charge.id for charge in charges])
    return [
        serialize_rent_charge(charge, session, include_payments=include_payments, include_reminder=include_reminder)
        for charge in charges
    ]


@app.get("/api/rent-charges")
def list_rent_charges(
    start: str | None = None,
    end: str | None = None,
    limit: int | None = None,
    offset: int = 0,
    session: Session = Depends(get_session),
) -> list[dict[str, Any]]:
    result = rent_charges_payload(
        start=start,
        end=end,
        session=session,
        include_payments=False,
        include_reminder=False,
        generate_missing=True,
        limit=limit,
        offset=offset,
    )
    session.commit()
    return result


@app.post("/api/rent-charges/generate")
def generate_charges(payload: dict[str, Any] | None = None, session: Session = Depends(get_session)) -> dict[str, Any]:
    until = parse_date((payload or {}).get("until"), date.today() + timedelta(days=90))
    created = generate_rent_charges(session, until=until)
    session.commit()
    return {"created": created}


@app.post("/api/rent-charges/{charge_id}/payments")
def add_rent_payment(charge_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    charge = session.get(RentCharge, charge_id)
    if not charge:
        raise HTTPException(404, "начисление не найдено")
    channel = payload.get("channel")
    if channel not in {"ip", "personal", "expense_fund"}:
        raise HTTPException(400, "канал платежа должен быть ip, personal или expense_fund")
    amount = float(payload.get("amount") or 0)
    if amount <= 0:
        raise HTTPException(400, "сумма платежа должна быть больше нуля")

    receipts = create_rent_receipts(
        session,
        charge.lease,
        channel,
        amount,
        paid_at=datetime.fromisoformat(payload["paid_at"]) if payload.get("paid_at") else utc_now(),
        source=payload.get("source") or "manual",
        status=payload.get("status") or "accepted",
        recipient_name=payload.get("recipient_name") or "",
        recipient_details=payload.get("recipient_details") or "",
        notes=payload.get("notes") or "",
        exact_only=False,
        cutoff=allocation_cutoff_date(session),
    )
    sync_expense_fund_receipts(session, receipts)
    generate_rent_charges(session)
    session.commit()
    session.refresh(charge)
    return serialize_rent_charge(charge, session)


@app.get("/api/leases/{lease_id}/payment-history")
def lease_payment_history(lease_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    lease = session.get(Lease, lease_id)
    if not lease:
        raise HTTPException(404, "аренда не найдена")
    receipts = session.scalars(
        select(PaymentReceipt)
        .where(PaymentReceipt.lease_id == lease_id)
        .order_by(PaymentReceipt.paid_at.desc(), PaymentReceipt.id.desc())
    ).all()
    return {
        "lease_id": lease.id,
        "tenant": lease.tenant.full_name,
        "apartment": lease.apartment.name,
        "current_year": date.today().year,
        "targets": lease_payment_target_options(session, lease),
        "receipts": [serialize_payment_receipt(receipt, session) for receipt in receipts],
    }


def ensure_rent_charge_for_month(session: Session, lease: Lease, year: int, month: int) -> RentCharge:
    if month < 1 or month > 12:
        raise HTTPException(400, "Месяц зачёта должен быть от 1 до 12")
    start, end = month_range(year, month)
    if start < RENT_GENERATION_START:
        raise HTTPException(400, "Начисления раньше января 2025 не создаются")
    generate_rent_charges(session, until=end + timedelta(days=40))
    session.flush()
    charge = session.scalar(
        select(RentCharge)
        .where(
            RentCharge.lease_id == lease.id,
            RentCharge.due_date >= start,
            RentCharge.due_date <= end,
        )
        .order_by(RentCharge.due_date, RentCharge.id)
        .limit(1)
    )
    if not charge:
        raise HTTPException(400, f"Для {MONTH_NAMES[month - 1]} {year} нет начисления аренды")
    return charge


def ensure_utility_line_target(session: Session, lease: Lease, line_id: int) -> UtilityBillLine:
    line = session.get(UtilityBillLine, line_id)
    if not line or line.lease_id != lease.id:
        raise HTTPException(404, "Коммунальный счёт не найден")
    return line


def apply_manual_debt_payload(debt: ManualDebt, lease: Lease, payload: dict[str, Any]) -> ManualDebt:
    today_value = date.today()
    existing = debt.id is not None
    kind = (payload.get("kind") or (debt.kind if existing else "other")).strip()
    if kind not in {"rent", "utility", "other"}:
        raise HTTPException(400, "назначение долга должно быть: rent, utility или other")
    channel = (payload.get("channel") or (debt.channel if existing else "")).strip()
    if kind == "rent" and channel not in {"ip", "personal", "expense_fund"}:
        raise HTTPException(400, "для арендного долга нужен канал ИП/по номеру/мне на расходы")

    raw_amount = payload.get("amount")
    amount = float(raw_amount if raw_amount not in {None, ""} else debt.amount if existing else 0)
    if amount <= 0:
        raise HTTPException(400, "сумма долга должна быть больше нуля")
    raw_paid = payload.get("paid_amount")
    paid_amount = float(raw_paid if raw_paid not in {None, ""} else debt.paid_amount if existing else 0)

    period_start = debt.period_start if existing else None
    period_end = debt.period_end if existing else None
    if "period_start" in payload:
        raw_start = str(payload.get("period_start") or "").strip()
        period_start = parse_date(raw_start) if raw_start else None
    if "period_end" in payload:
        raw_end = str(payload.get("period_end") or "").strip()
        period_end = parse_date(raw_end) if raw_end else None
    if kind == "rent":
        month = int(payload.get("target_month") or payload.get("month") or 0)
        year = int(payload.get("target_year") or payload.get("year") or today_value.year)
        if month:
            period_start, period_end = month_range(year, month)
    elif period_start and not period_end:
        period_end = period_start

    due_date = debt.due_date if existing else None
    if "due_date" in payload:
        raw_due = str(payload.get("due_date") or "").strip()
        due_date = parse_date(raw_due) if raw_due else None
    if due_date is None:
        due_date = period_end or period_start or today_value

    debt.lease_id = lease.id
    debt.apartment_id = lease.apartment_id
    debt.kind = kind
    debt.channel = channel if kind == "rent" else ""
    debt.title = (payload.get("title") if "title" in payload else debt.title) or manual_debt_kind_label(kind)
    debt.title = debt.title.strip()
    debt.period_start = period_start
    debt.period_end = period_end
    debt.due_date = due_date
    debt.amount = money(amount)
    debt.paid_amount = money(paid_amount)
    if "notes" in payload:
        debt.notes = (payload.get("notes") or "").strip()
    elif not existing:
        debt.notes = ""
    debt.active = True
    update_manual_debt_status(debt)
    return debt


@app.post("/api/payment-receipts/manual")
def create_manual_payment(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    lease = session.get(Lease, int(payload.get("lease_id") or 0))
    if not lease:
        raise HTTPException(404, "аренда не найдена")
    amount = float(payload.get("amount") or 0)
    if amount <= 0:
        raise HTTPException(400, "сумма платежа должна быть больше нуля")

    kind = (payload.get("kind") or "rent").strip()
    source = (payload.get("source") or "manual").strip() or "manual"
    paid_at = datetime.fromisoformat(payload["paid_at"]) if payload.get("paid_at") else utc_now()
    notes = (payload.get("notes") or "").strip()

    if kind == "rent":
        channel = (payload.get("channel") or "").strip()
        if channel not in {"ip", "personal", "expense_fund"}:
            raise HTTPException(400, "для аренды нужен канал ip, personal или expense_fund")
        target_charge_id = int(payload.get("rent_charge_id") or 0)
        target_month = int(payload.get("target_month") or 0)
        target_year = int(payload.get("target_year") or date.today().year)
        if target_charge_id:
            target_charge = session.get(RentCharge, target_charge_id)
            if not target_charge or target_charge.lease_id != lease.id:
                raise HTTPException(404, "Арендный месяц не найден")
            receipt = PaymentReceipt(
                lease_id=lease.id,
                rent_charge_id=target_charge.id,
                apartment_id=lease.apartment_id,
                amount=amount,
                channel=channel,
                paid_at=paid_at,
                source=source,
                status="accepted",
                notes=notes or "ручной платёж",
            )
            session.add(receipt)
            session.flush()
            recalculate_lease_balances(session, lease.id)
            sync_expense_fund_receipt(session, receipt)
            receipts = [receipt]
        elif target_month:
            target_charge = ensure_rent_charge_for_month(session, lease, target_year, target_month)
            receipt = PaymentReceipt(
                lease_id=lease.id,
                rent_charge_id=target_charge.id,
                apartment_id=lease.apartment_id,
                amount=amount,
                channel=channel,
                paid_at=paid_at,
                source=source,
                status="accepted",
                notes=notes or "ручной платёж",
            )
            session.add(receipt)
            session.flush()
            recalculate_lease_balances(session, lease.id)
            sync_expense_fund_receipt(session, receipt)
            receipts = [receipt]
        else:
            receipts = create_rent_receipts(
                session,
                lease,
                channel,
                amount,
                paid_at=paid_at,
                source=source,
                status="accepted",
                notes=notes or "ручной платёж",
                exact_only=False,
                cutoff=allocation_cutoff_date(session),
            )
            sync_expense_fund_receipts(session, receipts)
    elif kind == "utility":
        target_line_id = int(payload.get("utility_line_id") or 0)
        if target_line_id:
            target_line = ensure_utility_line_target(session, lease, target_line_id)
            receipt = PaymentReceipt(
                lease_id=lease.id,
                utility_line_id=target_line.id,
                apartment_id=lease.apartment_id,
                amount=amount,
                channel=UTILITY_ADVANCE_CHANNEL if utility_line_is_advance(target_line) else "utilities",
                paid_at=paid_at,
                source=source,
                status="accepted",
                notes=notes or ("оплата аванса коммуналки" if utility_line_is_advance(target_line) else "ручной платёж"),
            )
            session.add(receipt)
            session.flush()
            sync_utility_advance_credit_for_receipt(session, receipt)
            recalculate_lease_balances(session, lease.id)
            receipts = [receipt]
        else:
            receipts = create_utility_receipts(
                session,
                lease,
                amount,
                paid_at=paid_at,
                source=source,
                status="accepted",
                notes=notes or "ручной платёж",
                exact_only=False,
            )
            sync_utility_advance_credits_for_receipts(session, receipts)
    elif kind == UTILITY_ADVANCE_CHANNEL:
        receipt = PaymentReceipt(
            lease_id=lease.id,
            apartment_id=lease.apartment_id,
            amount=amount,
            channel=UTILITY_ADVANCE_CHANNEL,
            paid_at=paid_at,
            source=source,
            status="accepted",
            notes=notes or "аванс коммуналки",
        )
        session.add(receipt)
        session.flush()
        receipts = [receipt]
    else:
        raise HTTPException(400, "тип ручного платежа должен быть rent, utility или utility_advance")

    generate_rent_charges(session)
    session.commit()
    return {
        "ok": True,
        "lease_id": lease.id,
        "created": [serialize_payment_receipt(receipt, session) for receipt in receipts],
    }


@app.get("/api/leases/{lease_id}/manual-debts")
def list_manual_debts_for_lease(lease_id: int, session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    lease = session.get(Lease, lease_id)
    if not lease:
        raise HTTPException(404, "аренда не найдена")
    debts = session.scalars(
        select(ManualDebt)
        .where(ManualDebt.lease_id == lease.id, ManualDebt.active.is_(True))
        .order_by(ManualDebt.due_date.desc(), ManualDebt.created_at.desc(), ManualDebt.id.desc())
    ).all()
    return [serialize_manual_debt(debt, session) for debt in debts]


@app.post("/api/manual-debts")
def create_manual_debt(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    lease = session.get(Lease, int(payload.get("lease_id") or 0))
    if not lease:
        raise HTTPException(404, "аренда не найдена")
    debt = apply_manual_debt_payload(ManualDebt(), lease, payload)
    session.add(debt)
    session.commit()
    return serialize_manual_debt(debt, session)


@app.patch("/api/manual-debts/{debt_id}")
def update_manual_debt(debt_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    debt = session.get(ManualDebt, debt_id)
    if not debt or not debt.active:
        raise HTTPException(404, "ручной долг не найден")
    lease = session.get(Lease, int(payload.get("lease_id") or debt.lease_id))
    if not lease:
        raise HTTPException(404, "аренда не найдена")
    apply_manual_debt_payload(debt, lease, payload)
    session.commit()
    return serialize_manual_debt(debt, session)


@app.post("/api/manual-debts/{debt_id}/payments")
def add_manual_debt_payment(debt_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    debt = session.get(ManualDebt, debt_id)
    if not debt or not debt.active:
        raise HTTPException(404, "ручной долг не найден")
    amount = float(payload.get("amount") or 0)
    if amount <= 0:
        raise HTTPException(400, "сумма платежа должна быть больше нуля")
    note = (payload.get("notes") or "").strip()
    paid_at = payload.get("paid_at") or utc_now().strftime("%d.%m.%Y %H:%M")
    debt.paid_amount = money(float(debt.paid_amount or 0) + amount)
    if note:
        debt.notes = "; ".join(part for part in [debt.notes, f"оплата {money_text(amount)} от {paid_at}: {note}"] if part)
    else:
        debt.notes = "; ".join(part for part in [debt.notes, f"оплата {money_text(amount)} от {paid_at}"] if part)
    update_manual_debt_status(debt)
    session.commit()
    return serialize_manual_debt(debt, session)


@app.delete("/api/manual-debts/{debt_id}")
def delete_manual_debt(debt_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    debt = session.get(ManualDebt, debt_id)
    if not debt:
        raise HTTPException(404, "ручной долг не найден")
    debt.active = False
    update_manual_debt_status(debt)
    session.commit()
    return {"ok": True}


@app.patch("/api/payment-receipts/{receipt_id}")
def update_payment_receipt(receipt_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    receipt = session.get(PaymentReceipt, receipt_id)
    if not receipt:
        raise HTTPException(404, "платёж не найден")
    old_lease_id = receipt.lease_id
    if "amount" in payload:
        amount = float(payload.get("amount") or 0)
        if amount <= 0:
            raise HTTPException(400, "сумма платежа должна быть больше нуля")
        receipt.amount = amount
    if payload.get("paid_at"):
        receipt.paid_at = datetime.fromisoformat(payload["paid_at"])
    if "notes" in payload:
        receipt.notes = str(payload.get("notes") or "")
    target_kind = (payload.get("target_kind") or "").strip()
    if target_kind:
        if target_kind == "rent":
            target_charge = session.get(RentCharge, int(payload.get("rent_charge_id") or 0))
            if not target_charge:
                raise HTTPException(404, "Арендный месяц не найден")
            channel = (payload.get("channel") or receipt.channel or "ip").strip()
            if channel not in {"ip", "personal", "expense_fund"}:
                raise HTTPException(400, "для аренды канал платежа должен быть ip, personal или expense_fund")
            receipt.rent_charge_id = target_charge.id
            receipt.utility_line_id = None
            receipt.lease_id = target_charge.lease_id
            receipt.apartment_id = target_charge.lease.apartment_id
            receipt.channel = channel
        elif target_kind == "utility":
            target_line = session.get(UtilityBillLine, int(payload.get("utility_line_id") or 0))
            if not target_line or not target_line.lease_id:
                raise HTTPException(404, "Коммунальный счёт не найден")
            receipt.rent_charge_id = None
            receipt.utility_line_id = target_line.id
            receipt.lease_id = target_line.lease_id
            receipt.apartment_id = target_line.apartment_id
            receipt.channel = "utilities"
        else:
            raise HTTPException(400, "неизвестная цель зачёта")
    elif payload.get("channel") and receipt.rent_charge_id:
        channel = payload.get("channel")
        if channel not in {"ip", "personal", "expense_fund"}:
            raise HTTPException(400, "для аренды канал платежа должен быть ip, personal или expense_fund")
        receipt.channel = channel
    if not target_kind and receipt.rent_charge_id and ("target_month" in payload or "target_year" in payload):
        lease = receipt.rent_charge.lease if receipt.rent_charge else session.get(Lease, receipt.lease_id)
        if not lease:
            raise HTTPException(400, "Не удалось определить аренду для этого платежа")
        target_month = int(payload.get("target_month") or 0)
        target_year = int(payload.get("target_year") or date.today().year)
        if not target_month:
            raise HTTPException(400, "Нужно выбрать месяц зачёта")
        target_charge = ensure_rent_charge_for_month(session, lease, target_year, target_month)
        receipt.rent_charge_id = target_charge.id
        receipt.lease_id = lease.id
        receipt.utility_line_id = None
        receipt.apartment_id = lease.apartment_id
    requested_status = (payload.get("status") or "").strip()
    if requested_status:
        if requested_status != "accepted":
            raise HTTPException(400, "через редактор платёж можно только перевести в статус «зачтён»")
        if not receipt.rent_charge_id and not receipt.utility_line_id and receipt.channel != UTILITY_ADVANCE_CHANNEL:
            raise HTTPException(400, "сначала укажите, за что зачесть платёж")
        receipt.status = "accepted"
    session.flush()
    sync_expense_fund_receipt(session, receipt)
    sync_utility_advance_credit_for_receipt(session, receipt)
    if receipt.status == "accepted":
        affected_ids = {item for item in [old_lease_id, receipt.lease_id] if item}
        for lease_id in affected_ids:
            recalculate_lease_balances(session, int(lease_id))
    generate_rent_charges(session)
    session.commit()
    return serialize_payment_receipt(receipt, session)


@app.delete("/api/payment-receipts/{receipt_id}")
def delete_payment_receipt(receipt_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    receipt = session.get(PaymentReceipt, receipt_id)
    if not receipt:
        raise HTTPException(404, "платёж не найден")
    lease_id = receipt.lease_id
    status = receipt.status
    linked_expense = linked_expense_fund_record(session, receipt.id)
    if linked_expense:
        session.delete(linked_expense)
    for entry in session.scalars(select(UtilityAdvanceLedger).where(UtilityAdvanceLedger.payment_receipt_id == receipt.id)).all():
        session.delete(entry)
    session.delete(receipt)
    session.flush()
    if lease_id and status == "accepted":
        recalculate_lease_balances(session, lease_id)
    generate_rent_charges(session)
    session.commit()
    return {"ok": True}


def suspicious_receipts_payload(session: Session) -> list[dict[str, Any]]:
    receipts = session.scalars(
        select(PaymentReceipt)
        .options(
            joinedload(PaymentReceipt.rent_charge)
            .joinedload(RentCharge.lease)
            .joinedload(Lease.apartment)
            .joinedload(Apartment.object),
            joinedload(PaymentReceipt.rent_charge).joinedload(RentCharge.lease).joinedload(Lease.tenant),
        )
        .where(PaymentReceipt.status == "suspicious")
        .order_by(PaymentReceipt.paid_at.desc(), PaymentReceipt.id.desc())
    ).all()
    return [serialize_payment_receipt(receipt, session) for receipt in receipts]


@app.get("/api/payment-receipts/suspicious")
def suspicious_receipts(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    return suspicious_receipts_payload(session)


@app.post("/api/payment-receipts/{receipt_id}/ignore")
def ignore_payment_receipt(receipt_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    receipt = session.get(PaymentReceipt, receipt_id)
    if not receipt:
        raise HTTPException(404, "платёж не найден")
    receipt.status = "ignored"
    receipt.notes = "; ".join(part for part in [receipt.notes, "скрыто из дашборда вручную"] if part)
    session.commit()
    return serialize_payment_receipt(receipt, session)


@app.post("/api/payment-receipts/{receipt_id}/moderate")
def moderate_payment_receipt(receipt_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    receipt = session.get(PaymentReceipt, receipt_id)
    if not receipt:
        raise HTTPException(404, "платёж не найден")
    action = (payload.get("action") or "").strip()
    note = (payload.get("note") or "").strip()
    parsed = parse_receipt_details(receipt)
    channel_override = (payload.get("channel") or "").strip()
    channel = channel_override if channel_override in {"ip", "personal", "expense_fund"} else receipt.channel if receipt.channel in {"ip", "personal", "expense_fund"} else detect_receipt_channel(parsed)
    lease = session.get(Lease, receipt.lease_id) if receipt.lease_id else None
    if action in {"accept_rent", "accept_utility"} and not lease:
        raise HTTPException(400, "нужна аренда для этой операции")

    moderation_note = f"модерация owner{': ' + note if note else ''}"
    if action == "accept_rent":
        receipts = create_rent_receipts(
            session,
            lease,
            channel if channel in {"ip", "personal", "expense_fund"} else "personal",
            float(receipt.amount or 0),
            paid_at=receipt.paid_at,
            source=receipt.source,
            status="accepted",
            recipient_name=receipt.recipient_name,
            recipient_details=receipt.recipient_details,
            notes=moderation_note,
            file_path=receipt.file_path,
            exact_only=False,
            prefer_document_month=receipt.source == "telegram",
            cutoff=allocation_cutoff_date(session),
        )
        sync_expense_fund_receipts(session, receipts)
        receipt.status = "moderated"
    elif action == "accept_utility":
        receipts = create_utility_receipts(
            session,
            lease,
            float(receipt.amount or 0),
            paid_at=receipt.paid_at,
            source=receipt.source,
            status="accepted",
            recipient_name=receipt.recipient_name,
            recipient_details=receipt.recipient_details,
            notes=moderation_note,
            file_path=receipt.file_path,
            exact_only=False,
        )
        sync_utility_advance_credits_for_receipts(session, receipts)
        receipt.status = "moderated"
    elif action == "reject":
        receipt.status = "rejected"
    else:
        raise HTTPException(400, "неизвестное действие модерации")
    receipt.notes = "; ".join(part for part in [receipt.notes, moderation_note] if part)
    generate_rent_charges(session)
    session.commit()
    return serialize_payment_receipt(receipt, session)


@app.post("/api/rent-charges/{charge_id}/defer")
def defer_rent_charge(charge_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    charge = session.get(RentCharge, charge_id)
    if not charge:
        raise HTTPException(404, "Начисление не найдено")
    try:
        charge.deferral_until = resolve_deferral_until(payload)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    charge.deferral_note = payload.get("deferral_note") or ""
    update_rent_charge_status(charge)
    situation = payment_situation(session, "rent", charge.id, charge.lease_id)
    situation.promise_date = charge.deferral_until
    situation.paused_until = charge.deferral_until
    situation.status = "promised"
    session.commit()
    return serialize_rent_charge(charge, session)


def resolve_deferral_until(payload: dict[str, Any], today: date | None = None) -> date:
    today = today or date.today()
    raw_days = payload.get("deferral_days", payload.get("days"))
    if raw_days not in (None, ""):
        try:
            days = int(raw_days)
        except (TypeError, ValueError) as exc:
            raise ValueError("Количество дней отсрочки должно быть целым числом") from exc
        if days <= 0:
            raise ValueError("Количество дней отсрочки должно быть больше нуля")
        return today + timedelta(days=days)

    if payload.get("deferral_until"):
        return parse_date(payload.get("deferral_until"))

    raise ValueError("Укажите количество дней отсрочки")


@app.get("/api/meters")
def list_meters(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    return [serialize_meter(meter) for meter in session.scalars(select(Meter).order_by(Meter.object_id, Meter.scope, Meter.name)).all()]


@app.patch("/api/utility-services/{service_id}")
def update_utility_service(service_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    service = session.get(UtilityService, service_id)
    if not service:
        raise HTTPException(404, "Коммунальная услуга не найдена")
    if "provider_reading_due_day" in payload:
        service.provider_reading_due_day = min(31, max(1, int(payload.get("provider_reading_due_day") or 20)))
    if "provider_due_day" in payload:
        service.provider_due_day = min(31, max(1, int(payload.get("provider_due_day") or 24)))
    if "resident_due_days" in payload:
        service.resident_due_days = min(31, max(1, int(payload.get("resident_due_days") or 7)))
    session.commit()
    session.refresh(service)
    return serialize_service(service)


@app.post("/api/meter-readings")
def save_meter_reading(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    meter = session.get(Meter, int(payload["meter_id"]))
    if not meter:
        raise HTTPException(404, "Счётчик не найден")
    reading_date = parse_date(payload.get("reading_date"), date.today())
    existing = session.scalar(select(MeterReading).where(MeterReading.meter_id == meter.id, MeterReading.reading_date == reading_date))
    if existing:
        existing.value = float(payload["value"])
        existing.note = payload.get("note") or ""
        reading = existing
    else:
        reading = MeterReading(
            meter_id=meter.id,
            reading_date=reading_date,
            value=float(payload["value"]),
            note=payload.get("note") or "",
        )
        session.add(reading)
    session.commit()
    return {
        "id": reading.id,
        "meter": meter.name,
        "reading_date": reading.reading_date.isoformat(),
        "value": reading.value,
        "note": reading.note,
    }


def upsert_meter_reading(session: Session, meter_id: int, reading_date: date, value: float, note: str = "") -> MeterReading:
    meter = session.get(Meter, meter_id)
    if not meter:
        raise HTTPException(404, f"Счётчик {meter_id} не найден")
    existing = session.scalar(select(MeterReading).where(MeterReading.meter_id == meter.id, MeterReading.reading_date == reading_date))
    if existing:
        existing.value = value
        existing.note = note
        return existing
    reading = MeterReading(meter_id=meter.id, reading_date=reading_date, value=value, note=note)
    session.add(reading)
    return reading


@app.post("/api/meter-readings/batch")
def save_meter_readings_batch(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    reading_date = parse_date(payload.get("reading_date"), date.today())
    readings = payload.get("readings") or []
    saved: list[MeterReading] = []
    for item in readings:
        raw_value = item.get("value")
        if raw_value in (None, ""):
            continue
        try:
            value = float(raw_value)
        except (TypeError, ValueError) as exc:
            raise HTTPException(400, "Показание должно быть числом") from exc
        meter_id = int(item["meter_id"])
        saved.append(upsert_meter_reading(session, meter_id, reading_date, value, item.get("note") or "быстрая передача"))
    session.commit()
    return {
        "saved": len(saved),
        "reading_date": reading_date.isoformat(),
    }


def tariffs_payload(session: Session) -> list[dict[str, Any]]:
    tariffs = session.scalars(
        select(Tariff)
        .options(joinedload(Tariff.service).joinedload(UtilityService.object))
        .order_by(Tariff.service_id, Tariff.starts_on.desc())
    ).all()
    return [
        {
            "id": tariff.id,
            "service_id": tariff.service_id,
            "kind": tariff.service.kind,
            "service": tariff.service.name,
            "object": tariff.service.object.name,
            "starts_on": tariff.starts_on.isoformat(),
            "name": tariff.name,
            "tiers": json.loads(tariff.tiers_json),
        }
        for tariff in tariffs
    ]


@app.get("/api/tariffs")
def list_tariffs(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    return tariffs_payload(session)


def parse_tiers(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        return value
    result = []
    raw = str(value or "").strip()
    if not raw:
        raise ValueError("Тарифные ступени пустые")
    normalized = re.sub(r",\s*(?=[\"'])", "; ", raw.replace("\n", ";"))
    for chunk in normalized.split(";"):
        chunk = chunk.strip()
        if not chunk:
            continue
        if ":" in chunk or "=" in chunk:
            separator = ":" if ":" in chunk else "="
            limit_raw, price_raw = [part.strip().strip('"').strip("'") for part in chunk.split(separator, 1)]
            limit = None if limit_raw in {"*", "∞", "inf", "null"} else float(limit_raw.replace(" ", ""))
        else:
            limit = None
            price_raw = chunk.strip().strip('"').strip("'")
        price = float(price_raw.replace(" ", "").replace(",", "."))
        result.append({"limit": limit, "price": price})
    if not result or result[-1]["limit"] is not None:
        result.append({"limit": None, "price": result[-1]["price"]})
    return result


@app.post("/api/tariffs")
def create_tariff(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    try:
        tiers = parse_tiers(payload.get("tiers"))
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    tariff = Tariff(
        service_id=int(payload["service_id"]),
        starts_on=parse_date(payload.get("starts_on"), date.today()),
        name=payload.get("name") or "Тариф",
        tiers_json=json.dumps(tiers, ensure_ascii=False),
    )
    session.add(tariff)
    session.commit()
    session.refresh(tariff)
    return {
        "id": tariff.id,
        "service_id": tariff.service_id,
        "starts_on": tariff.starts_on.isoformat(),
        "name": tariff.name,
        "tiers": tiers,
    }


def utility_bills_payload(
    session: Session,
    *,
    include_line_payments: bool = True,
    include_line_reminders: bool = True,
) -> list[dict[str, Any]]:
    bills = session.scalars(
        select(UtilityBill)
        .options(
            joinedload(UtilityBill.service).joinedload(UtilityService.object),
            selectinload(UtilityBill.lines).joinedload(UtilityBillLine.apartment),
            selectinload(UtilityBill.lines).joinedload(UtilityBillLine.lease).joinedload(Lease.tenant),
        )
        .order_by(UtilityBill.created_at.desc(), UtilityBill.id.desc())
    ).all()
    for bill in bills:
        for line in bill.lines:
            update_utility_line_status(line)
    bills = sorted(
        bills,
        key=lambda bill: (
            0
            if bill.status == "draft"
            or not bill.provider_paid
            or any(utility_line_debt_amount(line) > EPS and line.status in {"issued", "partial", "overdue"} for line in bill.lines)
            else 1,
            0 if bill.status == "draft" else 1,
            -(bill.created_at.timestamp() if bill.created_at else 0),
            -bill.id,
        ),
    )
    if include_line_reminders:
        line_ids = [line.id for bill in bills for line in bill.lines]
        prefetch_latest_message_logs(session, utility_line_ids=line_ids)
    return [
        serialize_bill(
            bill,
            session,
            include_line_payments=include_line_payments,
            include_line_reminders=include_line_reminders,
        )
        for bill in bills
    ]


@app.get("/api/utility-bills")
def list_utility_bills(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    return utility_bills_payload(session, include_line_payments=False, include_line_reminders=False)


def utility_timeline_payload(session: Session) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    object_readings = session.scalars(
        select(MeterReading)
        .options(joinedload(MeterReading.meter).joinedload(Meter.service).joinedload(UtilityService.object))
        .join(Meter)
        .where(Meter.scope == "object", Meter.active.is_(True))
        .order_by(MeterReading.reading_date.desc(), MeterReading.id.desc())
    ).all()
    for reading in object_readings:
        meter = reading.meter
        service = meter.service
        events.append(
            {
                "kind": "reading",
                "sort_at": f"{reading.reading_date.isoformat()}T00:00:00",
                "date": reading.reading_date.isoformat(),
                "object": service.object.name,
                "service": service.name,
                "service_id": service.id,
                "title": "Общедомовые показания",
                "detail": f"{meter.name}: {reading.value}",
            }
        )

    bills = session.scalars(
        select(UtilityBill)
        .options(joinedload(UtilityBill.service).joinedload(UtilityService.object), selectinload(UtilityBill.lines))
        .order_by(UtilityBill.period_end.desc(), UtilityBill.id.desc())
    ).all()
    for bill in bills:
        events.append(
            {
                "kind": "bill",
                "sort_at": f"{bill.period_end.isoformat()}T12:00:00",
                "date": bill.period_end.isoformat(),
                "object": bill.service.object.name,
                "service": bill.service.name,
                "service_id": bill.service_id,
                "bill_id": bill.id,
                "title": f"{bill.period_start:%d.%m.%Y} -> {bill.period_end:%d.%m.%Y} ({(bill.period_end - bill.period_start).days} дн.)",
                "detail": f"{(bill.period_end - bill.period_start).days} дн., жильцам {money_text(sum(line.total_amount for line in bill.lines))}, по дому {money_text(bill.total_cost)}",
                "status": bill.status,
                "provider_paid": bill.provider_paid,
            }
        )
        if bill.provider_paid_at:
            events.append(
                {
                    "kind": "provider_paid",
                    "sort_at": bill.provider_paid_at.isoformat(),
                    "date": bill.provider_paid_at.date().isoformat(),
                    "object": bill.service.object.name,
                    "service": bill.service.name,
                    "service_id": bill.service_id,
                    "bill_id": bill.id,
                    "title": "Поставщик оплачен",
                    "detail": f"Период {format_full_date(bill.period_start)} -> {format_full_date(bill.period_end)}",
                    "status": bill.status,
                    "provider_paid": True,
                }
            )

    utility_receipts = session.scalars(
        select(PaymentReceipt).where(PaymentReceipt.utility_line_id.is_not(None)).order_by(PaymentReceipt.paid_at.desc(), PaymentReceipt.id.desc())
    ).all()
    utility_line_ids = [receipt.utility_line_id for receipt in utility_receipts if receipt.utility_line_id]
    utility_lines_by_id = {
        line.id: line
        for line in session.scalars(
            select(UtilityBillLine)
            .options(
                joinedload(UtilityBillLine.apartment),
                joinedload(UtilityBillLine.lease).joinedload(Lease.tenant),
                joinedload(UtilityBillLine.bill).joinedload(UtilityBill.service).joinedload(UtilityService.object),
            )
            .where(UtilityBillLine.id.in_(utility_line_ids))
        ).all()
    } if utility_line_ids else {}
    for receipt in utility_receipts:
        line = utility_lines_by_id.get(receipt.utility_line_id) if receipt.utility_line_id else None
        if not line:
            continue
        bill = line.bill
        events.append(
            {
                "kind": "resident_payment",
                "sort_at": receipt.paid_at.isoformat(),
                "date": receipt.paid_at.date().isoformat(),
                "object": bill.service.object.name,
                "service": bill.service.name,
                "service_id": bill.service_id,
                "bill_id": bill.id,
                "title": f"Оплата жильца: {line.apartment.name}",
                "detail": f"{line.lease.tenant.full_name if line.lease else ''} — {money_text(receipt.amount)}",
                "status": receipt.status,
                "provider_paid": bill.provider_paid,
            }
        )

    events.sort(key=lambda item: item["sort_at"], reverse=True)
    return events


@app.get("/api/utilities/timeline")
def utility_timeline(session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    return utility_timeline_payload(session)


@app.post("/api/utility-bills/calculate")
def create_utility_bill(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    try:
        existing = session.scalar(
            select(UtilityBill).where(
                UtilityBill.service_id == int(payload["service_id"]),
                UtilityBill.period_start == parse_date(payload.get("period_start")),
                UtilityBill.period_end == parse_date(payload.get("period_end")),
                UtilityBill.status == "draft",
            )
        )
        if existing:
            raise ValueError("Черновик за этот период уже есть. Удалите его и пересоздайте заново, если нужен новый расчёт.")
        bill, warnings = calculate_utility_bill(
            session=session,
            service_id=int(payload["service_id"]),
            period_start=parse_date(payload.get("period_start")),
            period_end=parse_date(payload.get("period_end")),
            allow_estimate=bool(payload.get("allow_estimate")),
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    session.add(bill)
    session.flush()
    advance_bills = ensure_utility_advance_drafts_for_bills(session, [bill])
    session.commit()
    session.refresh(bill)
    result = serialize_bill(bill)
    result["advance_bills"] = [serialize_bill(advance_bill) for advance_bill in advance_bills]
    result["warnings"] = warnings
    return result


@app.post("/api/utility-bills/calculate-object")
def create_object_utility_bills(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    object_id = int(payload.get("object_id") or 0)
    rental_object = session.get(RentalObject, object_id)
    if not rental_object:
        raise HTTPException(404, "объект не найден")
    period_start = parse_date(payload.get("period_start"))
    period_end = parse_date(payload.get("period_end"))
    allow_estimate = bool(payload.get("allow_estimate"))
    services = active_services_for_object(session, object_id)
    if not services:
        raise HTTPException(400, "у объекта нет активных коммунальных услуг")

    created_bills: list[UtilityBill] = []
    warnings: list[str] = []
    skipped: list[str] = []
    errors: list[str] = []
    for service in services:
        existing = session.scalar(
            select(UtilityBill).where(
                UtilityBill.service_id == service.id,
                UtilityBill.period_start == period_start,
                UtilityBill.period_end == period_end,
                UtilityBill.status == "draft",
            )
        )
        if existing:
            skipped.append(service.name)
            continue
        try:
            bill, bill_warnings = calculate_utility_bill(
                session=session,
                service_id=service.id,
                period_start=period_start,
                period_end=period_end,
                allow_estimate=allow_estimate,
            )
        except ValueError as exc:
            errors.append(f"{service.name}: {exc}")
            continue
        session.add(bill)
        created_bills.append(bill)
        warnings.extend(f"{service.name}: {warning}" for warning in bill_warnings)

    if not created_bills and errors:
        raise HTTPException(400, "\n".join(errors))
    if not created_bills and skipped:
        raise HTTPException(400, "Черновики за этот период уже есть: " + ", ".join(skipped))

    session.flush()
    advance_bills = ensure_utility_advance_drafts_for_bills(session, created_bills)
    session.commit()
    return {
        "object_id": object_id,
        "object": rental_object.name,
        "created": [serialize_bill(bill) for bill in created_bills],
        "advance_bills": [serialize_bill(bill) for bill in advance_bills],
        "warnings": warnings,
        "skipped": skipped,
        "errors": errors,
    }


@app.delete("/api/utility-bills/{bill_id}")
def delete_utility_bill(bill_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    bill = session.get(UtilityBill, bill_id)
    if not bill:
        raise HTTPException(404, "???????????? ???? ?? ??????")
    line_ids = [line.id for line in bill.lines]
    if any(money(line.paid_amount) > EPS for line in bill.lines):
        raise HTTPException(400, "?????? ??????? ????, ? ??????? ??? ???????? ??????. ??????? ???????? ??? ??????? ??? ???????.")
    linked_receipts = (
        session.scalars(select(PaymentReceipt).where(PaymentReceipt.utility_line_id.in_(line_ids))).all()
        if line_ids
        else []
    )
    if any(receipt.status == "accepted" and money(receipt.amount) > EPS for receipt in linked_receipts):
        raise HTTPException(400, "?????? ??????? ????, ?? ???????? ??? ??????? ??????. ??????? ????????? ??????????? ????.")
    linked_logs = (
        session.scalars(select(MessageLog).where(MessageLog.utility_line_id.in_(line_ids))).all()
        if line_ids
        else []
    )
    affected_lease_ids = {line.lease_id for line in bill.lines if line.lease_id}
    for receipt in linked_receipts:
        receipt.utility_line_id = None
    for log in linked_logs:
        log.utility_line_id = None
    deleted_status = bill.status
    session.delete(bill)
    session.flush()
    for lease_id in sorted(affected_lease_ids):
        recalculate_lease_balances(session, int(lease_id))
    session.commit()
    return {"ok": True, "bill_id": bill_id, "deleted_status": deleted_status}


def available_utility_advances(session: Session, lease_id: int) -> list[PaymentReceipt]:
    lease = session.get(Lease, lease_id)
    apartment_id = lease.apartment_id if lease else None
    return [
        receipt
        for receipt in session.scalars(
            select(PaymentReceipt)
            .where(
                PaymentReceipt.apartment_id == apartment_id if apartment_id else PaymentReceipt.lease_id == lease_id,
                PaymentReceipt.status == "accepted",
                PaymentReceipt.channel == UTILITY_ADVANCE_CHANNEL,
                PaymentReceipt.utility_line_id.is_(None),
                PaymentReceipt.rent_charge_id.is_(None),
            )
            .order_by(PaymentReceipt.paid_at, PaymentReceipt.id)
        ).all()
        if money(float(receipt.amount or 0)) > EPS
    ]


def append_receipt_note(receipt: PaymentReceipt, note: str) -> None:
    receipt.notes = "; ".join(part for part in [receipt.notes, note] if part)


def apply_utility_advances_to_bill(session: Session, bill: UtilityBill) -> float:
    applied_total = 0.0
    affected_lease_ids: set[int] = set()
    lines = sorted(
        [line for line in bill.lines if line.lease_id and not utility_line_is_advance(line)],
        key=lambda item: (item.lease_id or 0, item.due_date or bill.due_date or bill.period_end, item.id or 0),
    )
    for line in lines:
        line_debt = money(max(0.0, float(line.total_amount or 0) - float(line.paid_amount or 0)))
        if line_debt <= EPS:
            continue
        for advance in available_utility_advances(session, int(line.lease_id)):
            if line_debt <= EPS:
                break
            advance_amount = money(float(advance.amount or 0))
            if advance_amount <= EPS:
                continue
            portion = money(min(advance_amount, line_debt))
            note = f"зачтено из аванса в счёт {bill.service.name} за {bill.period_end:%m.%Y}"
            if advance_amount <= portion + EPS:
                advance.utility_line_id = line.id
                advance.channel = "utilities"
                advance.source = UTILITY_ADVANCE_SOURCE
                append_receipt_note(advance, note)
            else:
                advance.amount = money(advance_amount - portion)
                session.add(
                    PaymentReceipt(
                        lease_id=advance.lease_id,
                        utility_line_id=line.id,
                        apartment_id=advance.apartment_id,
                        amount=portion,
                        channel="utilities",
                        paid_at=advance.paid_at,
                        source=UTILITY_ADVANCE_SOURCE,
                        status="accepted",
                        recipient_name=advance.recipient_name,
                        recipient_details=advance.recipient_details,
                        file_path=advance.file_path,
                        notes=note,
                    )
                )
            applied_total = money(applied_total + portion)
            line_debt = money(line_debt - portion)
            affected_lease_ids.add(int(line.lease_id))
            session.flush()
        ledger_balance = utility_advance_ledger_balance(session, line.apartment_id)
        if line_debt > EPS and ledger_balance > EPS:
            portion = money(min(ledger_balance, line_debt))
            session.add(
                UtilityAdvanceLedger(
                    apartment_id=line.apartment_id,
                    lease_id=line.lease_id,
                    utility_line_id=line.id,
                    period_start=bill.period_start,
                    period_end=bill.period_end,
                    amount=-portion,
                    kind="advance_applied",
                    note=f"зачтено из аванса в счёт {bill.service.name} за {bill.period_end:%m.%Y}",
                )
            )
            applied_total = money(applied_total + portion)
            line_debt = money(line_debt - portion)
            affected_lease_ids.add(int(line.lease_id))
            session.flush()
    for lease_id in sorted(affected_lease_ids):
        recalculate_lease_balances(session, lease_id)
    return money(applied_total)


def active_services_for_object(session: Session, object_id: int) -> list[UtilityService]:
    return session.scalars(
        select(UtilityService)
        .where(UtilityService.object_id == object_id, UtilityService.active.is_(True))
        .order_by(UtilityService.id)
    ).all()


def advance_bill_for_object_period(
    session: Session,
    object_id: int,
    period_start: date,
    period_end: date,
) -> UtilityBill | None:
    return session.scalar(
        select(UtilityBill)
        .options(selectinload(UtilityBill.lines))
        .join(UtilityService)
        .where(
            UtilityService.object_id == object_id,
            UtilityBill.bill_type == UTILITY_BILL_ADVANCE,
            UtilityBill.period_start == period_start,
            UtilityBill.period_end == period_end,
        )
        .order_by(UtilityBill.id.desc())
        .limit(1)
    )


def append_line_note(line: UtilityBillLine, note: str) -> None:
    line.note = "; ".join(part for part in [line.note, note] if part)


def cancel_stale_unpaid_advance_lines(session: Session, object_id: int, before_start: date) -> int:
    bills = session.scalars(
        select(UtilityBill)
        .options(selectinload(UtilityBill.lines))
        .join(UtilityService)
        .where(
            UtilityService.object_id == object_id,
            UtilityBill.bill_type == UTILITY_BILL_ADVANCE,
            UtilityBill.period_start < before_start,
            UtilityBill.status.in_(["draft", "issued"]),
        )
        .order_by(UtilityBill.period_start, UtilityBill.id)
    ).all()
    cancelled = 0
    for bill in bills:
        for line in bill.lines:
            if not utility_line_is_advance(line) or line.status in {"paid", "paid_ahead", "cancelled"}:
                continue
            update_utility_line_status(line)
            debt = utility_line_debt_amount(line)
            if debt <= EPS:
                continue
            if money(line.paid_amount) > EPS:
                line.total_amount = money(line.paid_amount)
                update_utility_line_status(line)
                append_line_note(line, "остаток старого аванса отменён новым расчётным периодом")
            else:
                line.status = "cancelled"
                line.total_amount = 0
                append_line_note(line, "старый неоплаченный аванс отменён новым расчётным периодом")
            cancelled += 1
        if bill.lines and all(line.status == "cancelled" for line in bill.lines):
            bill.status = "cancelled"
    return cancelled


def utility_usage_amounts_by_apartment(bills: list[UtilityBill]) -> dict[int, float]:
    amounts: dict[int, float] = {}
    for bill in bills:
        if utility_bill_is_advance(bill):
            continue
        for line in bill.lines:
            if utility_line_is_advance(line):
                continue
            amounts[line.apartment_id] = money(amounts.get(line.apartment_id, 0.0) + float(line.total_amount or 0))
    return amounts


def build_utility_advance_line_payload(
    session: Session,
    apartment: Apartment,
    lease: Lease,
    advance_start: date,
    advance_end: date,
    line_start: date,
    line_end: date,
    *,
    current_amount: float,
) -> tuple[float, dict[str, Any]]:
    recommended, forecast_meta = estimated_apartment_utility_amount(
        session,
        apartment.id,
        advance_start,
        current_amount=current_amount,
    )
    full_days = max((advance_end - advance_start).days, 1)
    line_days = max((line_end - line_start).days, 0)
    period_factor = min(1.0, max(0.0, line_days / full_days))
    prorated_recommended = money(recommended * period_factor)
    balance_before = utility_advance_balance(session, apartment.id)
    amount = money(max(0.0, prorated_recommended - balance_before))
    metadata = {
        **forecast_meta,
        "advance_recommended_amount": recommended,
        "advance_prorated_amount": prorated_recommended,
        "advance_balance_before": balance_before,
        "advance_amount": amount,
        "advance_period_start": advance_start.isoformat(),
        "advance_period_end": advance_end.isoformat(),
        "line_period_start": line_start.isoformat(),
        "line_period_end": line_end.isoformat(),
        "period_factor": round(period_factor, 4),
        "calculated_at": utc_now().isoformat(),
    }
    return amount, metadata


def ensure_utility_advance_draft_for_object(
    session: Session,
    object_id: int,
    period_start: date,
    period_end: date,
    *,
    current_amounts_by_apartment: dict[int, float],
) -> UtilityBill | None:
    services = active_services_for_object(session, object_id)
    if not services:
        return None
    advance_start, advance_end = next_utility_advance_period(period_start, period_end)
    cancel_stale_unpaid_advance_lines(session, object_id, advance_start)

    existing = advance_bill_for_object_period(session, object_id, advance_start, advance_end)
    if existing and existing.status != "draft":
        return existing

    apartments = session.scalars(
        select(Apartment)
        .where(Apartment.object_id == object_id, Apartment.active.is_(True))
        .order_by(Apartment.sort_order, Apartment.name)
    ).all()
    desired: dict[int, tuple[Apartment, Lease, float, dict[str, Any], date, date]] = {}
    for apartment in apartments:
        lease = active_lease_for_apartment(session, apartment.id, period_start, period_end)
        if not lease or lease_ignored(session, lease.id):
            continue
        line_period = lease_utility_advance_period(lease, advance_start, advance_end)
        if not line_period:
            continue
        line_start, line_end = line_period
        amount, metadata = build_utility_advance_line_payload(
            session,
            apartment,
            lease,
            advance_start,
            advance_end,
            line_start,
            line_end,
            current_amount=current_amounts_by_apartment.get(apartment.id, 0.0),
        )
        if amount > EPS:
            desired[apartment.id] = (apartment, lease, amount, metadata, line_start, line_end)

    if not desired:
        if existing and not any(money(line.paid_amount) > EPS for line in existing.lines):
            session.delete(existing)
        return None

    bill = existing or UtilityBill(
        service_id=services[0].id,
        period_start=advance_start,
        period_end=advance_end,
        bill_type=UTILITY_BILL_ADVANCE,
        status="draft",
        total_consumption=0,
        apartment_consumption=0,
        odn_consumption=0,
        total_cost=0,
        average_unit_price=0,
        is_forecast=True,
        provider_paid=True,
        provider_paid_at=utc_now(),
        notes="Автоматический аванс коммуналки на следующий период.",
    )
    if not existing:
        session.add(bill)
        session.flush()

    existing_lines_by_apartment = {
        line.apartment_id: line
        for line in bill.lines
        if utility_line_is_advance(line)
    }
    for apartment_id, (apartment, lease, amount, metadata, line_start, line_end) in desired.items():
        line = existing_lines_by_apartment.pop(apartment_id, None)
        if line is None:
            line = UtilityBillLine(
                apartment_id=apartment.id,
                lease_id=lease.id,
                line_type=UTILITY_LINE_ADVANCE,
                personal_consumption=0,
                odn_consumption=0,
                status="draft",
            )
            bill.lines.append(line)
        line.lease_id = lease.id
        line.total_amount = amount
        line.paid_amount = money(line.paid_amount)
        line.note = f"Аванс коммуналки {line_start:%d.%m.%Y} -> {line_end:%d.%m.%Y}"
        set_utility_line_metadata(line, metadata)
        update_utility_line_status(line)
        if line.status == "paid" and utility_line_debt_amount(line) > EPS:
            line.status = "draft"
    for stale_line in existing_lines_by_apartment.values():
        if money(stale_line.paid_amount) > EPS:
            stale_line.total_amount = money(stale_line.paid_amount)
            update_utility_line_status(stale_line)
            append_line_note(stale_line, "новый расчёт аванса больше не требует доплаты")
        else:
            bill.lines.remove(stale_line)
            session.delete(stale_line)

    bill.total_cost = money(sum(line.total_amount for line in bill.lines if utility_line_is_advance(line)))
    bill.total_consumption = 0
    bill.apartment_consumption = 0
    bill.odn_consumption = 0
    return bill


def ensure_utility_advance_drafts_for_bills(session: Session, bills: list[UtilityBill]) -> list[UtilityBill]:
    groups: dict[tuple[int, date, date], list[UtilityBill]] = {}
    for bill in bills:
        if utility_bill_is_advance(bill) or not bill.service:
            continue
        groups.setdefault((bill.service.object_id, bill.period_start, bill.period_end), []).append(bill)
    result: list[UtilityBill] = []
    for (object_id, period_start, period_end), grouped_bills in groups.items():
        advance_bill = ensure_utility_advance_draft_for_object(
            session,
            object_id,
            period_start,
            period_end,
            current_amounts_by_apartment=utility_usage_amounts_by_apartment(grouped_bills),
        )
        if advance_bill:
            result.append(advance_bill)
    return result


def refreshed_utility_issue_group_bills(session: Session, bill: UtilityBill) -> list[UtilityBill]:
    bills = utility_issue_group_bills(session, bill)
    usage_bills = [
        item
        for item in bills
        if not utility_bill_is_advance(item) and item.status == "draft"
    ]
    if not usage_bills:
        return bills
    ensure_utility_advance_drafts_for_bills(session, usage_bills)
    session.flush()
    return utility_issue_group_bills(session, bill)


def exception_detail_text(exc: BaseException) -> str:
    if isinstance(exc, HTTPException):
        detail = exc.detail
    else:
        detail = str(exc)
    if isinstance(detail, str):
        return detail
    try:
        return json.dumps(detail, ensure_ascii=False)
    except TypeError:
        return str(detail)


def utility_issue_apartment_label(apartment: Apartment) -> str:
    object_name = apartment.object.name if apartment.object else ""
    return f"{object_name}, {apartment.name}" if object_name else apartment.name


def utility_issue_failed_recipient(lease: Lease, detail: str) -> dict[str, Any]:
    return {
        "lease_id": lease.id,
        "tenant": lease.tenant.full_name if lease.tenant else "",
        "apartment": utility_issue_apartment_label(lease.apartment) if lease.apartment else "",
        "detail": detail,
    }


@app.post("/api/utility-bills/{bill_id}/issue")
def issue_utility_bill(bill_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    bill = session.get(UtilityBill, bill_id)
    if not bill:
        raise HTTPException(404, "Коммунальный счёт не найден")
    bills = refreshed_utility_issue_group_bills(session, bill)
    previews = utility_issue_targets_for_bills(session, bills)
    due = resident_due_date_for_bills(bills)
    for item in bills:
        item.status = "issued"
        item.due_date = due
        if utility_bill_is_advance(item):
            item.provider_paid = True
            item.provider_paid_at = item.provider_paid_at or utc_now()
        for line in item.lines:
            if line.status == "cancelled":
                continue
            line.status = "issued"
            line.issued_at = utc_now()
            line.due_date = due
    session.flush()
    applied_advances = money(sum(apply_utility_advances_to_bill(session, item) for item in bills if not utility_bill_is_advance(item)))
    sent = 0
    skipped_unlinked = 0
    linked_targets = sum(1 for preview in previews if preview["linked"])
    failed_recipients: list[dict[str, Any]] = []
    for preview in previews:
        lease = session.get(Lease, preview["lease_id"])
        if not lease:
            continue
        if not preview["linked"]:
            skipped_unlinked += 1
            continue
        primary_line = session.get(UtilityBillLine, preview["line_ids"][0]) if preview["line_ids"] else None
        situation = (
            payment_situation(session, "utility", primary_line.id, int(primary_line.lease_id))
            if primary_line and primary_line.lease_id
            else None
        )
        try:
            send_tenant_message(
                session,
                lease,
                "message_utility_bill",
                line=primary_line,
                note="utility-issue",
                utility_lines_override=[line for line_id in preview["all_line_ids"] if (line := session.get(UtilityBillLine, line_id))],
                utility_due_date_override=due,
                keyboard=tenant_situation_keyboard(situation) if situation else None,
            )
            sent += 1
        except HTTPException as exc:
            failed_recipients.append(utility_issue_failed_recipient(lease, exception_detail_text(exc)))
    if linked_targets > 0 and sent == 0 and failed_recipients:
        first_failure = failed_recipients[0]
        raise HTTPException(
            400,
            f"Не отправлено ни одного сообщения. {first_failure['apartment']} {first_failure['tenant']}: {first_failure['detail']}",
        )
    session.commit()
    session.refresh(bill)
    payload = serialize_bill(bill)
    payload["bill_ids"] = [item.id for item in bills]
    payload["sent"] = sent
    payload["skipped_unlinked"] = skipped_unlinked
    payload["failed"] = len(failed_recipients)
    payload["failed_recipients"] = failed_recipients
    payload["applied_advances"] = applied_advances
    return payload


@app.get("/api/utility-bills/{bill_id}/issue-preview")
def preview_issue_utility_bill(bill_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    bill = session.get(UtilityBill, bill_id)
    if not bill:
        raise HTTPException(404, "Коммунальный счёт не найден")
    bills = refreshed_utility_issue_group_bills(session, bill)
    object_name = bill.service.object.name
    services = [utility_bill_service_label(item) for item in bills]
    period_label = (
        f"{bill.period_start:%d.%m.%Y} -> {bill.period_end:%d.%m.%Y}"
        if not utility_bill_is_advance(bill)
        else f"{bill.period_start:%d.%m.%Y} -> {bill.period_end:%d.%m.%Y}"
    )
    return {
        "bill_id": bill.id,
        "bill_ids": [item.id for item in bills],
        "object": object_name,
        "service": ", ".join(services),
        "period_label": period_label,
        "due_date": resident_due_date_for_bills(bills).isoformat(),
        "targets": utility_issue_targets_for_bills(session, bills),
    }


@app.post("/api/utility-bills/{bill_id}/provider-paid")
def mark_provider_paid(bill_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    bill = session.get(UtilityBill, bill_id)
    if not bill:
        raise HTTPException(404, "Коммунальный счёт не найден")
    bill.provider_paid = True
    bill.provider_paid_at = utc_now()
    session.commit()
    return serialize_bill(bill)


@app.post("/api/utility-lines/{line_id}/payments")
def add_utility_payment(line_id: int, payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    line = session.get(UtilityBillLine, line_id)
    if not line:
        raise HTTPException(404, "строка коммуналки не найдена")
    amount = float(payload.get("amount") or 0)
    if amount <= 0:
        raise HTTPException(400, "сумма платежа должна быть больше нуля")
    if not line.lease:
        raise HTTPException(400, "у этой строки нет привязанной аренды")
    paid_at = datetime.fromisoformat(payload["paid_at"]) if payload.get("paid_at") else utc_now()
    if utility_line_is_advance(line):
        receipt = PaymentReceipt(
            lease_id=line.lease_id,
            utility_line_id=line.id,
            apartment_id=line.apartment_id,
            amount=money(amount),
            channel=UTILITY_ADVANCE_CHANNEL,
            paid_at=paid_at,
            source=payload.get("source") or "manual",
            status=payload.get("status") or "accepted",
            recipient_name=payload.get("recipient_name") or "",
            recipient_details=payload.get("recipient_details") or "",
            notes=payload.get("notes") or "оплата аванса коммуналки",
        )
        session.add(receipt)
        session.flush()
        sync_utility_advance_credit_for_receipt(session, receipt)
        recalculate_lease_balances(session, int(line.lease_id))
    else:
        receipts = create_utility_receipts(
            session,
            line.lease,
            amount,
            paid_at=paid_at,
            source=payload.get("source") or "manual",
            status=payload.get("status") or "accepted",
            recipient_name=payload.get("recipient_name") or "",
            recipient_details=payload.get("recipient_details") or "",
            notes=payload.get("notes") or "",
            exact_only=False,
        )
        sync_utility_advance_credits_for_receipts(session, receipts)
    session.commit()
    return serialize_bill_line(line, session)


def expenses_payload(session: Session, *, limit: int | None = None, offset: int = 0) -> list[dict[str, Any]]:
    query = (
        select(Expense)
        .options(joinedload(Expense.object), joinedload(Expense.apartment))
        .order_by(Expense.expense_date.desc(), Expense.id.desc())
    )
    if offset:
        query = query.offset(max(0, int(offset)))
    if limit is not None:
        query = query.limit(max(1, min(int(limit), 500)))
    expenses = session.scalars(query).all()
    return [serialize_expense(expense) for expense in expenses]


@app.get("/api/expenses")
def list_expenses(limit: int | None = None, offset: int = 0, session: Session = Depends(get_session)) -> list[dict[str, Any]]:
    return expenses_payload(session, limit=limit, offset=offset)


@app.post("/api/expenses")
def create_expense(payload: dict[str, Any], session: Session = Depends(get_session)) -> dict[str, Any]:
    expense = Expense(
        expense_date=parse_date(payload.get("expense_date"), date.today()),
        object_id=int(payload["object_id"]) if payload.get("object_id") else None,
        apartment_id=int(payload["apartment_id"]) if payload.get("apartment_id") else None,
        category=payload.get("category") or "",
        amount=float(payload.get("amount") or 0),
        source_funds=payload.get("source_funds") or "personal",
        payment_method=payload.get("payment_method") or "",
        description=payload.get("description") or "",
        compensation_status="pending" if payload.get("source_funds", "personal") == "personal" else "not_required",
        file_path=payload.get("file_path") or "",
        notes=payload.get("notes") or "",
    )
    if expense.amount <= 0:
        raise HTTPException(400, "Сумма должна быть больше нуля")
    session.add(expense)
    session.commit()
    session.refresh(expense)
    return serialize_expense(expense)


@app.post("/api/expenses/{expense_id}/compensate")
def compensate_expense(expense_id: int, session: Session = Depends(get_session)) -> dict[str, Any]:
    expense = session.get(Expense, expense_id)
    if not expense:
        raise HTTPException(404, "Расход не найден")
    expense.compensation_status = "compensated"
    expense.compensated_at = utc_now()
    session.commit()
    return serialize_expense(expense)


def setup_sheet(wb: Workbook, title: str, headers: list[str]):
    if wb.active.max_row == 1 and wb.active.max_column == 1 and wb.active["A1"].value is None and wb.active.title == "Sheet":
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title)
    ws.append(headers)
    fill = PatternFill("solid", fgColor="E8EEF8")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    return ws


REPORT_FILL_DEBT = PatternFill("solid", fgColor="F8D7DA")
REPORT_FILL_WARN = PatternFill("solid", fgColor="FFF3CD")
REPORT_FILL_OK = PatternFill("solid", fgColor="D1E7DD")


def fill_report_row(ws, row_index: int, fill: PatternFill) -> None:
    for cell in ws[row_index]:
        cell.fill = fill


def highlight_status_row(ws, row_index: int, status: str, debt: float = 0.0, attention: bool = False) -> None:
    if debt > EPS or status in {"overdue", "partial"}:
        fill_report_row(ws, row_index, REPORT_FILL_DEBT)
    elif attention or status in {"issued", "pending", "deferred"}:
        fill_report_row(ws, row_index, REPORT_FILL_WARN)
    elif status in {"paid", "paid_ahead", "compensated", "not_required"}:
        fill_report_row(ws, row_index, REPORT_FILL_OK)


def expense_balance_value(spent_total: float, income_total: float) -> float:
    return money(float(spent_total or 0) - float(income_total or 0))


def expense_balance_comment(balance: float) -> str:
    if balance > EPS:
        return "Расходов больше, чем поступлений"
    if balance < -EPS:
        return "Поступлений больше, чем расходов"
    return "Сошлось в ноль"


def expense_period_summary(session: Session, start_date: date, end_date: date) -> dict[str, float]:
    opening_income = money(
        session.scalar(
            select(func.coalesce(func.sum(Expense.amount), 0)).where(
                Expense.expense_date < start_date,
                Expense.source_funds == "expense_fund_income",
            )
        )
        or 0
    )
    opening_spent = money(
        session.scalar(
            select(func.coalesce(func.sum(Expense.amount), 0)).where(
                Expense.expense_date < start_date,
                Expense.source_funds != "expense_fund_income",
            )
        )
        or 0
    )
    period_income = money(
        session.scalar(
            select(func.coalesce(func.sum(Expense.amount), 0)).where(
                Expense.expense_date >= start_date,
                Expense.expense_date <= end_date,
                Expense.source_funds == "expense_fund_income",
            )
        )
        or 0
    )
    period_spent = money(
        session.scalar(
            select(func.coalesce(func.sum(Expense.amount), 0)).where(
                Expense.expense_date >= start_date,
                Expense.expense_date <= end_date,
                Expense.source_funds != "expense_fund_income",
            )
        )
        or 0
    )
    closing_income = money(opening_income + period_income)
    closing_spent = money(opening_spent + period_spent)
    opening_balance = expense_balance_value(opening_spent, opening_income)
    closing_balance = expense_balance_value(closing_spent, closing_income)
    return {
        "opening_income": opening_income,
        "opening_spent": opening_spent,
        "opening_balance": opening_balance,
        "period_income": period_income,
        "period_spent": period_spent,
        "closing_income": closing_income,
        "closing_spent": closing_spent,
        "closing_balance": closing_balance,
    }


def prepend_expense_summary(
    ws,
    start_date: date,
    end_date: date,
    summary: dict[str, float],
    merge_columns: int,
) -> None:
    ws.insert_rows(1, amount=9)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, merge_columns))
    ws["A1"] = f"Сводка по расходам: {start_date:%d.%m.%Y} - {end_date:%d.%m.%Y}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Плюс = расходов больше, минус = поступлений больше. Да, знак теперь честный."
    ws["A2"].font = Font(italic=True)

    rows = [
        ("Сальдо на начало периода", summary["opening_balance"], expense_balance_comment(summary["opening_balance"])),
        ("Поступило на расходы до начала периода", summary["opening_income"], ""),
        ("Израсходовано до начала периода", summary["opening_spent"], ""),
        ("Поступило на расходы за период", summary["period_income"], ""),
        ("Израсходовано за период", summary["period_spent"], ""),
        ("Сальдо на конец периода", summary["closing_balance"], expense_balance_comment(summary["closing_balance"])),
    ]
    for row_index, (label, value, comment) in enumerate(rows, start=3):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=value)
        if merge_columns >= 3 and comment:
            ws.cell(row=row_index, column=3, value=comment)
        fill_report_row(
            ws,
            row_index,
            REPORT_FILL_WARN if abs(float(value or 0)) > EPS and "Сальдо" in label else REPORT_FILL_OK,
        )
    ws.freeze_panes = "A10"


def finalize_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_index, column in enumerate(ws.columns, start=1):
            letter = get_column_letter(column_index)
            max_length = 0
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                for part in value.splitlines() or [""]:
                    max_length = max(max_length, len(part))
            ws.column_dimensions[letter].width = max(10, min(max_length + 3, 60))
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions


def workbook_response(wb: Workbook, filename: str) -> StreamingResponse:
    finalize_workbook(wb)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


STATUS_LABELS = {
    "pending": "ожидается",
    "overdue": "просрочено",
    "partial": "частично оплачено",
    "paid": "оплачено",
    "paid_ahead": "оплачено вперёд",
    "deferred": "отсрочка",
    "draft": "черновик",
    "issued": "выставлено",
    "cancelled": "отменено",
    "not_required": "не требуется",
    "compensated": "компенсировано",
}


def report_status_label(status: str, due_date: date | None = None, today: date | None = None) -> str:
    label = STATUS_LABELS.get(status, status)
    today = today or date.today()
    if status == "overdue" and due_date:
        days = max((today - due_date).days, 0)
        return f"{label}, {days} дн."
    return label


def accepted_receipts_for_charge(charge: RentCharge, channel: str) -> list[PaymentReceipt]:
    return [receipt for receipt in charge.receipts if receipt.status == "accepted" and receipt.channel == channel]


def accepted_receipt_amount(receipts: list[PaymentReceipt]) -> float:
    return money(sum(float(receipt.amount or 0) for receipt in receipts))


def charge_expense_fund_receipts(charge: RentCharge) -> list[PaymentReceipt]:
    return accepted_receipts_for_charge(charge, "expense_fund")


def charge_direct_ip_receipts(charge: RentCharge) -> list[PaymentReceipt]:
    return accepted_receipts_for_charge(charge, "ip")


def owner_expected_ip_for_charge(charge: RentCharge) -> float:
    return money(max(0.0, float(charge.ip_due or 0) - accepted_receipt_amount(charge_expense_fund_receipts(charge))))


def owner_direct_ip_paid_for_charge(charge: RentCharge) -> float:
    direct_paid = accepted_receipt_amount(charge_direct_ip_receipts(charge))
    if direct_paid > EPS:
        return direct_paid
    expense_paid = accepted_receipt_amount(charge_expense_fund_receipts(charge))
    return money(max(0.0, float(charge.ip_paid or 0) - expense_paid))


def owner_charge_status_label(charge: RentCharge) -> str:
    expense_paid = accepted_receipt_amount(charge_expense_fund_receipts(charge))
    direct_ip_paid = owner_direct_ip_paid_for_charge(charge)
    adjusted_expected_ip = owner_expected_ip_for_charge(charge)
    total_owner_paid = money(direct_ip_paid + expense_paid)
    if expense_paid > EPS and total_owner_paid + EPS >= float(charge.ip_due or 0):
        if direct_ip_paid > EPS and adjusted_expected_ip <= EPS:
            return "оплачено, часть ушла на расходы"
        if adjusted_expected_ip <= EPS:
            return "оплачено, ушло на расходы"
    if adjusted_expected_ip <= EPS:
        return "оплачено"
    if total_owner_paid <= EPS:
        return report_status_label("overdue" if charge.due_date < date.today() else "pending", charge.due_date)
    if total_owner_paid + EPS < adjusted_expected_ip:
        return report_status_label("partial", charge.due_date)
    if total_owner_paid > adjusted_expected_ip + EPS:
        return report_status_label("paid_ahead", charge.due_date)
    return report_status_label("paid", charge.due_date)


def receipt_amounts_text(receipts: list[PaymentReceipt]) -> str:
    if not receipts:
        return ""
    return "; ".join(f"{receipt.paid_at:%d.%m.%Y} — {money_text(float(receipt.amount or 0))}" for receipt in sorted(receipts, key=lambda item: (item.paid_at, item.id)))


def receipt_senders_text(receipts: list[PaymentReceipt]) -> str:
    senders = []
    for receipt in receipts:
        sender = receipt_sender_name(parse_receipt_details(receipt))
        if sender not in senders:
            senders.append(sender)
    return ", ".join(senders)


def apartment_month_state(apartment: Apartment, start_date: date, end_date: date) -> dict[str, Any]:
    total_days = (end_date - start_date).days + 1
    occupied_days: set[date] = set()
    overlaps: list[Lease] = []
    for lease in sorted(apartment.leases, key=lambda item: (item.start_date, item.id)):
        if IGNORE_LEASE_MARK in (lease.notes or ""):
            continue
        overlap_start = max(lease.start_date, start_date)
        overlap_end = min(lease.end_date or end_date, end_date)
        if overlap_start > overlap_end:
            continue
        overlaps.append(lease)
        for day_index in range((overlap_end - overlap_start).days + 1):
            occupied_days.add(overlap_start + timedelta(days=day_index))

    tenant_names: list[str] = []
    for lease in overlaps:
        if lease.tenant.full_name not in tenant_names:
            tenant_names.append(lease.tenant.full_name)

    vacant_days = max(total_days - len(occupied_days), 0)
    flags: list[str] = []
    if not overlaps:
        flags.append("пустует весь месяц")
    else:
        if len(tenant_names) > 1:
            flags.append("смена жильца")
        if vacant_days > 0:
            flags.append(f"пустует {vacant_days} дн.")
        if not flags:
            flags.append("без смены")
    return {
        "state": "; ".join(flags),
        "tenant_names": " -> ".join(tenant_names) if tenant_names else "нет жильца",
        "vacant_days": vacant_days,
        "changed": len(tenant_names) > 1,
        "vacant_full": not overlaps,
    }


@app.get("/api/reports/rent.xlsx")
def rent_report(start: str | None = None, end: str | None = None, session: Session = Depends(get_session)) -> StreamingResponse:
    start_date, end_date = current_month_range()
    if start:
        start_date = parse_date(start)
    if end:
        end_date = parse_date(end)
    charges = session.scalars(
        select(RentCharge).where(RentCharge.due_date >= start_date, RentCharge.due_date <= end_date).order_by(RentCharge.due_date)
    ).all()
    charges = [charge for charge in charges if not lease_ignored(session, charge.lease_id)]
    wb = Workbook()
    ws = setup_sheet(
        wb,
        "Аренда",
        [
            "Дата",
            "Период",
            "Объект",
            "Квартира",
            "Жилец",
            "Статус квартиры в периоде",
            "ИП начислено",
            "ИП оплачено",
            "ИП платежи",
            "ИП отправители",
            "Личный начислено",
            "Личный оплачено",
            "Долг",
            "Статус",
        ],
    )
    for charge in charges:
        data = serialize_rent_charge(charge)
        ip_receipts = accepted_receipts_for_charge(charge, "ip")
        apartment_state = apartment_month_state(charge.lease.apartment, charge.period_start, charge.period_end)
        ws.append(
            [
                data["due_date"],
                f'{data["period_start"]} - {data["period_end"]}',
                data["object"],
                data["apartment"],
                data["tenant"],
                apartment_state["state"],
                data["ip_due"],
                data["ip_paid"],
                receipt_amounts_text(ip_receipts),
                receipt_senders_text(ip_receipts),
                data["personal_due"],
                data["personal_paid"],
                data["debt"],
                report_status_label(data["status"], charge.due_date),
            ]
        )
        highlight_status_row(ws, ws.max_row, data["status"], data["debt"], apartment_state["changed"] or apartment_state["vacant_days"] > 0)
    return workbook_response(wb, f"rent-{start_date}-{end_date}.xlsx")


@app.get("/api/reports/utilities.xlsx")
def utilities_report(start: str | None = None, end: str | None = None, session: Session = Depends(get_session)) -> StreamingResponse:
    start_date, end_date = current_month_range()
    if start:
        start_date = parse_date(start)
    if end:
        end_date = parse_date(end)
    bills = session.scalars(
        select(UtilityBill)
        .where(UtilityBill.period_start >= start_date, UtilityBill.period_end <= end_date)
        .order_by(UtilityBill.period_start)
    ).all()
    wb = Workbook()
    ws = setup_sheet(
        wb,
        "Коммуналка",
        ["Период", "Объект", "Услуга", "Квартира", "Жилец", "Личный расход", "ОДН", "Сумма", "Оплачено", "Долг", "Статус"],
    )
    for bill in bills:
        for line in bill.lines:
            if line.lease_id and lease_ignored(session, line.lease_id):
                continue
            data = serialize_bill_line(line)
            ws.append(
                [
                    f"{bill.period_start} - {bill.period_end}",
                    bill.service.object.name,
                    bill.service.name,
                    data["apartment"],
                    data["tenant"],
                    data["personal_consumption"],
                    data["odn_consumption"],
                    data["total_amount"],
                    data["paid_amount"],
                    data["debt"],
                    report_status_label(data["status"], line.due_date),
                ]
            )
            highlight_status_row(ws, ws.max_row, data["status"], data["debt"], data["status"] == "issued")
    return workbook_response(wb, f"utilities-{start_date}-{end_date}.xlsx")


@app.get("/api/reports/debts.xlsx")
def debts_report(session: Session = Depends(get_session)) -> StreamingResponse:
    wb = Workbook()
    ws = setup_sheet(wb, "Долги", ["Тип", "Дата", "Объект", "Квартира", "Жилец", "Сумма", "Комментарий"])
    for charge in session.scalars(select(RentCharge).order_by(RentCharge.due_date)).all():
        data = serialize_rent_charge(charge)
        if data["debt"] > 0 and data["status"] in {"overdue", "partial", "deferred"}:
            ws.append(["Аренда", data["due_date"], data["object"], data["apartment"], data["tenant"], data["debt"], report_status_label(data["status"], charge.due_date)])
            highlight_status_row(ws, ws.max_row, data["status"], data["debt"])
    for line in session.scalars(select(UtilityBillLine)).all():
        data = serialize_bill_line(line)
        if data["debt"] > 0 and data["status"] in {"overdue", "partial", "issued"}:
            ws.append(["Коммуналка", data["due_date"], line.bill.service.object.name, data["apartment"], data["tenant"], data["debt"], report_status_label(data["status"], line.due_date)])
            highlight_status_row(ws, ws.max_row, data["status"], data["debt"], data["status"] == "issued")
    for debt in outstanding_manual_debts(session, cutoff=configured_notification_cutoff_date(session)):
        data = serialize_manual_debt(debt, session)
        ws.append(["Ручной долг", data["due_date"], data["object"], data["apartment"], data["tenant"], data["debt"], f"{data['title']} ({data['kind_label']})"])
        highlight_status_row(ws, ws.max_row, data["status"], data["debt"], True)
    return workbook_response(wb, "debts.xlsx")


@app.get("/api/reports/expenses.xlsx")
def expenses_report(start: str | None = None, end: str | None = None, session: Session = Depends(get_session)) -> StreamingResponse:
    start_date, end_date = current_month_range()
    if start:
        start_date = parse_date(start)
    if end:
        end_date = parse_date(end)
    expense_summary = expense_period_summary(session, start_date, end_date)
    expenses = session.scalars(
        select(Expense).where(Expense.expense_date >= start_date, Expense.expense_date <= end_date).order_by(Expense.expense_date)
    ).all()
    wb = Workbook()
    ws = setup_sheet(wb, "Расходы", ["Дата", "Объект", "Квартира", "Категория", "Сумма", "Источник", "Способ", "Компенсация", "Описание"])
    prepend_expense_summary(ws, start_date, end_date, expense_summary, 9)
    for expense in expenses:
        data = serialize_expense(expense)
        ws.append(
            [
                data["expense_date"],
                data["object"],
                data["apartment"],
                data["category"],
                data["amount"],
                data["source_funds"],
                data["payment_method"],
                report_status_label(data["compensation_status"]),
                data["description"],
            ]
        )
        highlight_status_row(ws, ws.max_row, data["compensation_status"], 0, data["compensation_status"] == "pending")
    return workbook_response(wb, f"expenses-{start_date}-{end_date}.xlsx")


@app.get("/api/reports/monthly.xlsx")
def monthly_report(year: int, month: int, session: Session = Depends(get_session)) -> StreamingResponse:
    try:
        start_date, end_date = month_range(year, month)
    except ValueError as exc:
        raise HTTPException(400, "Некорректный месяц") from exc

    summary = monthly_report_status(session, year, month)
    rent_charges = session.scalars(
        select(RentCharge).where(RentCharge.due_date >= start_date, RentCharge.due_date <= end_date).order_by(RentCharge.due_date)
    ).all()
    rent_charges = [charge for charge in rent_charges if not lease_ignored(session, charge.lease_id)]
    bills = session.scalars(
        select(UtilityBill)
        .where(UtilityBill.period_start >= start_date, UtilityBill.period_start <= end_date)
        .order_by(UtilityBill.period_start)
    ).all()
    apartments = session.scalars(select(Apartment).where(Apartment.active.is_(True)).order_by(Apartment.object_id, Apartment.sort_order, Apartment.name)).all()

    utility_lines = [line for bill in bills for line in bill.lines if not line.lease_id or not lease_ignored(session, line.lease_id)]
    utility_debt_by_apartment: dict[int, float] = {}
    for line in utility_lines:
        utility_debt_by_apartment[line.apartment_id] = utility_debt_by_apartment.get(line.apartment_id, 0.0) + max(0.0, float(line.total_amount or 0) - float(line.paid_amount or 0))

    charge_by_apartment = {charge.lease.apartment_id: charge for charge in rent_charges}
    apartment_rows: list[dict[str, Any]] = []
    for apartment in apartments:
        state = apartment_month_state(apartment, start_date, end_date)
        charge = charge_by_apartment.get(apartment.id)
        utility_debt = money(utility_debt_by_apartment.get(apartment.id, 0.0))
        apartment_rows.append(
            {
                "object": apartment.object.name,
                "apartment": apartment.name,
                "tenant_names": state["tenant_names"],
                "state": state["state"],
                "changed": state["changed"],
                "vacant_full": state["vacant_full"],
                "rent_status": report_status_label(charge.status, charge.due_date) if charge else "нет начисления",
                "rent_debt": money(max(0.0, float(charge.ip_due + charge.personal_due) - float(charge.ip_paid + charge.personal_paid))) if charge else 0.0,
                "utility_debt": utility_debt,
            }
        )

    occupied_rows = [row for row in apartment_rows if not row["vacant_full"]]
    settled_rows = [row for row in occupied_rows if row["rent_debt"] <= EPS and row["utility_debt"] <= EPS]
    ip_receipts = [receipt for charge in rent_charges for receipt in accepted_receipts_for_charge(charge, "ip")]
    ip_income = money(sum(float(receipt.amount or 0) for receipt in ip_receipts))

    wb = Workbook()
    ws = setup_sheet(wb, "Сводка", ["Показатель", "Значение"])
    ws.append(["Период", f"{start_date} - {end_date}"])
    ws.append(["Статус", summary["label"]])
    ws.append(["Всего проблем", summary["issue_count"]])
    ws.append(["Критических", summary["danger_count"]])
    ws.append(["Предупреждений", summary["warning_count"]])
    ws.append(["Квартир всего", len(apartment_rows)])
    ws.append(["Занятых квартир", len(occupied_rows)])
    ws.append(["Полностью рассчитались", len(settled_rows)])
    ws.append(["Квартир с долгом по аренде", sum(1 for row in occupied_rows if row["rent_debt"] > EPS)])
    ws.append(["Квартир с долгом по коммуналке", sum(1 for row in occupied_rows if row["utility_debt"] > EPS)])
    ws.append(["Квартир пустовало весь месяц", sum(1 for row in apartment_rows if row["vacant_full"])])
    ws.append(["Квартир со сменой жильца", sum(1 for row in apartment_rows if row["changed"])])
    ws.append(["Поступления на ИП, зачтённые за месяц", ip_income])

    issues_ws = setup_sheet(wb, "Проблемы", ["Важность", "Проблема", "Количество", "Комментарий"])
    for issue in summary["issues"]:
        issues_ws.append([issue["severity"], issue["title"], issue["count"], issue["detail"]])
        fill_report_row(issues_ws, issues_ws.max_row, REPORT_FILL_DEBT if issue["severity"] == "danger" else REPORT_FILL_WARN)

    apartments_ws = setup_sheet(
        wb,
        "Квартиры",
        ["Объект", "Квартира", "Жильцы за месяц", "Статус месяца", "Аренда", "Долг аренды", "Долг коммуналки"],
    )
    for row in apartment_rows:
        apartments_ws.append(
            [
                row["object"],
                row["apartment"],
                row["tenant_names"],
                row["state"],
                row["rent_status"],
                row["rent_debt"],
                row["utility_debt"],
            ]
        )
        highlight_status_row(apartments_ws, apartments_ws.max_row, "overdue" if row["rent_debt"] > EPS or row["utility_debt"] > EPS else "paid", row["rent_debt"] + row["utility_debt"], row["changed"] or row["vacant_full"])

    rent_ws = setup_sheet(
        wb,
        "Аренда",
        [
            "Дата",
            "Период",
            "Объект",
            "Квартира",
            "Жилец",
            "Статус квартиры в месяце",
            "ИП начислено",
            "ИП оплачено",
            "ИП платежи",
            "ИП отправители",
            "Личный начислено",
            "Личный оплачено",
            "Долг",
            "Статус",
        ],
    )
    for charge in rent_charges:
        data = serialize_rent_charge(charge)
        ip_charge_receipts = accepted_receipts_for_charge(charge, "ip")
        apartment_state = apartment_month_state(charge.lease.apartment, start_date, end_date)
        rent_ws.append(
            [
                data["due_date"],
                f'{data["period_start"]} - {data["period_end"]}',
                data["object"],
                data["apartment"],
                data["tenant"],
                apartment_state["state"],
                data["ip_due"],
                data["ip_paid"],
                receipt_amounts_text(ip_charge_receipts),
                receipt_senders_text(ip_charge_receipts),
                data["personal_due"],
                data["personal_paid"],
                data["debt"],
                report_status_label(data["status"], charge.due_date),
            ]
        )
        highlight_status_row(rent_ws, rent_ws.max_row, data["status"], data["debt"], apartment_state["changed"] or apartment_state["vacant_days"] > 0)

    ip_ws = setup_sheet(
        wb,
        "ИП платежи",
        ["Объект", "Квартира", "Жилец", "Месяц зачёта", "Оплачен по чеку", "Добавлен в систему", "Сумма", "Отправитель", "Источник"],
    )
    for charge in rent_charges:
        for receipt in accepted_receipts_for_charge(charge, "ip"):
            parsed = parse_receipt_details(receipt)
            ip_ws.append(
                [
                    charge.lease.apartment.object.name,
                    charge.lease.apartment.name,
                    charge.lease.tenant.full_name,
                    debt_month_heading(charge.due_date),
                    receipt.paid_at.strftime("%d.%m.%Y %H:%M"),
                    receipt.created_at.strftime("%d.%m.%Y %H:%M"),
                    money(receipt.amount),
                    receipt_sender_name(parsed),
                    payment_source_label(receipt),
                ]
            )

    utilities_ws = setup_sheet(
        wb,
        "Коммуналка",
        ["Период", "Объект", "Услуга", "Квартира", "Жилец", "Сумма", "Оплачено", "Долг", "Статус", "Поставщик оплачен"],
    )
    for bill in bills:
        for line in bill.lines:
            data = serialize_bill_line(line)
            utilities_ws.append(
                [
                    f"{bill.period_start} - {bill.period_end}",
                    bill.service.object.name,
                    bill.service.name,
                    data["apartment"],
                    data["tenant"],
                    data["total_amount"],
                    data["paid_amount"],
                    data["debt"],
                    report_status_label(data["status"], line.due_date),
                    "да" if bill.provider_paid else "нет",
                ]
            )
            highlight_status_row(utilities_ws, utilities_ws.max_row, data["status"], data["debt"], not bill.provider_paid or data["status"] == "issued")

    expenses_ws = setup_sheet(wb, "Расходы", ["Дата", "Объект", "Квартира", "Категория", "Сумма", "Источник", "Способ", "Компенсация", "Описание"])
    expenses = session.scalars(
        select(Expense).where(Expense.expense_date >= start_date, Expense.expense_date <= end_date).order_by(Expense.expense_date)
    ).all()
    expense_summary = expense_period_summary(session, start_date, end_date)
    prepend_expense_summary(expenses_ws, start_date, end_date, expense_summary, 9)
    for expense in expenses:
        data = serialize_expense(expense)
        expenses_ws.append(
            [
                data["expense_date"],
                data["object"],
                data["apartment"],
                data["category"],
                data["amount"],
                data["source_funds"],
                data["payment_method"],
                report_status_label(data["compensation_status"]),
                data["description"],
            ]
        )
        highlight_status_row(expenses_ws, expenses_ws.max_row, data["compensation_status"], 0, data["compensation_status"] == "pending")

    filename = f"monthly-{year}-{month:02d}.xlsx"
    return workbook_response(wb, filename)


@app.post("/api/reports/monthly/{year}/{month}/accept")
def accept_monthly_report(year: int, month: int, payload: dict[str, Any] | None = None, session: Session = Depends(get_session)) -> dict[str, Any]:
    kind = ((payload or {}).get("kind") or "full").strip()
    if kind not in {"full", "preliminary"}:
        kind = "full"
    month_range(year, month)
    mark_monthly_report_accepted(session, year, month, kind)
    session.commit()
    return {"ok": True, "key": monthly_report_key(year, month, kind)}


def rent_period_short(charge: RentCharge | None, start_date: date, end_date: date) -> str:
    if not charge:
        return f"{start_date:%d.%m.%y} - {end_date:%d.%m.%y}"
    return f"{charge.period_start:%d.%m.%y} - {charge.period_end:%d.%m.%y}"


def prior_ip_debt_for_lease(lease: Lease, before_date: date, cutoff: date | None = None) -> float:
    total = 0.0
    for charge in lease.rent_charges:
        if charge.due_date >= before_date:
            continue
        if cutoff and charge.due_date < cutoff:
            continue
        update_rent_charge_status(charge)
        total += max(0.0, owner_expected_ip_for_charge(charge) - owner_direct_ip_paid_for_charge(charge))
    return money(total)


@app.get("/api/reports/owner.xlsx")
def owner_report(start: str | None = None, end: str | None = None, session: Session = Depends(get_session)) -> StreamingResponse:
    start_date, end_date = current_month_range()
    if start:
        start_date = parse_date(start)
    if end:
        end_date = parse_date(end)
    cutoff = configured_notification_cutoff_date(session)

    rent_charges = session.scalars(
        select(RentCharge).where(RentCharge.due_date >= start_date, RentCharge.due_date <= end_date).order_by(RentCharge.due_date)
    ).all()
    rent_charges = [charge for charge in rent_charges if not lease_ignored(session, charge.lease_id)]
    for charge in rent_charges:
        update_rent_charge_status(charge)

    expected_ip = money(sum(owner_expected_ip_for_charge(charge) for charge in rent_charges))
    ip_receipts = [receipt for charge in rent_charges for receipt in charge_direct_ip_receipts(charge)]
    actual_ip = money(sum(float(receipt.amount or 0) for receipt in ip_receipts))
    expense_fund_receipts = [receipt for charge in rent_charges for receipt in charge_expense_fund_receipts(charge)]
    expense_fund_rent_income = money(sum(float(receipt.amount or 0) for receipt in expense_fund_receipts))
    occupied_count = len({charge.lease.apartment_id for charge in rent_charges if charge.ip_due > EPS or charge.personal_due > EPS})
    has_rent_debt = any(max(0.0, charge.ip_due - charge.ip_paid) + max(0.0, charge.personal_due - charge.personal_paid) > EPS for charge in rent_charges)
    period_expenses = session.scalars(
        select(Expense).where(Expense.expense_date >= start_date, Expense.expense_date <= end_date).order_by(Expense.expense_date, Expense.id)
    ).all()
    expense_income_total = money(sum(float(expense.amount or 0) for expense in period_expenses if expense.source_funds == "expense_fund_income"))
    expense_spent_total = money(sum(float(expense.amount or 0) for expense in period_expenses if expense.source_funds != "expense_fund_income"))
    expense_balance = money(expense_income_total - expense_spent_total)
    expense_balance_label = (
        f"в расходном фонде осталось {money_text(expense_balance)}"
        if expense_balance > EPS
        else f"объекты должны хозяину {money_text(abs(expense_balance))}"
        if expense_balance < -EPS
        else "сошлось без хвостов"
    )

    wb = Workbook()
    ws = setup_sheet(wb, "Для хозяина", ["Показатель", "Значение"])
    ws.insert_rows(1)
    ws["A1"] = f"Отчёт для хозяина: {start_date:%d.%m.%Y} - {end_date:%d.%m.%Y}"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.append(["Заселённых квартир к оплате", occupied_count])
    ws.append(["Ожидалось на ИП", expected_ip])
    ws.append(["Получено на ИП", actual_ip])
    ws.append(["Получено на расходы вместо ИП", expense_fund_rent_income])
    ws.append(["Расходный фонд за период", expense_balance_label])
    ws.append(["Общий статус аренды", "есть долги" if has_rent_debt else "в порядке"])
    expense_summary = expense_period_summary(session, start_date, end_date)
    expense_income_total = expense_summary["period_income"]
    expense_spent_total = expense_summary["period_spent"]
    expense_balance = expense_summary["closing_balance"]
    expense_balance_label = expense_balance_comment(expense_balance)
    ws.append(["Сальдо расходного фонда на начало", expense_summary["opening_balance"]])
    ws.append(["Поступления на расходы за период", expense_income_total])
    ws.append(["Расходы за период", expense_spent_total])
    ws.append(["Сальдо расходного фонда на конец", expense_balance])
    ws.append(["Комментарий по сальдо", expense_balance_label])
    if has_rent_debt:
        fill_report_row(ws, ws.max_row, REPORT_FILL_DEBT)
    else:
        fill_report_row(ws, ws.max_row, REPORT_FILL_OK)

    rows_ws = setup_sheet(
        wb,
        "Квартиры",
        [
            "Период аренды",
            "Квартира",
            "Жилец",
            "Статус квартиры",
            "ИП оплачено",
            "На расходы",
            "Дата платежа",
            "Отправитель",
            "Долг ИП с прошлыми месяцами",
            "Статус оплаты",
        ],
    )
    apartments = session.scalars(select(Apartment).where(Apartment.active.is_(True)).order_by(Apartment.object_id, Apartment.sort_order, Apartment.name)).all()
    charges_by_apartment: dict[int, list[RentCharge]] = {}
    for charge in rent_charges:
        charges_by_apartment.setdefault(charge.lease.apartment_id, []).append(charge)
    for apartment in apartments:
        state = apartment_month_state(apartment, start_date, end_date)
        charges = charges_by_apartment.get(apartment.id, [])
        if not charges:
            rows_ws.append([f"{start_date:%d.%m.%y} - {end_date:%d.%m.%y}", apartment.name, state["tenant_names"], state["state"], 0, "", "", 0, "нет начисления"])
            highlight_status_row(rows_ws, rows_ws.max_row, "pending", 0, True)
            continue
        for charge in charges:
            receipts = charge_direct_ip_receipts(charge)
            expense_receipts = charge_expense_fund_receipts(charge)
            current_ip_debt = money(max(0.0, owner_expected_ip_for_charge(charge) - owner_direct_ip_paid_for_charge(charge)))
            total_ip_debt = money(current_ip_debt + prior_ip_debt_for_lease(charge.lease, start_date, cutoff))
            rows_ws.append(
                [
                    rent_period_short(charge, start_date, end_date),
                    apartment.name,
                    charge.lease.tenant.full_name,
                    state["state"],
                    owner_direct_ip_paid_for_charge(charge),
                    accepted_receipt_amount(expense_receipts),
                    "; ".join(bit for bit in [receipt_amounts_text(receipts), receipt_amounts_text(expense_receipts)] if bit),
                    "; ".join(bit for bit in [receipt_senders_text(receipts), receipt_senders_text(expense_receipts)] if bit),
                    total_ip_debt,
                    owner_charge_status_label(charge),
                ]
            )
            highlight_status_row(rows_ws, rows_ws.max_row, charge.status, total_ip_debt, state["changed"] or state["vacant_days"] > 0)

    utility_ws = setup_sheet(wb, "Коммуналка", ["Объект", "Услуга", "Период", "Сумма", "Оплачено поставщику", "Статус"])
    bills = session.scalars(
        select(UtilityBill)
        .where(UtilityBill.period_end >= start_date, UtilityBill.period_start <= end_date)
        .order_by(UtilityBill.period_start, UtilityBill.service_id)
    ).all()
    for bill in bills:
        status = "paid" if bill.provider_paid else "overdue"
        utility_ws.append(
            [
                bill.service.object.name,
                bill.service.name,
                f"{bill.period_start:%d.%m.%Y} - {bill.period_end:%d.%m.%Y}",
                money(bill.total_cost),
                money(bill.total_cost) if bill.provider_paid else 0,
                "оплачено" if bill.provider_paid else "не оплачено",
            ]
        )
        highlight_status_row(utility_ws, utility_ws.max_row, status, 0 if bill.provider_paid else money(bill.total_cost))

    expense_ws = setup_sheet(
        wb,
        "Расходы и зачёт",
        ["Дата", "Объект", "Квартира", "Категория", "Сумма", "Источник", "Способ", "Статус", "Описание"],
    )
    expense_ws.append(["Итог", "", "", "Поступило на расходы", expense_income_total, "", "", "", ""])
    prepend_expense_summary(expense_ws, start_date, end_date, expense_summary, 9)
    fill_report_row(expense_ws, expense_ws.max_row, REPORT_FILL_OK)
    expense_ws.append(["Итог", "", "", "Израсходовано", expense_spent_total, "", "", "", ""])
    fill_report_row(expense_ws, expense_ws.max_row, REPORT_FILL_WARN if expense_spent_total > EPS else REPORT_FILL_OK)
    expense_ws.append(["Итог", "", "", "Взаимозачёт", expense_balance, "", "", "", expense_balance_label])
    fill_report_row(
        expense_ws,
        expense_ws.max_row,
        REPORT_FILL_WARN if abs(expense_balance) > EPS else REPORT_FILL_OK,
    )
    for expense in period_expenses:
        data = serialize_expense(expense)
        expense_ws.append(
            [
                data["expense_date"],
                data["object"],
                data["apartment"],
                data["category"],
                data["amount"],
                data["source_funds"],
                data["payment_method"],
                report_status_label(data["compensation_status"]),
                data["description"],
            ]
        )
        highlight_status_row(expense_ws, expense_ws.max_row, data["compensation_status"], 0, data["source_funds"] == "expense_fund_income")

    return workbook_response(wb, f"owner-{start_date}-{end_date}.xlsx")


@app.get("/api/reports/history.xlsx")
def history_report(
    apartment_id: int | None = None,
    tenant_id: int | None = None,
    session: Session = Depends(get_session),
) -> StreamingResponse:
    wb = Workbook()
    ws = setup_sheet(wb, "История", ["Тип", "Дата", "Объект", "Квартира", "Жилец", "Сумма", "Статус", "Комментарий"])

    lease_ids = []
    if apartment_id:
        lease_ids = [lease.id for lease in session.scalars(select(Lease).where(Lease.apartment_id == apartment_id)).all()]
    if tenant_id:
        lease_ids = [lease.id for lease in session.scalars(select(Lease).where(Lease.tenant_id == tenant_id)).all()]
    if not lease_ids:
        lease_ids = [lease.id for lease in session.scalars(select(Lease)).all()]

    for charge in session.scalars(select(RentCharge).where(RentCharge.lease_id.in_(lease_ids)).order_by(RentCharge.due_date)).all():
        data = serialize_rent_charge(charge)
        ws.append(["Аренда", data["due_date"], data["object"], data["apartment"], data["tenant"], data["total_due"], report_status_label(data["status"], charge.due_date), ""])

    for line in session.scalars(select(UtilityBillLine).where(UtilityBillLine.lease_id.in_(lease_ids))).all():
        data = serialize_bill_line(line)
        ws.append(["Коммуналка", data["due_date"], line.bill.service.object.name, data["apartment"], data["tenant"], data["total_amount"], report_status_label(data["status"], line.due_date), ""])
        highlight_status_row(ws, ws.max_row, data["status"], data["debt"], data["status"] == "issued")

    for debt in session.scalars(select(ManualDebt).where(ManualDebt.lease_id.in_(lease_ids), ManualDebt.active.is_(True))).all():
        data = serialize_manual_debt(debt, session)
        ws.append(["Ручной долг", data["due_date"], data["object"], data["apartment"], data["tenant"], data["amount"], report_status_label(data["status"], debt.due_date), data["notes"]])
        highlight_status_row(ws, ws.max_row, data["status"], data["debt"], True)

    return workbook_response(wb, "history.xlsx")
